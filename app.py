from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify
import pandas as pd
import sqlite3
import os
from io import BytesIO
from datetime import datetime
from flask import session
from functools import wraps
from flask import session, redirect, url_for, flash
from flask import Flask, render_template, request, redirect, url_for, flash, session, get_flashed_messages
from functools import wraps
from flask import session, redirect, url_for, flash
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import base64
from io import BytesIO
import pandas as pd
from flask import send_file
from werkzeug.security import generate_password_hash, check_password_hash
import secrets
import string
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import timedelta
import requests
from notifications_service import notification_service

def send_email_via_resend(recipient_email, otp_code, username, api_key, sender_email):
    """
    Send OTP email via Resend API
    """
    try:
        url = "https://api.resend.com/emails"
        headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json"
        }
        
        html_body = f"""
        <html>
            <body style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
                <div style="background-color: #f8f9fa; padding: 20px; border-radius: 5px;">
                    <h2 style="color: #333;">Password Reset Request</h2>
                    <p>Hello <strong>{username}</strong>,</p>
                    <p>You have requested to reset your password. Please use the following One-Time Password (OTP) to continue:</p>
                    
                    <div style="background-color: #007bff; color: white; padding: 15px; text-align: center; border-radius: 5px; margin: 20px 0;">
                        <h1 style="margin: 0; font-size: 36px; letter-spacing: 5px;">{otp_code}</h1>
                    </div>
                    
                    <p style="color: #dc3545;"><strong>⚠️ Security Notice:</strong></p>
                    <ul style="color: #666;">
                        <li>This OTP will expire in <strong>15 minutes</strong></li>
                        <li>Do not share this code with anyone</li>
                        <li>If you didn't request this, please ignore this email</li>
                    </ul>
                    
                    <hr style="border: none; border-top: 1px solid #ddd; margin: 20px 0;">
                    <p style="color: #999; font-size: 12px;">
                        This is an automated email from the CRM System. Please do not reply to this email.
                    </p>
                </div>
            </body>
        </html>
        """
        
        data = {
            "from": sender_email,
            "to": [recipient_email],
            "subject": "Password Reset - OTP Code",
            "html": html_body
        }
        
        response = requests.post(url, json=data, headers=headers)
        
        if response.status_code == 200:
            print(f"✅ OTP email sent successfully via Resend to {recipient_email}")
            return True
        else:
            print(f"❌ Resend API error: {response.status_code} - {response.text}")
            print(f"Debug: OTP for {username} ({recipient_email}): {otp_code}")
            return False
            
    except Exception as e:
        print(f"Error sending OTP email via Resend: {e}")
        print(f"Debug: OTP for {username} ({recipient_email}): {otp_code}")
        return False

def generate_otp(length=6):
    """Generate a random OTP code"""
    return ''.join(secrets.choice(string.digits) for _ in range(length))

def send_otp_email(recipient_email, otp_code, username):
    """
    Send OTP email to user for password reset
    Supports multiple email providers through environment variables
    
    Required environment variables:
    For SMTP (Gmail, SendGrid, etc.):
    - EMAIL_PROVIDER: 'smtp' (default)
    - SMTP_HOST: SMTP server host (e.g., smtp.gmail.com)
    - SMTP_PORT: SMTP port (usually 587 for TLS)
    - SMTP_USERNAME: Your email/username
    - SMTP_PASSWORD: Your email password or app password
    - SENDER_EMAIL: Email address to send from
    
    For Resend:
    - EMAIL_PROVIDER: 'resend'
    - RESEND_API_KEY: Your Resend API key
    - SENDER_EMAIL: Email address to send from
    """
    try:
        # Email configuration from environment variables
        email_provider = os.getenv('EMAIL_PROVIDER', 'smtp').lower()
        sender_email = os.getenv('SENDER_EMAIL')
        
        # Check configuration based on provider
        if email_provider == 'resend':
            resend_api_key = os.getenv('RESEND_API_KEY')
            if not resend_api_key or not sender_email:
                print("WARNING: Resend credentials not configured. OTP email not sent.")
                print(f"Debug: OTP for {username} ({recipient_email}): {otp_code}")
                return False
            
            # Send email via Resend API
            return send_email_via_resend(recipient_email, otp_code, username, resend_api_key, sender_email)
        else:
            # SMTP configuration
            smtp_host = os.getenv('SMTP_HOST', 'smtp.gmail.com')
            smtp_port = int(os.getenv('SMTP_PORT', 587))
            smtp_username = os.getenv('SMTP_USERNAME')
            smtp_password = os.getenv('SMTP_PASSWORD')
            
            if not all([smtp_username, smtp_password]):
                print("WARNING: SMTP credentials not configured. OTP email not sent.")
                print(f"Debug: OTP for {username} ({recipient_email}): {otp_code}")
                return False
        
        # Create email message
        msg = MIMEMultipart('alternative')
        msg['Subject'] = 'Password Reset - OTP Code'
        msg['From'] = sender_email
        msg['To'] = recipient_email
        
        # Email body
        html_body = f"""
        <html>
            <body style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
                <div style="background-color: #f8f9fa; padding: 20px; border-radius: 5px;">
                    <h2 style="color: #333;">Password Reset Request</h2>
                    <p>Hello <strong>{username}</strong>,</p>
                    <p>You have requested to reset your password. Please use the following One-Time Password (OTP) to continue:</p>
                    
                    <div style="background-color: #007bff; color: white; padding: 15px; text-align: center; border-radius: 5px; margin: 20px 0;">
                        <h1 style="margin: 0; font-size: 36px; letter-spacing: 5px;">{otp_code}</h1>
                    </div>
                    
                    <p style="color: #dc3545;"><strong>⚠️ Security Notice:</strong></p>
                    <ul style="color: #666;">
                        <li>This OTP will expire in <strong>15 minutes</strong></li>
                        <li>Do not share this code with anyone</li>
                        <li>If you didn't request this, please ignore this email</li>
                    </ul>
                    
                    <hr style="border: none; border-top: 1px solid #ddd; margin: 20px 0;">
                    <p style="color: #999; font-size: 12px;">
                        This is an automated email from the CRM System. Please do not reply to this email.
                    </p>
                </div>
            </body>
        </html>
        """
        
        text_body = f"""
        Password Reset Request
        
        Hello {username},
        
        You have requested to reset your password. Please use the following One-Time Password (OTP) to continue:
        
        OTP CODE: {otp_code}
        
        Security Notice:
        - This OTP will expire in 15 minutes
        - Do not share this code with anyone
        - If you didn't request this, please ignore this email
        
        This is an automated email from the CRM System.
        """
        
        part1 = MIMEText(text_body, 'plain')
        part2 = MIMEText(html_body, 'html')
        msg.attach(part1)
        msg.attach(part2)
        
        # Send email via SMTP
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls()
            server.login(smtp_username, smtp_password)
            server.send_message(msg)
        
        print(f"OTP email sent successfully to {recipient_email}")
        return True
        
    except Exception as e:
        print(f"Error sending OTP email: {e}")
        print(f"Debug: OTP for {username} ({recipient_email}): {otp_code}")
        return False

def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                flash('You need to log in first!', 'danger')
                return redirect(url_for('login'))  # Redirect to the login page

            # Check if the user's role matches any of the required roles
            conn = sqlite3.connect('ProjectStatus.db')
            c = conn.cursor()
            c.execute("SELECT role FROM users WHERE id=?", (session['user_id'],))
            user_role = c.fetchone()
            conn.close()

            # If the user role is None or not in the list of allowed roles, redirect
            if user_role is None or user_role[0] not in roles:
                flash('You do not have permission to access this page.', 'danger')
                return redirect(url_for('index'))  # Redirect to a safe page

            return f(*args, **kwargs)

        return decorated_function

    return decorator

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', os.urandom(24).hex())

# Sample lists of names for engineers
PRESALE_ENGINEERS = ["m.saleh", "R.Elnaggar", "S.Hussin", "m.fakhrany",]
SALES_ENGINEERS = ["A.Manea", "m.fakhrany", "A.zain", "Isaac", "Jack"]

# Global variable to store the last uploaded cost sheet path
last_cost_sheet_path = ""

# ======================================================================
# AI-POWERED DEAL VALUE CALCULATOR
# Automatically calculates deal_value from quotation_selling_price
# Intelligently handles quote revisions to avoid duplication
# ======================================================================

def parse_quote_reference(quote_ref):
    """
    AI-based parser to extract base reference and revision number from quote
    
    Examples:
    - QT-ACS-Sup-BEDROCK-18425-R02 -> (QT-ACS-Sup-BEDROCK-18425, 2)
    - CS-EJT-LC-AHHOBP-13725-R07 -> (CS-EJT-LC-AHHOBP-13725, 7)
    """
    if not quote_ref:
        return (None, 0)
    
    # Pattern 1: Standard format with -R## at the end
    pattern1 = r'^(.+)-R(\d+)$'
    match = re.match(pattern1, quote_ref.strip())
    if match:
        base_ref = match.group(1)
        revision = int(match.group(2))
        return (base_ref, revision)
    
    # Pattern 2: Alternative formats (R## without dash)
    pattern2 = r'^(.+)R(\d+)$'
    match = re.match(pattern2, quote_ref.strip())
    if match:
        base_ref = match.group(1)
        revision = int(match.group(2))
        return (base_ref, revision)
    
    # Pattern 3: No revision number - treat as revision 0 (original)
    return (quote_ref.strip(), 0)

def get_latest_quote_revisions(quotes):
    """
    AI logic: From a list of quotes, return only the latest revision of each base quote
    
    Args:
        quotes: List of tuples (quote_ref, price, registered_date, system)
    
    Returns:
        List of latest quote dictionaries
    """
    # Group quotes by base reference
    quote_groups = defaultdict(list)
    
    for quote_ref, price, registered_date, system in quotes:
        base_ref, revision = parse_quote_reference(quote_ref)
        if base_ref:
            quote_groups[base_ref].append({
                'quote_ref': quote_ref,
                'base_ref': base_ref,
                'revision': revision,
                'price': price or 0.0,
                'registered_date': registered_date or '',
                'system': system or ''
            })
    
    # Select latest revision for each base reference
    latest_quotes = []
    for base_ref, quote_list in quote_groups.items():
        # Sort by revision number (descending), then by date (descending)
        sorted_quotes = sorted(
            quote_list,
            key=lambda x: (x['revision'], x['registered_date']),
            reverse=True
        )
        latest_quotes.append(sorted_quotes[0])
    
    return latest_quotes

def filter_duplicates_by_system(quotes):
    """
    Advanced AI logic: Detect and remove duplicate quotes for the same system
    If multiple quotes exist for the same system, keep only the latest
    """
    system_groups = defaultdict(list)
    
    for quote in quotes:
        system = quote.get('system', '').strip()
        if system:
            system_groups[system].append(quote)
    
    # For each system, keep only the latest quote
    filtered_quotes = []
    processed_systems = set()
    
    for quote in quotes:
        system = quote.get('system', '').strip()
        
        # If no system specified, include it
        if not system:
            filtered_quotes.append(quote)
            continue
        
        # If this system hasn't been processed yet
        if system not in processed_systems:
            # Find all quotes for this system
            system_quotes = system_groups[system]
            
            # If multiple quotes for same system, take the one with highest revision
            if len(system_quotes) > 1:
                latest = max(system_quotes, key=lambda x: (x['revision'], x['registered_date']))
                filtered_quotes.append(latest)
            else:
                filtered_quotes.append(quote)
            
            processed_systems.add(system)
    
    return filtered_quotes

def quote_matches_project(quote_ref, project_name):
    """
    AI-based intelligent matching: Determine if a quote belongs to a specific project
    by analyzing identifiers in the quote reference
    
    Critical for cases where one project name contains multiple sub-projects
    Example: "Qalam School 'Wadi'" should only include Wadi quotes, not Qalam quotes
    
    Args:
        quote_ref: Quote reference string (e.g., QT-MF0380225,SchlWadi,H3C-R09)
        project_name: Project name (e.g., "Qalam School 'Wadi'")
    
    Returns:
        Boolean: True if quote belongs to this project, False otherwise
    """
    if not quote_ref or not project_name:
        return True  # Default to include if missing data
    
    quote_lower = quote_ref.lower()
    project_lower = project_name.lower()
    
    # Define school/location identifiers and their keywords
    # Format: (identifier_in_quote, keyword_in_project_name)
    identifier_mappings = [
        ('schlqlm', 'qalam'),
        ('schlwadi', 'wadi'),
        ('schltawn', "ta'won"),
        ('schltawon', "ta'won"),
        ('tawon', "ta'won"),
        ('tawn', "ta'won"),
        ('almajd', 'majd'),
        ('soh', 'soho'),
        ('bedrock', 'bedrock'),
        ('kfsh', 'faisal'),
        ('kucct1', 'tender 1'),
        ('kucct2', 'tender 2'),
        ('nasr', 'nasr'),
    ]
    
    # Check if quote contains any specific identifier
    found_identifiers = []
    for quote_id, project_keyword in identifier_mappings:
        if quote_id in quote_lower:
            found_identifiers.append((quote_id, project_keyword))
    
    # If no specific identifiers found, include the quote (default behavior)
    if not found_identifiers:
        return True
    
    # ENHANCED AI LOGIC: Extract quoted portion from project name for priority matching
    # Example: "Qalam School 'Wadi'" → prioritize matching 'Wadi'
    import re
    quoted_match = re.search(r"['\"]([^'\"]+)['\"]", project_name)
    
    if quoted_match:
        # Project has quoted text - prioritize matching that
        quoted_text = quoted_match.group(1).lower()
        
        # Check if any identifier matches the quoted portion
        for quote_id, project_keyword in found_identifiers:
            if project_keyword in quoted_text:
                return True  # Match found in quoted portion
        
        # If we found identifiers but none match the quoted portion, exclude
        return False
    
    # No quoted text - use standard matching
    # Check if any found identifier matches the project name
    for quote_id, project_keyword in found_identifiers:
        if project_keyword in project_lower:
            return True
    
    # Identifier found in quote but doesn't match project name - exclude it
    return False

def calculate_deal_value_for_project(project_name, conn=None):
    """
    Calculate the deal value for a specific project from its quotations
    Uses AI logic to handle revisions and avoid duplicates
    
    ENHANCED: Now includes intelligent quote-to-project matching to handle
    cases where multiple sub-projects share similar names
    
    Args:
        project_name: Name of the project
        conn: Optional database connection (creates new one if None)
    
    Returns:
        Tuple of (total_deal_value, list_of_included_quotes)
    """
    close_conn = False
    if conn is None:
        conn = sqlite3.connect('ProjectStatus.db')
        close_conn = True
    
    cursor = conn.cursor()
    
    try:
        # Fetch all quotations for this project
        cursor.execute("""
            SELECT quote_ref, quotation_selling_price, registered_date, system
            FROM projects
            WHERE project_name = ?
            AND quotation_selling_price IS NOT NULL
            AND quotation_selling_price > 0
        """, (project_name,))
        
        quotes = cursor.fetchall()
        
        if not quotes:
            return 0.0, []
        
        # ENHANCED AI: Filter quotes to only include those that actually belong to this project
        # This prevents mixing quotes from different sub-projects (e.g., Qalam vs Wadi)
        matched_quotes = [
            q for q in quotes 
            if quote_matches_project(q[0], project_name)  # q[0] is quote_ref
        ]
        
        if not matched_quotes:
            return 0.0, []
        
        # Apply AI logic to get latest revisions only
        latest_quotes = get_latest_quote_revisions(matched_quotes)
        
        # Further filter by system to avoid duplicates
        filtered_quotes = filter_duplicates_by_system(latest_quotes)
        
        # Calculate total deal value
        total_deal_value = sum(q['price'] for q in filtered_quotes if q['price'])
        
        return total_deal_value, filtered_quotes
        
    finally:
        if close_conn:
            conn.close()

def init_db():
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS projects (
            id INTEGER PRIMARY KEY,
            project_name TEXT,
            quote_ref TEXT UNIQUE,
            presale_eng TEXT,
            sales_eng TEXT,
            system TEXT,
            sow TEXT,
            quotation BLOB,
            cost_sheet BLOB,
            status TEXT,
            quotation_note TEXT,
            feedback TEXT,
            registered_date TEXT,
            quarter TEXT,
            quotation_cost REAL,
            quotation_selling_price REAL,
            margin REAL,
            progress TEXT,
            rfq_reference TEXT,
            registered_by TEXT,
            updated_time TEXT,
            updated_by TEXT
        )
    ''')
    c.execute('''CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY,
            username TEXT UNIQUE,
            password TEXT,
            role TEXT,
            email TEXT
        )''')
    
    # Add email column to existing users table if it doesn't exist
    try:
        c.execute("ALTER TABLE users ADD COLUMN email TEXT")
    except sqlite3.OperationalError:
        pass  # Column already exists
    c.execute('''CREATE TABLE IF NOT EXISTS engineers (
            id INTEGER PRIMARY KEY,
            name TEXT,
            engineer_code TEXT UNIQUE,
            phone TEXT,
            email TEXT,
            role TEXT,
            username TEXT UNIQUE,
            password TEXT
        )''')
    
    # Registration requests table for public registration with admin approval
    c.execute('''CREATE TABLE IF NOT EXISTS registration_requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            requested_role TEXT NOT NULL,
            email TEXT,
            full_name TEXT,
            reason TEXT,
            status TEXT DEFAULT 'pending',
            requested_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            reviewed_by TEXT,
            reviewed_at TIMESTAMP
        )''')
    
    # Password reset tokens table for OTP-based password reset
    c.execute('''CREATE TABLE IF NOT EXISTS password_reset_tokens (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            email TEXT NOT NULL,
            otp_code TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            expires_at TIMESTAMP NOT NULL,
            used INTEGER DEFAULT 0,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )''')
    
    # PO Requests table for Purchase Order request workflow with approval
    c.execute('''CREATE TABLE IF NOT EXISTS po_requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            po_request_reference TEXT UNIQUE NOT NULL,
            quote_ref TEXT NOT NULL,
            project_name TEXT NOT NULL,
            system TEXT,
            presale_engineer TEXT,
            project_manager TEXT,
            vendor_id INTEGER,
            vendor_name TEXT,
            distributor_id INTEGER,
            distributor_name TEXT,
            notes TEXT,
            request_status TEXT DEFAULT 'Pending Approval',
            requested_by_id INTEGER NOT NULL,
            requested_by_name TEXT NOT NULL,
            requested_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            approved_by_id INTEGER,
            approved_by_name TEXT,
            decision_time TIMESTAMP,
            rejection_reason TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (requested_by_id) REFERENCES users(id)
        )''')
    
    # Create indexes for po_requests table for better dashboard and query performance
    c.execute('CREATE INDEX IF NOT EXISTS idx_po_requests_quote_ref ON po_requests(quote_ref)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_po_requests_status ON po_requests(request_status)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_po_requests_requested_by ON po_requests(requested_by_id)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_po_requests_created ON po_requests(created_at)')
    
    # Add supplier_quotation_id column to po_requests table if it doesn't exist
    try:
        c.execute("ALTER TABLE po_requests ADD COLUMN supplier_quotation_id INTEGER")
    except sqlite3.OperationalError:
        pass  # Column already exists
    
    # Add project_coordinator column to po_requests table if it doesn't exist
    try:
        c.execute("ALTER TABLE po_requests ADD COLUMN project_coordinator TEXT")
    except sqlite3.OperationalError:
        pass  # Column already exists
    
    # Supplier Quotations table for tracking supplier quotation PDFs with distributor/vendor
    c.execute('''CREATE TABLE IF NOT EXISTS supplier_quotations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            quote_ref TEXT NOT NULL,
            distributor_id INTEGER,
            distributor_name TEXT,
            vendor_id INTEGER,
            vendor_name TEXT,
            supplier_type TEXT,
            quotation_file BLOB NOT NULL,
            filename TEXT NOT NULL,
            uploaded_by TEXT NOT NULL,
            uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            notes TEXT,
            FOREIGN KEY (quote_ref) REFERENCES projects(quote_ref),
            FOREIGN KEY (distributor_id) REFERENCES distributors(id),
            FOREIGN KEY (vendor_id) REFERENCES vendors(id)
        )''')
    
    # Create indexes for supplier_quotations table
    c.execute('CREATE INDEX IF NOT EXISTS idx_supplier_quotations_quote_ref ON supplier_quotations(quote_ref)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_supplier_quotations_uploaded_at ON supplier_quotations(uploaded_at)')
    
    # Add system column to supplier_quotations table if it doesn't exist
    try:
        c.execute("ALTER TABLE supplier_quotations ADD COLUMN system TEXT")
    except sqlite3.OperationalError:
        pass  # Column already exists
    
    # Quotation Products table for storing products extracted from supplier quotations
    c.execute('''CREATE TABLE IF NOT EXISTS quotation_products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            supplier_quotation_id INTEGER NOT NULL,
            part_number TEXT NOT NULL,
            description TEXT NOT NULL,
            unit_price REAL NOT NULL,
            quantity INTEGER,
            currency TEXT DEFAULT 'USD',
            notes TEXT,
            quote_ref TEXT,
            supplier_name TEXT,
            supplier_type TEXT,
            system TEXT,
            added_by TEXT NOT NULL,
            added_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (supplier_quotation_id) REFERENCES supplier_quotations(id) ON DELETE CASCADE,
            FOREIGN KEY (quote_ref) REFERENCES projects(quote_ref)
        )''')
    
    # Create indexes for quotation_products table
    c.execute('CREATE INDEX IF NOT EXISTS idx_quotation_products_supplier_quotation ON quotation_products(supplier_quotation_id)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_quotation_products_part_number ON quotation_products(part_number)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_quotation_products_quote_ref ON quotation_products(quote_ref)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_quotation_products_added_at ON quotation_products(added_at)')
    
    # ============ PERMISSION SYSTEM TABLES ============
    # Permissions: Master list of all available permissions/pages
    c.execute('''CREATE TABLE IF NOT EXISTS permissions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT UNIQUE NOT NULL,
            label TEXT NOT NULL,
            category TEXT,
            description TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
    
    # Role Permissions: Default permissions for each role
    c.execute('''CREATE TABLE IF NOT EXISTS role_permissions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            role TEXT NOT NULL,
            permission_id INTEGER NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (permission_id) REFERENCES permissions(id),
            UNIQUE(role, permission_id)
        )''')
    
    # User Permissions: User-specific overrides
    c.execute('''CREATE TABLE IF NOT EXISTS user_permissions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            permission_id INTEGER NOT NULL,
            grant_type TEXT DEFAULT 'allow' CHECK(grant_type IN ('allow', 'deny')),
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_by TEXT,
            FOREIGN KEY (user_id) REFERENCES users(id),
            FOREIGN KEY (permission_id) REFERENCES permissions(id),
            UNIQUE(user_id, permission_id)
        )''')

    conn.commit()
    conn.close()

    # Create uploads directory if it doesn't exist
    if not os.path.exists('uploads'):
        os.makedirs('uploads')


def seed_permissions():
    """Seed permissions table with all available pages/features"""
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    
    # Check if permissions already exist
    c.execute("SELECT COUNT(*) FROM permissions")
    if c.fetchone()[0] > 0:
        conn.close()
        return  # Already seeded
    
    # Define all permissions with categories
    permissions = [
        # Dashboard
        ('view_dashboard', 'Dashboard', 'Dashboard', 'View main dashboard'),
        
        # CRM
        ('view_projects', 'Projects & Pipeline', 'CRM', 'View projects and pipeline'),
        ('view_pipeline_analysis', 'Pipeline Analysis', 'CRM', 'View pipeline analysis'),
        ('view_sales_performance', 'Sales Performance', 'CRM', 'View sales performance'),
        ('view_aging_dashboard', 'Aging Dashboard', 'CRM', 'View aging dashboard'),
        ('view_end_users', 'End Users', 'CRM', 'View end users'),
        ('view_contractors', 'Contractors', 'CRM', 'View contractors'),
        ('view_consultants', 'Consultants', 'CRM', 'View consultants'),
        
        # Sales
        ('view_presales', 'Presales Performance', 'Sales', 'View presales performance'),
        ('view_rfq', 'RFQ', 'Sales', 'View and manage RFQs'),
        ('view_rfts', 'RFTS', 'Sales', 'View technical support requests'),
        ('view_reports', 'Reports', 'Sales', 'View reports'),
        ('view_products', 'Products', 'Sales', 'View product catalog'),
        ('view_solution_builder', 'Solution Builder', 'Sales', 'Access solution builder'),
        ('view_quotation_builder', 'Quotation Builder', 'Sales', 'Access quotation builder'),
        
        # Purchasing
        ('view_po_status', 'PO Status', 'Purchasing', 'View purchase order status'),
        
        # Task Management
        ('view_tasks', 'Tasks', 'Task Management', 'View and manage tasks'),
        
        # SRM
        ('view_vendors', 'Vendors', 'SRM', 'View and manage vendors'),
        ('view_distributors', 'Distributors', 'SRM', 'View and manage distributors'),
        
        # Registration
        ('view_registration_hub', 'Registration Hub', 'Registration', 'Access registration hub'),
        
        # Administration
        ('manage_users', 'Manage Users', 'Administration', 'Add, edit, and delete users'),
        ('view_pending_registrations', 'Pending Registrations', 'Administration', 'Review registration requests'),
        ('view_password_reset_otps', 'Password Reset OTPs', 'Administration', 'View and manage OTP requests'),
        ('manage_permissions', 'Access Control', 'Administration', 'Manage user permissions'),
    ]
    
    for code, label, category, description in permissions:
        c.execute("""
            INSERT OR IGNORE INTO permissions (code, label, category, description)
            VALUES (?, ?, ?, ?)
        """, (code, label, category, description))
    
    conn.commit()
    conn.close()


def seed_default_role_permissions():
    """Set up default permissions for each role"""
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    
    # Check if role permissions already exist
    c.execute("SELECT COUNT(*) FROM role_permissions")
    if c.fetchone()[0] > 0:
        conn.close()
        return  # Already seeded
    
    # Get all permission IDs
    c.execute("SELECT id, code FROM permissions")
    perms = {row['code']: row['id'] for row in c.fetchall()}
    
    # Define default permissions for each role
    role_defaults = {
        'General Manager': list(perms.values()),  # Full access to everything
        'Technical Team Leader': list(perms.values()),  # Full access to everything
        'Presale Engineer': [
            perms['view_dashboard'],
            perms['view_projects'],
            perms['view_pipeline_analysis'],
            perms['view_sales_performance'],
            perms['view_aging_dashboard'],
            perms['view_end_users'],
            perms['view_contractors'],
            perms['view_consultants'],
            perms['view_presales'],
            perms['view_rfq'],
            perms['view_rfts'],
            perms['view_reports'],
            perms['view_products'],
            perms['view_solution_builder'],
            perms['view_quotation_builder'],
            perms['view_tasks'],
            perms['view_vendors'],
            perms['view_distributors'],
            perms['view_registration_hub'],
        ],
        'Sales Engineer': [
            perms['view_dashboard'],
            perms['view_projects'],
            perms['view_pipeline_analysis'],
            perms['view_sales_performance'],
            perms['view_aging_dashboard'],
            perms['view_end_users'],
            perms['view_contractors'],
            perms['view_consultants'],
            perms['view_rfq'],
            perms['view_products'],
            perms['view_quotation_builder'],
            perms['view_tasks'],
            perms['view_vendors'],
            perms['view_distributors'],
            perms['view_registration_hub'],
        ],
    }
    
    # Insert role permissions
    for role, permission_ids in role_defaults.items():
        for perm_id in permission_ids:
            c.execute("""
                INSERT OR IGNORE INTO role_permissions (role, permission_id)
                VALUES (?, ?)
            """, (role, perm_id))
    
    conn.commit()
    conn.close()


def get_user_permissions(user_id, user_role):
    """
    Get all effective permissions for a user
    Returns a set of permission codes
    
    Permission evaluation order:
    1. Explicit deny in user_permissions (highest priority)
    2. Explicit allow in user_permissions
    3. Role default permissions
    """
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    
    permissions_set = set()
    denied_permissions = set()
    
    # Get explicit user permissions (allow and deny)
    c.execute("""
        SELECT p.code, up.grant_type
        FROM user_permissions up
        JOIN permissions p ON up.permission_id = p.id
        WHERE up.user_id = ?
    """, (user_id,))
    
    for row in c.fetchall():
        if row['grant_type'] == 'deny':
            denied_permissions.add(row['code'])
        else:  # allow
            permissions_set.add(row['code'])
    
    # Get role default permissions
    c.execute("""
        SELECT p.code
        FROM role_permissions rp
        JOIN permissions p ON rp.permission_id = p.id
        WHERE rp.role = ?
    """, (user_role,))
    
    for row in c.fetchall():
        permissions_set.add(row['code'])
    
    # Remove denied permissions (explicit deny overrides everything)
    permissions_set -= denied_permissions
    
    conn.close()
    return permissions_set


def user_has_permission(permission_code):
    """Check if the current logged-in user has a specific permission"""
    if 'user_id' not in session:
        return False
    
    # Check session cache first
    if 'permissions' not in session:
        # Load permissions into session
        permissions = get_user_permissions(session['user_id'], session['user_role'])
        session['permissions'] = list(permissions)
    
    return permission_code in session['permissions']


def refresh_user_permissions():
    """Refresh permissions in session (call after permission changes)"""
    if 'user_id' in session:
        permissions = get_user_permissions(session['user_id'], session['user_role'])
        session['permissions'] = list(permissions)


def permission_required(*permission_codes):
    """
    Decorator to require specific permissions for a route
    Usage: @permission_required('view_projects', 'manage_users')
    """
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                flash('Please log in to access this page!', 'danger')
                return redirect(url_for('login'))
            
            # Check if user has at least one of the required permissions
            has_permission = False
            for perm_code in permission_codes:
                if user_has_permission(perm_code):
                    has_permission = True
                    break
            
            if not has_permission:
                flash('You do not have permission to access this page!', 'danger')
                return redirect(url_for('index'))
            
            return f(*args, **kwargs)
        return decorated_function
    return decorator


@app.context_processor
def inject_permissions():
    """Make user_has_permission available in all templates"""
    return dict(user_has_permission=user_has_permission)


@app.template_filter('b64encode')
def b64encode_filter(data):
    """Encode data to Base64."""
    return base64.b64encode(data).decode('utf-8') if data else ''


#########
def get_notifications(user_id):
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    c.execute('''
        SELECT * FROM notifications WHERE user_id = ? ORDER BY created_at DESC
    ''', (user_id,))
    notifications = c.fetchall()
    conn.close()
    return notifications
####

#########33
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        conn = sqlite3.connect('ProjectStatus.db')
        c = conn.cursor()
        # The users table schema is: (id, username, password, role)
        c.execute("SELECT * FROM users WHERE username=?", (username,))
        user = c.fetchone()
        
        # Check password (support both hashed and legacy plaintext passwords)
        password_valid = False
        if user:
            stored_password = user[2]
            # Try hashed password first (new format)
            try:
                if check_password_hash(stored_password, password):
                    password_valid = True
            except ValueError:
                # Not a valid hash format, try plaintext comparison
                pass
            
            # Fallback: check if it's a legacy plaintext password
            if not password_valid and stored_password == password:
                password_valid = True
                # Auto-upgrade to hashed password
                hashed_password = generate_password_hash(password)
                c.execute("UPDATE users SET password=? WHERE username=?", (hashed_password, username))
                # Also update engineers table if user is an engineer
                c.execute("UPDATE engineers SET password=? WHERE username=?", (hashed_password, username))
                conn.commit()
        
        conn.close()

        if password_valid:
            session['user_id'] = user[0]
            session['username'] = user[1]
            session['user_role'] = user[3]  # Save user's role in the session
            
            # Load user permissions into session
            refresh_user_permissions()

            flash('Login successful!', 'success')
            return redirect(url_for('index'))
        else:
            flash('Invalid username or password!', 'danger')

    return render_template('login.html')
###
from functools import wraps # Make sure this is imported at the top of app.py

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to access this page.', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function
################3
@app.route('/logout')
#@role_required('editor')
def logout():
    session.pop('user_id', None)  # Remove user_id from the session
    session.pop('username', None)  # Remove username from the session
    flash('You have been logged out.', 'success')
    return redirect(url_for('index'))  # Redirect to the main page

################################3
@app.route('/')
@login_required
@permission_required('view_dashboard')
def index():
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()

    # --- KPI Card Calculations ---
    current_year = str(datetime.now().year)
    pipeline_stages = ('Lead', 'Proposal Prep', 'Proposal Sent', 'Customer Pending', 'Technical Discussion', 'Negotiation')

    # 1. Pipeline Deals Count
    c.execute(f"SELECT COUNT(*) FROM register_project WHERE stage IN {pipeline_stages}")
    pipeline_deals_count = c.fetchone()[0]

    # 2. Pipeline Deals Amount
    c.execute(f"SELECT SUM(deal_value) FROM register_project WHERE stage IN {pipeline_stages}")
    pipeline_deals_amount = c.fetchone()[0] or 0

    # 3. Won Projects Count (This Year) - NOW FROM 'projects' TABLE
    c.execute("SELECT COUNT(*) FROM projects WHERE status = 'WON' AND strftime('%Y', registered_date) = ?",
              (current_year,))
    won_projects_count = c.fetchone()[0]

    # 4. Won Projects Amount (This Year) - NOW FROM 'projects' TABLE
    c.execute(
        "SELECT SUM(quotation_selling_price) FROM projects WHERE status = 'WON' AND strftime('%Y', registered_date) = ?",
        (current_year,))
    won_projects_amount = c.fetchone()[0] or 0

    # 5. Lost Projects Count (This Year)
    c.execute("SELECT COUNT(*) FROM register_project WHERE stage = 'Closed Lost' AND strftime('%Y', registered_date) = ?", (current_year,))
    lost_projects_count = c.fetchone()[0]

    # 6. Lost Projects Amount (This Year)
    c.execute("SELECT SUM(deal_value) FROM register_project WHERE stage = 'Closed Lost' AND strftime('%Y', registered_date) = ?", (current_year,))
    lost_projects_amount = c.fetchone()[0] or 0


    # --- Chart Calculation: Active Deals by Stage ---
    c.execute(
        "SELECT stage, COUNT(*) FROM register_project WHERE stage NOT IN ('Closed Won', 'Closed Lost', 'Cancelled') GROUP BY stage")
    deals_by_stage_raw = c.fetchall()

    conn.close()

    # Prepare data for chart
    deals_stage_labels = [row[0] for row in deals_by_stage_raw]
    deals_stage_data = [row[1] for row in deals_by_stage_raw]


    return render_template('index.html',
                           pipeline_deals_count=pipeline_deals_count,
                           pipeline_deals_amount=pipeline_deals_amount,
                           won_projects_count=won_projects_count,
                           won_projects_amount=won_projects_amount,
                           lost_projects_count=lost_projects_count,
                           lost_projects_amount=lost_projects_amount,
                           deals_stage_labels=deals_stage_labels,
                           deals_stage_data=deals_stage_data
                           )
##########3
import random
@app.route('/register_engineer', methods=['GET', 'POST'])
@role_required('editor','General Manager','Technical Team Leader')
def register_engineer():
    if request.method == 'POST':
        name = request.form['name']
        phone = request.form['phone']
        email = request.form['email']
        role = request.form['role']
        username = request.form['username']
        password = request.form['password']
        hashed_password = generate_password_hash(password)

        # Generate engineer code
        engineer_code = generate_engineer_code(name)
        if engineer_code is None:
            flash('Please provide a full name with at least three parts!', 'danger')
            return redirect(url_for('register_engineer'))

        conn = sqlite3.connect('ProjectStatus.db')
        c = conn.cursor()
        try:
            # Insert into engineers table
            c.execute('''INSERT INTO engineers (name, engineer_code, phone, email, role, username, password)
                         VALUES (?, ?, ?, ?, ?, ?, ?)''',
                      (name, engineer_code, phone, email, role, username, hashed_password))

            # Insert into users table
            c.execute('''INSERT INTO users (username, password, role)
                         VALUES (?, ?, ?)''', (username, hashed_password, role))

            conn.commit()
            flash('Engineer registered successfully!', 'success')
        except sqlite3.IntegrityError:
            flash('Error: Username or engineer code must be unique!', 'danger')
        finally:
            conn.close()

        return redirect(url_for('register_engineer'))

    return render_template('register_engineer.html')

def generate_engineer_code(name):
    parts = name.split()
    if len(parts) < 3:
        return None  # Ensure there are at least three parts for first, middle, last names
    initials = ''.join(part[0].upper() for part in parts[:3])  # First letters of first, middle, and last names
    code = f"{initials}{str(random.randint(100, 999))}"  # Append a random three-digit number
    return code
#######################3
@app.route('/upload', methods=['GET', 'POST'])
@role_required('Technical Team Leader', 'Presale Engineer', 'Sales Engineer', 'editor')
def upload():
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()

    if request.method == 'POST':
        # Get all data from the form, including auto-filled and manual fields
        project_name = request.form['project_name']
        quote_ref = request.form['quote_ref']
        presale_eng = request.form.get('presale_eng')
        sales_eng = request.form.get('sales_eng')
        system = request.form['system']
        sow = request.form['sow']
        status = request.form['status']
        progress = request.form.get('progress')
        rfq_reference = request.form.get('rfq_reference')
        quotation_cost = request.form['quotation_cost']
        quotation_selling_price = request.form['quotation_selling_price']
        quotation_note = request.form.get('quotation_note')
        feedback = request.form.get('feedback')
        quotation_file = request.files['quotation']
        cost_sheet_file = request.files['cost_sheet']

        # Automatic calculation for Quarter and Registered Date
        registered_date = datetime.now()
        month = registered_date.month
        quarter = f"Q{(month - 1) // 3 + 1}"
        registered_date_str = registered_date.strftime('%Y-%m-%d %H:%M:%S')

        # Calculate Margin
        margin = 0
        if float(quotation_selling_price) > 0:
            margin = (1 - (float(quotation_cost) / float(quotation_selling_price))) * 100

        # File handling
        quotation_data = quotation_file.read()
        cost_sheet_data = cost_sheet_file.read()

        try:
            c.execute('''INSERT INTO projects 
                         (project_name, quote_ref, presale_eng, sales_eng, system, sow, status, quarter, 
                          quotation_cost, quotation_selling_price, margin, progress, registered_date, 
                          quotation_note, feedback, quotation, cost_sheet, rfq_reference, registered_by, updated_time)
                         VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                      (project_name, quote_ref, presale_eng, sales_eng, system, sow, status, quarter,
                       quotation_cost, quotation_selling_price, margin, progress, registered_date_str,
                       quotation_note, feedback, quotation_data, cost_sheet_data, rfq_reference,
                       session.get('username'), registered_date_str))

            # If this quotation was generated from an RFQ, update the RFQ's status
            if rfq_reference:
                c.execute(
                    "UPDATE rfq_requests SET rfq_status = 'Quoted', quotation_status = 'Quotation Sent' WHERE rfq_reference = ?",
                    (rfq_reference,))

            conn.commit()
            flash('Quotation submitted successfully!', 'success')
        except sqlite3.IntegrityError:
            flash('Error: That Quote Reference already exists. Please use a unique reference.', 'danger')
        except Exception as e:
            flash(f'An error occurred: {e}', 'danger')
        finally:
            conn.close()

        return redirect(url_for('project_summary'))

    # --- GET request logic ---

    # Get pre-filled data from the URL query (from RFQ page)
    prefill_data = {
        'project_name': request.args.get('project_name'),
        'rfq_reference': request.args.get('rfq_ref'),
        'presale_engineer': request.args.get('presale_engineer'),
        'sales_engineer': request.args.get('sales_engineer')
    }

    # Fetch data for dropdowns (for when the form is NOT pre-filled)
    c.execute("SELECT project_name FROM register_project")
    projects = [row[0] for row in c.fetchall()]
    c.execute("SELECT username FROM engineers WHERE role IN ('Presale Engineer', 'Technical Team Leader')")
    presale_engineers = [row[0] for row in c.fetchall()]
    c.execute("SELECT username FROM engineers WHERE role IN ('Sales Engineer', 'Technical Team Leader')")
    sales_engineers = [row[0] for row in c.fetchall()]
    conn.close()

    status_options = ['Quotation Sent', 'Customer Pending', 'Technical Discussion', 'Negotiation', 'Closed Won',
                      'Closed Lost', 'Cancelled']

    return render_template('submit_quotation.html',
                           projects=projects,
                           presale_engineers=presale_engineers,
                           sales_engineers=sales_engineers,
                           status_options=status_options,
                           prefill_data=prefill_data)
##################3
@app.route('/search_quote', methods=['POST'])
#@role_required('editor')
def search_quote():
    quote_ref = request.form['quote_ref']
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    c.execute("SELECT * FROM projects WHERE quote_ref=?", (quote_ref,))
    project = c.fetchone()
    conn.close()

    if project:
        return render_template('update_project.html', project=project)
    else:
        flash('No project found with that quote reference!', 'danger')
        return redirect(url_for('index'))


@app.route('/update_project/<quote_ref>', methods=['POST'])
#@role_required('editor')
def update_project(quote_ref):
    project_name = request.form['project_name']
    presale_eng = request.form['presale_eng']
    sales_eng = request.form['sales_eng']
    system = request.form['system']
    sow = request.form['sow']
    status = request.form['status']
    quotation_note = request.form['quotation_note']
    feedback = request.form['feedback']
    quarter = request.form['quarter']
    quotation_cost = request.form['quotation_cost']
    quotation_selling_price = request.form['quotation_selling_price']  # Get the updated selling price
    margin = (1 - (float(quotation_cost) / float(quotation_selling_price))) * 100 if float(
        quotation_selling_price) > 0 else 0
    progress = request.form.get('progress')  # Use .get()

    updated_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    updated_by = session.get('username')  # Get the signed-in user's username

    updated_quotation_file = request.files.get('updated_quotation')
    updated_cost_sheet_file = request.files.get('updated_cost_sheet')

    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()

    # Prepare the update query
    update_query = '''
        UPDATE projects
        SET project_name=?, presale_eng=?, sales_eng=?, system=?, sow=?, status=?, 
            quotation_note=?, feedback=?, quarter=?, quotation_cost=?,quotation_selling_price=?, margin=?, progress=?, updated_time=?, updated_by=?
        WHERE quote_ref=?
    '''
    c.execute(update_query, (project_name, presale_eng, sales_eng, system, sow, status,
                             quotation_note, feedback, quarter, quotation_cost,quotation_selling_price, margin, progress, updated_time, updated_by, quote_ref))

    # Handle updated files if they are provided
    if updated_quotation_file:
        quotation_data = updated_quotation_file.read()
        c.execute("UPDATE projects SET quotation=? WHERE quote_ref=?", (quotation_data, quote_ref))

    if updated_cost_sheet_file:
        cost_sheet_data = updated_cost_sheet_file.read()
        c.execute("UPDATE projects SET cost_sheet=? WHERE quote_ref=?", (cost_sheet_data, quote_ref))

    conn.commit()
    conn.close()

    flash('Project updated successfully!', 'success')
    return redirect(url_for('project_summary'))

##########33
##########
@app.route('/project_summary', methods=['GET', 'POST'])
#@role_required('editor')
def project_summary():
    conn = sqlite3.connect('ProjectStatus.db')  # Connect to the database
    c = conn.cursor()

    # Fetch presale engineers
    c.execute("SELECT username FROM engineers WHERE role='Presale Engineer'")
    presale_engineers = [engineer[0] for engineer in c.fetchall()]  # Extract only names

    # Fetch sales engineers or team leaders
    c.execute(
        "SELECT username FROM engineers WHERE role IN ('Sales Engineer', 'Technical Team Leader','General Manager','Project Manager','Implementation Engineer')")
    sales_engineers = [engineer[0] for engineer in c.fetchall()]  # Extract only names

    # Initialize filter parameters
    presale_eng_filter = request.args.get('presale_eng', default=None)  # Get presale engineer filter
    sales_eng_filter = request.args.get('sales_eng', default=None)  # Get sales engineer filter
    status_filter = request.args.get('status', default=None)  # Get status filter
    quarter_filter = request.args.get('quarter', default=None)  # Get quarter filter
    reg_year_filter = request.args.get('reg_year', default=None)  # Get registration year filter
    upd_year_filter = request.args.get('upd_year', default=None)  # Get update year filter

    # Build the query based on filter parameters
    query = "SELECT project_name, presale_eng, sales_eng, status, quarter FROM projects WHERE 1=1"
    params = []  # List to hold parameters for the query

    # Append conditions for filters if they are provided
    if presale_eng_filter:
        query += " AND presale_eng = ?"
        params.append(presale_eng_filter)
    if sales_eng_filter:
        query += " AND sales_eng = ?"
        params.append(sales_eng_filter)
    if status_filter:
        query += " AND status = ?"
        params.append(status_filter)
    if quarter_filter:
        query += " AND quarter = ?"
        params.append(quarter_filter)

    if reg_year_filter:
        query += " AND strftime('%Y', registered_date) = ?"
        params.append(reg_year_filter)
    if upd_year_filter:
        query += " AND strftime('%Y', updated_time) = ?"
        params.append(upd_year_filter)

    c.execute(query, params)  # Execute the constructed query
    projec = c.fetchall()  # Fetch all projects that match the filters

    # Fetch distinct quarters for the dropdown
    c.execute("SELECT DISTINCT quarter FROM projects")
    quarters = [row[0] for row in c.fetchall()]  # List of unique quarters

    # Remove duplicates from the project list
    unique_projects = {}
    for project in projec:
        unique_projects[project[0]] = project  # Assuming project[0] is the unique identifier (project_name)

    # Convert back to a list
    projec = list(unique_projects.values())

    # Fetch status counts for pie chart data
    c.execute("""SELECT status, COUNT(*) FROM projects GROUP BY status""")
    status_counts = c.fetchall()  # Get counts of each status

    # Prepare data for the pie chart
    statuses = [row[0] for row in status_counts]  # List of statuses
    counts = [row[1] for row in status_counts]  # Corresponding counts

    conn.close()  # Close the database connection
    current_year = datetime.now().year  # Get the current year
    return render_template('project_summary.html',
                           quarters=quarters,
                           projec=projec,
                           projects=statuses,
                           counts=counts,
                           presale_engineers=presale_engineers,
                           sales_engineers=sales_engineers,
                           current_year=current_year)

##
@app.route('/project_history/<project_name>', methods=['GET'])
@login_required
def show_project_history(project_name):
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    # Fetch all quotation attempts for the given project name
    c.execute("SELECT * FROM projects WHERE project_name=?", (project_name,))
    projects = c.fetchall()
    conn.close()

    # Always render the history page and let the template decide what to show
    return render_template('project_history.html', projects=projects, project_name=project_name)

@app.route('/project_detail/<int:project_id>', methods=['GET'])
@login_required
def project_detail(project_id):
    """Comprehensive project profile showing all related data"""
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    # Get project details
    cursor.execute("""
        SELECT 
            rp.*,
            eu.name as end_user_name,
            c.name as contractor_name,
            cons.name as consultant_name,
            e.name as sales_engineer_name
        FROM register_project rp
        LEFT JOIN end_users eu ON rp.end_user_id = eu.id
        LEFT JOIN contractors c ON rp.contractor_id = c.id
        LEFT JOIN consultants cons ON rp.consultant_id = cons.id
        LEFT JOIN engineers e ON CAST(rp.sales_engineer_id AS TEXT) = CAST(e.id AS TEXT)
        WHERE rp.id = ?
    """, (project_id,))
    project = cursor.fetchone()
    
    if not project:
        flash('Project not found!', 'danger')
        conn.close()
        return redirect(url_for('view_projects'))
    
    # Get all quotations for this project
    cursor.execute("""
        SELECT * FROM projects
        WHERE project_name = ?
        ORDER BY registered_date DESC
    """, (project['project_name'],))
    quotations = cursor.fetchall()
    
    # Get all RFQs for this project  
    cursor.execute("""
        SELECT * FROM rfq_requests
        WHERE project_name = ?
        ORDER BY requested_time DESC
    """, (project['project_name'],))
    rfqs = cursor.fetchall()
    
    # Get all Purchase Orders for this project
    cursor.execute("""
        SELECT 
            po.*,
            d.name as distributor_name,
            v.name as vendor_name
        FROM purchase_orders po
        LEFT JOIN distributors d ON CAST(po.distributor AS TEXT) = CAST(d.id AS TEXT)
        LEFT JOIN vendors v ON CAST(po.vendor AS TEXT) = CAST(v.id AS TEXT)
        WHERE CAST(po.project_name AS TEXT) = CAST(? AS TEXT)
        ORDER BY po.created_at DESC
    """, (project_id,))
    purchase_orders = cursor.fetchall()
    
    # Get all PO Requests for this project
    cursor.execute("""
        SELECT 
            rfpo.*,
            d.name as distributor_name,
            v.name as vendor_name,
            rfpo.requested_by_name as requester_name,
            sq.filename as supplier_quotation_filename
        FROM po_requests rfpo
        LEFT JOIN distributors d ON CAST(rfpo.distributor_id AS TEXT) = CAST(d.id AS TEXT)
        LEFT JOIN vendors v ON rfpo.vendor_id = v.id
        LEFT JOIN supplier_quotations sq ON rfpo.supplier_quotation_id = sq.id
        WHERE rfpo.project_name = ?
        ORDER BY rfpo.requested_time DESC
    """, (project['project_name'],))
    po_requests = cursor.fetchall()
    
    # Get all supplier quotations for quotations in this project
    quote_refs = [q['quote_ref'] for q in quotations]
    supplier_quotations = []
    if quote_refs:
        placeholders = ','.join('?' * len(quote_refs))
        cursor.execute(f"""
            SELECT * FROM supplier_quotations
            WHERE quote_ref IN ({placeholders})
            ORDER BY uploaded_at DESC
        """, quote_refs)
        supplier_quotations = cursor.fetchall()
    
    # Get all distributors and vendors for the upload form
    cursor.execute("SELECT id, name FROM distributors ORDER BY name")
    distributors = cursor.fetchall()
    
    cursor.execute("SELECT id, name FROM vendors ORDER BY name")
    vendors = cursor.fetchall()
    
    # Calculate statistics
    total_quotation_value = sum(q['quotation_selling_price'] or 0 for q in quotations)
    total_po_value = sum(po['total_amount'] or 0 for po in purchase_orders)
    
    conn.close()
    
    return render_template('project_detail.html',
                         project=project,
                         quotations=quotations,
                         rfqs=rfqs,
                         purchase_orders=purchase_orders,
                         po_requests=po_requests,
                         supplier_quotations=supplier_quotations,
                         distributors=distributors,
                         vendors=vendors,
                         total_quotation_value=total_quotation_value,
                         total_po_value=total_po_value)

@app.route('/download_project_data_excel/<int:project_id>')
@login_required
def download_project_data_excel(project_id):
    """Export all project data (quotations, RFQs, POs) to Excel"""
    conn = sqlite3.connect('ProjectStatus.db')
    
    # Get project name for filename
    cursor = conn.cursor()
    cursor.execute("SELECT project_name FROM register_project WHERE id = ?", (project_id,))
    project_result = cursor.fetchone()
    project_name = project_result[0] if project_result else "Project"
    
    # Query to fetch quotations
    quotations_query = '''
        SELECT 
            quote_ref AS "Quote Reference",
            system AS "System",
            presale_eng AS "Presale Engineer",
            sales_eng AS "Sales Engineer",
            quotation_cost AS "Cost (SAR)",
            quotation_selling_price AS "Selling Price (SAR)",
            margin AS "Margin %",
            status AS "Status",
            registered_date AS "Registered Date",
            updated_time AS "Updated Date",
            sow AS "Scope of Work",
            quotation_note AS "Notes",
            feedback AS "Feedback"
        FROM projects
        WHERE project_name = ?
        ORDER BY registered_date DESC
    '''
    df_quotations = pd.read_sql_query(quotations_query, conn, params=(project_name,))
    
    # Query to fetch RFQs
    rfqs_query = '''
        SELECT 
            rfq_reference AS "RFQ Reference",
            project_status AS "Project Status",
            priority AS "Priority",
            sales_engineer_presale AS "Presale Engineer",
            sales_engineer_sales AS "Sales Engineer",
            rfq_status AS "RFQ Status",
            quotation_status AS "Quotation Status",
            deadline AS "Deadline",
            requested_time AS "Requested Date",
            note AS "Notes"
        FROM rfq_requests
        WHERE project_name = ?
        ORDER BY requested_time DESC
    '''
    df_rfqs = pd.read_sql_query(rfqs_query, conn, params=(project_name,))
    
    # Query to fetch Purchase Orders
    pos_query = '''
        SELECT 
            po.po_request_number AS "PO Request Number",
            po.po_number AS "PO Number",
            d.name AS "Distributor",
            v.name AS "Vendor",
            po.system AS "System/Scope",
            po.total_amount AS "Total Amount (SAR)",
            po.po_approval_status AS "Approval Status",
            po.po_delivery_status AS "Delivery Status",
            eng.username AS "Presale Engineer",
            pmeng.username AS "Project Manager",
            po.created_at AS "Created Date",
            po.po_notes_vendor AS "Vendor Notes",
            po.po_notes_client AS "Client Notes"
        FROM purchase_orders po
        LEFT JOIN distributors d ON CAST(po.distributor AS TEXT) = CAST(d.id AS TEXT)
        LEFT JOIN vendors v ON CAST(po.vendor AS TEXT) = CAST(v.id AS TEXT)
        LEFT JOIN engineers eng ON po.presale_engineer = eng.id
        LEFT JOIN engineers pmeng ON po.project_manager = pmeng.id
        WHERE CAST(po.project_name AS TEXT) = CAST(? AS TEXT)
        ORDER BY po.created_at DESC
    '''
    df_pos = pd.read_sql_query(pos_query, conn, params=(project_id,))
    
    conn.close()
    
    # Create Excel file with multiple sheets
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df_quotations.to_excel(writer, sheet_name='Quotations', index=False)
        df_rfqs.to_excel(writer, sheet_name='RFQs', index=False)
        df_pos.to_excel(writer, sheet_name='Purchase Orders', index=False)
        
        # Auto-adjust column widths for all sheets
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            if sheet_name == 'Quotations':
                df = df_quotations
            elif sheet_name == 'RFQs':
                df = df_rfqs
            else:
                df = df_pos
                
            for idx, col in enumerate(df.columns):
                max_length = max(
                    df[col].astype(str).map(len).max(),
                    len(str(col))
                )
                worksheet.set_column(idx, idx, min(max_length + 2, 50))
    
    output.seek(0)
    filename = f'{project_name}_Project_Data.xlsx'
    
    return send_file(output, 
                     download_name=filename, 
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
############333
@app.route('/charts')
#@role_required('editor')
def charts():
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()

    # Fetch total won cost by summing all quotation costs
    c.execute("""
        SELECT SUM(quotation_cost) FROM projects WHERE status='WON'
    """)
    total_won_cost = c.fetchone()[0] or 0  # Default to 0 if None

    # Fetch all projects with their quote references, costs, and engineers
    c.execute("""
        SELECT quote_ref, quotation_cost, sales_eng, presale_eng
        FROM projects
        WHERE status='WON'
    """)
    won_projects = c.fetchall()
    conn.close()

    # Prepare data for the chart
    labels = [row[0] for row in won_projects]
    data = [row[1] for row in won_projects]
    sales_eng = [row[2] for row in won_projects]
    presale_eng = [row[3] for row in won_projects]

    return render_template('charts.html', labels=labels, data=data,
                           sales_eng=sales_eng, presale_eng=presale_eng,
                           total_won_cost=total_won_cost)
##########3
@app.route('/export_project_history/<project_name>')
#@role_required('editor')
def export_project_history(project_name):
    conn = sqlite3.connect('ProjectStatus.db')

    # Fetch data without quotation and cost sheet columns
    df = pd.read_sql_query(
        "SELECT id, project_name, quote_ref, presale_eng, sales_eng, system, sow, status, quotation_note, feedback,registered_date "
        "FROM projects WHERE project_name=?", conn, params=(project_name,)
    )
    conn.close()

    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Project History')

    output.seek(0)
    filename = f'{project_name}_project_history.xlsx'
    return send_file(output, download_name=filename, as_attachment=True)
###############
@app.route('/download_project_history/<project_name>')
#@role_required('editor')
def download_project_history(project_name):
    conn = sqlite3.connect('ProjectStatus.db')

    # Fetch data without quotation and cost sheet columns
    df = pd.read_sql_query(
        "SELECT id, project_name, quote_ref, presale_eng, sales_eng, system, sow, status, quotation_note, feedback,quarter,quotation_cost,registered_date,progress,registered_by,updated_by,updated_time "
        "FROM projects WHERE project_name=?", conn, params=(project_name,)
    )
    conn.close()

    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Project History')

    output.seek(0)
    # Set filename to the same name as the last cost sheet uploaded
    if last_cost_sheet_path:
        filename = os.path.basename(last_cost_sheet_path).replace('.xlsx', '_history.xlsx')
    else:
        filename = f'{project_name}_project_history.xlsx'

    return send_file(output, download_name=filename, as_attachment=True)

@app.route('/download_quotation', methods=['POST'])
#@role_required('editor')
def download_quotation():
    quote_ref = request.form['quote_ref']
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    c.execute("SELECT quotation FROM projects WHERE quote_ref=?", (quote_ref,))

    result = c.fetchone()
    conn.close()

    if result:
        quotation_data = result[0]
        output = BytesIO(quotation_data)
        return send_file(output, download_name=f'{quote_ref}.xlsx', as_attachment=True)
    else:
        flash('Quotation not found!', 'danger')
        return redirect(url_for('index'))

@app.route('/download_cost_sheet', methods=['POST'])
#@role_required('editor')
def download_cost_sheet():
    quote_ref = request.form['quote_ref']
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    c.execute("SELECT cost_sheet FROM projects WHERE quote_ref=?", (quote_ref,))

    result = c.fetchone()
    conn.close()

    if result:
        cost_sheet_data = result[0]
        output = BytesIO(cost_sheet_data)
        return send_file(output, download_name=f'{quote_ref}cost_sheet.xlsx', as_attachment=True)
    else:
        flash('Cost sheet not found!', 'danger')
        return redirect(url_for('index'))

@app.route('/upload_supplier_quotation', methods=['POST'])
@login_required
def upload_supplier_quotation():
    """Upload supplier quotation PDF for a specific quotation with distributor/vendor tracking"""
    try:
        quote_ref = request.form.get('quote_ref')
        distributor_id_str = request.form.get('distributor_id')
        vendor_id_str = request.form.get('vendor_id')
        system = request.form.get('system', '')
        notes = request.form.get('notes', '')
        
        if not quote_ref:
            flash('Quote reference is required!', 'danger')
            return redirect(request.referrer or url_for('index'))
        
        # Validate at least one supplier is selected
        if not distributor_id_str and not vendor_id_str:
            flash('Please select at least one distributor or vendor!', 'danger')
            return redirect(request.referrer or url_for('index'))
        
        # Get the uploaded file
        quotation_file = request.files.get('supplier_quotation_file')
        if not quotation_file or quotation_file.filename == '':
            flash('Please select a PDF file to upload!', 'danger')
            return redirect(request.referrer or url_for('index'))
        
        # Validate file type
        if not quotation_file.filename.lower().endswith('.pdf'):
            flash('Only PDF files are allowed!', 'danger')
            return redirect(request.referrer or url_for('index'))
        
        # Read file data
        quotation_data = quotation_file.read()
        filename = quotation_file.filename
        
        conn = sqlite3.connect('ProjectStatus.db')
        cursor = conn.cursor()
        
        # Get distributor name if selected
        distributor_id = None
        distributor_name = None
        if distributor_id_str:
            distributor_id = int(distributor_id_str)
            cursor.execute("SELECT name FROM distributors WHERE id = ?", (distributor_id,))
            result = cursor.fetchone()
            distributor_name = result[0] if result else None
        
        # Get vendor name if selected
        vendor_id = None
        vendor_name = None
        if vendor_id_str:
            vendor_id = int(vendor_id_str)
            cursor.execute("SELECT name FROM vendors WHERE id = ?", (vendor_id,))
            result = cursor.fetchone()
            vendor_name = result[0] if result else None
        
        # Insert the supplier quotation
        cursor.execute("""
            INSERT INTO supplier_quotations 
            (quote_ref, distributor_id, distributor_name, vendor_id, vendor_name, 
             system, quotation_file, filename, uploaded_by, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (quote_ref, distributor_id, distributor_name, vendor_id, vendor_name,
              system, quotation_data, filename, session['username'], notes))
        
        conn.commit()
        conn.close()
        
        # Build success message
        supplier_names = []
        if distributor_name:
            supplier_names.append(f"Distributor: {distributor_name}")
        if vendor_name:
            supplier_names.append(f"Vendor: {vendor_name}")
        
        flash(f'Supplier quotation uploaded successfully for {", ".join(supplier_names)}!', 'success')
        return redirect(request.referrer or url_for('index'))
        
    except Exception as e:
        flash(f'Error uploading supplier quotation: {str(e)}', 'danger')
        return redirect(request.referrer or url_for('index'))

@app.route('/download_supplier_quotation/<int:quotation_id>', methods=['GET'])
@login_required
def download_supplier_quotation(quotation_id):
    """View or Download a specific supplier quotation PDF"""
    try:
        action = request.args.get('action', 'download')  # Default to download
        
        conn = sqlite3.connect('ProjectStatus.db')
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT quotation_file, filename, distributor_name, vendor_name 
            FROM supplier_quotations 
            WHERE id = ?
        """, (quotation_id,))
        
        result = cursor.fetchone()
        conn.close()
        
        if result and result[0]:
            quotation_data = result[0]
            filename = result[1]
            supplier_name = result[2] or result[3]
            
            output = BytesIO(quotation_data)
            
            # If action is 'view', display inline; otherwise download
            if action == 'view':
                return send_file(output, mimetype='application/pdf', as_attachment=False, download_name=filename)
            else:
                return send_file(output, download_name=filename, as_attachment=True, mimetype='application/pdf')
        else:
            flash('Supplier quotation not found!', 'danger')
            return redirect(request.referrer or url_for('index'))
            
    except Exception as e:
        flash(f'Error accessing supplier quotation: {str(e)}', 'danger')
        return redirect(request.referrer or url_for('index'))


# Add Product from Supplier Quotation
@app.route('/add_product_from_quotation/<int:quotation_id>', methods=['POST'])
@login_required
def add_product_from_quotation(quotation_id):
    """Add a product extracted from a supplier quotation"""
    try:
        part_number = request.form.get('part_number', '').strip()
        description = request.form.get('description', '').strip()
        unit_price = request.form.get('unit_price', '').strip()
        quantity = request.form.get('quantity', '').strip()
        currency = request.form.get('currency', 'USD').strip()
        notes = request.form.get('notes', '').strip()
        
        # Validate required fields
        if not part_number or not description or not unit_price:
            flash('Part Number, Description, and Unit Price are required!', 'danger')
            return redirect(request.referrer or url_for('index'))
        
        # Validate unit price is a number
        try:
            unit_price_float = float(unit_price)
        except ValueError:
            flash('Unit Price must be a valid number!', 'danger')
            return redirect(request.referrer or url_for('index'))
        
        # Validate quantity if provided
        quantity_int = None
        if quantity:
            try:
                quantity_int = int(quantity)
            except ValueError:
                flash('Quantity must be a valid integer!', 'danger')
                return redirect(request.referrer or url_for('index'))
        
        conn = sqlite3.connect('ProjectStatus.db')
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        # Get supplier quotation details for metadata
        cursor.execute("""
            SELECT quote_ref, distributor_name, vendor_name, system, 
                   distributor_id, vendor_id
            FROM supplier_quotations 
            WHERE id = ?
        """, (quotation_id,))
        quotation = cursor.fetchone()
        
        if not quotation:
            flash('Supplier quotation not found!', 'danger')
            conn.close()
            return redirect(request.referrer or url_for('index'))
        
        # Determine supplier name and type
        supplier_name = quotation['distributor_name'] or quotation['vendor_name'] or 'Unknown'
        supplier_type = 'distributor' if quotation['distributor_name'] else 'vendor'
        
        # Insert the product
        cursor.execute("""
            INSERT INTO quotation_products 
            (supplier_quotation_id, part_number, description, unit_price, quantity, 
             currency, notes, quote_ref, supplier_name, supplier_type, system, added_by,
             vendor_id, distributor_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (quotation_id, part_number, description, unit_price_float, quantity_int,
              currency, notes, quotation['quote_ref'], supplier_name, supplier_type,
              quotation['system'], session.get('username'),
              quotation['vendor_id'], quotation['distributor_id']))
        
        conn.commit()
        conn.close()
        
        flash(f'Product "{part_number}" added successfully!', 'success')
        return redirect(request.referrer or url_for('index'))
        
    except Exception as e:
        flash(f'Error adding product: {str(e)}', 'danger')
        return redirect(request.referrer or url_for('index'))


# Quotation Products Dashboard
@app.route('/quotation_products')
@login_required
def quotation_products_dashboard():
    """Dashboard showing all products extracted from supplier quotations"""
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    # Get filter parameters
    search_query = request.args.get('search_query', '').strip()
    supplier_filter = request.args.get('supplier_filter', '').strip()
    system_filter = request.args.get('system_filter', '').strip()
    start_date = request.args.get('start_date', '').strip()
    end_date = request.args.get('end_date', '').strip()
    
    # Build the query with filters
    query = """
        SELECT 
            qp.*,
            sq.filename as quotation_filename,
            p.project_name,
            v.name as vendor_name,
            d.name as distributor_name
        FROM quotation_products qp
        LEFT JOIN supplier_quotations sq ON qp.supplier_quotation_id = sq.id
        LEFT JOIN projects p ON qp.quote_ref = p.quote_ref
        LEFT JOIN vendors v ON qp.vendor_id = v.id
        LEFT JOIN distributors d ON qp.distributor_id = d.id
        WHERE 1=1
    """
    params = []
    
    if search_query:
        query += " AND (qp.part_number LIKE ? OR qp.description LIKE ?)"
        params.extend([f'%{search_query}%', f'%{search_query}%'])
    
    if supplier_filter:
        query += " AND qp.supplier_name LIKE ?"
        params.append(f'%{supplier_filter}%')
    
    if system_filter:
        query += " AND qp.system = ?"
        params.append(system_filter)
    
    # Date range filtering
    if start_date:
        query += " AND DATE(qp.added_at) >= ?"
        params.append(start_date)
    
    if end_date:
        query += " AND DATE(qp.added_at) <= ?"
        params.append(end_date)
    
    query += " ORDER BY qp.added_at DESC"
    
    cursor.execute(query, params)
    products = cursor.fetchall()
    
    # Get unique suppliers for filter dropdown
    cursor.execute("""
        SELECT DISTINCT supplier_name 
        FROM quotation_products 
        WHERE supplier_name IS NOT NULL
        ORDER BY supplier_name
    """)
    suppliers = [row['supplier_name'] for row in cursor.fetchall()]
    
    # Get unique systems for filter dropdown
    cursor.execute("""
        SELECT DISTINCT system 
        FROM quotation_products 
        WHERE system IS NOT NULL AND system != ''
        ORDER BY system
    """)
    systems = [row['system'] for row in cursor.fetchall()]
    
    conn.close()
    
    return render_template('quotation_products.html',
                         products=products,
                         suppliers=suppliers,
                         systems=systems,
                         search_query=search_query,
                         supplier_filter=supplier_filter,
                         system_filter=system_filter,
                         start_date=start_date,
                         end_date=end_date)


# Delete Product
@app.route('/delete_quotation_product/<int:product_id>', methods=['POST'])
@login_required
def delete_quotation_product(product_id):
    """Delete a product from quotation products"""
    try:
        conn = sqlite3.connect('ProjectStatus.db')
        cursor = conn.cursor()
        
        cursor.execute("DELETE FROM quotation_products WHERE id = ?", (product_id,))
        conn.commit()
        conn.close()
        
        flash('Product deleted successfully!', 'success')
    except Exception as e:
        flash(f'Error deleting product: {str(e)}', 'danger')
    
    return redirect(request.referrer or url_for('quotation_products_dashboard'))


###################
@app.route('/download_files', methods=['GET', 'POST'])
#@role_required('editor')
def download_files():
    return render_template('download_files.html')
###################
@app.route('/edit_project/<quote_ref>', methods=['GET', 'POST'])
@role_required('Technical Team Leader', 'Sales Engineer', 'Presale Engineer', 'editor')
def edit_project(quote_ref):
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    if request.method == 'POST':
        # First, fetch the existing project to get the rfq_reference and old file data
        c.execute("SELECT quotation, cost_sheet, rfq_reference FROM projects WHERE quote_ref = ?", (quote_ref,))
        existing_project_data = c.fetchone()
        if not existing_project_data:
            flash('Quotation not found!', 'danger')
            conn.close()
            return redirect(url_for('registered_quotations'))

        # Get text/select data from the form
        project_name = request.form['project_name']
        presale_eng = request.form['presale_eng']
        sales_eng = request.form['sales_eng']
        system = request.form['system']
        sow = request.form['sow']
        status = request.form['status']  # This is the new status
        quotation_note = request.form['quotation_note']
        feedback = request.form['feedback']
        progress = request.form.get('progress')
        quotation_cost = request.form['quotation_cost']
        quotation_selling_price = request.form['quotation_selling_price']

        # Get optional new files from the form
        updated_quotation_file = request.files.get('updated_quotation')
        updated_cost_sheet_file = request.files.get('updated_cost_sheet')

        # Decide which file data to use
        quotation_data = updated_quotation_file.read() if updated_quotation_file and updated_quotation_file.filename != '' else \
        existing_project_data['quotation']
        cost_sheet_data = updated_cost_sheet_file.read() if updated_cost_sheet_file and updated_cost_sheet_file.filename != '' else \
        existing_project_data['cost_sheet']

        margin = (1 - (float(quotation_cost) / float(quotation_selling_price))) * 100 if float(
            quotation_selling_price) > 0 else 0
        updated_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        updated_by = session.get('username')

        try:
            # 1. Update the 'projects' table (the quotation itself)
            c.execute("""
                UPDATE projects SET 
                project_name=?, presale_eng=?, sales_eng=?, system=?, sow=?, status=?, 
                quotation_note=?, feedback=?, progress=?, quotation_cost=?, 
                quotation_selling_price=?, margin=?, updated_time=?, updated_by=?,
                quotation=?, cost_sheet=?
                WHERE quote_ref=?
            """, (project_name, presale_eng, sales_eng, system, sow, status, quotation_note,
                  feedback, progress, quotation_cost, quotation_selling_price, margin,
                  updated_time, updated_by, quotation_data, cost_sheet_data, quote_ref))

            # 2. If this quotation is linked to an RFQ, update the RFQ's status as well
            rfq_ref_to_update = existing_project_data['rfq_reference']
            if rfq_ref_to_update:
                c.execute("UPDATE rfq_requests SET quotation_status = ? WHERE rfq_reference = ?",
                          (status, rfq_ref_to_update))

            conn.commit()
            flash('Project and linked RFQ status updated successfully!', 'success')
        except Exception as e:
            flash(f"An error occurred: {e}", 'danger')
        finally:
            conn.close()

        return redirect(url_for('project_summary'))

    # --- GET request logic (remains the same) ---
    c.execute('SELECT * FROM projects WHERE quote_ref = ?', (quote_ref,))
    project = c.fetchone()
    c.execute("SELECT username FROM engineers WHERE role IN ('Presale Engineer', 'Technical Team Leader')")
    presale_engineers = c.fetchall()
    c.execute("SELECT username FROM engineers WHERE role IN ('Sales Engineer', 'Technical Team Leader')")
    sales_engineers = c.fetchall()
    conn.close()

    status_options = ['Quotation Sent', 'Customer Pending', 'Technical Discussion', 'Negotiation', 'Closed Won',
                      'Closed Lost', 'Cancelled']

    if project is None:
        flash('Quotation not found!', 'danger')
        return redirect(url_for('registered_quotations'))

    return render_template('update_project.html',
                           project=project,
                           presale_engineers=presale_engineers,
                           sales_engineers=sales_engineers,
                           status_options=status_options)
#############
@app.route('/registered_quotations')
@permission_required('view_reports')
def registered_quotations():
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()

    # Get filter values from URL
    filters = {
        'start_registered': request.args.get('start_registered_date'),
        'end_registered': request.args.get('end_registered_date'),
        'start_updated': request.args.get('start_updated_date'),
        'end_updated': request.args.get('end_updated_date'),
        'sales_eng': request.args.get('sales_eng'),
        'presale_eng': request.args.get('presale_eng')
    }

    # Base query updated to select quotation_cost
    query = """
        SELECT id, project_name, quote_ref, presale_eng, sales_eng, status, 
               quotation_selling_price, margin, registered_date, updated_time,
               quotation_cost 
        FROM projects 
        WHERE 1=1
    """
    params = []

    # Apply filters to the query
    if filters['start_registered']: query += " AND date(registered_date) >= ?"; params.append(
        filters['start_registered'])
    if filters['end_registered']: query += " AND date(registered_date) <= ?"; params.append(filters['end_registered'])
    if filters['start_updated']: query += " AND date(updated_time) >= ?"; params.append(filters['start_updated'])
    if filters['end_updated']: query += " AND date(updated_time) <= ?"; params.append(filters['end_updated'])
    if filters['sales_eng']: query += " AND sales_eng = ?"; params.append(filters['sales_eng'])
    if filters['presale_eng']: query += " AND presale_eng = ?"; params.append(filters['presale_eng'])

    query += " ORDER BY registered_date DESC"
    c.execute(query, params)
    quotations = c.fetchall()

    # Fetch engineers for filter dropdowns
    c.execute("SELECT DISTINCT username FROM engineers WHERE role IN ('Sales Engineer', 'Technical Team Leader')")
    sales_engineers = [row[0] for row in c.fetchall()]
    c.execute("SELECT DISTINCT username FROM engineers WHERE role IN ('Presale Engineer', 'Technical Team Leader')")
    presale_engineers = [row[0] for row in c.fetchall()]
    conn.close()

    return render_template('registered_quotations.html',
                           quotations=quotations,
                           presale_engineers=presale_engineers,
                           sales_engineers=sales_engineers,
                           current_filters=filters)


@app.route('/download_quotations_excel')
@login_required
def download_quotations_excel():
    conn = sqlite3.connect('ProjectStatus.db')

    # Replicate the filtering logic from the main page
    filters = {
        'start_registered': request.args.get('start_registered_date'),
        'end_registered': request.args.get('end_registered_date'),
        'start_updated': request.args.get('start_updated_date'),
        'end_updated': request.args.get('end_updated_date'),
        'sales_eng': request.args.get('sales_eng'),
        'presale_eng': request.args.get('presale_eng')
    }

    query = "SELECT * FROM projects WHERE 1=1"  # Select all columns for the report
    params = []

    if filters['start_registered']: query += " AND date(registered_date) >= ?"; params.append(
        filters['start_registered'])
    if filters['end_registered']: query += " AND date(registered_date) <= ?"; params.append(filters['end_registered'])
    if filters['start_updated']: query += " AND date(updated_time) >= ?"; params.append(filters['start_updated'])
    if filters['end_updated']: query += " AND date(updated_time) <= ?"; params.append(filters['end_updated'])
    if filters['sales_eng']: query += " AND sales_eng = ?"; params.append(filters['sales_eng'])
    if filters['presale_eng']: query += " AND presale_eng = ?"; params.append(filters['presale_eng'])

    # Use pandas to read data and create an Excel file
    df = pd.read_sql_query(query, conn, params=params)
    conn.close()

    # Remove binary data columns before exporting
    df = df.drop(columns=['quotation', 'cost_sheet'], errors='ignore')

    output = BytesIO()
    df.to_excel(output, index=False, sheet_name='Quotations Report')
    output.seek(0)

    return send_file(output, download_name='quotations_report.xlsx', as_attachment=True)

#################
@app.route('/aging_dashboard')
@permission_required('view_aging_dashboard')
def aging_dashboard():
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()

    today = datetime.now()

    # Define the unique aging threshold for each stage (in days)
    stage_thresholds = {
        'Lead': 180,
        'Proposal Prep': 5,
        'Proposal Sent': 30,
        'Customer Pending': 60,
        'Technical Discussion': 30,
        'Negotiation': 30
    }

    c.execute("""
        SELECT 
            rp.project_name, rp.stage, en.username AS sales_engineer,
            rp.updated_time, rp.registered_date
        FROM register_project rp
        LEFT JOIN engineers en ON rp.sales_engineer_id = en.id
        WHERE rp.stage NOT IN ('Closed Won', 'Closed Lost', 'Cancelled')
    """)
    projects_raw = c.fetchall()
    conn.close()

    aging_deals = []
    stalled_deals_count = 0
    total_age_days = 0
    deals_with_age = 0

    for project in projects_raw:
        project_name, stage, sales_engineer, updated_time, registered_date = project

        last_action_date_str = updated_time if updated_time else registered_date

        age_days = None
        if last_action_date_str:
            try:
                last_action_date = datetime.strptime(last_action_date_str, '%Y-%m-%d %H:%M:%S')
                age_days = (today - last_action_date).days

                total_age_days += age_days
                deals_with_age += 1

                # Check if the deal is stalled based on its stage's unique threshold
                threshold = stage_thresholds.get(stage, 30)  # Default to 30 if stage not in map
                if age_days > threshold:
                    stalled_deals_count += 1

            except (ValueError, TypeError):
                age_days = None

        aging_deals.append({
            'name': project_name, 'stage': stage,
            'sales_engineer': sales_engineer, 'age': age_days
        })

    aging_deals.sort(key=lambda x: x['age'] if isinstance(x['age'], int) else -1, reverse=True)
    average_age = total_age_days / deals_with_age if deals_with_age > 0 else 0

    return render_template('aging_dashboard.html',
                           aging_deals=aging_deals,
                           stalled_deals_count=stalled_deals_count,
                           average_age=average_age,
                           stage_thresholds=stage_thresholds)  # Pass thresholds to the template
####################
@app.route('/upload_documents', methods=['GET', 'POST'])
#@role_required('editor')
def upload_documents():
    if request.method == 'POST':
        project_id = request.form['project_id']
        system_name = request.form['system']
        sld_file = request.files['sld']
        technical_submittal_file = request.files['technical_submittal']
        other_document_file = request.files.get('other_document')

        # Read the files as binary
        sld_data = sld_file.read() if sld_file else None
        tech_data = technical_submittal_file.read() if technical_submittal_file else None
        other_data = other_document_file.read() if other_document_file else None

        # Store the documents in the database
        conn = sqlite3.connect('ProjectStatus.db')
        c = conn.cursor()
        try:
            c.execute('''INSERT INTO project_documents (project_id, system_name, sld, technical_submittal, other_document)
                         VALUES (?, ?, ?, ?, ?)''',
                         (project_id, system_name, sld_data, tech_data, other_data))
            conn.commit()
            flash('Documents uploaded successfully!', 'success')
        except Exception as e:
            flash(f'Error: {str(e)}', 'danger')
        finally:
            conn.close()

        return redirect(url_for('upload_documents'))

    # Fetch project IDs from the database for the dropdown
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    c.execute("SELECT id, project_name FROM projects")
    projects = c.fetchall()
    conn.close()

    return render_template('upload_documents.html', projects=projects)
######3
@app.route('/view_documents', methods=['GET'])
#@role_required('editor')
def view_documents():
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()

    # Fetch project names for the dropdown
    c.execute('SELECT project_name FROM projects')
    project_names = [row[0] for row in c.fetchall()]

    project_name_filter = request.args.get('project_name')  # Get the project name from the query

    if project_name_filter:
        c.execute('''
            SELECT pd.id, p.project_name, pd.system_name
            FROM project_documents pd
            JOIN projects p ON pd.project_id = p.id
            WHERE p.project_name LIKE ?
        ''', ('%' + project_name_filter + '%',))
    else:
        c.execute('''
            SELECT pd.id, p.project_name, pd.system_name
            FROM project_documents pd
            JOIN projects p ON pd.project_id = p.id
        ''')

    documents = c.fetchall()
    conn.close()

    # Prepare documents data for rendering
    documents_data = [{'id': doc[0], 'project_name': doc[1], 'system_name': doc[2]} for doc in documents]

    return render_template('view_documents.html', documents=documents_data, projects=project_names)
############3
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify, make_response
@app.route('/download/<doc_type>/<int:doc_id>', methods=['GET'])
@login_required
def download_file(doc_type, doc_id):
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    c.execute(f'SELECT {doc_type} FROM project_documents WHERE id = ?', (doc_id,))
    file_data = c.fetchone()
    conn.close()

    if file_data and file_data[0]:
        response = make_response(file_data[0])
        response.headers.set('Content-Type', 'application/pdf')
        response.headers.set('Content-Disposition', f'attachment; filename={doc_type}.pdf')
        return response
    else:
        flash('File not found!', 'danger')
        return redirect(url_for('view_documents'))
##############
@app.route('/products')
@login_required
def products_page():
    return render_template('products_page.html')
#############
@app.route('/register_vendor', methods=['GET', 'POST'])
#@role_required('editor')
def register_vendor():
    if request.method == 'POST':
        name = request.form['name']
        address = request.form['address']
        contact_person = request.form['contact_person']
        phone = request.form['phone']
        email = request.form['email']

        conn = sqlite3.connect('ProjectStatus.db')
        c = conn.cursor()
        try:
            c.execute('''INSERT INTO vendors (name, address, contact_person, phone, email)
                         VALUES (?, ?, ?, ?, ?)''', (name, address, contact_person, phone, email))
            conn.commit()
            flash('Vendor registered successfully!', 'success')
        except sqlite3.IntegrityError:
            flash('Error: Vendor name must be unique!', 'danger')
        finally:
            conn.close()

        return redirect(url_for('register_vendor'))

    return render_template('register_vendor.html')

############################3333
@app.route('/register_distributor', methods=['GET', 'POST'])
@login_required
def register_distributor():
    """Register a new distributor with full SRM support"""
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    if request.method == 'POST':
        name = request.form['name']
        address = request.form.get('address', '')
        contact_person = request.form.get('contact_person', '')
        phone = request.form.get('phone', '')
        email = request.form.get('email', '')
        status = request.form.get('status', 'active')
        category = request.form.get('category', '')
        website = request.form.get('website', '')
        notes = request.form.get('notes', '')
        selected_vendors = request.form.getlist('vendors')

        try:
            # Insert distributor with all SRM fields
            cursor.execute('''
                INSERT INTO distributors 
                (name, address, contact_person, phone, email, status, category, website, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (name, address, contact_person, phone, email, status, category, website, notes))
            
            distributor_id = cursor.lastrowid

            # Link to selected vendors
            for vendor_id in selected_vendors:
                cursor.execute('''
                    INSERT INTO vendor_distributor (vendor_id, distributor_id)
                    VALUES (?, ?)
                ''', (vendor_id, distributor_id))

            # Log activity in SRM activity log
            cursor.execute("""
                INSERT INTO srm_activity_log (entity_type, entity_id, activity_type, description, user_id)
                VALUES ('distributor', ?, 'distributor_created', ?, ?)
            """, (distributor_id, f'Created new distributor: {name}', session.get('user_id')))

            conn.commit()
            flash(f'Distributor "{name}" registered successfully!', 'success')
            
            # Redirect to distributor detail page
            return redirect(url_for('distributor_detail', distributor_id=distributor_id))
            
        except sqlite3.IntegrityError:
            flash('Error: Distributor name must be unique!', 'danger')
            return redirect(url_for('register_distributor'))
        except Exception as e:
            conn.rollback()
            flash(f'An error occurred: {e}', 'danger')
            return redirect(url_for('register_distributor'))
        finally:
            conn.close()

    # Fetch all vendors for the form
    cursor.execute("SELECT id, name FROM vendors ORDER BY name")
    vendors = cursor.fetchall()
    conn.close()
    return render_template('register_distributor.html', vendors=vendors)
###################3
# NEW SRM-ENABLED DISTRIBUTORS ROUTE
@app.route('/distributors')
@login_required
@permission_required('view_distributors')
def show_distributors():
    """Display all distributors with SRM enhancements (associated vendors)"""
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    search_query = request.args.get('search_query', '')

    # Query distributors with optional search
    query = "SELECT * FROM distributors"
    params = []
    if search_query:
        query += " WHERE name LIKE ?"
        params.append(f'%{search_query}%')
    query += " ORDER BY name"

    cursor.execute(query, params)
    distributors = cursor.fetchall()

    # Enrich with associated vendors
    distributors_data = []
    for distributor in distributors:
        # Get associated vendors for this distributor
        cursor.execute("""
            SELECT v.* 
            FROM vendors v
            JOIN vendor_distributor vd ON v.id = vd.vendor_id
            WHERE vd.distributor_id = ?
            ORDER BY v.name
        """, (distributor['id'],))
        vendors = cursor.fetchall()
        
        distributors_data.append({
            'distributor': distributor,
            'vendors': vendors
        })

    conn.close()
    return render_template('distributors.html', distributors_data=distributors_data, search_query=search_query)

##########
# Add this new route anywhere in your app.py file

@app.route('/add_distributor_ajax', methods=['POST'])
@login_required
def add_distributor_ajax():
    """Handles new distributor registration from a modal form with full SRM support."""
    data = request.get_json()
    name = data.get('name')

    # Basic validation
    if not name:
        return jsonify({'status': 'error', 'message': 'Distributor name is required.'}), 400

    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    try:
        # Insert the new distributor with full SRM fields
        c.execute('''INSERT INTO distributors (name, address, contact_person, phone, email, status, category, website, notes)
                     VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                  (name, 
                   data.get('address', ''), 
                   data.get('contact_person', ''), 
                   data.get('phone', ''),
                   data.get('email', ''),
                   data.get('status', 'active'),
                   data.get('category', ''),
                   data.get('website', ''),
                   'Quick-added from PO Request form'))

        new_distributor_id = c.lastrowid
        
        # Log activity in SRM activity log
        c.execute("""
            INSERT INTO srm_activity_log (entity_type, entity_id, activity_type, description, user_id)
            VALUES ('distributor', ?, 'distributor_created', ?, ?)
        """, (new_distributor_id, f'Quick-registered distributor: {name}', session.get('user_id')))
        
        conn.commit()

        # Return a success response with the new distributor's details
        return jsonify({
            'status': 'success',
            'id': new_distributor_id,
            'name': name
        })

    except sqlite3.IntegrityError:
        # This error occurs if the distributor name is not unique
        return jsonify({'status': 'error', 'message': 'A distributor with this name already exists.'}), 409
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        conn.close()
@app.route('/add_vendor_ajax', methods=['POST'])
@login_required
def add_vendor_ajax():
    """Handles new vendor registration from a modal form with full SRM support."""
    data = request.get_json()
    name = data.get('name')

    # Basic validation
    if not name:
        return jsonify({'status': 'error', 'message': 'Vendor name is required.'}), 400

    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    try:
        # Insert the new vendor with full SRM fields
        c.execute('''INSERT INTO vendors (name, address, contact_person, phone, email, status, category, website, notes)
                     VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                  (name, 
                   data.get('address', ''), 
                   data.get('contact_person', ''), 
                   data.get('phone', ''),
                   data.get('email', ''),
                   data.get('status', 'active'),
                   data.get('category', ''),
                   data.get('website', ''),
                   'Quick-added from PO Request form'))

        new_vendor_id = c.lastrowid
        
        # Log activity in SRM activity log
        c.execute("""
            INSERT INTO srm_activity_log (entity_type, entity_id, activity_type, description, user_id)
            VALUES ('vendor', ?, 'vendor_created', ?, ?)
        """, (new_vendor_id, f'Quick-registered vendor: {name}', session.get('user_id')))
        
        conn.commit()

        # Return a success response with the new vendor's details
        return jsonify({
            'status': 'success',
            'id': new_vendor_id,
            'name': name
        })

    except sqlite3.IntegrityError:
        # This error occurs if the vendor name is not unique
        return jsonify({'status': 'error', 'message': 'A vendor with this name already exists.'}), 409
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        conn.close()

############################
# Add this new route anywhere in your app.py file

@app.route('/get_distributor_details/<int:distributor_id>')
@login_required
def get_distributor_details(distributor_id):
    """Fetches distributor details and contacts to auto-populate the PO form."""
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    # Fetch main distributor contact
    c.execute("SELECT contact_person, phone, email FROM distributors WHERE id = ?", (distributor_id,))
    distributor = c.fetchone()
    
    # Fetch all contacts for this distributor from SRM contacts table
    c.execute("""
        SELECT id, name, role, phone, email, is_primary
        FROM distributor_contacts
        WHERE distributor_id = ?
        ORDER BY is_primary DESC, name ASC
    """, (distributor_id,))
    contacts = c.fetchall()
    
    conn.close()

    if distributor:
        # Build contacts list
        contacts_list = []
        for contact in contacts:
            contacts_list.append({
                'id': contact['id'],
                'name': contact['name'],
                'role': contact['role'] or '',
                'phone': contact['phone'] or '',
                'email': contact['email'] or '',
                'is_primary': bool(contact['is_primary'])
            })
        
        return jsonify({
            'contact_person': distributor['contact_person'],
            'phone': distributor['phone'],
            'email': distributor['email'],
            'contacts': contacts_list
        })
    else:
        return jsonify({'error': 'Distributor not found'}), 404
#################
from flask import Flask, render_template, request, redirect, url_for, flash, session
import sqlite3

# ... [rest of your existing imports and code] ...

@app.route('/edit_distributor/<int:distributor_id>', methods=['GET', 'POST'])
@login_required
def edit_distributor(distributor_id):
    """Edit distributor with full SRM support"""
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    if request.method == 'POST':
        name = request.form['name']
        address = request.form.get('address', '')
        contact_person = request.form.get('contact_person', '')
        phone = request.form.get('phone', '')
        email = request.form.get('email', '')
        status = request.form.get('status', 'active')
        category = request.form.get('category', '')
        website = request.form.get('website', '')
        notes = request.form.get('notes', '')

        try:
            # Update distributor with all SRM fields
            cursor.execute('''
                UPDATE distributors SET
                name = ?, address = ?, contact_person = ?, phone = ?, email = ?,
                status = ?, category = ?, website = ?, notes = ?
                WHERE id = ?
            ''', (name, address, contact_person, phone, email, status, category, website, notes, distributor_id))
            
            # Log activity in SRM activity log
            cursor.execute("""
                INSERT INTO srm_activity_log (entity_type, entity_id, activity_type, description, user_id)
                VALUES ('distributor', ?, 'distributor_updated', ?, ?)
            """, (distributor_id, f'Updated distributor details', session.get('user_id')))
            
            conn.commit()
            flash('Distributor updated successfully!', 'success')
            
            # Redirect to distributor detail page
            return redirect(url_for('distributor_detail', distributor_id=distributor_id))
            
        except Exception as e:
            flash(f'An error occurred: {e}', 'danger')
            conn.rollback()
        finally:
            conn.close()

        return redirect(url_for('show_distributors'))

    # Fetch the distributor for GET request
    cursor.execute('SELECT * FROM distributors WHERE id = ?', (distributor_id,))
    distributor = cursor.fetchone()
    conn.close()

    if distributor:
        return render_template('edit_distributor.html', distributor=distributor)
    else:
        flash('Distributor not found!', 'danger')
        return redirect(url_for('show_distributors'))


# DELETE DISTRIBUTOR ROUTE
@app.route('/delete_distributor/<int:distributor_id>', methods=['POST'])
@login_required
@role_required('General Manager', 'Technical Team Leader')
def delete_distributor(distributor_id):
    """Delete a distributor (Admin only) with SRM activity logging"""
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    try:
        # Get distributor details before deletion
        cursor.execute("SELECT name FROM distributors WHERE id = ?", (distributor_id,))
        distributor = cursor.fetchone()
        
        if not distributor:
            flash('Distributor not found.', 'danger')
            return redirect(url_for('show_distributors'))
        
        distributor_name = distributor['name']
        
        # Delete related records first (cascade delete)
        cursor.execute("DELETE FROM distributor_contacts WHERE distributor_id = ?", (distributor_id,))
        cursor.execute("DELETE FROM vendor_distributor WHERE distributor_id = ?", (distributor_id,))
        cursor.execute("DELETE FROM account_manager_assignments WHERE entity_type = 'distributor' AND entity_id = ?", (distributor_id,))
        cursor.execute("DELETE FROM performance_metrics WHERE entity_type = 'distributor' AND entity_id = ?", (distributor_id,))
        cursor.execute("DELETE FROM srm_documents WHERE entity_type = 'distributor' AND entity_id = ?", (distributor_id,))
        cursor.execute("DELETE FROM srm_activity_log WHERE entity_type = 'distributor' AND entity_id = ?", (distributor_id,))
        
        # Finally delete the distributor
        cursor.execute("DELETE FROM distributors WHERE id = ?", (distributor_id,))
        
        conn.commit()
        flash(f'Distributor "{distributor_name}" has been successfully deleted!', 'success')
        
    except Exception as e:
        conn.rollback()
        flash(f'Error deleting distributor: {e}', 'danger')
    finally:
        conn.close()
    
    return redirect(url_for('show_distributors'))


# =================================================================
# ================ ENHANCED SRM ROUTES (OPTION B) =================
# =================================================================

# Vendor Detail Page with All Information
@app.route('/vendor_detail/<int:vendor_id>')
@login_required
@permission_required('view_vendors')
def vendor_detail(vendor_id):
    """Comprehensive vendor detail page with contacts, performance, documents, POs"""
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    # Get vendor basic info
    cursor.execute("SELECT * FROM vendors WHERE id = ?", (vendor_id,))
    vendor = cursor.fetchone()
    
    if not vendor:
        flash('Vendor not found', 'danger')
        return redirect(url_for('show_vendors'))
    
    # Get contacts
    cursor.execute("""
        SELECT * FROM vendor_contacts 
        WHERE vendor_id = ? 
        ORDER BY is_primary DESC, name
    """, (vendor_id,))
    contacts = cursor.fetchall()
    
    # Get account manager
    cursor.execute("""
        SELECT ama.*, u.username, u.email as user_email
        FROM account_manager_assignments ama
        JOIN users u ON ama.user_id = u.id
        WHERE ama.entity_type = 'vendor' AND ama.entity_id = ?
        ORDER BY ama.assigned_date DESC
        LIMIT 1
    """, (vendor_id,))
    account_manager = cursor.fetchone()
    
    # Get performance metrics
    cursor.execute("""
        SELECT * FROM performance_metrics 
        WHERE entity_type = 'vendor' AND entity_id = ?
        ORDER BY updated_at DESC
        LIMIT 1
    """, (vendor_id,))
    performance = cursor.fetchone()
    
    # Get documents
    cursor.execute("""
        SELECT sd.*, u.username as uploaded_by_name
        FROM srm_documents sd
        LEFT JOIN users u ON sd.uploaded_by = u.id
        WHERE sd.entity_type = 'vendor' AND sd.entity_id = ?
        ORDER BY sd.uploaded_at DESC
    """, (vendor_id,))
    documents = cursor.fetchall()
    
    # Get purchase orders: both directly linked to vendor AND through distributors
    cursor.execute("""
        SELECT DISTINCT po.*, d.name as distributor_name,
               CASE 
                   WHEN CAST(po.vendor AS TEXT) = CAST(? AS TEXT) THEN 'Direct'
                   ELSE 'Via Distributor'
               END as po_source
        FROM purchase_orders po
        LEFT JOIN vendor_distributor vd ON CAST(po.distributor AS TEXT) = CAST(vd.distributor_id AS TEXT)
        LEFT JOIN distributors d ON CAST(po.distributor AS TEXT) = CAST(d.id AS TEXT)
        WHERE CAST(po.vendor AS TEXT) = CAST(? AS TEXT) OR vd.vendor_id = ?
        ORDER BY po.created_at DESC
    """, (vendor_id, vendor_id, vendor_id))
    purchase_orders = cursor.fetchall()
    
    # Calculate total spending: both direct and through distributors
    cursor.execute("""
        SELECT COALESCE(SUM(total_amount), 0) as total_spending
        FROM purchase_orders po
        LEFT JOIN vendor_distributor vd ON CAST(po.distributor AS TEXT) = CAST(vd.distributor_id AS TEXT)
        WHERE CAST(po.vendor AS TEXT) = CAST(? AS TEXT) OR vd.vendor_id = ?
    """, (vendor_id, vendor_id))
    spending_result = cursor.fetchone()
    total_spending = spending_result['total_spending']
    
    # Get recent activity
    cursor.execute("""
        SELECT sal.*, u.username
        FROM srm_activity_log sal
        LEFT JOIN users u ON sal.user_id = u.id
        WHERE sal.entity_type = 'vendor' AND sal.entity_id = ?
        ORDER BY sal.created_at DESC
        LIMIT 10
    """, (vendor_id,))
    activities = cursor.fetchall()
    
    # Get all users for account manager dropdown
    cursor.execute("SELECT id, username, role FROM users ORDER BY username")
    all_users = cursor.fetchall()
    
    # Get supplier quotations for this vendor with project names
    cursor.execute("""
        SELECT sq.*, p.project_name
        FROM supplier_quotations sq
        LEFT JOIN projects p ON sq.quote_ref = p.quote_ref
        WHERE sq.vendor_id = ?
        ORDER BY sq.uploaded_at DESC
    """, (vendor_id,))
    supplier_quotations = cursor.fetchall()
    
    conn.close()
    
    return render_template('vendor_detail.html',
                         vendor=vendor,
                         contacts=contacts,
                         account_manager=account_manager,
                         performance=performance,
                         documents=documents,
                         purchase_orders=purchase_orders,
                         total_spending=total_spending,
                         activities=activities,
                         all_users=all_users,
                         supplier_quotations=supplier_quotations)


# Distributor Detail Page
@app.route('/distributor_detail/<int:distributor_id>')
@login_required
@permission_required('view_distributors')
def distributor_detail(distributor_id):
    """Comprehensive distributor detail page"""
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    # Get distributor basic info
    cursor.execute("SELECT * FROM distributors WHERE id = ?", (distributor_id,))
    distributor = cursor.fetchone()
    
    if not distributor:
        flash('Distributor not found', 'danger')
        return redirect(url_for('show_distributors'))
    
    # Get contacts
    cursor.execute("""
        SELECT * FROM distributor_contacts 
        WHERE distributor_id = ? 
        ORDER BY is_primary DESC, name
    """, (distributor_id,))
    contacts = cursor.fetchall()
    
    # Get account manager
    cursor.execute("""
        SELECT ama.*, u.username, u.email as user_email
        FROM account_manager_assignments ama
        JOIN users u ON ama.user_id = u.id
        WHERE ama.entity_type = 'distributor' AND ama.entity_id = ?
        ORDER BY ama.assigned_date DESC
        LIMIT 1
    """, (distributor_id,))
    account_manager = cursor.fetchone()
    
    # Get performance metrics
    cursor.execute("""
        SELECT * FROM performance_metrics 
        WHERE entity_type = 'distributor' AND entity_id = ?
        ORDER BY updated_at DESC
        LIMIT 1
    """, (distributor_id,))
    performance = cursor.fetchone()
    
    # Get documents
    cursor.execute("""
        SELECT sd.*, u.username as uploaded_by_name
        FROM srm_documents sd
        LEFT JOIN users u ON sd.uploaded_by = u.id
        WHERE sd.entity_type = 'distributor' AND sd.entity_id = ?
        ORDER BY sd.uploaded_at DESC
    """, (distributor_id,))
    documents = cursor.fetchall()
    
    # Get associated vendors
    cursor.execute("""
        SELECT v.*
        FROM vendors v
        JOIN vendor_distributor vd ON v.id = vd.vendor_id
        WHERE vd.distributor_id = ?
    """, (distributor_id,))
    vendors = cursor.fetchall()
    
    # Get purchase orders for this distributor
    cursor.execute("""
        SELECT * FROM purchase_orders 
        WHERE CAST(distributor AS TEXT) = CAST(? AS TEXT)
        ORDER BY created_at DESC
    """, (distributor_id,))
    purchase_orders = cursor.fetchall()
    
    # Calculate total spending from POs
    cursor.execute("""
        SELECT COALESCE(SUM(total_amount), 0) as total_spending
        FROM purchase_orders 
        WHERE CAST(distributor AS TEXT) = CAST(? AS TEXT)
    """, (distributor_id,))
    total_spending = cursor.fetchone()['total_spending']
    
    # Get recent activity
    cursor.execute("""
        SELECT sal.*, u.username
        FROM srm_activity_log sal
        LEFT JOIN users u ON sal.user_id = u.id
        WHERE sal.entity_type = 'distributor' AND sal.entity_id = ?
        ORDER BY sal.created_at DESC
        LIMIT 10
    """, (distributor_id,))
    activities = cursor.fetchall()
    
    # Get all users for account manager dropdown
    cursor.execute("SELECT id, username, role FROM users ORDER BY username")
    all_users = cursor.fetchall()
    
    # Get supplier quotations for this distributor with project names
    cursor.execute("""
        SELECT sq.*, p.project_name
        FROM supplier_quotations sq
        LEFT JOIN projects p ON sq.quote_ref = p.quote_ref
        WHERE sq.distributor_id = ?
        ORDER BY sq.uploaded_at DESC
    """, (distributor_id,))
    supplier_quotations = cursor.fetchall()
    
    conn.close()
    
    return render_template('distributor_detail.html',
                         distributor=distributor,
                         contacts=contacts,
                         account_manager=account_manager,
                         performance=performance,
                         documents=documents,
                         vendors=vendors,
                         purchase_orders=purchase_orders,
                         total_spending=total_spending,
                         activities=activities,
                         all_users=all_users,
                         supplier_quotations=supplier_quotations)


# Add Contact (Vendor)
@app.route('/add_vendor_contact/<int:vendor_id>', methods=['POST'])
@login_required
def add_vendor_contact_new(vendor_id):
    """Add new contact to vendor"""
    conn = sqlite3.connect('ProjectStatus.db')
    cursor = conn.cursor()
    
    name = request.form.get('name')
    role = request.form.get('role')
    phone = request.form.get('phone')
    email = request.form.get('email')
    notes = request.form.get('notes')
    is_primary = 1 if request.form.get('is_primary') else 0
    
    # If setting as primary, unset others
    if is_primary:
        cursor.execute("UPDATE vendor_contacts SET is_primary = 0 WHERE vendor_id = ?", (vendor_id,))
    
    cursor.execute("""
        INSERT INTO vendor_contacts (vendor_id, name, role, phone, email, is_primary, notes)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (vendor_id, name, role, phone, email, is_primary, notes))
    
    # Log activity
    cursor.execute("""
        INSERT INTO srm_activity_log (entity_type, entity_id, activity_type, description, user_id)
        VALUES ('vendor', ?, 'contact_added', ?, ?)
    """, (vendor_id, f"Added contact: {name} ({role})", session.get('user_id')))
    
    conn.commit()
    conn.close()
    
    flash(f'Contact {name} added successfully!', 'success')
    return redirect(url_for('vendor_detail', vendor_id=vendor_id))


# Add Contact (Distributor)
@app.route('/add_distributor_contact/<int:distributor_id>', methods=['POST'])
@login_required
def add_distributor_contact_new(distributor_id):
    """Add new contact to distributor"""
    conn = sqlite3.connect('ProjectStatus.db')
    cursor = conn.cursor()
    
    name = request.form.get('name')
    role = request.form.get('role')
    phone = request.form.get('phone')
    email = request.form.get('email')
    notes = request.form.get('notes')
    is_primary = 1 if request.form.get('is_primary') else 0
    
    # If setting as primary, unset others
    if is_primary:
        cursor.execute("UPDATE distributor_contacts SET is_primary = 0 WHERE distributor_id = ?", (distributor_id,))
    
    cursor.execute("""
        INSERT INTO distributor_contacts (distributor_id, name, role, phone, email, is_primary, notes)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (distributor_id, name, role, phone, email, is_primary, notes))
    
    # Log activity
    cursor.execute("""
        INSERT INTO srm_activity_log (entity_type, entity_id, activity_type, description, user_id)
        VALUES ('distributor', ?, 'contact_added', ?, ?)
    """, (distributor_id, f"Added contact: {name} ({role})", session.get('user_id')))
    
    conn.commit()
    conn.close()
    
    flash(f'Contact {name} added successfully!', 'success')
    return redirect(url_for('distributor_detail', distributor_id=distributor_id))


# Delete Contact
@app.route('/delete_contact/<entity_type>/<int:contact_id>', methods=['POST'])
@login_required
def delete_contact(entity_type, contact_id):
    """Delete a contact"""
    conn = sqlite3.connect('ProjectStatus.db')
    cursor = conn.cursor()
    
    if entity_type == 'vendor':
        cursor.execute("SELECT vendor_id, name FROM vendor_contacts WHERE id = ?", (contact_id,))
        result = cursor.fetchone()
        if result:
            entity_id, name = result
            cursor.execute("DELETE FROM vendor_contacts WHERE id = ?", (contact_id,))
            cursor.execute("""
                INSERT INTO srm_activity_log (entity_type, entity_id, activity_type, description, user_id)
                VALUES ('vendor', ?, 'contact_deleted', ?, ?)
            """, (entity_id, f"Deleted contact: {name}", session.get('user_id')))
    else:
        cursor.execute("SELECT distributor_id, name FROM distributor_contacts WHERE id = ?", (contact_id,))
        result = cursor.fetchone()
        if result:
            entity_id, name = result
            cursor.execute("DELETE FROM distributor_contacts WHERE id = ?", (contact_id,))
            cursor.execute("""
                INSERT INTO srm_activity_log (entity_type, entity_id, activity_type, description, user_id)
                VALUES ('distributor', ?, 'contact_deleted', ?, ?)
            """, (entity_id, f"Deleted contact: {name}", session.get('user_id')))
    
    conn.commit()
    conn.close()
    
    flash('Contact deleted successfully!', 'success')
    return redirect(request.referrer or url_for('show_vendors'))


# Assign Account Manager
@app.route('/assign_account_manager', methods=['POST'])
@login_required
def assign_account_manager():
    """Assign account manager to vendor or distributor"""
    conn = sqlite3.connect('ProjectStatus.db')
    cursor = conn.cursor()
    
    entity_type = request.form.get('entity_type')
    entity_id = int(request.form.get('entity_id'))
    user_id = int(request.form.get('user_id'))
    notes = request.form.get('notes', '')
    
    # Check if already assigned
    cursor.execute("""
        DELETE FROM account_manager_assignments 
        WHERE entity_type = ? AND entity_id = ?
    """, (entity_type, entity_id))
    
    # Create new assignment
    cursor.execute("""
        INSERT INTO account_manager_assignments (entity_type, entity_id, user_id, notes)
        VALUES (?, ?, ?, ?)
    """, (entity_type, entity_id, user_id, notes))
    
    # Get username for log
    cursor.execute("SELECT username FROM users WHERE id = ?", (user_id,))
    username = cursor.fetchone()[0]
    
    # Log activity
    cursor.execute("""
        INSERT INTO srm_activity_log (entity_type, entity_id, activity_type, description, user_id)
        VALUES (?, ?, 'account_manager_assigned', ?, ?)
    """, (entity_type, entity_id, f"Assigned account manager: {username}", session.get('user_id')))
    
    conn.commit()
    conn.close()
    
    flash(f'Account manager assigned successfully!', 'success')
    return redirect(request.referrer)


# Update Performance
@app.route('/update_performance', methods=['POST'])
@login_required
def update_performance():
    """Update performance metrics"""
    conn = sqlite3.connect('ProjectStatus.db')
    cursor = conn.cursor()
    
    entity_type = request.form.get('entity_type')
    entity_id = int(request.form.get('entity_id'))
    delivery_score = int(request.form.get('delivery_score', 0))
    quality_score = int(request.form.get('quality_score', 0))
    response_score = int(request.form.get('response_score', 0))
    notes = request.form.get('notes', '')
    
    # Calculate overall rating
    overall_rating = (delivery_score + quality_score + response_score) / 3
    
    # Check if performance record exists
    cursor.execute("""
        SELECT id FROM performance_metrics 
        WHERE entity_type = ? AND entity_id = ?
    """, (entity_type, entity_id))
    existing = cursor.fetchone()
    
    if existing:
        cursor.execute("""
            UPDATE performance_metrics 
            SET delivery_score = ?, quality_score = ?, response_score = ?, 
                overall_rating = ?, notes = ?, updated_at = CURRENT_TIMESTAMP,
                updated_by = ?
            WHERE entity_type = ? AND entity_id = ?
        """, (delivery_score, quality_score, response_score, overall_rating, notes,
              session.get('user_id'), entity_type, entity_id))
    else:
        cursor.execute("""
            INSERT INTO performance_metrics 
            (entity_type, entity_id, delivery_score, quality_score, response_score, 
             overall_rating, notes, updated_by)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (entity_type, entity_id, delivery_score, quality_score, response_score,
              overall_rating, notes, session.get('user_id')))
    
    # Log activity
    cursor.execute("""
        INSERT INTO srm_activity_log (entity_type, entity_id, activity_type, description, user_id)
        VALUES (?, ?, 'performance_updated', ?, ?)
    """, (entity_type, entity_id, f"Updated performance: {overall_rating:.1f}/10", session.get('user_id')))
    
    conn.commit()
    conn.close()
    
    flash('Performance metrics updated successfully!', 'success')
    return redirect(request.referrer)


# Upload Document
@app.route('/upload_srm_document', methods=['POST'])
@login_required
def upload_srm_document():
    """Upload document for vendor/distributor"""
    from werkzeug.utils import secure_filename
    
    entity_type = request.form.get('entity_type')
    entity_id = int(request.form.get('entity_id'))
    document_type = request.form.get('document_type')
    expiry_date = request.form.get('expiry_date')
    notes = request.form.get('notes', '')
    
    if 'document_file' not in request.files:
        flash('No file uploaded', 'danger')
        return redirect(request.referrer)
    
    file = request.files['document_file']
    if file.filename == '':
        flash('No file selected', 'danger')
        return redirect(request.referrer)
    
    # Create uploads directory if not exists
    upload_dir = 'static/srm_documents'
    os.makedirs(upload_dir, exist_ok=True)
    
    # Save file
    filename = secure_filename(file.filename)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"{entity_type}_{entity_id}_{timestamp}_{filename}"
    file_path = os.path.join(upload_dir, filename)
    file.save(file_path)
    file_size = os.path.getsize(file_path)
    
    # Save to database
    conn = sqlite3.connect('ProjectStatus.db')
    cursor = conn.cursor()
    
    cursor.execute("""
        INSERT INTO srm_documents 
        (entity_type, entity_id, document_name, document_type, file_path, 
         file_size, expiry_date, uploaded_by, notes)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (entity_type, entity_id, file.filename, document_type, file_path,
          file_size, expiry_date, session.get('user_id'), notes))
    
    # Log activity
    cursor.execute("""
        INSERT INTO srm_activity_log (entity_type, entity_id, activity_type, description, user_id)
        VALUES (?, ?, 'document_uploaded', ?, ?)
    """, (entity_type, entity_id, f"Uploaded document: {file.filename}", session.get('user_id')))
    
    conn.commit()
    conn.close()
    
    flash('Document uploaded successfully!', 'success')
    return redirect(request.referrer)


# Delete Document
@app.route('/delete_srm_document/<int:document_id>', methods=['POST'])
@login_required
def delete_srm_document(document_id):
    """Delete a document"""
    conn = sqlite3.connect('ProjectStatus.db')
    cursor = conn.cursor()
    
    cursor.execute("SELECT entity_type, entity_id, file_path, document_name FROM srm_documents WHERE id = ?", (document_id,))
    result = cursor.fetchone()
    
    if result:
        entity_type, entity_id, file_path, doc_name = result
        
        # Delete file
        if os.path.exists(file_path):
            os.remove(file_path)
        
        # Delete from database
        cursor.execute("DELETE FROM srm_documents WHERE id = ?", (document_id,))
        
        # Log activity
        cursor.execute("""
            INSERT INTO srm_activity_log (entity_type, entity_id, activity_type, description, user_id)
            VALUES (?, ?, 'document_deleted', ?, ?)
        """, (entity_type, entity_id, f"Deleted document: {doc_name}", session.get('user_id')))
        
        conn.commit()
        flash('Document deleted successfully!', 'success')
    else:
        flash('Document not found', 'danger')
    
    conn.close()
    return redirect(request.referrer)


# SRM Analytics Dashboard
@app.route('/srm_analytics')
@login_required
@role_required('General Manager', 'Technical Team Leader', 'Presale Engineer')
def srm_analytics():
    """SRM Analytics Dashboard with charts and metrics"""
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    # Get total counts
    cursor.execute("SELECT COUNT(*) as count FROM vendors WHERE status = 'active'")
    active_vendors = cursor.fetchone()['count']
    
    cursor.execute("SELECT COUNT(*) as count FROM distributors WHERE status = 'active'")
    active_distributors = cursor.fetchone()['count']
    
    # Get vendors with account managers
    cursor.execute("""
        SELECT COUNT(DISTINCT entity_id) as count
        FROM account_manager_assignments 
        WHERE entity_type = 'vendor'
    """)
    vendors_with_managers = cursor.fetchone()['count']
    
    # Get total spending
    cursor.execute("SELECT SUM(CAST(total_price AS REAL)) as total FROM purchase_orders")
    total_spending_result = cursor.fetchone()
    total_spending = total_spending_result['total'] if total_spending_result['total'] else 0
    
    # Top performing vendors
    cursor.execute("""
        SELECT v.name, pm.overall_rating, pm.delivery_score, pm.quality_score
        FROM vendors v
        JOIN performance_metrics pm ON v.id = pm.entity_id AND pm.entity_type = 'vendor'
        ORDER BY pm.overall_rating DESC
        LIMIT 10
    """)
    top_vendors = cursor.fetchall()
    
    # Spending by vendor
    cursor.execute("""
        SELECT vendor_name, SUM(CAST(total_price AS REAL)) as total_spent
        FROM purchase_orders
        GROUP BY vendor_name
        ORDER BY total_spent DESC
        LIMIT 10
    """)
    spending_by_vendor = cursor.fetchall()
    
    # Documents expiring soon (next 30 days)
    cursor.execute("""
        SELECT sd.*, v.name as entity_name
        FROM srm_documents sd
        LEFT JOIN vendors v ON sd.entity_id = v.id AND sd.entity_type = 'vendor'
        LEFT JOIN distributors d ON sd.entity_id = d.id AND sd.entity_type = 'distributor'
        WHERE sd.expiry_date IS NOT NULL 
        AND sd.expiry_date BETWEEN date('now') AND date('now', '+30 days')
        AND sd.status = 'active'
        ORDER BY sd.expiry_date
    """)
    expiring_documents = cursor.fetchall()
    
    # Recent activities
    cursor.execute("""
        SELECT sal.*, u.username, 
               CASE 
                   WHEN sal.entity_type = 'vendor' THEN v.name
                   WHEN sal.entity_type = 'distributor' THEN d.name
               END as entity_name
        FROM srm_activity_log sal
        LEFT JOIN users u ON sal.user_id = u.id
        LEFT JOIN vendors v ON sal.entity_id = v.id AND sal.entity_type = 'vendor'
        LEFT JOIN distributors d ON sal.entity_id = d.id AND sal.entity_type = 'distributor'
        ORDER BY sal.created_at DESC
        LIMIT 20
    """)
    recent_activities = cursor.fetchall()
    
    conn.close()
    
    return render_template('srm_analytics.html',
                         active_vendors=active_vendors,
                         active_distributors=active_distributors,
                         vendors_with_managers=vendors_with_managers,
                         total_spending=total_spending,
                         top_vendors=top_vendors,
                         spending_by_vendor=spending_by_vendor,
                         expiring_documents=expiring_documents,
                         recent_activities=recent_activities)


# Secure Document Download Route
@app.route('/download_srm_document/<int:document_id>')
@login_required
def download_srm_document(document_id):
    """Secure document download with authorization"""
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    # Get document info
    cursor.execute("SELECT * FROM srm_documents WHERE id = ?", (document_id,))
    document = cursor.fetchone()
    conn.close()
    
    if not document:
        flash('Document not found', 'danger')
        return redirect(url_for('index'))
    
    # Check if file exists
    if not os.path.exists(document['file_path']):
        flash('Document file not found on server', 'danger')
        return redirect(request.referrer or url_for('index'))
    
    # Send file
    return send_file(document['file_path'], 
                     as_attachment=True,
                     download_name=document['document_name'])


# =================================================================
# =================== END OF ENHANCED SRM ROUTES ==================
# =================================================================

# ... [rest of your existing routes and code] ...
###########
@app.route('/fetch_cctv_products', methods=['GET'])
#@role_required('editor')  # Adjust if necessary
def fetch_cctv_products():
    # Redirect API calls to new system
    return jsonify({
        'message': 'This API has been deprecated. Please use /cctv_smart_selector for product browsing.',
        'redirect': '/cctv_smart_selector'
    })

##########
import base64
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify
import sqlite3


@app.route('/cctv_products', methods=['GET', 'POST'])
@login_required
def cctv_products():
    """CCTV Products Page with Advanced Filtering"""
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    
    # Get filter parameters from request
    vendor_name_filter = request.args.get('vendor_name')
    model_number_filter = request.args.get('model_number')
    camera_type_filter = request.args.get('camera_type')
    image_sensor_filter = request.args.get('image_sensor')
    max_resolution_filter = request.args.get('max_resolution')
    min_illumination_filter = request.args.get('min_illumination')
    wdr_filter = request.args.get('wdr')
    built_in_mic_filter = request.args.get('built_in_Mic')
    supplement_light_range_filter = request.args.get('supplement_light_range')
    lens_type_filter = request.args.get('lens_type')
    focal_length_filter = request.args.get('focal_length')
    
    # Build query with filters
    query = """
        SELECT id, vendor_name, model_number, camera_image, price, datasheet_url,
               camera_type, lens_type, max_resolution, image_sensor
        FROM cctv_products
        WHERE 1=1
    """
    params = []
    
    if vendor_name_filter:
        query += " AND vendor_name = ?"
        params.append(vendor_name_filter)
    if model_number_filter:
        query += " AND model_number = ?"
        params.append(model_number_filter)
    if camera_type_filter:
        query += " AND camera_type = ?"
        params.append(camera_type_filter)
    if image_sensor_filter:
        query += " AND image_sensor = ?"
        params.append(image_sensor_filter)
    if max_resolution_filter:
        query += " AND max_resolution = ?"
        params.append(max_resolution_filter)
    if min_illumination_filter:
        query += " AND min_illumination = ?"
        params.append(min_illumination_filter)
    if wdr_filter:
        query += " AND wdr = ?"
        params.append(wdr_filter)
    if built_in_mic_filter:
        query += " AND built_in_mic = ?"
        params.append(built_in_mic_filter)
    if supplement_light_range_filter:
        query += " AND supplement_light_range = ?"
        params.append(supplement_light_range_filter)
    if lens_type_filter:
        query += " AND lens_type = ?"
        params.append(lens_type_filter)
    if focal_length_filter:
        query += " AND focal_length = ?"
        params.append(focal_length_filter)
    
    query += " ORDER BY vendor_name, model_number"
    
    c.execute(query, params)
    products_raw = c.fetchall()
    
    # Convert images to base64
    products = []
    for product in products_raw:
        product_list = list(product)
        if product_list[3]:  # If camera_image exists (index 3)
            # Check if it's already a string (base64 encoded) or bytes
            if isinstance(product_list[3], bytes):
                product_list[3] = base64.b64encode(product_list[3]).decode('utf-8')
            # If it's already a string, leave it as is
        products.append(product_list)
    
    # Get distinct values for filter dropdowns
    c.execute('SELECT DISTINCT vendor_name FROM cctv_products WHERE vendor_name IS NOT NULL ORDER BY vendor_name')
    vendors = [row[0] for row in c.fetchall()]
    
    c.execute('SELECT DISTINCT model_number FROM cctv_products WHERE model_number IS NOT NULL ORDER BY model_number')
    models = [row[0] for row in c.fetchall()]
    
    c.execute('SELECT DISTINCT camera_type FROM cctv_products WHERE camera_type IS NOT NULL ORDER BY camera_type')
    camera_types = [row[0] for row in c.fetchall()]
    
    c.execute('SELECT DISTINCT image_sensor FROM cctv_products WHERE image_sensor IS NOT NULL ORDER BY image_sensor')
    image_sensors = [row[0] for row in c.fetchall()]
    
    c.execute('SELECT DISTINCT max_resolution FROM cctv_products WHERE max_resolution IS NOT NULL ORDER BY max_resolution')
    max_resolutions = [row[0] for row in c.fetchall()]
    
    c.execute('SELECT DISTINCT min_illumination FROM cctv_products WHERE min_illumination IS NOT NULL ORDER BY min_illumination')
    min_illuminations = [row[0] for row in c.fetchall()]
    
    c.execute('SELECT DISTINCT wdr FROM cctv_products WHERE wdr IS NOT NULL ORDER BY wdr')
    wdr_options = [row[0] for row in c.fetchall()]
    
    c.execute('SELECT DISTINCT built_in_mic FROM cctv_products WHERE built_in_mic IS NOT NULL ORDER BY built_in_mic')
    built_in_mic_options = [row[0] for row in c.fetchall()]
    
    c.execute('SELECT DISTINCT supplement_light_range FROM cctv_products WHERE supplement_light_range IS NOT NULL ORDER BY supplement_light_range')
    supplement_light_ranges = [row[0] for row in c.fetchall()]
    
    c.execute('SELECT DISTINCT lens_type FROM cctv_products WHERE lens_type IS NOT NULL ORDER BY lens_type')
    lens_types = [row[0] for row in c.fetchall()]
    
    c.execute('SELECT DISTINCT focal_length FROM cctv_products WHERE focal_length IS NOT NULL ORDER BY focal_length')
    focal_lengths = [row[0] for row in c.fetchall()]
    
    conn.close()
    
    return render_template('cctv_products.html',
                         products=products,
                         vendors=vendors,
                         models=models,
                         camera_types=camera_types,
                         image_sensors=image_sensors,
                         max_resolutions=max_resolutions,
                         min_illuminations=min_illuminations,
                         wdr_options=wdr_options,
                         built_in_mic_options=built_in_mic_options,
                         supplement_light_ranges=supplement_light_ranges,
                         lens_types=lens_types,
                         focal_lengths=focal_lengths)
########
@app.route('/register_product', methods=['GET', 'POST'])
#@role_required('editor')  # Adjust if necessary
def register_product():
    if request.method == 'POST':
        # Check if the request contains the file part
        if 'camera_image' not in request.files:
            flash('No file part', 'danger')
            return redirect(request.url)

        camera_image = request.files['camera_image']

        # Check if the user has selected a file
        if camera_image.filename == '':
            flash('No selected file', 'danger')
            return redirect(request.url)

        # Read the image data
        camera_image_data = camera_image.read()

        # Get data from the form
        vendor_name = request.form.get('vendor_name')
        model_number = request.form.get('model_number')
        image_sensor = request.form.get('image_sensor')
        max_resolution = request.form.get('max_resolution')
        min_illumination = request.form.get('min_illumination')
        lens_type = request.form.get('lens_type')
        focal_length = request.form.get('focal_length')
        iris_type = request.form.get('iris_type')
        supplement_light_type = request.form.get('supplement_light_type')
        supplement_light_range = request.form.get('supplement_light_range')
        built_in_mic = request.form.get('built_in_mic')
        wdr = request.form.get('wdr')
        price = request.form.get('price')  # New Price field
        camera_type = request.form.get('camera_type')  # New Camera Type field
        datasheet_url = request.form.get('datasheet_url')  # New Datasheet URL field

        # Insert the data into the database
        conn = sqlite3.connect('ProjectStatus.db')
        c = conn.cursor()
        c.execute('''
            INSERT INTO cctv_products (
                vendor_name, model_number, image_sensor, max_resolution, min_illumination,
                lens_type, focal_length, iris_type, supplement_light_type,
                supplement_light_range, built_in_mic, wdr, camera_image, price, camera_type, datasheet_url
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            vendor_name, model_number, image_sensor, max_resolution, min_illumination,
            lens_type, focal_length, iris_type, supplement_light_type,
            supplement_light_range, built_in_mic, wdr, camera_image_data,
            price, camera_type, datasheet_url
        ))

        conn.commit()
        conn.close()

        flash('Product registered successfully!', 'success')
        return redirect(url_for('cctv_products'))

    return render_template('register_product.html')


@app.route('/import_products_excel', methods=['POST'])
@login_required
def import_products_excel():
    """Bulk import CCTV products from Excel file"""
    import pandas as pd
    
    if 'excel_file' not in request.files:
        flash('No file uploaded', 'danger')
        return redirect(url_for('cctv_products'))
    
    excel_file = request.files['excel_file']
    
    if excel_file.filename == '':
        flash('No file selected', 'danger')
        return redirect(url_for('cctv_products'))
    
    if not excel_file.filename.endswith(('.xlsx', '.xls')):
        flash('Please upload an Excel file (.xlsx or .xls)', 'danger')
        return redirect(url_for('cctv_products'))
    
    try:
        # Read Excel file
        df = pd.read_excel(excel_file)
        
        conn = sqlite3.connect('ProjectStatus.db')
        cursor = conn.cursor()
        
        imported_count = 0
        skipped_count = 0
        
        for idx, row in df.iterrows():
            try:
                vendor_name = row.get('vendor_name', '')
                model_number = row.get('model_number', '')
                camera_type = row.get('camera_type', '')
                image_sensor = row.get('image_sensor', '')
                max_resolution = row.get('max_resolution', '')
                
                # Combine color and B/W illumination if separate columns exist
                min_illum_color = row.get('min_illumination_color', '')
                min_illum_bw = row.get('min_illumination_bw', '')
                min_illumination = row.get('min_illumination', '')
                
                if not min_illumination and (min_illum_color or min_illum_bw):
                    min_illumination = f"{min_illum_color} / {min_illum_bw}" if min_illum_color and min_illum_bw else (min_illum_color or min_illum_bw)
                
                lens_type = row.get('lens_type', '')
                focal_length = row.get('focal_length', '')
                iris_type = row.get('iris_type', '')
                supplement_light_type = row.get('illumination_type', row.get('supplement_light_type', ''))
                supplement_light_range = row.get('illumination_range', row.get('supplement_light_range', ''))
                built_in_mic = row.get('built_in_mic', '')
                wdr = row.get('wdr', '')
                price = row.get('price') if pd.notna(row.get('price')) else None
                datasheet_url = row.get('Data sheet', row.get('datasheet_url', ''))
                
                # Check if product already exists
                cursor.execute("SELECT id FROM cctv_products WHERE model_number = ?", (model_number,))
                existing = cursor.fetchone()
                
                if existing:
                    skipped_count += 1
                    continue
                
                # Insert into database (no image for now)
                cursor.execute('''
                    INSERT INTO cctv_products (
                        vendor_name, model_number, image_sensor, max_resolution, min_illumination,
                        lens_type, focal_length, iris_type, supplement_light_type,
                        supplement_light_range, built_in_mic, wdr, camera_image, price, camera_type, datasheet_url
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    vendor_name, model_number, image_sensor, max_resolution, min_illumination,
                    lens_type, focal_length, iris_type, supplement_light_type,
                    supplement_light_range, built_in_mic, wdr, None,
                    price, camera_type, datasheet_url
                ))
                
                imported_count += 1
                
            except Exception as e:
                print(f"Error importing row {idx}: {e}")
                continue
        
        conn.commit()
        conn.close()
        
        flash(f'Successfully imported {imported_count} products! ({skipped_count} duplicates skipped)', 'success')
        
    except Exception as e:
        flash(f'Error processing Excel file: {str(e)}', 'danger')
    
    return redirect(url_for('cctv_products'))


@app.route('/view_product_details/<int:product_id>')
@login_required
def view_product_details(product_id):
    """View detailed specifications of a CCTV product"""
    conn = sqlite3.connect('ProjectStatus.db')
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT id, vendor_name, model_number, camera_type, image_sensor, max_resolution,
               min_illumination, lens_type, focal_length, iris_type, supplement_light_type,
               supplement_light_range, built_in_mic, wdr, camera_image, price, datasheet_url
        FROM cctv_products
        WHERE id = ?
    ''', (product_id,))
    
    product = cursor.fetchone()
    conn.close()
    
    if not product:
        flash('Product not found', 'danger')
        return redirect(url_for('cctv_products'))
    
    # Convert image to base64 if exists
    camera_image_encoded = None
    if product[14]:
        if isinstance(product[14], bytes):
            camera_image_encoded = base64.b64encode(product[14]).decode('utf-8')
        else:
            camera_image_encoded = product[14]  # Already encoded
    
    product_data = {
        'id': product[0],
        'vendor_name': product[1],
        'model_number': product[2],
        'camera_type': product[3],
        'image_sensor': product[4],
        'max_resolution': product[5],
        'min_illumination': product[6],
        'lens_type': product[7],
        'focal_length': product[8],
        'iris_type': product[9],
        'supplement_light_type': product[10],
        'supplement_light_range': product[11],
        'built_in_mic': product[12],
        'wdr': product[13],
        'camera_image': camera_image_encoded,
        'price': product[15],
        'datasheet_url': product[16]
    }
    
    return render_template('view_product_details.html', product=product_data)


@app.route('/edit_product/<int:product_id>', methods=['GET', 'POST'])
@login_required
def edit_product(product_id):
    """Edit CCTV product specifications"""
    conn = sqlite3.connect('ProjectStatus.db')
    cursor = conn.cursor()
    
    if request.method == 'POST':
        # Get form data
        vendor_name = request.form.get('vendor_name')
        model_number = request.form.get('model_number')
        image_sensor = request.form.get('image_sensor')
        max_resolution = request.form.get('max_resolution')
        min_illumination = request.form.get('min_illumination')
        lens_type = request.form.get('lens_type')
        focal_length = request.form.get('focal_length')
        iris_type = request.form.get('iris_type')
        supplement_light_type = request.form.get('supplement_light_type')
        supplement_light_range = request.form.get('supplement_light_range')
        built_in_mic = request.form.get('built_in_mic')
        wdr = request.form.get('wdr')
        price = request.form.get('price')
        camera_type = request.form.get('camera_type')
        datasheet_url = request.form.get('datasheet_url')
        
        # Check if new image uploaded
        camera_image_data = None
        if 'camera_image' in request.files and request.files['camera_image'].filename != '':
            camera_image = request.files['camera_image']
            camera_image_data = camera_image.read()
        
        # Update database
        if camera_image_data:
            cursor.execute('''
                UPDATE cctv_products
                SET vendor_name=?, model_number=?, image_sensor=?, max_resolution=?, min_illumination=?,
                    lens_type=?, focal_length=?, iris_type=?, supplement_light_type=?,
                    supplement_light_range=?, built_in_mic=?, wdr=?, camera_image=?, price=?, camera_type=?, datasheet_url=?
                WHERE id=?
            ''', (vendor_name, model_number, image_sensor, max_resolution, min_illumination,
                  lens_type, focal_length, iris_type, supplement_light_type,
                  supplement_light_range, built_in_mic, wdr, camera_image_data, price, camera_type, datasheet_url, product_id))
        else:
            cursor.execute('''
                UPDATE cctv_products
                SET vendor_name=?, model_number=?, image_sensor=?, max_resolution=?, min_illumination=?,
                    lens_type=?, focal_length=?, iris_type=?, supplement_light_type=?,
                    supplement_light_range=?, built_in_mic=?, wdr=?, price=?, camera_type=?, datasheet_url=?
                WHERE id=?
            ''', (vendor_name, model_number, image_sensor, max_resolution, min_illumination,
                  lens_type, focal_length, iris_type, supplement_light_type,
                  supplement_light_range, built_in_mic, wdr, price, camera_type, datasheet_url, product_id))
        
        conn.commit()
        conn.close()
        
        # Send notification to admins
        try:
            actor_id = session.get('user_id')
            actor_name = session.get('username', 'Unknown User')
            
            # Get admin recipients
            recipients = notification_service.get_admin_recipients()
            
            # Create notification for product edit
            if recipients:
                notification_service.notify_activity(
                    event_code='product.edited',
                    recipient_ids=recipients,
                    actor_id=actor_id,
                    context={
                        'actor_name': actor_name,
                        'product_vendor': vendor_name,
                        'product_model': model_number,
                        'product_id': product_id
                    },
                    url=url_for('view_product_details', product_id=product_id)
                )
        except Exception as e:
            print(f"Notification error: {e}")
        
        flash('Product updated successfully!', 'success')
        return redirect(url_for('cctv_products'))
    
    # GET request - fetch product data
    cursor.execute('''
        SELECT id, vendor_name, model_number, camera_type, image_sensor, max_resolution,
               min_illumination, lens_type, focal_length, iris_type, supplement_light_type,
               supplement_light_range, built_in_mic, wdr, camera_image, price, datasheet_url
        FROM cctv_products
        WHERE id = ?
    ''', (product_id,))
    
    product = cursor.fetchone()
    conn.close()
    
    if not product:
        flash('Product not found', 'danger')
        return redirect(url_for('cctv_products'))
    
    # Convert image to base64 if exists
    camera_image_encoded = None
    if product[14]:
        if isinstance(product[14], bytes):
            camera_image_encoded = base64.b64encode(product[14]).decode('utf-8')
        else:
            camera_image_encoded = product[14]  # Already encoded
    
    product_data = {
        'id': product[0],
        'vendor_name': product[1],
        'model_number': product[2],
        'camera_type': product[3],
        'image_sensor': product[4],
        'max_resolution': product[5],
        'min_illumination': product[6],
        'lens_type': product[7],
        'focal_length': product[8],
        'iris_type': product[9],
        'supplement_light_type': product[10],
        'supplement_light_range': product[11],
        'built_in_mic': product[12],
        'wdr': product[13],
        'camera_image': camera_image_encoded,
        'price': product[15],
        'datasheet_url': product[16]
    }
    
    return render_template('edit_product.html', product=product_data)


@app.route('/delete_product/<int:product_id>', methods=['POST'])
@login_required
@role_required('General Manager', 'Technical Team Leader')
def delete_product(product_id):
    """Delete a CCTV product (Admin only)"""
    conn = sqlite3.connect('ProjectStatus.db')
    cursor = conn.cursor()
    
    # First, get the product details for confirmation message
    cursor.execute('SELECT vendor_name, model_number FROM cctv_products WHERE id = ?', (product_id,))
    product = cursor.fetchone()
    
    if not product:
        flash('Product not found', 'danger')
        conn.close()
        return redirect(url_for('cctv_products'))
    
    # Store product details before deletion
    vendor_name = product[0]
    model_number = product[1]
    
    # Delete the product
    cursor.execute('DELETE FROM cctv_products WHERE id = ?', (product_id,))
    conn.commit()
    conn.close()
    
    # Send notification to admins
    try:
        actor_id = session.get('user_id')
        actor_name = session.get('username', 'Unknown User')
        
        # Get admin recipients
        recipients = notification_service.get_admin_recipients()
        
        # Create notification for product deletion
        if recipients:
            notification_service.notify_activity(
                event_code='product.deleted',
                recipient_ids=recipients,
                actor_id=actor_id,
                context={
                    'actor_name': actor_name,
                    'product_vendor': vendor_name,
                    'product_model': model_number,
                    'product_id': product_id
                },
                url=url_for('cctv_products')
            )
    except Exception as e:
        print(f"Notification error: {e}")
    
    flash(f'Product "{vendor_name} {model_number}" has been successfully deleted!', 'success')
    return redirect(url_for('cctv_products'))


@app.route('/compare_products', methods=['GET'])
@login_required
def compare_products():
    """Compare selected CCTV products side-by-side"""
    product_ids = request.args.get('ids', '')
    
    if not product_ids:
        flash('No products selected for comparison', 'warning')
        return redirect(url_for('cctv_products'))
    
    # Parse product IDs from comma-separated string
    try:
        ids = [int(id.strip()) for id in product_ids.split(',')]
    except ValueError:
        flash('Invalid product IDs', 'danger')
        return redirect(url_for('cctv_products'))
    
    if len(ids) < 2:
        flash('Please select at least 2 products to compare', 'warning')
        return redirect(url_for('cctv_products'))
    
    if len(ids) > 4:
        flash('You can compare up to 4 products at a time', 'warning')
        return redirect(url_for('cctv_products'))
    
    # Fetch products from database
    conn = sqlite3.connect('ProjectStatus.db')
    cursor = conn.cursor()
    
    placeholders = ','.join(['?' for _ in ids])
    query = f"""
        SELECT id, vendor_name, model_number, camera_type, image_sensor, max_resolution, 
               min_illumination, lens_type, focal_length, iris_type, supplement_light_type,
               supplement_light_range, built_in_mic, wdr, camera_image, price, datasheet_url
        FROM cctv_products
        WHERE id IN ({placeholders})
    """
    
    cursor.execute(query, ids)
    products_raw = cursor.fetchall()
    conn.close()
    
    if not products_raw:
        flash('No products found for comparison', 'warning')
        return redirect(url_for('cctv_products'))
    
    # Convert images to base64
    products = []
    for product in products_raw:
        product_list = list(product)
        if product_list[14]:  # camera_image
            # Handle both bytes and pre-encoded strings
            image_data = product_list[14]
            if isinstance(image_data, bytes):
                product_list[14] = base64.b64encode(image_data).decode('utf-8')
        products.append(product_list)
    
    return render_template('compare_products.html', products=products)


#####33

@app.route('/fire_alarm_products', methods=['GET'])
#@role_required('editor')  # Adjust role as necessary
def fire_alarm_products():
    return render_template('fire_alarm_products.html')

###########33
@app.route('/view_detectors_products', methods=['GET'])
#@role_required('editor')
def view_detectors_products():
    # Connect to the database
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()

    # Initialize filter parameters
    vendor_name_filter = request.args.get('vendor_name', default=None)
    model_number_filter = request.args.get('model_number', default=None)
    detector_type_filter = request.args.get('detector_type', default=None)

    # Base query
    query = "SELECT vendor_name, model_number, detector_type, price, detector_image, max_area_coverage FROM detectors_products WHERE 1=1"
    params = []  # List to hold parameters for filtering

    # Append conditions for filters if they are provided
    if vendor_name_filter:
        query += " AND vendor_name = ?"
        params.append(vendor_name_filter)
    if model_number_filter:
        query += " AND model_number = ?"
        params.append(model_number_filter)
    if detector_type_filter:
        query += " AND detector_type = ?"
        params.append(detector_type_filter)

    try:
        # Execute the constructed query
        c.execute(query, params)
        products = c.fetchall()

        # Process products to include base64 encoded images
        for i in range(len(products)):
            if products[i][4]:  # Check if detector_image is not None
                products[i] = list(products[i])  # Convert tuple to list to modify
                products[i][4] = base64.b64encode(products[i][4]).decode('utf-8')  # Encode the image to base64

        # Get distinct values for filters
        c.execute('SELECT DISTINCT vendor_name FROM detectors_products')
        vendors = [row[0] for row in c.fetchall()]

        c.execute('SELECT DISTINCT model_number FROM detectors_products')
        models = [row[0] for row in c.fetchall()]

        c.execute('SELECT DISTINCT detector_type FROM detectors_products')
        detector_types = [row[0] for row in c.fetchall()]

        return render_template('detectors_products.html', products=products, vendors=vendors, models=models, detector_types=detector_types)

    except Exception as e:
        flash(f'An error occurred: {e}', 'danger')
    finally:
        conn.close()

    return render_template('detectors_products.html', products=[], vendors=[], models=[], detector_types=[])

@app.route('/register_detector_product', methods=['GET', 'POST'])
#@role_required('editor')  # Adjust role as necessary
def register_detector_product():
    if request.method == 'POST':
        # Retrieve form data
        vendor_name = request.form['vendor_name']
        model_number = request.form['model_number']
        detector_type = request.form['detector_type']
        operating_voltage = request.form['operating_voltage']
        max_area_coverage = request.form.get('max_area_coverage', type=float)
        temperature_range = request.form['temperature_range']
        price = request.form.get('price', type=float)

        # Handle the uploaded image
        detector_image = request.files['detector_image']
        if detector_image and detector_image.filename:
            detector_image_data = detector_image.read()
        else:
            flash('Please upload a valid detector image.', 'danger')
            return redirect(url_for('register_detector_product'))

        # Database connection
        conn = sqlite3.connect('ProjectStatus.db')
        c = conn.cursor()

        try:
            # Insert data into the database
            c.execute('''INSERT INTO detectors_products (vendor_name, model_number, detector_type,
                        operating_voltage, max_area_coverage, temperature_range, price, detector_image)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
                      (vendor_name, model_number, detector_type, operating_voltage,
                       max_area_coverage, temperature_range, price, detector_image_data))
            conn.commit()
            flash('Detector product registered successfully!', 'success')
        except sqlite3.IntegrityError:
            flash('Error: Model number must be unique!', 'danger')
        except Exception as e:
            flash(f'An error occurred: {e}', 'danger')
        finally:
            conn.close()

        return redirect(url_for('view_detectors_products'))  # Redirect to the product list or another page

    return render_template('register_detector_product.html')  # Render the registration form


@app.route('/view_manual_call_point')
#@role_required('editor')
def view_manual_call_point():
    return render_template('view_manual_call_point.html')  # Create this HTML file

@app.route('/view_manual_call_point_products', methods=['GET'])
#@role_required('editor')
def view_manual_call_point_products():
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()

    # Initialize filter parameters
    vendor_name_filter = request.args.get('vendor_name', default=None)
    model_number_filter = request.args.get('model_number', default=None)
    manual_call_point_type_filter = request.args.get('manual_call_point_type', default=None)

    # Base query
    query = "SELECT vendor_name, model_number, manual_call_point_type, price, image FROM manual_call_point_products WHERE 1=1"
    params = []  # List to hold parameters for filtering

    # Append conditions for filters if they are provided
    if vendor_name_filter:
        query += " AND vendor_name = ?"
        params.append(vendor_name_filter)
    if model_number_filter:
        query += " AND model_number = ?"
        params.append(model_number_filter)
    if manual_call_point_type_filter:
        query += " AND manual_call_point_type = ?"
        params.append(manual_call_point_type_filter)

    c.execute(query, params)
    products = c.fetchall()

    # Fetch distinct values for filters
    c.execute('SELECT DISTINCT vendor_name FROM manual_call_point_products')
    vendors = [row[0] for row in c.fetchall()]

    c.execute('SELECT DISTINCT model_number FROM manual_call_point_products')
    models = [row[0] for row in c.fetchall()]

    c.execute('SELECT DISTINCT manual_call_point_type FROM manual_call_point_products')
    manual_call_point_types = [row[0] for row in c.fetchall()]

    conn.close()

    return render_template('manual_call_point_products.html', products=products,
                           vendors=vendors, models=models, manual_call_point_types=manual_call_point_types)
@app.route('/register_manual_call_point_product_page', methods=['GET'])
#@role_required('editor')  # Adjust role as needed
def register_manual_call_point_product_page():
    return render_template('register_manual_call_point_product.html')

@app.route('/register_manual_call_point_product', methods=['GET', 'POST'])
#@role_required('editor')
def register_manual_call_point_product():
    if request.method == 'POST':
        # Retrieve form data
        vendor_name = request.form['vendor_name']
        model_number = request.form['model_number']
        manual_call_point_type = request.form['manual_call_point_type']
        price = request.form['price']
        description = request.form['description']

        # Handle the uploaded image
        image_file = request.files['image']
        if image_file and image_file.filename:
            image_data = image_file.read()
        else:
            flash('Please upload a valid image.', 'danger')
            return redirect(url_for('register_manual_call_point_product'))

        # Database connection
        conn = sqlite3.connect('ProjectStatus.db')
        c = conn.cursor()

        try:
            # Insert data into the database
            c.execute('''INSERT INTO manual_call_point_products (vendor_name, model_number, 
                        manual_call_point_type, price, description, image)
                        VALUES (?, ?, ?, ?, ?, ?)''',
                      (vendor_name, model_number, manual_call_point_type, price, description, image_data))
            conn.commit()
            flash('Product registered successfully!', 'success')
        except sqlite3.IntegrityError:
            flash('Error: Model number must be unique!', 'danger')
        except Exception as e:
            flash(f'An error occurred: {e}', 'danger')
        finally:
            conn.close()

        return redirect(url_for('view_manual_call_point_products'))

    return render_template('register_manual_call_point_product.html')

@app.route('/view_flasher_sounder')
#@role_required('editor')
def view_flasher_sounder():
    return render_template('view_flasher_sounder.html')  # Create this HTML file

@app.route('/view_control_monitor_module')
#@role_required('editor')
def view_control_monitor_module():
    return render_template('view_control_monitor_module.html')  # Create this HTML file

@app.route('/view_facp')
#@role_required('editor')
def view_facp():
    return render_template('view_facp.html')  # Create this HTML file
###################################3

@app.route('/register_consultant', methods=['GET', 'POST'])
#@role_required('editor')
def register_consultant():
    if request.method == 'POST':
        name = request.form['name']
        contact_person = request.form['contact_person']
        phone = request.form['phone']
        email = request.form['email']
        note = request.form['note']

        conn = sqlite3.connect('ProjectStatus.db')
        c = conn.cursor()
        try:
            c.execute('''INSERT INTO consultants (name, contact_person, phone, email, note)
                         VALUES (?, ?, ?, ?, ?)''', (name, contact_person, phone, email, note))
            conn.commit()
            flash('Consultant registered successfully!', 'success')
        except sqlite3.IntegrityError:
            flash('Error: Consultant name must be unique!', 'danger')
        finally:
            conn.close()

        return redirect(url_for('register_project'))

    return render_template('register_consultant.html')
###################################3
# Route for registering contractors
@app.route('/register_contractor', methods=['GET', 'POST'])
#@role_required('editor')
def register_contractor():
    if request.method == 'POST':
        name = request.form['name']
        contact_person = request.form['contact_person']
        phone = request.form['phone']
        email = request.form['email']
        note = request.form['note']

        conn = sqlite3.connect('ProjectStatus.db')
        c = conn.cursor()
        try:
            c.execute('''INSERT INTO contractors (name, contact_person, phone, email, note)
                         VALUES (?, ?, ?, ?, ?)''', (name, contact_person, phone, email, note))
            conn.commit()
            flash('Contractor registered successfully!', 'success')
        except sqlite3.IntegrityError:
            flash('Error: Contractor name must be unique!', 'danger')
        finally:
            conn.close()

        return redirect(url_for('register_project'))

    return render_template('register_contractor.html')  # Create this HTML template

# Route for registering end users
@app.route('/register_end_user', methods=['GET', 'POST'])
#@role_required('editor')
def register_end_user():
    if request.method == 'POST':
        name = request.form['name']
        contact_person = request.form['contact_person']
        phone = request.form['phone']
        email = request.form['email']
        note = request.form['note']

        conn = sqlite3.connect('ProjectStatus.db')
        c = conn.cursor()
        try:
            c.execute('''INSERT INTO end_users (name, contact_person, phone, email, note)
                         VALUES (?, ?, ?, ?, ?)''', (name, contact_person, phone, email, note))
            conn.commit()
            flash('End user registered successfully!', 'success')
        except sqlite3.IntegrityError:
            flash('Error: End user name must be unique!', 'danger')
        finally:
            conn.close()

        return redirect(url_for('register_project'))

    return render_template('register_end_user.html')  # Create this HTML template

##################################
@app.route('/register_project', methods=['GET', 'POST'])
@permission_required('view_projects')
def register_project():
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()

    if request.method == 'POST':
        project_name = request.form['project_name']
        stage = request.form['stage']
        probability = request.form['probability']
        # Deal value removed from form - will be auto-calculated by AI
        deal_value = 0.0
        expected_close_date = request.form['expected_close_date']
        sales_engineer_id = request.form.get('sales_engineer_id')
        end_user_id = request.form['end_user_id']
        contractor_id = request.form.get('contractor_id')
        consultant_id = request.form.get('consultant_id')
        scope_of_work = request.form['scope_of_work']
        note = request.form['note']
        client_type = request.form.get('client_type')

        # Automatically generate the current timestamp
        registered_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        # Get the username of who is registering the project
        updated_by = session.get('username', 'Unknown')

        try:
            # Updated INSERT statement with approval_status='Pending' for new projects
            c.execute('''INSERT INTO register_project 
                         (project_name, end_user_id, contractor_id, consultant_id, scope_of_work, note, 
                          stage, deal_value, expected_close_date, probability, sales_engineer_id, registered_date, approval_status, updated_by, client_type)
                         VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                      (project_name, end_user_id, contractor_id, consultant_id, scope_of_work, note,
                       stage, deal_value, expected_close_date, probability, sales_engineer_id, registered_date, 'Pending', updated_by, client_type))
            project_id = c.lastrowid
            conn.commit()
            flash('Project registered successfully! It is pending admin approval before appearing in the pipeline.', 'info')
            
            # Send notifications to admins only (project needs approval)
            try:
                actor_id = session.get('user_id')
                actor_name = session.get('username', 'Unknown')
                
                # Get admin recipients only (General Manager and Technical Team Leader)
                recipients = notification_service.get_admin_recipients()
                
                # Create notification for pending approval
                notification_service.notify_activity(
                    event_code='project.pending_approval',
                    recipient_ids=recipients,
                    actor_id=actor_id,
                    context={
                        'actor_name': actor_name,
                        'project_name': project_name,
                        'project_id': project_id,
                        'stage': stage
                    },
                    url=url_for('pending_project_approvals')
                )
            except Exception as e:
                print(f"Notification error: {e}")
        except sqlite3.IntegrityError:
            flash(f'Error: A project with the name "{project_name}" already exists.', 'danger')
        except Exception as e:
            flash(f'An error occurred: {e}', 'danger')
        finally:
            conn.close()
        return redirect(url_for('register_project'))

    # Fetch data for dropdowns
    stages = [
        'Lead', 'Proposal Prep', 'Proposal Sent',
        'Customer Pending', 'Technical Discussion', 'Negotiation', 'Closed Won', 'Closed Lost', 'Cancelled'
    ]
    c.execute('SELECT id, name FROM end_users')
    end_users = c.fetchall()
    c.execute('SELECT id, name FROM contractors')
    contractors = c.fetchall()
    c.execute('SELECT id, name FROM consultants')
    consultants = c.fetchall()
    c.execute("SELECT id, username FROM engineers WHERE role IN ('Sales Engineer', 'Technical Team Leader')")
    sales_engineers = c.fetchall()

    # Get current logged-in user's engineer ID (if they are a sales engineer)
    current_username = session.get('username')
    current_user_engineer_id = None
    if current_username:
        c.execute("SELECT id FROM engineers WHERE username = ?", (current_username,))
        result = c.fetchone()
        if result:
            current_user_engineer_id = result[0]

    conn.close()

    # Get the current date to display on the form
    current_date = datetime.now().strftime('%Y-%m-%d')

    return render_template('register_project.html', end_users=end_users, contractors=contractors,
                           consultants=consultants, stages=stages, sales_engineers=sales_engineers,
                           current_date=current_date, current_user_engineer_id=current_user_engineer_id)

####################################
@app.route('/project_pipeline')
@permission_required('view_projects')
def project_pipeline():
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()

    # Get filter values
    sales_engineer_filter = request.args.get('sales_engineer_id')
    start_date_filter = request.args.get('start_date')
    end_date_filter = request.args.get('end_date')

    # Fetch all sales engineers for the filter dropdown
    c.execute("SELECT id, name FROM engineers WHERE role IN ('Sales Engineer', 'Technical Team Leader')")
    sales_engineers = c.fetchall()

    stages_ordered = [
        'Lead', 'Proposal Prep', 'Proposal Sent',
        'Customer Pending', 'Technical Discussion', 'Negotiation', 'Closed Won', 'Closed Lost', 'Cancelled'
    ]
    pipeline_data = {stage: {'projects': [], 'count': 0, 'total_value': 0} for stage in stages_ordered}

    # Base query
    query = '''
        SELECT 
            rp.id, rp.project_name, rp.deal_value, rp.probability, 
            rp.expected_close_date, rp.stage, eu.name AS end_user_name,
            co.name AS contractor_name, rp.scope_of_work,
            en.name AS sales_engineer_name, rp.registered_date
        FROM register_project rp
        LEFT JOIN end_users eu ON rp.end_user_id = eu.id
        LEFT JOIN contractors co ON rp.contractor_id = co.id
        LEFT JOIN engineers en ON rp.sales_engineer_id = en.id
        WHERE 1=1
    '''
    params = []
    user_role = None

    # --- NEW ACCESS CONTROL LOGIC ---
    if 'user_id' in session:
        c.execute("SELECT role FROM users WHERE id = ?", (session['user_id'],))
        user_role_result = c.fetchone()
        if user_role_result:
            user_role = user_role_result[0]
            if user_role == 'Sales Engineer':
                c.execute("SELECT id FROM engineers WHERE username = ?", (session['username'],))
                engineer_id_result = c.fetchone()
                if engineer_id_result:
                    query += " AND rp.sales_engineer_id = ?"
                    params.append(engineer_id_result[0])

    # Add other filters
    if sales_engineer_filter and user_role != 'Sales Engineer':
        query += " AND rp.sales_engineer_id = ?"
        params.append(sales_engineer_filter)
    if start_date_filter:
        query += " AND date(rp.registered_date) >= ?"
        params.append(start_date_filter)
    if end_date_filter:
        query += " AND date(rp.registered_date) <= ?"
        params.append(end_date_filter)

    c.execute(query, params)
    projects = c.fetchall()
    conn.close()

    # Process and group projects... (this part remains the same)
    for project in projects:
        stage = project[5]
        if stage in pipeline_data:
            quarter = 'N/A'
            if project[10]:
                try:
                    dt_object = datetime.strptime(project[10], '%Y-%m-%d %H:%M:%S')
                    q = (dt_object.month - 1) // 3 + 1
                    quarter = f"Q{q}-{dt_object.year}"
                except (ValueError, TypeError):
                    quarter = 'N/A'
            pipeline_data[stage]['projects'].append({
                'id': project[0], 'name': project[1], 'value': project[2] or 0,
                'probability': project[3], 'close_date': project[4], 'end_user': project[6],
                'contractor': project[7], 'sow': project[8], 'sales_engineer': project[9],
                'quarter': quarter
            })
            pipeline_data[stage]['count'] += 1
            pipeline_data[stage]['total_value'] += (project[2] or 0)

    return render_template('project_pipeline.html',
                           pipeline_data=pipeline_data,
                           stages_ordered=stages_ordered,
                           sales_engineers=sales_engineers,
                           user_role=user_role,
                           current_filters={
                               'sales_engineer_id': sales_engineer_filter,
                               'start_date': start_date_filter,
                               'end_date': end_date_filter
                           })
#########################
@app.route('/edit_project_pipeline/<int:project_id>', methods=['GET', 'POST'])
@login_required
def edit_project_pipeline(project_id):
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    if request.method == 'POST':
        # Get data from the form
        project_name = request.form['project_name']
        stage = request.form['stage']
        probability = request.form['probability']
        # Deal value removed from form - will be auto-calculated by AI
        deal_value = 0.0
        expected_close_date = request.form['expected_close_date']
        sales_engineer_id = request.form.get('sales_engineer_id')
        end_user_id = request.form['end_user_id']
        contractor_id = request.form.get('contractor_id')
        consultant_id = request.form.get('consultant_id')
        client_type = request.form.get('client_type')
        scope_of_work = request.form['scope_of_work']
        note = request.form['note']

        # Automatically get the update time and user
        updated_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        updated_by = session.get('username', 'Unknown')  # Gets username from session

        # Update the database with the new fields
        c.execute('''
            UPDATE register_project SET
            project_name = ?, stage = ?, probability = ?, deal_value = ?,
            expected_close_date = ?, end_user_id = ?, contractor_id = ?,
            consultant_id = ?, client_type = ?, scope_of_work = ?, note = ?, sales_engineer_id = ?,
            updated_time = ?, updated_by = ?
            WHERE id = ?
        ''', (project_name, stage, probability, deal_value, expected_close_date,
              end_user_id, contractor_id, consultant_id, client_type, scope_of_work, note,
              sales_engineer_id, updated_time, updated_by, project_id))

        conn.commit()
        conn.close()
        flash('Project updated successfully!', 'success')
        
        # Send notifications to stakeholders
        try:
            actor_id = session.get('user_id')
            actor_name = session.get('username', 'Unknown')
            
            # Get recipients: sales engineer + admins
            recipients = notification_service.get_project_stakeholders(project_id)
            
            # Create notification
            notification_service.notify_activity(
                event_code='project.updated',
                recipient_ids=recipients,
                actor_id=actor_id,
                context={
                    'actor_name': actor_name,
                    'project_name': project_name,
                    'project_id': project_id,
                    'stage': stage
                },
                url=url_for('edit_project_pipeline', project_id=project_id)
            )
        except Exception as e:
            print(f"Notification error: {e}")
        return redirect(url_for('project_pipeline'))

    # GET request logic remains the same
    c.execute("SELECT * FROM register_project WHERE id = ?", (project_id,))
    project = c.fetchone()
    stages = [
        'Lead', 'Proposal Prep', 'Proposal Sent',
        'Customer Pending', 'Technical Discussion', 'Negotiation', 'Closed Won', 'Closed Lost', 'Cancelled'
    ]
    c.execute('SELECT id, name FROM end_users')
    end_users = c.fetchall()
    c.execute('SELECT id, name FROM contractors')
    contractors = c.fetchall()
    c.execute('SELECT id, name FROM consultants')
    consultants = c.fetchall()
    c.execute("SELECT id, username FROM engineers WHERE role IN ('Sales Engineer', 'Technical Team Leader')")
    sales_engineers = c.fetchall()
    conn.close()

    return render_template('edit_project_pipeline.html', project=project, stages=stages,
                           end_users=end_users, contractors=contractors,
                           consultants=consultants, sales_engineers=sales_engineers)

@app.route('/delete_project/<int:project_id>', methods=['POST'])
@login_required
@role_required('General Manager', 'Technical Team Leader')
def delete_project(project_id):
    """
    Delete a project from the pipeline (Admin only)
    """
    try:
        conn = sqlite3.connect('ProjectStatus.db')
        c = conn.cursor()
        
        # Get project name for confirmation
        c.execute("SELECT project_name FROM register_project WHERE id = ?", (project_id,))
        project = c.fetchone()
        
        if not project:
            conn.close()
            return jsonify({'success': False, 'message': 'Project not found'}), 404
        
        project_name = project[0]
        
        # Delete the project
        c.execute("DELETE FROM register_project WHERE id = ?", (project_id,))
        conn.commit()
        conn.close()
        
        # Send notifications to admins only
        try:
            actor_id = session.get('user_id')
            actor_name = session.get('username', 'Unknown')
            
            # Get admin recipients only
            recipients = notification_service.get_admin_recipients()
            
            # Create notification
            notification_service.notify_activity(
                event_code='project.deleted',
                recipient_ids=recipients,
                actor_id=actor_id,
                context={
                    'actor_name': actor_name,
                    'project_name': project_name,
                    'project_id': project_id
                },
                url=url_for('project_pipeline')
            )
        except Exception as e:
            print(f"Notification error: {e}")
        
        return jsonify({'success': True, 'message': f'Project "{project_name}" deleted successfully'})
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

###################3333
@app.route('/update_project_stage', methods=['POST'])
@login_required
def update_project_stage():
    data = request.json
    project_name = data.get('project_name')
    new_stage = data.get('new_stage')

    if not project_name or not new_stage:
        return jsonify({'status': 'error', 'message': 'Missing data'}), 400

    # Map stages to their corresponding probability
    stage_probability_map = {
        'Lead': 0.01,
        'Proposal Prep': 1.00,
        'Proposal Sent': 5.00,
        'Customer Pending': 2.00,
        'Technical Discussion': 50.00,
        'Negotiation': 70.00,
        'Closed Won': 100.00,
        'Closed Lost': 0.00,
        'Cancelled': 0.00
    }
    new_probability = stage_probability_map.get(new_stage)

    # Get the current timestamp to record the time of the update
    updated_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    try:
        conn = sqlite3.connect('ProjectStatus.db')
        c = conn.cursor()
        
        # Get current stage and project details before updating
        c.execute("""
            SELECT rp.stage, rp.id, rp.sales_engineer_id, e.username as sales_engineer_name
            FROM register_project rp
            LEFT JOIN engineers e ON rp.sales_engineer_id = e.id
            WHERE rp.project_name = ?
        """, (project_name,))
        project_info = c.fetchone()
        
        if not project_info:
            conn.close()
            return jsonify({'status': 'error', 'message': 'Project not found'}), 404
        
        old_stage = project_info[0]
        project_id = project_info[1]
        sales_engineer_id = project_info[2]
        sales_engineer_name = project_info[3]
        
        # Update the stage, probability, AND the updated_time
        c.execute("""
            UPDATE register_project 
            SET stage = ?, probability = ?, updated_time = ? 
            WHERE project_name = ?
        """, (new_stage, new_probability, updated_time, project_name))
        conn.commit()
        conn.close()
        
        # Send notifications for stage change
        try:
            # Get actor information
            actor_id = session.get('user_id')
            actor_name = session.get('username', 'Unknown User')
            
            # Get admin recipients
            admin_recipients = notification_service.get_admin_recipients()
            
            # Add sales engineer to recipients if exists
            recipients = admin_recipients.copy()
            if sales_engineer_id and sales_engineer_name:
                # Get user_id for this sales engineer from users table
                conn2 = sqlite3.connect('ProjectStatus.db')
                c2 = conn2.cursor()
                c2.execute("SELECT id FROM users WHERE username = ?", (sales_engineer_name,))
                sales_user = c2.fetchone()
                conn2.close()
                
                if sales_user and sales_user[0] not in recipients:
                    recipients.append(sales_user[0])
            
            # Create notification for stage change
            notification_service.notify_activity(
                event_code='project.stage_changed',
                recipient_ids=recipients,
                actor_id=actor_id,
                context={
                    'actor_name': actor_name,
                    'project_name': project_name,
                    'old_stage': old_stage,
                    'new_stage': new_stage,
                    'sales_engineer': sales_engineer_name or 'Not assigned'
                },
                url=url_for('project_pipeline')
            )
        except Exception as e:
            print(f"Notification error: {e}")

        return jsonify({'status': 'success', 'message': 'Project stage updated', 'newProbability': new_probability})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500
################3
@app.route('/pipeline_analysis')
@login_required
@permission_required('view_pipeline_analysis')
def pipeline_analysis():
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()

    current_year = datetime.now().year
    selected_year = request.args.get('year', default=current_year, type=int)
    sales_engineer_filter = request.args.get('sales_engineer_id')

    # Fetch all sales engineers for the filter dropdown
    c.execute("SELECT id, username FROM engineers WHERE role IN ('Sales Engineer', 'Technical Team Leader')")
    sales_engineers = c.fetchall()

    # Base query
    query = """
        SELECT stage, deal_value, probability, expected_close_date 
        FROM register_project rp 
        WHERE 1=1
    """
    params = []
    user_role = None

    if 'user_id' in session:
        c.execute("SELECT role FROM users WHERE id = ?", (session['user_id'],))
        user_role_result = c.fetchone()
        if user_role_result:
            user_role = user_role_result[0]
            if user_role == 'Sales Engineer':
                c.execute("SELECT id FROM engineers WHERE username = ?", (session['username'],))
                engineer_id_result = c.fetchone()
                if engineer_id_result:
                    query += " AND rp.sales_engineer_id = ?"
                    params.append(engineer_id_result[0])

    if sales_engineer_filter and user_role in ['General Manager', 'Technical Team Leader']:
        query += " AND rp.sales_engineer_id = ?"
        params.append(sales_engineer_filter)

    c.execute(query, params)
    projects = c.fetchall()
    conn.close()

    # --- CORRECTED CALCULATION LOGIC ---

    quarterly_expected_values = {'Q1': 0, 'Q2': 0, 'Q3': 0, 'Q4': 0}
    quarterly_won_values = {'Q1': 0, 'Q2': 0, 'Q3': 0, 'Q4': 0}
    quarterly_lost_values = {'Q1': 0, 'Q2': 0, 'Q3': 0, 'Q4': 0}
    deal_count = 0
    closed_stages = ['Closed Won', 'Closed Lost', 'Cancelled']

    for project in projects:
        stage, deal_value, probability, close_date_str = project
        deal_value = deal_value or 0

        if not (close_date_str and stage and probability is not None):
            continue
        try:
            close_date = datetime.strptime(close_date_str, '%Y-%m-%d')
            if close_date.year == selected_year:
                quarter = f"Q{(close_date.month - 1) // 3 + 1}"

                # ONLY calculate Expected Value if the deal is NOT in a closed state
                if stage not in closed_stages:
                    deal_count += 1
                    expected_value = (probability / 100) * deal_value
                    quarterly_expected_values[quarter] += expected_value

                # Separately, calculate the actual Won/Lost values
                if stage == 'Closed Won':
                    quarterly_won_values[quarter] += deal_value
                elif stage == 'Closed Lost':
                    quarterly_lost_values[quarter] += deal_value
        except (ValueError, TypeError):
            continue

    # The rest of the function remains the same
    total_expected_value = sum(quarterly_expected_values.values())
    total_won_value = sum(quarterly_won_values.values())
    total_lost_value = sum(quarterly_lost_values.values())
    labels = list(quarterly_expected_values.keys())
    expected_data = list(quarterly_expected_values.values())
    won_data = list(quarterly_won_values.values())
    lost_data = list(quarterly_lost_values.values())
    year_range = range(current_year - 5, current_year + 5)

    return render_template('pipeline_analysis.html',
                           labels=labels, expected_data=expected_data,
                           won_data=won_data, lost_data=lost_data,
                           selected_year=selected_year, year_range=year_range,
                           total_expected_value=total_expected_value,
                           total_won_value=total_won_value,
                           total_lost_value=total_lost_value,
                           deal_count=deal_count,
                           sales_engineers=sales_engineers,
                           user_role=user_role,
                           current_filters={'sales_engineer_id': sales_engineer_filter})
################
@app.route('/presales_performance')
@permission_required('view_presales')
def presales_performance():
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()

    # Get filter values from URL
    current_year = datetime.now().year
    selected_year = request.args.get('year', default=current_year, type=int)
    presale_filter = request.args.get('presale_engineer')

    # Fetch all presales engineers for the filter dropdown
    c.execute("SELECT username FROM engineers WHERE role IN ('Presale Engineer', 'Technical Team Leader')")
    presale_engineers = [row[0] for row in c.fetchall()]

    # --- Calculation 1: RFQs Requested vs. Quoted (for chart) ---
    # ... (This calculation remains the same)
    params_req = [str(selected_year)]
    query_req = "SELECT strftime('%m', requested_time) FROM rfq_requests WHERE strftime('%Y', requested_time) = ?"
    if presale_filter:
        query_req += " AND sales_engineer_presale = ?"
        params_req.append(presale_filter)
    c.execute(query_req, params_req)
    requested_rfqs = c.fetchall()
    quarterly_requests = {'Q1': 0, 'Q2': 0, 'Q3': 0, 'Q4': 0}
    for month in requested_rfqs:
        q = f"Q{(int(month[0]) - 1) // 3 + 1}"
        quarterly_requests[q] += 1

    params_quoted = [str(selected_year)]
    query_quoted = "SELECT strftime('%m', registered_date) FROM projects WHERE rfq_reference IS NOT NULL AND strftime('%Y', registered_date) = ?"
    if presale_filter:
        query_quoted += " AND presale_eng = ?"
        params_quoted.append(presale_filter)
    c.execute(query_quoted, params_quoted)
    quoted_rfqs = c.fetchall()
    quarterly_quoted = {'Q1': 0, 'Q2': 0, 'Q3': 0, 'Q4': 0}
    for month in quoted_rfqs:
        q = f"Q{(int(month[0]) - 1) // 3 + 1}"
        quarterly_quoted[q] += 1

    # --- Calculation 2: Average Turnaround Time (for chart) ---
    # ... (This calculation remains the same)
    turnaround_query = "SELECT r.sales_engineer_presale, julianday(p.registered_date) - julianday(r.requested_time) AS turnaround_days FROM rfq_requests r JOIN projects p ON r.rfq_reference = p.rfq_reference WHERE strftime('%Y', r.requested_time) = ?"
    turnaround_params = [str(selected_year)]
    if presale_filter:
        turnaround_query += " AND r.sales_engineer_presale = ?"
        turnaround_params.append(presale_filter)
    c.execute(turnaround_query, turnaround_params)
    turnaround_results = c.fetchall()
    engineer_times = {}
    for engineer, days in turnaround_results:
        if days is None: continue
        if engineer not in engineer_times: engineer_times[engineer] = []
        engineer_times[engineer].append(days)
    avg_turnaround_data = {engineer: sum(times) / len(times) for engineer, times in engineer_times.items()}
    sorted_turnaround = sorted(avg_turnaround_data.items(), key=lambda item: item[1], reverse=True)
    turnaround_labels = [item[0] for item in sorted_turnaround]
    turnaround_data = [item[1] for item in sorted_turnaround]

    # --- Calculation 3: Win Rate & Value Analysis ---
    value_query = """
        SELECT
            presale_eng,
            SUM(quotation_selling_price) AS total_quoted,
            SUM(CASE WHEN status = 'WON' THEN quotation_selling_price ELSE 0 END) AS total_won
        FROM projects
        WHERE strftime('%Y', registered_date) = ?
    """
    value_params = [str(selected_year)]
    if presale_filter:
        value_query += " AND presale_eng = ?"
        value_params.append(presale_filter)
    value_query += " GROUP BY presale_eng"
    c.execute(value_query, value_params)
    value_results = c.fetchall()

    # --- Calculation 4: Presales Leaderboard ---
    # ... (This calculation remains the same)
    leaderboard_query = "SELECT p.presale_eng, COUNT(p.id), SUM(CASE WHEN p.status = 'WON' THEN 1 ELSE 0 END), SUM(CASE WHEN p.status = 'WON' THEN p.quotation_selling_price ELSE 0 END), AVG(CASE WHEN p.status = 'WON' THEN p.margin ELSE NULL END) FROM projects p WHERE strftime('%Y', p.registered_date) = ?"
    leaderboard_params = [str(selected_year)]
    if presale_filter:
        leaderboard_query += " AND p.presale_eng = ?"
        leaderboard_params.append(presale_filter)
    leaderboard_query += " GROUP BY p.presale_eng ORDER BY SUM(CASE WHEN p.status = 'WON' THEN p.quotation_selling_price ELSE 0 END) DESC"
    c.execute(leaderboard_query, leaderboard_params)
    leaderboard_data = c.fetchall()

    conn.close()

    # --- Prepare all data for the template ---
    labels = list(quarterly_requests.keys())
    requested_data = list(quarterly_requests.values())
    quoted_data = list(quarterly_quoted.values())
    year_range = range(current_year - 5, current_year + 5)

    # Prepare data for the new value analysis chart
    value_labels = [row[0] for row in value_results]
    total_quoted_data = [row[1] for row in value_results]
    total_won_data = [row[2] for row in value_results]

    return render_template('presales_performance.html',
                           labels=labels, requested_data=requested_data, quoted_data=quoted_data,
                           selected_year=selected_year, year_range=year_range,
                           presale_engineers=presale_engineers, presale_filter=presale_filter,
                           turnaround_labels=turnaround_labels, turnaround_data=turnaround_data,
                           leaderboard_data=leaderboard_data,
                           value_labels=value_labels,
                           total_quoted_data=total_quoted_data,
                           total_won_data=total_won_data)
###############
@app.route('/sales_performance')
@login_required
@permission_required('view_sales_performance')
def sales_performance():
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()

    # Get filter values from URL
    current_year = datetime.now().year
    selected_year = request.args.get('year', default=current_year, type=int)
    sales_filter = request.args.get('sales_engineer')

    # Fetch SALES engineers for the filter dropdown
    c.execute("SELECT username FROM engineers WHERE role IN ('Sales Engineer', 'Technical Team Leader')")
    sales_engineers = [row[0] for row in c.fetchall()]

    # --- Calculation 1: Win Rate & Value Analysis (Grouped by SALES Engineer) ---
    value_query = """
        SELECT
            sales_eng,
            SUM(quotation_selling_price) AS total_quoted,
            SUM(CASE WHEN status = 'WON' THEN quotation_selling_price ELSE 0 END) AS total_won
        FROM projects
        WHERE strftime('%Y', registered_date) = ? AND sales_eng IS NOT NULL
    """
    value_params = [str(selected_year)]
    if sales_filter:
        value_query += " AND sales_eng = ?"
        value_params.append(sales_filter)
    value_query += " GROUP BY sales_eng ORDER BY total_won DESC"
    c.execute(value_query, value_params)
    value_results = c.fetchall()

    # --- Calculation 2: Sales Leaderboard (Grouped by SALES Engineer) ---
    leaderboard_query = """
        SELECT 
            p.sales_eng, 
            COUNT(p.id) AS total_quotes, 
            SUM(CASE WHEN p.status = 'WON' THEN 1 ELSE 0 END) AS won_deals,
            SUM(CASE WHEN p.status = 'WON' THEN p.quotation_selling_price ELSE 0 END) AS won_value,
            AVG(CASE WHEN p.status = 'WON' THEN p.margin ELSE NULL END) AS avg_margin
        FROM projects p 
        WHERE strftime('%Y', p.registered_date) = ? AND p.sales_eng IS NOT NULL
    """
    leaderboard_params = [str(selected_year)]
    if sales_filter:
        leaderboard_query += " AND p.sales_eng = ?"
        leaderboard_params.append(sales_filter)
    leaderboard_query += " GROUP BY p.sales_eng ORDER BY won_value DESC"
    c.execute(leaderboard_query, leaderboard_params)
    leaderboard_data = c.fetchall()
    conn.close()

    # Prepare data for the charts
    value_labels = [row[0] for row in value_results]
    total_quoted_data = [row[1] for row in value_results]
    total_won_data = [row[2] for row in value_results]
    year_range = range(current_year - 5, current_year + 5)

    return render_template('sales_performance.html',
                           selected_year=selected_year, year_range=year_range,
                           sales_engineers=sales_engineers, sales_filter=sales_filter,
                           leaderboard_data=leaderboard_data,
                           value_labels=value_labels,
                           total_quoted_data=total_quoted_data,
                           total_won_data=total_won_data)
##########################
@app.route('/comparison/<ref>', methods=['GET'])
#@role_required('editor')
def show_comparison(ref):
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()

    # Fetch the data for the given reference
    c.execute("SELECT quotation_cost, quotation_selling_price, margin FROM projects WHERE quote_ref=?", (ref,))
    comparison_data = c.fetchone()
    conn.close()

    if comparison_data:
        quotation_cost, quotation_selling_price, margin = comparison_data
        
        # Sanitize None values and prepare template-ready display values
        cost_value = quotation_cost if quotation_cost is not None else 0
        price_value = quotation_selling_price if quotation_selling_price is not None else 0
        margin_value = margin if margin is not None else None
        
        # Calculate markup safely
        markup_value = None
        if cost_value and cost_value > 0 and price_value is not None:
            markup_value = ((price_value - cost_value) / cost_value) * 100
        
        # Prepare formatted display strings
        margin_display = f"{margin_value:.2f}%" if margin_value is not None else "N/A"
        markup_display = f"{markup_value:.2f}%" if markup_value is not None else "N/A"

        # Prepare data for the chart
        chart_data = {
            'quotation_reference': ref,
            'quotation_cost': cost_value,
            'quotation_selling_price': price_value,
            'margin': margin_value,
            'margin_display': margin_display,
            'markup_display': markup_display,
        }

        return render_template('comparison_charts.html', data=chart_data)
    else:
        flash('No project found for this Quote Reference!', 'danger')
        return redirect(url_for('index'))
##############################

@app.route('/request_for_quotation', methods=['GET', 'POST'])
@login_required
def request_for_quotation():
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()

    if request.method == 'POST':
        # Get form data
        rfq_reference = request.form['rfq_reference']
        project_name = request.form['project_name']
        project_status = request.form['project_status']
        priority = request.form['priority']
        sales_engineer_presale = request.form['sales_engineer_presale']
        sales_engineer_sales = request.form['sales_engineer_sales']
        rfq_status = request.form['rfq_status']
        quotation_status = request.form['quotation_status']
        deadline = request.form['deadline']
        note = request.form['note']
        requested_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        try:
            # 1. Insert the new RFQ record
            c.execute('''INSERT INTO rfq_requests (rfq_reference, project_name, project_status, priority,
                        sales_engineer_presale, sales_engineer_sales, rfq_status, quotation_status,
                        deadline, note,requested_time)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?,?)''',
                      (rfq_reference, project_name, project_status, priority,
                       sales_engineer_presale, sales_engineer_sales, rfq_status, quotation_status,
                       deadline, note, requested_time))

            # 2. Update the project's stage from 'Qualification' to 'Proposal Prep'
            c.execute("""
                UPDATE register_project
                SET stage = ?
                WHERE project_name = ? AND stage = ?
            """, ('Proposal Prep', project_name, 'Qualification'))

            rfq_id = c.lastrowid
            conn.commit()
            flash('RFQ created and project stage updated successfully!', 'success')

            # Send notifications
            try:
                actor_id = session.get('user_id')
                actor_name = session.get('username', 'Unknown')
                
                # Build recipient list: all presales + admins + sales engineer
                recipients = []
                
                # Add all Presale Engineers
                recipients.extend(notification_service.get_presale_recipients())
                
                # Add General Managers and Technical Team Leaders
                recipients.extend(notification_service.get_admin_recipients())
                
                # Add the Sales Engineer from the RFQ
                if sales_engineer_sales:
                    sales_eng_user_id = notification_service.get_user_id_by_username(sales_engineer_sales)
                    if sales_eng_user_id and sales_eng_user_id not in recipients:
                        recipients.append(sales_eng_user_id)
                
                # Remove duplicates and current user
                recipients = list(set(recipients))
                if actor_id in recipients:
                    recipients.remove(actor_id)
                
                # Send notification with RFQ details
                if recipients:
                    notification_service.notify_activity(
                        event_code='rfq.created',
                        recipient_ids=recipients,
                        actor_id=actor_id,
                        context={
                            'actor_name': actor_name,
                            'project_name': project_name,
                            'rfq_reference': rfq_reference,
                            'priority': priority,
                            'rfq_status': rfq_status,
                            'quotation_status': quotation_status,
                            'presale_engineer': sales_engineer_presale,
                            'sales_engineer': sales_engineer_sales,
                            'deadline': deadline
                        },
                        url=url_for('rfq_summary')
                    )
            except Exception as e:
                print(f"Notification error: {e}")

        except Exception as e:
            flash(f'Error creating RFQ: {str(e)}', 'danger')
        finally:
            conn.close()

        return redirect(url_for('rfq_summary'))

    # --- GET request logic ---
    project_name_prefill = request.args.get('project_name')
    sales_engineer_prefill = request.args.get('sales_engineer')
    c.execute("SELECT username FROM engineers WHERE role IN ('Presale Engineer', 'Technical Team Leader')")
    presale_engineers = [engineer[0] for engineer in c.fetchall()]
    c.execute("SELECT username FROM engineers WHERE role IN ('Sales Engineer', 'Technical Team Leader')")
    sales_engineers = [engineer[0] for engineer in c.fetchall()]
    c.execute("SELECT project_name FROM register_project")
    projects = c.fetchall()
    conn.close()

    current_datetime = datetime.now().strftime('%Y%m%d%H%M%S')
    priority_options = ['High', 'Medium', 'Low']
    rfq_status_options = ['Queue', 'Studying', 'Pricing', 'Quoted', 'Cancelled']
    quotation_status_options = ['Queue','Quotation Sent', 'Customer Pending', 'Technical Discussion', 'Negotiation',
                                'Closed Won', 'Closed Lost', 'Cancelled']

    return render_template('request_for_quotation.html',
                           projects=projects, priority_options=priority_options,
                           rfq_status_options=rfq_status_options,
                           quotation_status_options=quotation_status_options,
                           presale_engineers=presale_engineers,
                           sales_engineers=sales_engineers,
                           current_datetime=current_datetime,
                           project_name_prefill=project_name_prefill,
                           sales_engineer_prefill=sales_engineer_prefill)

########################
@app.route('/rfq_summary', methods=['GET'])
@permission_required('view_rfq')
def rfq_summary():
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()

    # Get user role to conditionally show elements
    user_role = None
    if 'user_id' in session:
        c.execute("SELECT role FROM users WHERE id = ?", (session['user_id'],))
        user_role_result = c.fetchone()
        if user_role_result:
            user_role = user_role_result[0]

    # Get filter values from URL
    filters = {
        'presale_engineer': request.args.get('presale_engineer'),
        'sales_engineer': request.args.get('sales_engineer'),
        'rfq_status': request.args.get('rfq_status'),
        'quotation_status': request.args.get('quotation_status'),
        'start_date': request.args.get('start_date'),
        'end_date': request.args.get('end_date'),
    }

    # Fetch data for filter dropdowns
    c.execute("SELECT DISTINCT sales_engineer_presale FROM rfq_requests")
    presale_engineers = [row[0] for row in c.fetchall()]
    c.execute("SELECT DISTINCT sales_engineer_sales FROM rfq_requests")
    sales_engineers = [row[0] for row in c.fetchall()]
    c.execute("SELECT DISTINCT rfq_status FROM rfq_requests")
    rfq_status_options = [row[0] for row in c.fetchall()]
    c.execute("SELECT DISTINCT quotation_status FROM rfq_requests")
    quotation_status_options = [row[0] for row in c.fetchall()]

    # Base query with LEFT JOIN to get quotation reference
    query = """
        SELECT
            r.id, r.rfq_reference, r.project_name, r.project_status, r.priority,
            r.sales_engineer_presale, r.sales_engineer_sales, r.rfq_status,
            r.quotation_status, r.deadline, r.note, r.requested_time,
            p.quote_ref
        FROM rfq_requests r
        LEFT JOIN projects p ON r.rfq_reference = p.rfq_reference
        WHERE 1=1
    """
    params = []

    # Apply filters
    if filters['presale_engineer']: query += " AND r.sales_engineer_presale = ?"; params.append(filters['presale_engineer'])
    if filters['sales_engineer']: query += " AND r.sales_engineer_sales = ?"; params.append(filters['sales_engineer'])
    if filters['rfq_status']: query += " AND r.rfq_status = ?"; params.append(filters['rfq_status'])
    if filters['quotation_status']: query += " AND r.quotation_status = ?"; params.append(filters['quotation_status'])
    if filters['start_date']: query += " AND date(r.requested_time) >= ?"; params.append(filters['start_date'])
    if filters['end_date']: query += " AND date(r.requested_time) <= ?"; params.append(filters['end_date'])
    
    query += " ORDER BY r.requested_time DESC"

    c.execute(query, params)
    rfqs = c.fetchall()
    total_rfqs = len(rfqs)

    # Chart data logic that also respects filters
    chart_query = "SELECT rfq_status, COUNT(*) FROM rfq_requests r WHERE 1=1"
    # Re-apply the same parameter logic for the chart query
    if filters['presale_engineer']: chart_query += " AND r.sales_engineer_presale = ?"
    if filters['sales_engineer']: chart_query += " AND r.sales_engineer_sales = ?"
    if filters['rfq_status']: chart_query += " AND r.rfq_status = ?"
    if filters['quotation_status']: chart_query += " AND r.quotation_status = ?"
    if filters['start_date']: chart_query += " AND date(r.requested_time) >= ?"
    if filters['end_date']: chart_query += " AND date(r.requested_time) <= ?"
    chart_query += " GROUP BY r.rfq_status"
    c.execute(chart_query, params) # Use the same params list
    status_counts_raw = c.fetchall()
    conn.close()

    status_labels = [row[0] for row in status_counts_raw]
    status_counts = [row[1] for row in status_counts_raw]

    return render_template('rfq_summary.html',
                           rfqs=rfqs,
                           total_rfqs=total_rfqs,
                           user_role=user_role,
                           presale_engineers=presale_engineers,
                           sales_engineers=sales_engineers,
                           rfq_status_options=rfq_status_options,
                           quotation_status_options=quotation_status_options,
                           status_labels=status_labels,
                           status_counts=status_counts,
                           current_filters=filters)

###################################3
@app.route('/delete_rfq/<int:rfq_id>', methods=['POST'])
@role_required('Technical Team Leader') # Only allows this role to access the route
def delete_rfq(rfq_id):
    try:
        conn = sqlite3.connect('ProjectStatus.db')
        c = conn.cursor()
        c.execute("DELETE FROM rfq_requests WHERE id = ?", (rfq_id,))
        conn.commit()
        conn.close()
        flash('RFQ has been deleted successfully.', 'success')
    except Exception as e:
        flash(f'An error occurred while deleting the RFQ: {e}', 'danger')
    return redirect(url_for('rfq_summary'))
####################################
@app.route('/download_filtered_rfqs', methods=['GET'])
@login_required
def download_filtered_rfqs():
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()

    # Reconstruct the filtering logic including date range
    filters = {
        'presale_engineer': request.args.get('presale_engineer'),
        'sales_engineer': request.args.get('sales_engineer'),
        'rfq_status': request.args.get('rfq_status'),
        'quotation_status': request.args.get('quotation_status'),
        'start_date': request.args.get('start_date'),
        'end_date': request.args.get('end_date'),
    }

    query = """
        SELECT
            r.rfq_reference, r.project_name, r.project_status, r.priority,
            r.sales_engineer_presale, r.sales_engineer_sales, r.rfq_status,
            r.quotation_status, r.deadline, r.note, r.requested_time,
            p.quote_ref
        FROM rfq_requests r
        LEFT JOIN projects p ON r.rfq_reference = p.rfq_reference
        WHERE 1=1
    """
    params = []

    if filters['presale_engineer']:
        query += " AND r.sales_engineer_presale = ?"
        params.append(filters['presale_engineer'])
    if filters['sales_engineer']:
        query += " AND r.sales_engineer_sales = ?"
        params.append(filters['sales_engineer'])
    if filters['rfq_status']:
        query += " AND r.rfq_status = ?"
        params.append(filters['rfq_status'])
    if filters['quotation_status']:
        query += " AND r.quotation_status = ?"
        params.append(filters['quotation_status'])
    if filters['start_date']:
        query += " AND date(r.requested_time) >= ?"
        params.append(filters['start_date'])
    if filters['end_date']:
        query += " AND date(r.requested_time) <= ?"
        params.append(filters['end_date'])
    
    query += " ORDER BY r.requested_time DESC"

    c.execute(query, params)
    rfqs = c.fetchall()

    df = pd.DataFrame(rfqs, columns=[
        'RFQ Reference', 'Project Name', 'Project Status', 'Priority', 'Presale Engineer',
        'Sales Engineer', 'RFQ Status', 'Quotation Status', 'Deadline', 'Note',
        'Requested Time', 'Quotation Reference'
    ])

    output = BytesIO()
    df.to_excel(output, index=False, sheet_name='Filtered RFQs')
    output.seek(0)

    conn.close()
    return send_file(output, download_name='filtered_rfqs.xlsx', as_attachment=True)
##############3
@app.route('/edit_rfq/<int:rfq_id>', methods=['GET', 'POST'])
@role_required('Technical Team Leader', 'Presale Engineer')  # Define roles that can edit
def edit_rfq(rfq_id):
    conn = sqlite3.connect('ProjectStatus.db')
    # Use row_factory to easily access columns by name in the template
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    if request.method == 'POST':
        # Get RFQ reference before updating
        c.execute("SELECT rfq_reference FROM rfq_requests WHERE id = ?", (rfq_id,))
        rfq_ref_result = c.fetchone()
        rfq_reference = rfq_ref_result['rfq_reference'] if rfq_ref_result else 'Unknown'
        
        # Get all updated data from the form
        project_name = request.form['project_name']
        project_status = request.form['project_status']
        priority = request.form['priority']
        sales_engineer_presale = request.form['sales_engineer_presale']
        sales_engineer_sales = request.form['sales_engineer_sales']
        rfq_status = request.form['rfq_status']
        quotation_status = request.form['quotation_status']
        deadline = request.form['deadline']
        note = request.form['note']

        try:
            c.execute("""
                UPDATE rfq_requests SET
                project_name = ?, project_status = ?, priority = ?, sales_engineer_presale = ?,
                sales_engineer_sales = ?, rfq_status = ?, quotation_status = ?,
                deadline = ?, note = ?
                WHERE id = ?
            """, (project_name, project_status, priority, sales_engineer_presale,
                  sales_engineer_sales, rfq_status, quotation_status, deadline, note, rfq_id))
            conn.commit()
            flash('RFQ updated successfully!', 'success')
            
            # Send notifications
            try:
                actor_id = session.get('user_id')
                actor_name = session.get('username', 'User')
                
                # Build recipient list: admins + presales + sales engineer
                recipients = []
                
                # Add General Managers and Technical Team Leaders
                recipients.extend(notification_service.get_admin_recipients())
                
                # Add all Presale Engineers
                recipients.extend(notification_service.get_presale_recipients())
                
                # Add the Sales Engineer from the RFQ
                if sales_engineer_sales:
                    sales_eng_user_id = notification_service.get_user_id_by_username(sales_engineer_sales)
                    if sales_eng_user_id and sales_eng_user_id not in recipients:
                        recipients.append(sales_eng_user_id)
                
                # Remove duplicates and current user
                recipients = list(set(recipients))
                if actor_id in recipients:
                    recipients.remove(actor_id)
                
                # Send notification with updated data
                if recipients:
                    notification_service.notify_activity(
                        event_code='rfq.updated',
                        recipient_ids=recipients,
                        actor_id=actor_id,
                        context={
                            'actor_name': actor_name,
                            'rfq_reference': rfq_reference,
                            'project_name': project_name,
                            'priority': priority,
                            'rfq_status': rfq_status,
                            'quotation_status': quotation_status,
                            'presale_engineer': sales_engineer_presale,
                            'sales_engineer': sales_engineer_sales,
                            'deadline': deadline
                        },
                        url=url_for('rfq_summary')
                    )
            except Exception as e:
                print(f"Notification error: {e}")
                
        except Exception as e:
            flash(f'An error occurred: {e}', 'danger')
        finally:
            conn.close()

        return redirect(url_for('rfq_summary'))

    # For a GET request, fetch the data needed for the form
    c.execute("SELECT * FROM rfq_requests WHERE id = ?", (rfq_id,))
    rfq = c.fetchone()

    # Fetch data for dropdowns
    c.execute("SELECT project_name FROM register_project")
    projects = c.fetchall()
    c.execute("SELECT username FROM engineers WHERE role IN ('Presale Engineer', 'Technical Team Leader')")
    presale_engineers = c.fetchall()
    c.execute("SELECT username FROM engineers WHERE role IN ('Sales Engineer', 'Technical Team Leader')")
    sales_engineers = c.fetchall()
    conn.close()

    # Define options for static dropdowns
    priority_options = ['High', 'Medium', 'Low']
    rfq_status_options = ['Queue', 'Studying', 'Pricing', 'Quoted', 'Cancelled']

    # --- THIS IS THE UPDATED LIST ---
    quotation_status_options = [
        'Queue','Quotation Sent', 'Customer Pending', 'Technical Discussion',
        'Negotiation', 'Closed Won', 'Closed Lost', 'Cancelled'
    ]

    if rfq is None:
        flash('RFQ not found!', 'danger')
        return redirect(url_for('rfq_summary'))

    return render_template('edit_rfq.html',
                           rfq=rfq,
                           projects=projects,
                           presale_engineers=presale_engineers,
                           sales_engineers=sales_engineers,
                           priority_options=priority_options,
                           rfq_status_options=rfq_status_options,
                           quotation_status_options=quotation_status_options)

####
@app.route('/rfq_pipeline')
@permission_required('view_rfts')
def rfq_pipeline():
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    # Define the order of statuses for the pipeline columns
    status_ordered = ['Queue', 'Studying', 'Pricing', 'Quoted', 'Cancelled']

    # Initialize a dictionary to hold the structured pipeline data
    pipeline_data = {status: [] for status in status_ordered}

    # Query to get all RFQs ordered by newest first
    c.execute("SELECT * FROM rfq_requests ORDER BY requested_time DESC")
    rfqs = c.fetchall()
    conn.close()

    # Process and group the RFQs by their rfq_status
    for rfq in rfqs:
        status = rfq['rfq_status']
        if status in pipeline_data:
            pipeline_data[status].append(rfq)

    return render_template('rfq_pipeline.html',
                           pipeline_data=pipeline_data,
                           status_ordered=status_ordered)


@app.route('/update_rfq_status', methods=['POST'])
@role_required('Technical Team Leader', 'editor', 'Presale Engineer') # Define roles that can update
def update_rfq_status():
    data = request.json
    rfq_id = data.get('rfq_id')
    new_status = data.get('new_status')

    if not rfq_id or not new_status:
        return jsonify({'status': 'error', 'message': 'Missing data'}), 400

    try:
        conn = sqlite3.connect('ProjectStatus.db')
        c = conn.cursor()
        c.execute("UPDATE rfq_requests SET rfq_status = ? WHERE id = ?", (new_status, rfq_id))
        conn.commit()
        conn.close()
        return jsonify({'status': 'success', 'message': 'RFQ status updated successfully'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

######3
from datetime import datetime

@app.route('/request_technical_support', methods=['GET', 'POST'])
@login_required
def request_technical_support():
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    # Fetch presale engineers
    c.execute("SELECT username FROM engineers WHERE role='Presale Engineer'")
    presale_engineers = [engineer[0] for engineer in c.fetchall()]  # Extract only names

    # Fetch sales engineers or team leaders
    c.execute("SELECT username FROM engineers WHERE role IN ('Sales Engineer', 'Technical Team Leader','General Manager','Project Manager','Implementation Engineer')")
    sales_engineers = [engineer[0] for engineer in c.fetchall()]  # Extract only names

    if request.method == 'POST':
        # Process the form data
        request_type = request.form['request_type']
        project_name = request.form['project_name']
        priority = request.form['priority']
        presale_engineer = request.form['presale_engineer']
        sales_engineer = request.form['sales_engineer']
        request_status = request.form['request_status']
        request_result = request.form['request_result']
        deadline = request.form['deadline']

        # Automatically set Requested Time to the current time
        requested_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Generate automatic RFTS Reference
        c.execute("SELECT COUNT(*) FROM technical_support_requests")
        rfts_reference = f"RFTS-{c.fetchone()[0] + 1}"  # Incremental reference

        # Insert data into the database
        c.execute('''INSERT INTO technical_support_requests 
                     (rfts_reference, request_type, project_name, priority, 
                      presale_engineer, sales_engineer, request_status, 
                      request_result, deadline, note, requested_time) 
                     VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                  (rfts_reference, request_type, project_name, priority,
                   presale_engineer, sales_engineer, request_status,
                   request_result, deadline, request.form['note'], requested_time))

        conn.commit()
        conn.close()
        flash('Technical support request submitted successfully!', 'success')
        return redirect(url_for('technical_support_summary'))

    # Fetching options for the dropdowns
    c.execute("SELECT DISTINCT project_name FROM register_project")
    project_names = [row[0] for row in c.fetchall()]

    return render_template('request_technical_support.html',
                           project_names=project_names,
                           presale_engineers=presale_engineers,
                           sales_engineers=sales_engineers)
###########
@app.route('/technical_support_summary', methods=['GET'])
@login_required
def technical_support_summary():
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()

    # Get filter parameters from URL
    filters = {
        'request_type': request.args.get('request_type'),
        'request_status': request.args.get('request_status'),
        'priority': request.args.get('priority'),
        'presale_engineer': request.args.get('presale_engineer'),
        'sales_engineer': request.args.get('sales_engineer'),
    }

    # Build the query for requests
    query = "SELECT * FROM technical_support_requests WHERE 1=1"
    params = []
    # (The logic for adding filter conditions to the query remains the same)
    if filters['request_type']: query += " AND request_type = ?"; params.append(filters['request_type'])
    if filters['request_status']: query += " AND request_status = ?"; params.append(filters['request_status'])
    if filters['priority']: query += " AND priority = ?"; params.append(filters['priority'])
    if filters['presale_engineer']: query += " AND presale_engineer = ?"; params.append(filters['presale_engineer'])
    if filters['sales_engineer']: query += " AND sales_engineer = ?"; params.append(filters['sales_engineer'])

    c.execute(query, params)
    requests = c.fetchall()
    total_requests = len(requests)  # Calculate total number of requests

    # Prepare data for the pie chart
    c.execute("SELECT request_status, COUNT(*) FROM technical_support_requests GROUP BY request_status")
    status_counts_raw = c.fetchall()

    # Fetch distinct values for filtering
    c.execute("SELECT DISTINCT request_type FROM technical_support_requests")
    request_types = [row[0] for row in c.fetchall()]
    c.execute("SELECT DISTINCT request_status FROM technical_support_requests")
    request_statuses = [row[0] for row in c.fetchall()]
    c.execute("SELECT DISTINCT priority FROM technical_support_requests")
    priorities = [row[0] for row in c.fetchall()]
    c.execute("SELECT DISTINCT presale_engineer FROM technical_support_requests")
    presale_engineers = [row[0] for row in c.fetchall()]
    c.execute("SELECT DISTINCT sales_engineer FROM technical_support_requests")
    sales_engineers = [row[0] for row in c.fetchall()]

    conn.close()

    status_labels = [row[0] for row in status_counts_raw]
    status_counts = [row[1] for row in status_counts_raw]

    return render_template('technical_support_summary.html',
                           requests=requests,
                           total_requests=total_requests,  # Pass total to template
                           request_types=request_types,
                           request_statuses=request_statuses,
                           priorities=priorities,
                           presale_engineers=presale_engineers,
                           sales_engineers=sales_engineers,
                           status_labels=status_labels,
                           status_counts=status_counts)
######
@app.route('/download_filtered_technical_support', methods=['GET'])
#@role_required('editor')
def download_filtered_technical_support():
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()

    # Reconstruct the filtering logic based on the parameters passed from the request
    filters = {
        'presale_engineer': request.args.get('presale_engineer'),
        'sales_engineer': request.args.get('sales_engineer'),
        'request_type': request.args.get('request_type'),
        'request_status': request.args.get('request_status'),
        'priority': request.args.get('priority'),
    }

    query = "SELECT * FROM technical_support_requests WHERE 1=1"
    params = []

    # Append conditions based on the filters
    if filters['presale_engineer']:
        query += " AND presale_engineer = ?"
        params.append(filters['presale_engineer'])
    if filters['sales_engineer']:
        query += " AND sales_engineer = ?"
        params.append(filters['sales_engineer'])
    if filters['request_type']:
        query += " AND request_type = ?"
        params.append(filters['request_type'])
    if filters['request_status']:
        query += " AND request_status = ?"
        params.append(filters['request_status'])
    if filters['priority']:
        query += " AND priority = ?"
        params.append(filters['priority'])

    # Execute the filtered query
    c.execute(query, params)
    requests = c.fetchall()

    # Create a DataFrame from the requests
    df = pd.DataFrame(requests, columns=['id', 'rfts_reference', 'request_type', 'project_name',
                                         'priority', 'presale_engineer', 'sales_engineer',
                                         'request_status', 'request_result', 'deadline',
                                         'note', 'requested_time'])

    # Create an Excel file
    output = BytesIO()
    df.to_excel(output, index=False, sheet_name='Filtered Requests')  # Shortened name
    output.seek(0)

    # Send the file to the user
    return send_file(output, download_name='filtered_technical_support_requests.xlsx', as_attachment=True)
#########33
@app.route('/edit_request', methods=['GET', 'POST'])
@login_required
def edit_request():
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()

    if request.method == 'POST':
        request_id = request.form['request_id']
        new_status = request.form['request_status']
        new_result = request.form['request_result']

        # Update the request in the database
        c.execute("UPDATE technical_support_requests SET request_status = ?, request_result = ? WHERE id = ?",
                  (new_status, new_result, request_id))
        conn.commit()
        conn.close()
        return redirect(url_for('technical_support_summary'))

    # If GET, retrieve the request details to edit
    request_id = request.args.get('request_id')
    c.execute("SELECT * FROM technical_support_requests WHERE id = ?", (request_id,))
    request_details = c.fetchone()

    # Define selectable options
    request_status_options = ["Queue", "Studying", "Done", "Cancelled"]  # Adjust as needed
    request_result_options = ["Queue","Approved", "Dis Approved", "Cancelled"]  # Adjust as needed

    conn.close()

    return render_template('edit_request.html', request_details=request_details,
                           request_status_options=request_status_options,
                           request_result_options=request_result_options)
#######3
@app.route('/registration_page')
@permission_required('view_registration_hub')
def registration_page():
    return render_template('registration_page.html')
#######
@app.route('/view_summary')
#@role_required('editor')
def view_summary():
    return render_template('view_summary.html')
#######3
@app.route('/vd_page')
#@role_required('editor')
def vd_page():
    return render_template('vd_page.html')
####
@app.route('/documents_page')
#@role_required('editor')
def documents_page():
    return render_template('documents_page.html')
######333
@app.route('/view_consultants')
@permission_required('view_consultants')
def view_consultants():
    """
    View consultants page
    Sales Engineers see only their assigned consultants
    Admins (GM, TTL, Presale) see all consultants
    """
    search_query = request.args.get('search_query', '')

    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    
    # Get user role and engineer ID for filtering
    c.execute("SELECT role FROM users WHERE id = ?", (session['user_id'],))
    user_role_result = c.fetchone()
    user_role = user_role_result[0] if user_role_result else None
    
    # Get engineer ID if user is a Sales Engineer
    engineer_id = None
    if user_role == 'Sales Engineer':
        c.execute("SELECT id FROM engineers WHERE username = ?", (session['username'],))
        engineer_id_result = c.fetchone()
        if engineer_id_result:
            engineer_id = engineer_id_result[0]
    
    # Get all sales engineers for assignment dropdown
    c.execute("SELECT id, username FROM engineers WHERE role IN ('Sales Engineer', 'Technical Team Leader')")
    sales_engineers = c.fetchall()

    if search_query:
        # If a search query exists, filter the results by name or contact person
        query_param = f'%{search_query}%'
        if engineer_id:
            # Sales Engineer: only their assigned consultants
            c.execute("""
                SELECT con.*, e.username as assigned_engineer_name 
                FROM consultants con
                LEFT JOIN engineers e ON con.assigned_sales_engineer_id = e.id
                WHERE con.assigned_sales_engineer_id = ?
                AND (con.name LIKE ? OR con.contact_person LIKE ?)
            """, (engineer_id, query_param, query_param))
        else:
            # Admin: all consultants
            c.execute("""
                SELECT con.*, e.username as assigned_engineer_name 
                FROM consultants con
                LEFT JOIN engineers e ON con.assigned_sales_engineer_id = e.id
                WHERE con.name LIKE ? OR con.contact_person LIKE ?
            """, (query_param, query_param))
    else:
        # If there is no search query, fetch consultants
        if engineer_id:
            # Sales Engineer: only their assigned consultants
            c.execute("""
                SELECT con.*, e.username as assigned_engineer_name 
                FROM consultants con
                LEFT JOIN engineers e ON con.assigned_sales_engineer_id = e.id
                WHERE con.assigned_sales_engineer_id = ?
            """, (engineer_id,))
        else:
            # Admin: all consultants
            c.execute("""
                SELECT con.*, e.username as assigned_engineer_name 
                FROM consultants con
                LEFT JOIN engineers e ON con.assigned_sales_engineer_id = e.id
            """)

    consultants = c.fetchall()
    
    # Get project count for each consultant
    consultants_with_counts = []
    for consultant in consultants:
        c.execute("SELECT COUNT(*) FROM register_project WHERE consultant_id = ?", (consultant[0],))
        project_count = c.fetchone()[0]
        consultants_with_counts.append(consultant + (project_count,))
    
    conn.close()
    
    # Sort consultants by project count (descending - most projects first)
    # Index 9 is project_count (after id[0], name[1], contact[2], phone[3], email[4], note[5], is_client[6], assigned_id[7], assigned_name[8])
    consultants_with_counts.sort(key=lambda x: x[9], reverse=True)

    # Pass the list of consultants (with project counts), sales engineers, and the search query to the template
    return render_template('view_consultants.html', 
                         consultants=consultants_with_counts, 
                         sales_engineers=sales_engineers,
                         search_query=search_query)
##########33
@app.route('/consultant_projects/<int:consultant_id>')
@login_required
def consultant_projects(consultant_id):
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    
    # Get consultant details
    c.execute('SELECT id, name FROM consultants WHERE id = ?', (consultant_id,))
    consultant = c.fetchone()
    
    if not consultant:
        conn.close()
        flash('Consultant not found.', 'danger')
        return redirect(url_for('view_consultants'))
    
    # Get all projects for this consultant with additional details
    c.execute("""
        SELECT 
            rp.id, rp.project_name, eu.name AS end_user, rp.stage,
            rp.deal_value, rp.probability, rp.expected_close_date,
            en.username AS sales_engineer, rp.registered_date,
            rp.scope_of_work
        FROM register_project rp
        LEFT JOIN end_users eu ON rp.end_user_id = eu.id
        LEFT JOIN engineers en ON rp.sales_engineer_id = en.id
        WHERE rp.consultant_id = ?
        ORDER BY rp.registered_date DESC
    """, (consultant_id,))
    
    projects = c.fetchall()
    
    # Calculate summary statistics
    total_projects = len(projects)
    total_value = sum(p[4] if p[4] else 0 for p in projects)
    
    # Count by stage
    stage_counts = {}
    for project in projects:
        stage = project[3] if project[3] else 'Unknown'
        stage_counts[stage] = stage_counts.get(stage, 0) + 1
    
    conn.close()
    
    return render_template('consultant_projects.html', 
                         consultant=consultant, 
                         projects=projects,
                         total_projects=total_projects,
                         total_value=total_value,
                         stage_counts=stage_counts)

##########33
@app.route('/view_all_clients')
@login_required
def view_all_clients():
    """
    View all entities marked as clients (End Users, Contractors, Consultants)
    Sales Engineers see only their assigned clients
    Admins (GM, TTL, Presale) see all clients
    """
    search_query = request.args.get('search_query', '')
    
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    
    # Get user role and engineer ID for filtering
    c.execute("SELECT role FROM users WHERE id = ?", (session['user_id'],))
    user_role_result = c.fetchone()
    user_role = user_role_result[0] if user_role_result else None
    
    # Get engineer ID if user is a Sales Engineer
    engineer_id = None
    if user_role == 'Sales Engineer':
        c.execute("SELECT id FROM engineers WHERE username = ?", (session['username'],))
        engineer_id_result = c.fetchone()
        if engineer_id_result:
            engineer_id = engineer_id_result[0]
    
    # Get all sales engineers for assignment dropdown
    c.execute("SELECT id, username FROM engineers WHERE role IN ('Sales Engineer', 'Technical Team Leader')")
    sales_engineers = c.fetchall()
    
    all_clients = []
    
    # Fetch End Users marked as clients
    if search_query:
        query_param = f'%{search_query}%'
        if engineer_id:
            # Sales Engineer: only their assigned clients
            c.execute("""
                SELECT eu.*, e.username as assigned_engineer_name 
                FROM end_users eu
                LEFT JOIN engineers e ON eu.assigned_sales_engineer_id = e.id
                WHERE eu.is_client = 1 
                AND eu.assigned_sales_engineer_id = ?
                AND (eu.name LIKE ? OR eu.contact_person LIKE ?)
            """, (engineer_id, query_param, query_param))
        else:
            # Admin: all clients
            c.execute("""
                SELECT eu.*, e.username as assigned_engineer_name 
                FROM end_users eu
                LEFT JOIN engineers e ON eu.assigned_sales_engineer_id = e.id
                WHERE eu.is_client = 1 AND (eu.name LIKE ? OR eu.contact_person LIKE ?)
            """, (query_param, query_param))
    else:
        if engineer_id:
            # Sales Engineer: only their assigned clients
            c.execute("""
                SELECT eu.*, e.username as assigned_engineer_name 
                FROM end_users eu
                LEFT JOIN engineers e ON eu.assigned_sales_engineer_id = e.id
                WHERE eu.is_client = 1 AND eu.assigned_sales_engineer_id = ?
            """, (engineer_id,))
        else:
            # Admin: all clients
            c.execute("""
                SELECT eu.*, e.username as assigned_engineer_name 
                FROM end_users eu
                LEFT JOIN engineers e ON eu.assigned_sales_engineer_id = e.id
                WHERE eu.is_client = 1
            """)
    
    end_users = c.fetchall()
    for eu in end_users:
        # Count projects
        c.execute("SELECT COUNT(*) FROM register_project WHERE end_user_id = ?", (eu[0],))
        project_count = c.fetchone()[0]
        
        # Count contractors
        c.execute("""
            SELECT COUNT(DISTINCT contractor_id) 
            FROM register_project 
            WHERE end_user_id = ? AND contractor_id IS NOT NULL
        """, (eu[0],))
        contractor_count = c.fetchone()[0]
        
        all_clients.append({
            'id': eu[0],
            'name': eu[1],
            'type': 'End User',
            'contact_person': eu[2],
            'phone': eu[3],
            'email': eu[4],
            'note': eu[5],
            'assigned_engineer': eu[8] if len(eu) > 8 else None,
            'assigned_engineer_id': eu[7],
            'project_count': project_count,
            'contractor_count': contractor_count
        })
    
    # Fetch Contractors marked as clients
    if search_query:
        query_param = f'%{search_query}%'
        if engineer_id:
            # Sales Engineer: only their assigned clients
            c.execute("""
                SELECT c.*, e.username as assigned_engineer_name 
                FROM contractors c
                LEFT JOIN engineers e ON c.assigned_sales_engineer_id = e.id
                WHERE c.is_client = 1 
                AND c.assigned_sales_engineer_id = ?
                AND (c.name LIKE ? OR c.contact_person LIKE ?)
            """, (engineer_id, query_param, query_param))
        else:
            # Admin: all clients
            c.execute("""
                SELECT c.*, e.username as assigned_engineer_name 
                FROM contractors c
                LEFT JOIN engineers e ON c.assigned_sales_engineer_id = e.id
                WHERE c.is_client = 1 AND (c.name LIKE ? OR c.contact_person LIKE ?)
            """, (query_param, query_param))
    else:
        if engineer_id:
            # Sales Engineer: only their assigned clients
            c.execute("""
                SELECT c.*, e.username as assigned_engineer_name 
                FROM contractors c
                LEFT JOIN engineers e ON c.assigned_sales_engineer_id = e.id
                WHERE c.is_client = 1 AND c.assigned_sales_engineer_id = ?
            """, (engineer_id,))
        else:
            # Admin: all clients
            c.execute("""
                SELECT c.*, e.username as assigned_engineer_name 
                FROM contractors c
                LEFT JOIN engineers e ON c.assigned_sales_engineer_id = e.id
                WHERE c.is_client = 1
            """)
    
    contractors = c.fetchall()
    for con in contractors:
        # Count projects
        c.execute("SELECT COUNT(*) FROM register_project WHERE contractor_id = ?", (con[0],))
        project_count = c.fetchone()[0]
        
        all_clients.append({
            'id': con[0],
            'name': con[1],
            'type': 'Contractor',
            'contact_person': con[2],
            'phone': con[3],
            'email': con[4],
            'note': con[5],
            'assigned_engineer': con[8] if len(con) > 8 else None,
            'assigned_engineer_id': con[7],
            'project_count': project_count,
            'contractor_count': None
        })
    
    # Fetch Consultants marked as clients
    if search_query:
        query_param = f'%{search_query}%'
        if engineer_id:
            # Sales Engineer: only their assigned clients
            c.execute("""
                SELECT c.*, e.username as assigned_engineer_name 
                FROM consultants c
                LEFT JOIN engineers e ON c.assigned_sales_engineer_id = e.id
                WHERE c.is_client = 1 
                AND c.assigned_sales_engineer_id = ?
                AND (c.name LIKE ? OR c.contact_person LIKE ?)
            """, (engineer_id, query_param, query_param))
        else:
            # Admin: all clients
            c.execute("""
                SELECT c.*, e.username as assigned_engineer_name 
                FROM consultants c
                LEFT JOIN engineers e ON c.assigned_sales_engineer_id = e.id
                WHERE c.is_client = 1 AND (c.name LIKE ? OR c.contact_person LIKE ?)
            """, (query_param, query_param))
    else:
        if engineer_id:
            # Sales Engineer: only their assigned clients
            c.execute("""
                SELECT c.*, e.username as assigned_engineer_name 
                FROM consultants c
                LEFT JOIN engineers e ON c.assigned_sales_engineer_id = e.id
                WHERE c.is_client = 1 AND c.assigned_sales_engineer_id = ?
            """, (engineer_id,))
        else:
            # Admin: all clients
            c.execute("""
                SELECT c.*, e.username as assigned_engineer_name 
                FROM consultants c
                LEFT JOIN engineers e ON c.assigned_sales_engineer_id = e.id
                WHERE c.is_client = 1
            """)
    
    consultants = c.fetchall()
    for cons in consultants:
        # Count projects
        c.execute("SELECT COUNT(*) FROM register_project WHERE consultant_id = ?", (cons[0],))
        project_count = c.fetchone()[0]
        
        all_clients.append({
            'id': cons[0],
            'name': cons[1],
            'type': 'Consultant',
            'contact_person': cons[2],
            'phone': cons[3],
            'email': cons[4],
            'note': cons[5],
            'assigned_engineer': cons[8] if len(cons) > 8 else None,
            'assigned_engineer_id': cons[7],
            'project_count': project_count,
            'contractor_count': None
        })
    
    conn.close()
    
    # Sort by name
    all_clients.sort(key=lambda x: x['name'])
    
    return render_template('view_all_clients.html',
                           clients=all_clients,
                           sales_engineers=sales_engineers,
                           search_query=search_query)

##########33
@app.route('/view_contractors')
@permission_required('view_contractors')
def view_contractors():
    """
    View contractors page
    Sales Engineers see only their assigned contractors
    Admins (GM, TTL, Presale) see all contractors
    """
    search_query = request.args.get('search_query', '')

    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    
    # Get user role and engineer ID for filtering
    c.execute("SELECT role FROM users WHERE id = ?", (session['user_id'],))
    user_role_result = c.fetchone()
    user_role = user_role_result[0] if user_role_result else None
    
    # Get engineer ID if user is a Sales Engineer
    engineer_id = None
    if user_role == 'Sales Engineer':
        c.execute("SELECT id FROM engineers WHERE username = ?", (session['username'],))
        engineer_id_result = c.fetchone()
        if engineer_id_result:
            engineer_id = engineer_id_result[0]
    
    # Get all sales engineers for assignment dropdown
    c.execute("SELECT id, username FROM engineers WHERE role IN ('Sales Engineer', 'Technical Team Leader')")
    sales_engineers = c.fetchall()

    if search_query:
        # If a search query exists, filter the results
        query_param = f'%{search_query}%'
        if engineer_id:
            # Sales Engineer: only their assigned contractors
            c.execute("""
                SELECT c.*, e.username as assigned_engineer_name 
                FROM contractors c
                LEFT JOIN engineers e ON c.assigned_sales_engineer_id = e.id
                WHERE c.assigned_sales_engineer_id = ?
                AND (c.name LIKE ? OR c.contact_person LIKE ?)
            """, (engineer_id, query_param, query_param))
        else:
            # Admin: all contractors
            c.execute("""
                SELECT c.*, e.username as assigned_engineer_name 
                FROM contractors c
                LEFT JOIN engineers e ON c.assigned_sales_engineer_id = e.id
                WHERE c.name LIKE ? OR c.contact_person LIKE ?
            """, (query_param, query_param))
    else:
        # If there is no search query, fetch contractors
        if engineer_id:
            # Sales Engineer: only their assigned contractors
            c.execute("""
                SELECT c.*, e.username as assigned_engineer_name 
                FROM contractors c
                LEFT JOIN engineers e ON c.assigned_sales_engineer_id = e.id
                WHERE c.assigned_sales_engineer_id = ?
            """, (engineer_id,))
        else:
            # Admin: all contractors
            c.execute("""
                SELECT c.*, e.username as assigned_engineer_name 
                FROM contractors c
                LEFT JOIN engineers e ON c.assigned_sales_engineer_id = e.id
            """)

    contractors = c.fetchall()
    
    # Get project count for each contractor
    contractors_with_counts = []
    for contractor in contractors:
        c.execute("SELECT COUNT(*) FROM register_project WHERE contractor_id = ?", (contractor[0],))
        project_count = c.fetchone()[0]
        contractors_with_counts.append(contractor + (project_count,))
    
    conn.close()
    
    # Sort contractors by project count (descending - most projects first)
    # Index 9 is project_count (after id[0], name[1], contact[2], phone[3], email[4], note[5], is_client[6], assigned_id[7], assigned_name[8])
    contractors_with_counts.sort(key=lambda x: x[9], reverse=True)

    # Pass the list of contractors (with project counts), sales engineers, and the search query back to the template
    return render_template('view_contractors.html', 
                         contractors=contractors_with_counts, 
                         sales_engineers=sales_engineers,
                         search_query=search_query)
#############
@app.route('/contractor_projects/<int:contractor_id>')
@login_required
def contractor_projects(contractor_id):
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    
    # Get contractor details
    c.execute('SELECT id, name FROM contractors WHERE id = ?', (contractor_id,))
    contractor = c.fetchone()
    
    if not contractor:
        conn.close()
        flash('Contractor not found.', 'danger')
        return redirect(url_for('view_contractors'))
    
    # Get all projects for this contractor with additional details
    c.execute("""
        SELECT 
            rp.id, rp.project_name, eu.name AS end_user, rp.stage,
            rp.deal_value, rp.probability, rp.expected_close_date,
            en.username AS sales_engineer, rp.registered_date,
            rp.scope_of_work
        FROM register_project rp
        LEFT JOIN end_users eu ON rp.end_user_id = eu.id
        LEFT JOIN engineers en ON rp.sales_engineer_id = en.id
        WHERE rp.contractor_id = ?
        ORDER BY rp.registered_date DESC
    """, (contractor_id,))
    
    projects = c.fetchall()
    
    # Calculate summary statistics
    total_projects = len(projects)
    total_value = sum(p[4] if p[4] else 0 for p in projects)
    
    # Count by stage
    stage_counts = {}
    for project in projects:
        stage = project[3] if project[3] else 'Unknown'
        stage_counts[stage] = stage_counts.get(stage, 0) + 1
    
    conn.close()
    
    return render_template('contractor_projects.html', 
                         contractor=contractor, 
                         projects=projects,
                         total_projects=total_projects,
                         total_value=total_value,
                         stage_counts=stage_counts)

#############
@app.route('/edit_consultant/<int:consultant_id>', methods=['GET', 'POST'])
#@role_required('editor')
def edit_consultant(consultant_id):
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()

    if request.method == 'POST':
        # Get form data and update the consultant in the database
        name = request.form['name']
        contact_person = request.form['contact_person']
        phone = request.form['phone']
        email = request.form['email']
        note = request.form['note']

        c.execute('''
            UPDATE consultants
            SET name = ?, contact_person = ?, phone = ?, email = ?, note = ?
            WHERE id = ?
        ''', (name, contact_person, phone, email, note, consultant_id))

        conn.commit()
        conn.close()
        return redirect(url_for('view_consultants'))

    # Fetch existing consultant data for pre-filling the form
    c.execute('SELECT * FROM consultants WHERE id = ?', (consultant_id,))
    consultant = c.fetchone()
    conn.close()

    return render_template('edit_consultant.html', consultant=consultant)
############3
@app.route('/edit_contractor/<int:contractor_id>', methods=['GET', 'POST'])
#@role_required('editor')
def edit_contractor(contractor_id):
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()

    if request.method == 'POST':
        # Get form data and update the contractor in the database
        name = request.form['name']
        contact_person = request.form['contact_person']
        phone = request.form['phone']
        email = request.form['email']
        note = request.form['note']

        c.execute('''
            UPDATE contractors
            SET name = ?, contact_person = ?, phone = ?, email = ?, note = ?
            WHERE id = ?
        ''', (name, contact_person, phone, email, note, contractor_id))

        conn.commit()
        conn.close()
        return redirect(url_for('view_contractors'))

    # Fetch existing contractor data for pre-filling the form
    c.execute('SELECT * FROM contractors WHERE id = ?', (contractor_id,))
    contractor = c.fetchone()
    conn.close()

    return render_template('edit_contractor.html', contractor=contractor)
#########
@app.route('/view_projects')
@login_required
def view_projects():
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()

    # Get filter values
    sales_engineer_filter = request.args.get('sales_engineer_id')
    start_date_filter = request.args.get('start_date')
    end_date_filter = request.args.get('end_date')

    # Fetch all sales engineers for the filter dropdown
    c.execute("SELECT id, username FROM engineers WHERE role IN ('Sales Engineer', 'Technical Team Leader')")
    sales_engineers = c.fetchall()

    # Base query updated to fetch contractor and consultant names - only show approved projects
    query = '''
        SELECT 
            rp.id, rp.project_name, eu.name AS end_user, rp.stage,
            rp.deal_value, rp.probability, rp.expected_close_date,
            en.username AS sales_engineer, rp.registered_date,
            rp.scope_of_work, rp.note,
            co.name as contractor, cn.name as consultant, rp.client_type
        FROM register_project rp
        LEFT JOIN end_users eu ON rp.end_user_id = eu.id
        LEFT JOIN contractors co ON rp.contractor_id = co.id
        LEFT JOIN consultants cn ON rp.consultant_id = cn.id
        LEFT JOIN engineers en ON rp.sales_engineer_id = en.id
        WHERE rp.approval_status = 'Approved'
    '''
    params = []
    user_role = None

    # --- NEW ACCESS CONTROL LOGIC ---
    if 'user_id' in session:
        c.execute("SELECT role FROM users WHERE id = ?", (session['user_id'],))
        user_role_result = c.fetchone()
        if user_role_result:
            user_role = user_role_result[0]
            if user_role == 'Sales Engineer':
                # If user is a Sales Engineer, find their engineer ID
                c.execute("SELECT id FROM engineers WHERE username = ?", (session['username'],))
                engineer_id_result = c.fetchone()
                if engineer_id_result:
                    # Filter the main query by their ID
                    query += " AND rp.sales_engineer_id = ?"
                    params.append(engineer_id_result[0])

    # Add other filters if they exist
    if sales_engineer_filter and user_role != 'Sales Engineer':
        query += " AND rp.sales_engineer_id = ?"
        params.append(sales_engineer_filter)
    if start_date_filter:
        query += " AND date(rp.registered_date) >= ?"
        params.append(start_date_filter)
    if end_date_filter:
        query += " AND date(rp.registered_date) <= ?"
        params.append(end_date_filter)

    # Order by registered date - newest first
    query += " ORDER BY rp.registered_date DESC"
    
    c.execute(query, params)
    projects = c.fetchall()
    conn.close()

    return render_template('view_projects.html',
                           projects=projects,
                           sales_engineers=sales_engineers,
                           user_role=user_role,
                           current_filters={
                               'sales_engineer_id': sales_engineer_filter,
                               'start_date': start_date_filter,
                               'end_date': end_date_filter
                           })
###############3
@app.route('/view_end_users')
@permission_required('view_end_users')
def view_end_users():
    """
    View end users page
    Sales Engineers see only their assigned end users
    Admins (GM, TTL, Presale) see all end users
    """
    search_query = request.args.get('search_query', '')

    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    
    # Get user role and engineer ID for filtering
    c.execute("SELECT role FROM users WHERE id = ?", (session['user_id'],))
    user_role_result = c.fetchone()
    user_role = user_role_result[0] if user_role_result else None
    
    # Get engineer ID if user is a Sales Engineer
    engineer_id = None
    if user_role == 'Sales Engineer':
        c.execute("SELECT id FROM engineers WHERE username = ?", (session['username'],))
        engineer_id_result = c.fetchone()
        if engineer_id_result:
            engineer_id = engineer_id_result[0]
    
    # Get all sales engineers for assignment dropdown
    c.execute("SELECT id, username FROM engineers WHERE role IN ('Sales Engineer', 'Technical Team Leader')")
    sales_engineers = c.fetchall()

    if search_query:
        # If a search query exists, filter the results by name or contact person
        query_param = f'%{search_query}%'
        if engineer_id:
            # Sales Engineer: only their assigned end users
            c.execute("""
                SELECT eu.*, e.username as assigned_engineer_name 
                FROM end_users eu
                LEFT JOIN engineers e ON eu.assigned_sales_engineer_id = e.id
                WHERE eu.assigned_sales_engineer_id = ?
                AND (eu.name LIKE ? OR eu.contact_person LIKE ?)
            """, (engineer_id, query_param, query_param))
        else:
            # Admin: all end users
            c.execute("""
                SELECT eu.*, e.username as assigned_engineer_name 
                FROM end_users eu
                LEFT JOIN engineers e ON eu.assigned_sales_engineer_id = e.id
                WHERE eu.name LIKE ? OR eu.contact_person LIKE ?
            """, (query_param, query_param))
    else:
        # If there is no search query, fetch end users
        if engineer_id:
            # Sales Engineer: only their assigned end users
            c.execute("""
                SELECT eu.*, e.username as assigned_engineer_name 
                FROM end_users eu
                LEFT JOIN engineers e ON eu.assigned_sales_engineer_id = e.id
                WHERE eu.assigned_sales_engineer_id = ?
            """, (engineer_id,))
        else:
            # Admin: all end users
            c.execute("""
                SELECT eu.*, e.username as assigned_engineer_name 
                FROM end_users eu
                LEFT JOIN engineers e ON eu.assigned_sales_engineer_id = e.id
            """)

    end_users = c.fetchall()
    
    # Get project count and contractor count for each end user
    end_users_with_counts = []
    for end_user in end_users:
        # Count projects
        c.execute("SELECT COUNT(*) FROM register_project WHERE end_user_id = ?", (end_user[0],))
        project_count = c.fetchone()[0]
        
        # Count unique contractors associated with this end user's projects
        c.execute("""
            SELECT COUNT(DISTINCT contractor_id) 
            FROM register_project 
            WHERE end_user_id = ? AND contractor_id IS NOT NULL
        """, (end_user[0],))
        contractor_count = c.fetchone()[0]
        
        end_users_with_counts.append(end_user + (project_count, contractor_count))
    
    conn.close()
    
    # Sort end users by project count (descending - most projects first)
    # Index 9 is project_count (after id[0], name[1], contact[2], phone[3], email[4], note[5], is_client[6], assigned_id[7], assigned_name[8])
    end_users_with_counts.sort(key=lambda x: x[9], reverse=True)

    # Pass the list of end users (with project counts), sales engineers, and the search query to the template
    return render_template('view_end_users.html', 
                         end_users=end_users_with_counts, 
                         sales_engineers=sales_engineers,
                         search_query=search_query)
####################
@app.route('/toggle_client_status', methods=['POST'])
@login_required
def toggle_client_status():
    # Check if user is admin
    if session.get('user_role') not in ['General Manager', 'Technical Team Leader']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    entity_type = request.form.get('entity_type')  # 'end_user', 'contractor', 'consultant'
    entity_id = request.form.get('entity_id')
    is_client = request.form.get('is_client')  # '1' or '0'
    
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    
    try:
        if entity_type == 'end_user':
            c.execute("UPDATE end_users SET is_client = ? WHERE id = ?", (is_client, entity_id))
        elif entity_type == 'contractor':
            c.execute("UPDATE contractors SET is_client = ? WHERE id = ?", (is_client, entity_id))
        elif entity_type == 'consultant':
            c.execute("UPDATE consultants SET is_client = ? WHERE id = ?", (is_client, entity_id))
        
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 500

####################
@app.route('/assign_sales_engineer', methods=['POST'])
@login_required
def assign_sales_engineer():
    # Check if user is admin
    if session.get('user_role') not in ['General Manager', 'Technical Team Leader']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    entity_type = request.form.get('entity_type')
    entity_id = request.form.get('entity_id')
    sales_engineer_id = request.form.get('sales_engineer_id')
    
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    
    try:
        if entity_type == 'end_user':
            c.execute("UPDATE end_users SET assigned_sales_engineer_id = ? WHERE id = ?", (sales_engineer_id, entity_id))
        elif entity_type == 'contractor':
            c.execute("UPDATE contractors SET assigned_sales_engineer_id = ? WHERE id = ?", (sales_engineer_id, entity_id))
        elif entity_type == 'consultant':
            c.execute("UPDATE consultants SET assigned_sales_engineer_id = ? WHERE id = ?", (sales_engineer_id, entity_id))
        
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 500

####################
@app.route('/manage_clients')
@login_required
def manage_clients():
    # Check if user is admin
    if session.get('user_role') not in ['General Manager', 'Technical Team Leader']:
        flash('Access denied. Admin privileges required.', 'danger')
        return redirect(url_for('index'))
    
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    
    # Get all engineers for assignment dropdown (Sales Engineer, Project Manager, General Manager, Technical Team Leader)
    c.execute("SELECT id, username, role FROM engineers WHERE role IN ('Sales Engineer', 'Project Manager', 'General Manager', 'Technical Team Leader') ORDER BY role, username")
    engineers = c.fetchall()
    
    # Get ALL End Users with client status and assigned sales engineer
    c.execute("""
        SELECT eu.id, eu.name, eu.contact_person, eu.phone, eu.email, 
               eu.is_client, eu.assigned_sales_engineer_id, e.username as assigned_engineer_name,
               COUNT(DISTINCT rp.id) as project_count
        FROM end_users eu
        LEFT JOIN engineers e ON eu.assigned_sales_engineer_id = e.id
        LEFT JOIN register_project rp ON eu.id = rp.end_user_id
        GROUP BY eu.id, eu.name, eu.contact_person, eu.phone, eu.email, 
                 eu.is_client, eu.assigned_sales_engineer_id, e.username
        ORDER BY eu.is_client DESC, project_count DESC
    """)
    all_end_users = c.fetchall()
    
    # Get ALL Contractors with client status and assigned sales engineer
    c.execute("""
        SELECT c.id, c.name, c.contact_person, c.phone, c.email, 
               c.is_client, c.assigned_sales_engineer_id, e.username as assigned_engineer_name,
               COUNT(DISTINCT rp.id) as project_count
        FROM contractors c
        LEFT JOIN engineers e ON c.assigned_sales_engineer_id = e.id
        LEFT JOIN register_project rp ON c.id = rp.contractor_id
        GROUP BY c.id, c.name, c.contact_person, c.phone, c.email, 
                 c.is_client, c.assigned_sales_engineer_id, e.username
        ORDER BY c.is_client DESC, project_count DESC
    """)
    all_contractors = c.fetchall()
    
    # Get ALL Consultants with client status and assigned sales engineer
    c.execute("""
        SELECT con.id, con.name, con.contact_person, con.phone, con.email, 
               con.is_client, con.assigned_sales_engineer_id, e.username as assigned_engineer_name,
               COUNT(DISTINCT rp.id) as project_count
        FROM consultants con
        LEFT JOIN engineers e ON con.assigned_sales_engineer_id = e.id
        LEFT JOIN register_project rp ON con.id = rp.consultant_id
        GROUP BY con.id, con.name, con.contact_person, con.phone, con.email, 
                 con.is_client, con.assigned_sales_engineer_id, e.username
        ORDER BY con.is_client DESC, project_count DESC
    """)
    all_consultants = c.fetchall()
    
    conn.close()
    
    # Count clients vs non-clients
    client_end_users = [eu for eu in all_end_users if eu[5] == 1]
    client_contractors = [c for c in all_contractors if c[5] == 1]
    client_consultants = [con for con in all_consultants if con[5] == 1]
    
    return render_template('manage_clients.html',
                         all_end_users=all_end_users,
                         all_contractors=all_contractors,
                         all_consultants=all_consultants,
                         client_end_users=client_end_users,
                         client_contractors=client_contractors,
                         client_consultants=client_consultants,
                         engineers=engineers)

####################
# PROJECT APPROVAL SYSTEM
####################
@app.route('/pending_project_approvals')
@login_required
def pending_project_approvals():
    # Check if user is admin
    if session.get('user_role') not in ['General Manager', 'Technical Team Leader']:
        flash('Access denied. Admin privileges required.', 'danger')
        return redirect(url_for('index'))
    
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    
    # Get all pending projects with full details
    c.execute("""
        SELECT 
            rp.id,
            rp.project_name,
            rp.stage,
            rp.deal_value,
            rp.probability,
            rp.expected_close_date,
            rp.scope_of_work,
            rp.note,
            rp.registered_date,
            eu.name as end_user_name,
            c.name as contractor_name,
            con.name as consultant_name,
            e.username as sales_engineer_name,
            reg_user.username as registered_by,
            rp.client_type
        FROM register_project rp
        LEFT JOIN end_users eu ON rp.end_user_id = eu.id
        LEFT JOIN contractors c ON rp.contractor_id = c.id
        LEFT JOIN consultants con ON rp.consultant_id = con.id
        LEFT JOIN engineers e ON rp.sales_engineer_id = e.id
        LEFT JOIN users reg_user ON rp.updated_by = reg_user.username
        WHERE rp.approval_status = 'Pending'
        ORDER BY rp.registered_date DESC
    """)
    pending_projects = c.fetchall()
    
    conn.close()
    
    return render_template('pending_project_approvals.html', pending_projects=pending_projects)

@app.route('/approve_project/<int:project_id>', methods=['POST'])
@login_required
def approve_project(project_id):
    # Check if user is admin
    if session.get('user_role') not in ['General Manager', 'Technical Team Leader']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    approval_notes = request.form.get('approval_notes', '')
    approved_by_id = session.get('user_id')
    approved_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    
    try:
        # Get project details before updating
        c.execute("""
            SELECT project_name, updated_by, sales_engineer_id
            FROM register_project
            WHERE id = ?
        """, (project_id,))
        project_data = c.fetchone()
        
        if not project_data:
            return jsonify({'success': False, 'error': 'Project not found'}), 404
        
        project_name = project_data[0]
        registered_by_username = project_data[1]
        sales_engineer_id = project_data[2]
        
        # Update project approval status
        c.execute("""
            UPDATE register_project 
            SET approval_status = 'Approved',
                approved_by_id = ?,
                approved_at = ?,
                approval_notes = ?
            WHERE id = ?
        """, (approved_by_id, approved_at, approval_notes, project_id))
        
        conn.commit()
        flash('Project approved successfully!', 'success')
        
        # Send notifications
        try:
            actor_id = session.get('user_id')
            actor_name = session.get('username', 'Admin')
            
            # Build recipient list: admins + registrar + sales engineer
            recipients = notification_service.get_admin_recipients()
            
            # Add the user who registered the project
            if registered_by_username:
                registrar_user_id = notification_service.get_user_id_by_username(registered_by_username)
                if registrar_user_id and registrar_user_id not in recipients:
                    recipients.append(registrar_user_id)
            
            # Add the sales engineer
            if sales_engineer_id:
                # Get user_id from engineers table
                c.execute("SELECT username FROM engineers WHERE id = ?", (sales_engineer_id,))
                engineer_result = c.fetchone()
                if engineer_result:
                    engineer_username = engineer_result[0]
                    engineer_user_id = notification_service.get_user_id_by_username(engineer_username)
                    if engineer_user_id and engineer_user_id not in recipients:
                        recipients.append(engineer_user_id)
            
            # Send notification
            notification_service.notify_activity(
                event_code='project.approved',
                recipient_ids=recipients,
                actor_id=actor_id,
                context={
                    'actor_name': actor_name,
                    'project_name': project_name,
                    'project_id': project_id,
                    'approval_notes': approval_notes
                },
                url=url_for('project_pipeline')
            )
        except Exception as e:
            print(f"Notification error: {e}")
        
        return jsonify({'success': True})
    except Exception as e:
        import traceback
        print(f"ERROR in approve_project: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()

@app.route('/reject_project/<int:project_id>', methods=['POST'])
@login_required
def reject_project(project_id):
    # Check if user is admin
    if session.get('user_role') not in ['General Manager', 'Technical Team Leader']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    rejection_reason = request.form.get('rejection_reason', 'No reason provided')
    approved_by_id = session.get('user_id')
    approved_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    
    try:
        # Get project details before updating
        c.execute("""
            SELECT project_name, updated_by, sales_engineer_id
            FROM register_project
            WHERE id = ?
        """, (project_id,))
        project_data = c.fetchone()
        
        if not project_data:
            return jsonify({'success': False, 'error': 'Project not found'}), 404
        
        project_name = project_data[0]
        registered_by_username = project_data[1]
        sales_engineer_id = project_data[2]
        
        # Update project rejection status - set to Approved with Cancelled stage so it appears in pipeline
        c.execute("""
            UPDATE register_project 
            SET approval_status = 'Approved',
                stage = 'Cancelled',
                approved_by_id = ?,
                approved_at = ?,
                approval_notes = ?
            WHERE id = ?
        """, (approved_by_id, approved_at, rejection_reason, project_id))
        
        conn.commit()
        flash('Project rejected and moved to Cancelled stage in pipeline.', 'warning')
        
        # Send notifications
        try:
            actor_id = session.get('user_id')
            actor_name = session.get('username', 'Admin')
            
            # Build recipient list: admins + registrar + sales engineer
            recipients = notification_service.get_admin_recipients()
            
            # Add the user who registered the project
            if registered_by_username:
                registrar_user_id = notification_service.get_user_id_by_username(registered_by_username)
                if registrar_user_id and registrar_user_id not in recipients:
                    recipients.append(registrar_user_id)
            
            # Add the sales engineer
            if sales_engineer_id:
                # Get user_id from engineers table
                c.execute("SELECT username FROM engineers WHERE id = ?", (sales_engineer_id,))
                engineer_result = c.fetchone()
                if engineer_result:
                    engineer_username = engineer_result[0]
                    engineer_user_id = notification_service.get_user_id_by_username(engineer_username)
                    if engineer_user_id and engineer_user_id not in recipients:
                        recipients.append(engineer_user_id)
            
            # Send notification
            notification_service.notify_activity(
                event_code='project.rejected',
                recipient_ids=recipients,
                actor_id=actor_id,
                context={
                    'actor_name': actor_name,
                    'project_name': project_name,
                    'project_id': project_id,
                    'rejection_reason': rejection_reason
                },
                url=url_for('view_rejection_details', project_id=project_id)
            )
        except Exception as e:
            print(f"Notification error: {e}")
        
        return jsonify({'success': True})
    except Exception as e:
        import traceback
        print(f"ERROR in reject_project: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()

@app.route('/view_rejection_details/<int:project_id>')
@login_required
def view_rejection_details(project_id):
    """
    Show rejection details for a project
    """
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    
    # Get project details including rejection reason
    c.execute("""
        SELECT 
            rp.id,
            rp.project_name,
            rp.stage,
            rp.approval_notes as rejection_reason,
            rp.approved_at as rejected_at,
            u.username as rejected_by,
            eu.name as end_user_name,
            cont.name as contractor_name,
            cons.name as consultant_name,
            e.username as sales_engineer_name
        FROM register_project rp
        LEFT JOIN users u ON rp.approved_by_id = u.id
        LEFT JOIN end_users eu ON rp.end_user_id = eu.id
        LEFT JOIN contractors cont ON rp.contractor_id = cont.id
        LEFT JOIN consultants cons ON rp.consultant_id = cons.id
        LEFT JOIN engineers e ON rp.sales_engineer_id = e.id
        WHERE rp.id = ? AND rp.stage = 'Cancelled'
    """, (project_id,))
    
    project = c.fetchone()
    conn.close()
    
    if not project:
        flash('Project not found or not rejected.', 'danger')
        return redirect(url_for('project_pipeline'))
    
    return render_template('rejection_details.html', project=project)

@app.route('/api/pending_approvals_count')
@login_required
def pending_approvals_count():
    # Only admins can check pending count
    if session.get('user_role') not in ['General Manager', 'Technical Team Leader']:
        return jsonify({'count': 0})
    
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    
    c.execute("SELECT COUNT(*) FROM register_project WHERE approval_status = 'Pending'")
    count = c.fetchone()[0]
    
    conn.close()
    
    return jsonify({'count': count})

####################
@app.route('/end_user_contractors/<int:end_user_id>')
@login_required
def end_user_contractors(end_user_id):
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    
    # Get end user details
    c.execute('SELECT id, name FROM end_users WHERE id = ?', (end_user_id,))
    end_user = c.fetchone()
    
    if not end_user:
        conn.close()
        flash('End user not found.', 'danger')
        return redirect(url_for('view_end_users'))
    
    # Get all unique contractors for this end user's projects
    c.execute("""
        SELECT DISTINCT c.id, c.name, c.contact_person, c.phone, c.email, c.note,
               COUNT(rp.id) as project_count
        FROM contractors c
        INNER JOIN register_project rp ON c.id = rp.contractor_id
        WHERE rp.end_user_id = ?
        GROUP BY c.id, c.name, c.contact_person, c.phone, c.email, c.note
        ORDER BY project_count DESC
    """, (end_user_id,))
    
    contractors = c.fetchall()
    conn.close()
    
    return render_template('end_user_contractors.html', end_user=end_user, contractors=contractors)

####################
@app.route('/end_user_projects/<int:end_user_id>')
@login_required
def end_user_projects(end_user_id):
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    
    # Get end user details
    c.execute('SELECT id, name FROM end_users WHERE id = ?', (end_user_id,))
    end_user = c.fetchone()
    
    if not end_user:
        conn.close()
        flash('End user not found.', 'danger')
        return redirect(url_for('view_end_users'))
    
    # Get all projects for this end user with additional details
    c.execute("""
        SELECT 
            rp.id,
            rp.project_name,
            rp.stage,
            rp.deal_value,
            rp.probability,
            rp.expected_close_date,
            rp.registered_date,
            e.username as sales_engineer,
            c.name as contractor,
            con.name as consultant,
            rp.scope_of_work
        FROM register_project rp
        LEFT JOIN engineers e ON rp.sales_engineer_id = e.id
        LEFT JOIN contractors c ON rp.contractor_id = c.id
        LEFT JOIN consultants con ON rp.consultant_id = con.id
        WHERE rp.end_user_id = ?
        ORDER BY rp.registered_date DESC
    """, (end_user_id,))
    
    projects = c.fetchall()
    
    # Calculate summary statistics
    total_projects = len(projects)
    total_value = sum(p[3] if p[3] else 0 for p in projects)
    
    # Count by stage
    stage_counts = {}
    for project in projects:
        stage = project[2] if project[2] else 'Unknown'
        stage_counts[stage] = stage_counts.get(stage, 0) + 1
    
    conn.close()
    
    return render_template('end_user_projects.html', 
                         end_user=end_user, 
                         projects=projects,
                         total_projects=total_projects,
                         total_value=total_value,
                         stage_counts=stage_counts)

####################
@app.route('/edit_end_user/<int:end_user_id>', methods=['GET', 'POST'])
#@role_required('editor')
def edit_end_user(end_user_id):
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()

    if request.method == 'POST':
        # Get form data and update the end user in the database
        name = request.form['name']
        contact_person = request.form['contact_person']
        phone = request.form['phone']
        email = request.form['email']
        note = request.form['note']

        c.execute('''
            UPDATE end_users
            SET name = ?, contact_person = ?, phone = ?, email = ?, note = ?
            WHERE id = ?
        ''', (name, contact_person, phone, email, note, end_user_id))

        conn.commit()
        conn.close()
        return redirect(url_for('view_end_users'))

    # Fetch existing end user data for pre-filling the form
    c.execute('SELECT * FROM end_users WHERE id = ?', (end_user_id,))
    end_user = c.fetchone()
    conn.close()

    return render_template('edit_end_user.html', end_user=end_user)
######3
@app.route('/crm_page')
#@role_required('editor')
def crm_page():
    return render_template('crm_page.html')
####################
@app.route('/view_engineers')
@role_required('editor', 'General Manager', 'Technical Team Leader')  # Adjust roles as needed
def view_engineers():
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()

    # Fetch all engineers from the engineers table
    c.execute("SELECT id, name, phone, email, role, username FROM engineers")
    engineers = c.fetchall()
    conn.close()

    return render_template('view_engineers.html', engineers=engineers)
####################
@app.route('/delete_engineer/<int:engineer_id>', methods=['POST'])
@role_required('editor', 'General Manager', 'Technical Team Leader')  # Adjust roles as needed
def delete_engineer(engineer_id):
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()

    try:
        # First, delete the engineer from the users table
        c.execute("DELETE FROM users WHERE username IN (SELECT username FROM engineers WHERE id=?)", (engineer_id,))
        # Then, delete the engineer from the engineers table
        c.execute("DELETE FROM engineers WHERE id=?", (engineer_id,))
        conn.commit()
        flash('Engineer deleted successfully!', 'success')
    except Exception as e:
        flash(f'Error deleting engineer: {str(e)}', 'danger')
    finally:
        conn.close()

    return redirect(url_for('view_engineers'))
############################
@app.route('/edit_engineer/<int:engineer_id>', methods=['GET', 'POST'])
@role_required('editor', 'General Manager', 'Technical Team Leader')
def edit_engineer(engineer_id):
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()

    if request.method == 'POST':
        # First, get the engineer's original username to find them in the users table
        c.execute("SELECT username FROM engineers WHERE id = ?", (engineer_id,))
        old_username_result = c.fetchone()
        if not old_username_result:
            flash('Engineer not found!', 'danger')
            conn.close()
            return redirect(url_for('view_engineers'))
        old_username = old_username_result[0]

        # Get updated data from the form
        name = request.form['name']
        phone = request.form['phone']
        email = request.form['email']
        role = request.form['role']
        new_username = request.form['username']
        new_password = request.form['password']  # This can be empty

        try:
            # --- Update the engineers table ---
            if new_password:
                # If a new password is provided, update it
                hashed_password = generate_password_hash(new_password)
                c.execute('''UPDATE engineers SET name=?, phone=?, email=?, role=?, username=?, password=? 
                             WHERE id=?''',
                          (name, phone, email, role, new_username, hashed_password, engineer_id))
            else:
                # If password field is blank, keep the old password (don't update it)
                c.execute('''UPDATE engineers SET name=?, phone=?, email=?, role=?, username=? 
                             WHERE id=?''',
                          (name, phone, email, role, new_username, engineer_id))

            # --- Update the corresponding users table record ---
            if new_password:
                # If a new password is provided, update it here as well
                hashed_password = generate_password_hash(new_password)
                c.execute('''UPDATE users SET username=?, password=?, role=?
                             WHERE username=?''',
                          (new_username, hashed_password, role, old_username))
            else:
                # If password field is blank, keep the old password
                c.execute('''UPDATE users SET username=?, role=?
                             WHERE username=?''',
                          (new_username, role, old_username))

            conn.commit()
            flash('Engineer and user profiles updated successfully!', 'success')

        except sqlite3.IntegrityError:
            flash('Error: That username is already in use by another user.', 'danger')
        except Exception as e:
            flash(f'An error occurred: {e}', 'danger')
        finally:
            conn.close()

        return redirect(url_for('view_engineers'))

    # --- GET request logic (to show the form) ---
    c.execute('SELECT * FROM engineers WHERE id=?', (engineer_id,))
    engineer = c.fetchone()
    conn.close()

    return render_template('edit_engineer.html', engineer=engineer)
##########
@app.route('/register_po', methods=['GET', 'POST'])
#@role_required('editor', 'General Manager', 'Technical Team Leader','Project Coordinator')  # Adjust roles as needed
def register_po():
    if request.method == 'POST':
        # Collect form data from the submitted request
        po_request_number = request.form['po_request_number']
        project_name = request.form['project_name']
        system = request.form['system']
        presale_engineer = request.form['presale_engineer']
        project_manager = request.form['project_manager']
        
        # Process the uploaded files first
        quotation = request.files['quotation']
        po_document = request.files.get('po_document')
        quotation_data = quotation.read()
        po_document_data = po_document.read() if po_document else None
        
        distributor_engineer = request.form['distributor_engineer']
        distributor_contact = request.form['distributor_contact']
        distributor_email = request.form['distributor_email']
        po_number = request.form['po_number']
        total_amount = request.form['total_amount']
        po_approval_status = request.form['po_approval_status']
        po_delivery_status = request.form['po_delivery_status']
        po_notes_vendor = request.form['po_notes_vendor']
        po_notes_client = request.form['po_notes_client']

        # Store the data in the database
        conn = sqlite3.connect('ProjectStatus.db')
        conn.row_factory = sqlite3.Row
        c = conn.cursor()
        
        # Convert vendor and distributor IDs to names
        vendor_id = request.form.get('vendor') or None
        distributor_id = request.form['distributor']
        
        # Look up vendor name from ID
        if vendor_id:
            c.execute("SELECT name FROM vendors WHERE id = ?", (vendor_id,))
            vendor_result = c.fetchone()
            vendor = vendor_result['name'] if vendor_result else None
        else:
            vendor = None
        
        # Look up distributor name from ID
        c.execute("SELECT name FROM distributors WHERE id = ?", (distributor_id,))
        dist_result = c.fetchone()
        if not dist_result:
            flash('Invalid distributor selected!', 'danger')
            conn.close()
            return redirect(url_for('register_po'))
        distributor = dist_result['name']
        try:
            c.execute('''INSERT INTO purchase_orders (
                po_request_number, project_name, system, presale_engineer, project_manager,
                vendor, distributor, distributor_engineer, distributor_contact, distributor_email,
                quotation, po_document, po_number, total_amount, po_approval_status,
                po_delivery_status, po_notes_vendor, po_notes_client
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''', (
                po_request_number, project_name, system, presale_engineer, project_manager,
                vendor, distributor, distributor_engineer, distributor_contact, distributor_email,
                quotation_data, po_document_data, po_number, total_amount, po_approval_status,
                po_delivery_status, po_notes_vendor, po_notes_client
            ))
            po_id = c.lastrowid
            conn.commit()
            flash('Purchase Order registered successfully!', 'success')
            
            # Send notifications to stakeholders
            try:
                actor_id = session.get('user_id')
                actor_name = session.get('username', 'Unknown')
                
                # Notify presale engineer, project manager, and admins
                recipients = []
                presale_user_id = notification_service.get_user_id_by_username(presale_engineer)
                pm_user_id = notification_service.get_user_id_by_username(project_manager)
                if presale_user_id:
                    recipients.append(presale_user_id)
                if pm_user_id:
                    recipients.append(pm_user_id)
                recipients.extend(notification_service.get_admin_recipients())
                recipients = list(set(recipients))
                
                notification_service.notify_activity(
                    event_code='po.created',
                    recipient_ids=recipients,
                    actor_id=actor_id,
                    context={
                        'actor_name': actor_name,
                        'po_number': po_number,
                        'project_name': project_name,
                        'total_amount': total_amount
                    },
                    url=url_for('view_po_status')
                )
            except Exception as e:
                print(f"Notification error: {e}")
            
            return redirect(url_for('register_po'))  # Redirect to the view page for POs
        except sqlite3.IntegrityError as e:
            if 'UNIQUE constraint failed: purchase_orders.po_number' in str(e):
                flash('Error: PO Number must be unique!', 'danger')
            else:
                flash('An error occurred while registering the PO.', 'danger')
            return redirect(url_for('register_po'))  # Redirect back to the form for correction
        finally:
            conn.close()

    # Handle GET request: Fetch data for dropdowns
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    c.execute("SELECT id, project_name FROM register_project")
    projects = c.fetchall()
    c.execute("SELECT id, name FROM engineers WHERE role IN ('Presale Engineer', 'Technical Team Leader')")
    presale_engineers = c.fetchall()
    c.execute("SELECT id, name FROM engineers WHERE role IN ('Implementation Engineer', 'Project Manager')")
    project_managers = c.fetchall()
    c.execute("SELECT id, name FROM vendors ORDER BY name")
    vendors = c.fetchall()
    c.execute("SELECT id, name FROM distributors")
    distributors = c.fetchall()
    conn.close()

    return render_template('register_po.html', projects=projects, presale_engineers=presale_engineers, project_managers=project_managers, vendors=vendors, distributors=distributors)
#######################333
@app.route('/view_po_status', methods=['GET', 'POST'])
@permission_required('view_po_status')
def view_po_status():
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row  # Enable named field access
    c = conn.cursor()

    # Fetch distinct presale engineers and Technical Team Leaders using JOIN on username
    c.execute('''
        SELECT DISTINCT e.username
        FROM purchase_orders po
        LEFT JOIN engineers e ON CAST(po.presale_engineer AS TEXT) = e.username
        WHERE e.username IS NOT NULL
        ORDER BY e.username
    ''')
    presale_engineers = c.fetchall()

    # Fetch distinct project managers using JOIN on username
    c.execute('''
        SELECT DISTINCT e.username
        FROM purchase_orders po
        LEFT JOIN engineers e ON CAST(po.project_manager AS TEXT) = e.username
        WHERE e.username IS NOT NULL
        ORDER BY e.username
    ''')
    project_managers = c.fetchall()

    # Fetch distinct distributors using JOIN on names
    c.execute('''
        SELECT DISTINCT d.name
        FROM purchase_orders po
        LEFT JOIN distributors d ON CAST(po.distributor AS TEXT) = d.name
        WHERE d.name IS NOT NULL
        ORDER BY d.name
    ''')
    distributors = c.fetchall()
    
    # Fetch distinct vendors using JOIN on names
    c.execute('''
        SELECT DISTINCT v.name
        FROM purchase_orders po
        LEFT JOIN vendors v ON CAST(po.vendor AS TEXT) = v.name
        WHERE v.name IS NOT NULL
        ORDER BY v.name
    ''')
    vendors = c.fetchall()
    
    # Fetch distinct project names
    c.execute('''
        SELECT DISTINCT rp.id, rp.project_name
        FROM purchase_orders po
        JOIN register_project rp ON po.project_name = rp.id
    ''')
    projects = c.fetchall()
    
    # Fetch ALL vendors and distributors for inline editing dropdowns
    c.execute('SELECT id, name FROM vendors ORDER BY name')
    all_vendors = c.fetchall()
    
    c.execute('SELECT id, name FROM distributors ORDER BY name')
    all_distributors = c.fetchall()

    # Initialize filters
    presale_engineer_filter = request.form.get('presale_engineer', '')
    project_manager_filter = request.form.get('project_manager', '')
    distributor_filter = request.form.get('distributor', '')
    vendor_filter = request.form.get('vendor', '')
    po_delivery_status_filter = request.form.get('po_delivery_status', '')
    po_approval_status_filter = request.form.get('po_approval_status', '')
    project_name_filter = request.form.get('project_name', '')

    # Calculate statistics with filters applied - JOIN on names/usernames not IDs
    stats_query = '''
    SELECT 
        COUNT(*) as total_pos,
        SUM(CASE WHEN po.po_approval_status = 'Approved' THEN 1 ELSE 0 END) as approved_count,
        SUM(CASE WHEN po.po_approval_status = 'Pending' THEN 1 ELSE 0 END) as pending_count,
        SUM(CASE WHEN po.po_approval_status = 'Rejected' THEN 1 ELSE 0 END) as rejected_count,
        SUM(CASE WHEN po.po_delivery_status LIKE '%Delivered%' THEN 1 ELSE 0 END) as delivered_count,
        SUM(CASE WHEN po.po_delivery_status = 'Order Placed' THEN 1 ELSE 0 END) as in_progress_count,
        COALESCE(SUM(po.total_amount), 0) as total_amount,
        COALESCE(AVG(po.total_amount), 0) as avg_amount
    FROM purchase_orders po
    LEFT JOIN register_project rp ON po.project_name = rp.id
    LEFT JOIN distributors d ON CAST(po.distributor AS TEXT) = d.name
    LEFT JOIN vendors v ON CAST(po.vendor AS TEXT) = v.name
    LEFT JOIN engineers eng ON CAST(po.presale_engineer AS TEXT) = eng.username
    LEFT JOIN engineers pmeng ON CAST(po.project_manager AS TEXT) = pmeng.username
    WHERE 1=1
    '''
    stats_params = []
    
    if presale_engineer_filter:
        stats_query += " AND eng.username = ?"
        stats_params.append(presale_engineer_filter)
    
    if project_manager_filter:
        stats_query += " AND pmeng.username = ?"
        stats_params.append(project_manager_filter)
    
    if distributor_filter:
        stats_query += " AND d.name = ?"
        stats_params.append(distributor_filter)
    
    if vendor_filter:
        stats_query += " AND v.name = ?"
        stats_params.append(vendor_filter)
    
    if po_delivery_status_filter:
        stats_query += " AND po.po_delivery_status = ?"
        stats_params.append(po_delivery_status_filter)
    
    if po_approval_status_filter:
        stats_query += " AND po.po_approval_status = ?"
        stats_params.append(po_approval_status_filter)
        
    if project_name_filter:
        stats_query += " AND rp.id = ?"
        stats_params.append(project_name_filter)
    
    c.execute(stats_query, stats_params)
    stats = c.fetchone()

    # Build the main query with vendor information - JOIN on names/usernames not IDs
    query = '''
    SELECT 
        po.id AS po_id,
        po.po_request_number, 
        COALESCE(rp.project_name, 'Project ID: ' || po.project_name) AS project_name,
        d.id AS distributor_id,
        COALESCE(d.name, po.distributor) AS distributor_name,
        v.id AS vendor_id,
        COALESCE(v.name, po.vendor) AS vendor_name,
        po.po_approval_status, 
        po.po_delivery_status, 
        COALESCE(eng.username, po.presale_engineer) AS presale_engineer_name,
        COALESCE(pmeng.username, po.project_manager) AS project_manager_name,
        po.po_number,
        po.po_notes_vendor,
        po.po_notes_client,
        po.total_amount,
        po.system,
        po.project_name AS project_id,
        po.presale_engineer AS presale_engineer_id,
        po.project_manager AS project_manager_id,
        po.distributor AS distributor_raw_id
    FROM purchase_orders po
    LEFT JOIN register_project rp ON po.project_name = rp.id
    LEFT JOIN distributors d ON CAST(po.distributor AS TEXT) = d.name
    LEFT JOIN vendors v ON CAST(po.vendor AS TEXT) = v.name
    LEFT JOIN engineers eng ON CAST(po.presale_engineer AS TEXT) = eng.username
    LEFT JOIN engineers pmeng ON CAST(po.project_manager AS TEXT) = pmeng.username
    WHERE 1=1
    '''
    params = []

    if presale_engineer_filter:
        query += " AND eng.username = ?"
        params.append(presale_engineer_filter)

    if project_manager_filter:
        query += " AND pmeng.username = ?"
        params.append(project_manager_filter)

    if distributor_filter:
        query += " AND d.name = ?"
        params.append(distributor_filter)
    
    if vendor_filter:
        query += " AND v.name = ?"
        params.append(vendor_filter)

    if po_delivery_status_filter:
        query += " AND po.po_delivery_status = ?"
        params.append(po_delivery_status_filter)

    if po_approval_status_filter:
        query += " AND po.po_approval_status = ?"
        params.append(po_approval_status_filter)
        
    if project_name_filter:
        query += " AND rp.id = ?"
        params.append(project_name_filter)
    
    query += " ORDER BY po.id DESC"

    c.execute(query, params)
    purchase_orders = c.fetchall()
    conn.close()

    return render_template('view_po_status.html',
                           purchase_orders=purchase_orders,
                           presale_engineers=presale_engineers,
                           project_managers=project_managers,
                           distributors=distributors,
                           vendors=vendors,
                           projects=projects,
                           all_vendors=all_vendors,
                           all_distributors=all_distributors,
                           stats=stats,
                           project_name_filter=project_name_filter,
                           presale_engineer_filter=presale_engineer_filter,
                           project_manager_filter=project_manager_filter,
                           distributor_filter=distributor_filter,
                           vendor_filter=vendor_filter,
                           po_delivery_status_filter=po_delivery_status_filter,
                           po_approval_status_filter=po_approval_status_filter)

@app.route('/delete_po/<int:po_id>', methods=['POST'])
@login_required
@role_required('General Manager', 'Technical Team Leader')
def delete_po(po_id):
    """Delete Purchase Order and all associated data (Admin only)"""
    try:
        conn = sqlite3.connect('ProjectStatus.db')
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        # Get PO details for logging
        cursor.execute("SELECT po_request_number, po_number FROM purchase_orders WHERE id = ?", (po_id,))
        po_data = cursor.fetchone()
        
        if not po_data:
            return jsonify({'success': False, 'message': 'Purchase Order not found'}), 404
        
        po_request_number = po_data['po_request_number']
        po_number = po_data['po_number']
        
        # Delete all related data in correct order (foreign key constraints)
        # 1. Delete delivery notes
        cursor.execute("DELETE FROM delivery_notes WHERE po_number = ?", (po_number,))
        
        # 2. Delete PO items
        cursor.execute("DELETE FROM po_items WHERE po_number = ?", (po_number,))
        
        # 3. Delete supplier quotations associated with this PO
        cursor.execute("DELETE FROM supplier_quotations WHERE po_request_number = ?", (po_request_number,))
        
        # 4. Finally delete the purchase order itself
        cursor.execute("DELETE FROM purchase_orders WHERE id = ?", (po_id,))
        
        conn.commit()
        conn.close()
        
        flash(f'Purchase Order {po_request_number} and all associated data deleted successfully!', 'success')
        return jsonify({'success': True, 'message': 'Purchase Order deleted successfully'})
        
    except Exception as e:
        if conn:
            conn.close()
        return jsonify({'success': False, 'message': f'Error deleting PO: {str(e)}'}), 500

######################
# RFQ Comments Routes
######################
@app.route('/add_rfq_comment/<int:rfq_id>', methods=['POST'])
@login_required
def add_rfq_comment(rfq_id):
    comment_text = request.form.get('comment', '').strip()
    
    if not comment_text:
        flash('Comment cannot be empty!', 'danger')
        return redirect(url_for('rfq_summary'))
    
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    
    try:
        # Get RFQ details for notification
        c.execute("SELECT rfq_reference, project_name FROM rfq_requests WHERE id = ?", (rfq_id,))
        rfq_data = c.fetchone()
        rfq_reference = rfq_data[0] if rfq_data else 'Unknown'
        project_name = rfq_data[1] if rfq_data else 'Unknown'
        
        # Insert comment
        c.execute('''
            INSERT INTO rfq_comments (rfq_id, user_id, username, comment, created_at)
            VALUES (?, ?, ?, ?, datetime('now', 'localtime'))
        ''', (rfq_id, session['user_id'], session['username'], comment_text))
        conn.commit()
        flash('Comment added successfully!', 'success')
        
        # Send notifications to all presale engineers and technical team leaders
        try:
            actor_id = session.get('user_id')
            actor_name = session.get('username', 'User')
            
            # Build recipient list: all presales + technical team leaders
            recipients = []
            
            # Add all Presale Engineers
            recipients.extend(notification_service.get_presale_recipients())
            
            # Add all Technical Team Leaders and General Managers
            recipients.extend(notification_service.get_admin_recipients())
            
            # Remove duplicates and current user
            recipients = list(set(recipients))
            if actor_id in recipients:
                recipients.remove(actor_id)
            
            # Send notification
            if recipients:
                notification_service.notify_activity(
                    event_code='rfq.comment_added',
                    recipient_ids=recipients,
                    actor_id=actor_id,
                    context={
                        'actor_name': actor_name,
                        'rfq_reference': rfq_reference,
                        'project_name': project_name,
                        'comment_preview': comment_text[:100]
                    },
                    url=url_for('rfq_summary')
                )
        except Exception as e:
            print(f"Notification error: {e}")
            
    except Exception as e:
        flash(f'Error adding comment: {str(e)}', 'danger')
    finally:
        conn.close()
    
    return redirect(url_for('rfq_summary'))

@app.route('/get_rfq_comments/<int:rfq_id>', methods=['GET'])
@login_required
def get_rfq_comments(rfq_id):
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    
    c.execute('''
        SELECT id, username, comment, created_at
        FROM rfq_comments
        WHERE rfq_id = ?
        ORDER BY created_at DESC
    ''', (rfq_id,))
    
    comments = c.fetchall()
    conn.close()
    
    comments_list = [{
        'id': c[0],
        'username': c[1],
        'comment': c[2],
        'created_at': c[3]
    } for c in comments]
    
    return jsonify({'comments': comments_list})

######################
# PO Comments Routes
######################
@app.route('/add_po_comment/<int:po_id>', methods=['POST'])
@login_required
def add_po_comment(po_id):
    comment_text = request.form.get('comment', '').strip()
    
    if not comment_text:
        flash('Comment cannot be empty!', 'danger')
        return redirect(url_for('view_po_status'))
    
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    
    try:
        c.execute('''
            INSERT INTO po_comments (po_id, user_id, username, comment, created_at)
            VALUES (?, ?, ?, ?, datetime('now', 'localtime'))
        ''', (po_id, session['user_id'], session['username'], comment_text))
        conn.commit()
        flash('Comment added successfully!', 'success')
    except Exception as e:
        flash(f'Error adding comment: {str(e)}', 'danger')
    finally:
        conn.close()
    
    return redirect(url_for('view_po_status'))

@app.route('/get_po_comments/<int:po_id>', methods=['GET'])
@login_required
def get_po_comments(po_id):
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    
    c.execute('''
        SELECT id, username, comment, created_at
        FROM po_comments
        WHERE po_id = ?
        ORDER BY created_at DESC
    ''', (po_id,))
    
    comments = c.fetchall()
    conn.close()
    
    comments_list = [{
        'id': c[0],
        'username': c[1],
        'comment': c[2],
        'created_at': c[3]
    } for c in comments]
    
    return jsonify({'comments': comments_list})

######################
@app.route('/download_filtered_po_excel', methods=['GET'])
#@role_required('editor', 'General Manager', 'Technical Team Leader','Project Coordinator')  # Adjust roles as needed
def download_filtered_po_excel():
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()

    # Initialize filter parameters
    presale_engineer_filter = request.args.get('presale_engineer', '')
    project_manager_filter = request.args.get('project_manager', '')
    distributor_filter = request.args.get('distributor', '')
    po_delivery_status_filter = request.args.get('po_delivery_status', '')
    po_approval_status_filter = request.args.get('po_approval_status', '')

    # Build the query based on filters
    query = '''
        SELECT 
            po.po_request_number,
            po.po_number, 
            rp.project_name, 
            po.system,
            po.total_amount,
            d.name AS distributor_name, 
            po.distributor_engineer,
            po.distributor_contact,
            po.distributor_email,
            po.po_approval_status, 
            po.po_delivery_status, 
            eng.username AS presale_engineer_name,
            pmeng.username AS project_manager_name, 
            po.po_notes_vendor,
            po.po_notes_client
        FROM purchase_orders po
        JOIN register_project rp ON po.project_name = rp.id
        JOIN distributors d ON po.distributor = d.id
        JOIN engineers eng ON po.presale_engineer = eng.id
        JOIN engineers pmeng ON po.project_manager = pmeng.id
        WHERE 1=1
        '''
    params = []

    if presale_engineer_filter:
        query += " AND presale_engineer_name = ?"
        params.append(presale_engineer_filter)

    if project_manager_filter:
        query += " AND project_manager_name = ?"
        params.append(project_manager_filter)

    if distributor_filter:
        query += " AND distributor_name = ?"
        params.append(distributor_filter)

    if po_delivery_status_filter:
        query += " AND po.po_delivery_status = ?"
        params.append(po_delivery_status_filter)

    if po_approval_status_filter:
        query += " AND po.po_approval_status = ?"
        params.append(po_approval_status_filter)

    # Execute the query to fetch filtered purchase orders
    c.execute(query, params)
    purchase_orders = c.fetchall()
    conn.close()

    # Create a DataFrame from the results
    df = pd.DataFrame(purchase_orders, columns=[
        'PO Request Number', 'PO Number','Project Name','System','PO Total Amount SAR', 'Distributor','Distributor Engineer','Disriributer_contact',
        'Distributor Email','PO Approval Status', 'PO Delivery Status', 'Presale Engineer',
        'Project Manager', 'Vendor Notes', 'Client Notes'
    ])

    # Create an Excel file
    output = BytesIO()
    df.to_excel(output, index=False, sheet_name='Filtered Purchase Orders')
    output.seek(0)

    # Send the file to the user
    return send_file(output, download_name='filtered_purchase_orders.xlsx', as_attachment=True)

##########################
# Download Vendor Purchase Orders as Excel
@app.route('/download_vendor_pos_excel/<int:vendor_id>', methods=['GET'])
@login_required
@permission_required('view_vendors')
def download_vendor_pos_excel(vendor_id):
    """Export all purchase orders for a specific vendor to Excel"""
    conn = sqlite3.connect('ProjectStatus.db')
    
    # Get vendor name for filename
    cursor = conn.cursor()
    cursor.execute("SELECT name FROM vendors WHERE id = ?", (vendor_id,))
    vendor_result = cursor.fetchone()
    vendor_name = vendor_result[0] if vendor_result else "Vendor"
    
    # Query to fetch purchase orders through vendor-distributor relationships
    query = '''
        SELECT DISTINCT
            po.po_request_number AS "PO Request Number",
            po.po_number AS "PO Number",
            rp.project_name AS "Project Name",
            po.system AS "System/Scope",
            po.total_amount AS "Total Amount (SAR)",
            d.name AS "Distributor",
            po.distributor_engineer AS "Distributor Engineer",
            po.distributor_contact AS "Distributor Contact",
            po.distributor_email AS "Distributor Email",
            po.po_approval_status AS "Approval Status",
            po.po_delivery_status AS "Delivery Status",
            eng.username AS "Presale Engineer",
            pmeng.username AS "Project Manager",
            po.created_at AS "Created Date",
            po.po_notes_vendor AS "Vendor Notes",
            po.po_notes_client AS "Client Notes"
        FROM purchase_orders po
        LEFT JOIN register_project rp ON po.project_name = rp.id
        LEFT JOIN distributors d ON CAST(po.distributor AS TEXT) = CAST(d.id AS TEXT)
        LEFT JOIN engineers eng ON po.presale_engineer = eng.id
        LEFT JOIN engineers pmeng ON po.project_manager = pmeng.id
        LEFT JOIN vendor_distributor vd ON CAST(po.distributor AS TEXT) = CAST(vd.distributor_id AS TEXT)
        WHERE CAST(po.vendor AS TEXT) = CAST(? AS TEXT) OR vd.vendor_id = ?
        ORDER BY po.created_at DESC
    '''
    
    df = pd.read_sql_query(query, conn, params=(vendor_id, vendor_id))
    conn.close()
    
    # Create an Excel file
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Purchase Orders')
        
        # Auto-adjust column widths
        worksheet = writer.sheets['Purchase Orders']
        for idx, col in enumerate(df.columns):
            max_len = max(df[col].astype(str).apply(len).max(), len(col)) + 2
            worksheet.set_column(idx, idx, min(max_len, 50))
    
    output.seek(0)
    
    # Clean filename
    safe_vendor_name = "".join(c for c in vendor_name if c.isalnum() or c in (' ', '_')).strip()
    filename = f'{safe_vendor_name}_Purchase_Orders.xlsx'
    
    return send_file(output, download_name=filename, as_attachment=True)


# Download Distributor Purchase Orders as Excel
@app.route('/download_distributor_pos_excel/<int:distributor_id>', methods=['GET'])
@login_required
@permission_required('view_distributors')
def download_distributor_pos_excel(distributor_id):
    """Export all purchase orders for a specific distributor to Excel"""
    conn = sqlite3.connect('ProjectStatus.db')
    
    # Get distributor name for filename
    cursor = conn.cursor()
    cursor.execute("SELECT name FROM distributors WHERE id = ?", (distributor_id,))
    distributor_result = cursor.fetchone()
    distributor_name = distributor_result[0] if distributor_result else "Distributor"
    
    # Query to fetch purchase orders for this distributor
    query = '''
        SELECT
            po.po_request_number AS "PO Request Number",
            po.po_number AS "PO Number",
            rp.project_name AS "Project Name",
            po.system AS "System/Scope",
            po.total_amount AS "Total Amount (SAR)",
            po.distributor_engineer AS "Distributor Engineer",
            po.distributor_contact AS "Distributor Contact",
            po.distributor_email AS "Distributor Email",
            po.po_approval_status AS "Approval Status",
            po.po_delivery_status AS "Delivery Status",
            eng.username AS "Presale Engineer",
            pmeng.username AS "Project Manager",
            po.created_at AS "Created Date",
            po.po_notes_vendor AS "Vendor Notes",
            po.po_notes_client AS "Client Notes"
        FROM purchase_orders po
        LEFT JOIN register_project rp ON po.project_name = rp.id
        LEFT JOIN engineers eng ON po.presale_engineer = eng.id
        LEFT JOIN engineers pmeng ON po.project_manager = pmeng.id
        WHERE CAST(po.distributor AS TEXT) = CAST(? AS TEXT)
        ORDER BY po.created_at DESC
    '''
    
    df = pd.read_sql_query(query, conn, params=(distributor_id,))
    conn.close()
    
    # Create an Excel file
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Purchase Orders')
        
        # Auto-adjust column widths
        worksheet = writer.sheets['Purchase Orders']
        for idx, col in enumerate(df.columns):
            max_len = max(df[col].astype(str).apply(len).max(), len(col)) + 2
            worksheet.set_column(idx, idx, min(max_len, 50))
    
    output.seek(0)
    
    # Clean filename
    safe_distributor_name = "".join(c for c in distributor_name if c.isalnum() or c in (' ', '_')).strip()
    filename = f'{safe_distributor_name}_Purchase_Orders.xlsx'
    
    return send_file(output, download_name=filename, as_attachment=True)

##########################
# Inline Edit PO Field
##########################
@app.route('/update_po_field', methods=['POST'])
@login_required
def update_po_field():
    """Handle inline editing of PO fields from view_po_status page"""
    try:
        data = request.get_json()
        po_id = data.get('po_id')
        field = data.get('field')
        value = data.get('value')
        
        if not po_id or not field:
            return jsonify({'success': False, 'message': 'Missing required parameters'}), 400
        
        conn = sqlite3.connect('ProjectStatus.db')
        c = conn.cursor()
        
        # Map frontend field names to database columns
        field_mapping = {
            'vendor': 'vendor',
            'distributor': 'distributor',
            'delivery_status': 'po_delivery_status'
        }
        
        if field not in field_mapping:
            return jsonify({'success': False, 'message': 'Invalid field'}), 400
        
        db_field = field_mapping[field]
        
        # Handle empty vendor (optional field)
        if field == 'vendor' and value == '':
            value = None
        
        # Validate required fields
        if field in ['distributor', 'delivery_status'] and not value:
            return jsonify({'success': False, 'message': f'{field.replace("_", " ").title()} is required'}), 400
        
        # Validate foreign keys and convert IDs to names
        if field == 'vendor' and value:
            c.execute("SELECT name FROM vendors WHERE id = ?", (value,))
            vendor_result = c.fetchone()
            if not vendor_result:
                return jsonify({'success': False, 'message': 'Invalid vendor'}), 400
            # Store the NAME not the ID
            value = vendor_result[0]
        
        if field == 'distributor':
            c.execute("SELECT name FROM distributors WHERE id = ?", (value,))
            dist_result = c.fetchone()
            if not dist_result:
                return jsonify({'success': False, 'message': 'Invalid distributor'}), 400
            # Store the NAME not the ID
            value = dist_result[0]
        
        # Update the field (value is now NAME for vendor/distributor)
        query = f"UPDATE purchase_orders SET {db_field} = ? WHERE id = ?"
        c.execute(query, (value, po_id))
        conn.commit()
        
        if c.rowcount == 0:
            conn.close()
            return jsonify({'success': False, 'message': 'Purchase order not found'}), 404
        
        conn.close()
        return jsonify({'success': True, 'message': 'Updated successfully'}), 200
        
    except Exception as e:
        print(f"Error updating PO field: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

##########################
@app.route('/edit_po/<int:po_id>', methods=['GET', 'POST'])
#@role_required('editor', 'General Manager', 'Technical Team Leader','Project Coordinator')  # Adjust roles as needed
def edit_po(po_id):
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    if request.method == 'POST':
        # Collect form data from the submitted request
        po_request_number = request.form['po_request_number']
        project_name = request.form['project_name']
        system = request.form['system']
        presale_engineer = request.form['presale_engineer']
        project_manager = request.form['project_manager']
        
        # Convert vendor and distributor IDs to names
        vendor_id = request.form.get('vendor') or None
        distributor_id = request.form['distributor']
        
        # Look up vendor name from ID
        if vendor_id:
            c.execute("SELECT name FROM vendors WHERE id = ?", (vendor_id,))
            vendor_result = c.fetchone()
            vendor = vendor_result['name'] if vendor_result else None
        else:
            vendor = None
        
        # Look up distributor name from ID
        c.execute("SELECT name FROM distributors WHERE id = ?", (distributor_id,))
        dist_result = c.fetchone()
        if not dist_result:
            flash('Invalid distributor selected!', 'danger')
            conn.close()
            return redirect(url_for('edit_po', po_id=po_id))
        distributor = dist_result['name']
        
        distributor_engineer = request.form['distributor_engineer']
        distributor_contact = request.form['distributor_contact']
        distributor_email = request.form['distributor_email']
        po_number = request.form['po_number']
        total_amount = request.form['total_amount']
        po_approval_status = request.form['po_approval_status']
        po_delivery_status = request.form['po_delivery_status']
        po_notes_vendor = request.form['po_notes_vendor']
        po_notes_client = request.form['po_notes_client']

        # Fetch the current purchase order details to retrieve existing documents
        c.execute("SELECT quotation, po_document FROM purchase_orders WHERE id=?", (po_id,))
        existing_documents = c.fetchone()
        existing_quotation = existing_documents['quotation']
        existing_po_document = existing_documents['po_document']

        # Process the uploaded files
        quotation_data = request.files['quotation'].read() if 'quotation' in request.files and request.files['quotation'].filename else existing_quotation
        po_document_data = request.files['po_document'].read() if 'po_document' in request.files and request.files['po_document'].filename else existing_po_document

        try:
            c.execute('''UPDATE purchase_orders SET
                         po_request_number=?, project_name=?, system=?, presale_engineer=?, 
                         project_manager=?, vendor=?, distributor=?, distributor_engineer=?, 
                         distributor_contact=?, distributor_email=?, quotation=?, 
                         po_document=?, po_number=?, total_amount=?, 
                         po_approval_status=?, po_delivery_status=?, 
                         po_notes_vendor=?, po_notes_client=?
                         WHERE id=?''', (
                po_request_number, project_name, system, presale_engineer,
                project_manager, vendor, distributor, distributor_engineer,
                distributor_contact, distributor_email, quotation_data,
                po_document_data, po_number, total_amount,
                po_approval_status, po_delivery_status,
                po_notes_vendor, po_notes_client, po_id))
            conn.commit()
            flash('Purchase Order updated successfully!', 'success')
            
            # Send notifications to stakeholders
            try:
                actor_id = session.get('user_id')
                actor_name = session.get('username', 'Unknown')
                
                # Notify presale engineer, project manager, and admins
                recipients = []
                presale_user_id = notification_service.get_user_id_by_username(presale_engineer)
                pm_user_id = notification_service.get_user_id_by_username(project_manager)
                if presale_user_id:
                    recipients.append(presale_user_id)
                if pm_user_id:
                    recipients.append(pm_user_id)
                recipients.extend(notification_service.get_admin_recipients())
                recipients = list(set(recipients))
                
                notification_service.notify_activity(
                    event_code='po.updated',
                    recipient_ids=recipients,
                    actor_id=actor_id,
                    context={
                        'actor_name': actor_name,
                        'po_number': po_number,
                        'project_name': project_name,
                        'total_amount': total_amount
                    },
                    url=url_for('view_po_status')
                )
            except Exception as e:
                print(f"Notification error: {e}")
            
            conn.close()
            return redirect(url_for('view_po_status'))
        except sqlite3.IntegrityError as e:
            flash('Error updating Purchase Order: {}'.format(e), 'danger')
            conn.close()
            return redirect(url_for('edit_po', po_id=po_id))

    # Fetch the current purchase order details
    c.execute("SELECT * FROM purchase_orders WHERE id=?", (po_id,))
    purchase_order = c.fetchone()

    # Fetch projects, presale engineers, project managers, distributors, and vendors
    c.execute("SELECT id, project_name FROM register_project")
    projects = c.fetchall()
    c.execute("SELECT username, name FROM engineers WHERE role IN ('Presale Engineer', 'Technical Team Leader')")
    presale_engineers = c.fetchall()
    c.execute("SELECT username, name FROM engineers WHERE role IN ('Implementation Engineer', 'Project Manager', 'Technical Team Leader')")
    project_managers = c.fetchall()
    c.execute("SELECT id, name FROM distributors")
    distributors = c.fetchall()
    c.execute("SELECT id, name FROM vendors")
    vendors = c.fetchall()
    
    conn.close()
    
    return render_template('edit_po.html', purchase_order=purchase_order,
                           projects=projects, presale_engineers=presale_engineers,
                           project_managers=project_managers, distributors=distributors,
                           vendors=vendors)
#####################
#####################
@app.route('/download_po/<po_number>', methods=['GET'], endpoint='download_po_endpoint')
@role_required('editor', 'General Manager', 'Technical Team Leader','Project Coordinator')  # Adjust roles as needed
def download_po(po_number):
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    c.execute("SELECT po_document FROM purchase_orders WHERE po_request_number=?", (po_number,))
    result = c.fetchone()
    c.execute("SELECT po_number FROM purchase_orders WHERE po_request_number=?", (po_number,))
    ponumber=c.fetchone()
    conn.close()

    if result and result[0]:
        output = BytesIO(result[0])  # Assuming PO document is stored as binary data
        return send_file(output, download_name=f'{ponumber}_PO.pdf', as_attachment=True, mimetype='application/pdf')
    else:
        flash('PO not found!', 'danger')
        return redirect(url_for('view_po_status'))

##############################3
@app.route('/download_quotation/<po_number>', methods=['GET'], endpoint='download_quotation_endpoint')
@role_required('editor', 'General Manager', 'Technical Team Leader','Project Coordinator')  # Adjust roles as needed
def download_quotation(po_number):
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    c.execute("SELECT quotation FROM purchase_orders WHERE po_request_number=?", (po_number,))
    result = c.fetchone()
    c.execute("SELECT po_number FROM purchase_orders WHERE po_request_number=?", (po_number,))
    ponumber=c.fetchone()
    conn.close()

    if result and result[0]:
        output = BytesIO(result[0])  # Assuming the quotation is stored as binary PDF data
        return send_file(output, download_name=f'{ponumber}_quotation.pdf', as_attachment=True)
    else:
        flash('Quotation not found!', 'danger')
        return redirect(url_for('view_po_status'))
#####################################################
@app.route('/view_po_details/<po_number>', methods=['GET'])
@login_required
def view_po_details(po_number):
    conn = sqlite3.connect('ProjectStatus.db')
    # Use row_factory to access columns by name
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    # Fetch all follow-up entries for the selected PO number
    c.execute("SELECT * FROM purchase_order_monitoring WHERE po_number = ? ORDER BY order_date DESC", (po_number,))
    purchase_order_details = c.fetchall()

    # Fetch the total amount from the main purchase_orders table
    c.execute("SELECT total_amount FROM purchase_orders WHERE po_number = ?", (po_number,))
    po_main = c.fetchone()
    total_amount = po_main['total_amount'] if po_main else 0

    conn.close()

    # The template will handle the case where no details are found
    return render_template('po_details.html',
                           purchase_order_details=purchase_order_details,
                           po_number=po_number,
                           total_amount=total_amount)

@app.route('/po_profile/<po_request_number>', methods=['GET'])
@login_required
def po_profile(po_request_number):
    """Comprehensive PO Profile page with delivery tracking"""
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT 
            po.*,
            rp.project_name as project_name_actual,
            d.name as distributor_name,
            v.name as vendor_name,
            e.name as presale_engineer_name,
            pm.name as project_manager_name
        FROM purchase_orders po
        LEFT JOIN register_project rp ON CAST(po.project_name AS INTEGER) = rp.id
        LEFT JOIN distributors d ON CAST(po.distributor AS INTEGER) = d.id
        LEFT JOIN vendors v ON CAST(po.vendor AS INTEGER) = v.id
        LEFT JOIN engineers e ON po.presale_engineer = e.username
        LEFT JOIN engineers pm ON po.project_manager = pm.username
        WHERE po.po_request_number = ?
    """, (po_request_number,))
    po = cursor.fetchone()
    
    if not po:
        flash('Purchase Order not found!', 'danger')
        conn.close()
        return redirect(url_for('view_po_status'))
    
    po_number = po['po_number']
    
    cursor.execute("""
        SELECT * FROM po_items 
        WHERE po_number = ? 
        ORDER BY item_number ASC
    """, (po_number,))
    po_items = cursor.fetchall()
    
    cursor.execute("""
        SELECT * FROM purchase_order_monitoring 
        WHERE (po_number = ? OR po_number = ?) AND deleted_at IS NULL
        ORDER BY order_date DESC
    """, (po_number, po_request_number))
    delivery_notes = cursor.fetchall()
    
    cursor.execute("""
        SELECT * FROM vat_invoices 
        WHERE po_request_number = ?
        ORDER BY uploaded_at DESC
    """, (po_request_number,))
    vat_invoices = cursor.fetchall()
    
    delivered_count = sum(1 for item in po_items if item['delivery_status'] == 'Delivered')
    partial_count = sum(1 for item in po_items if item['delivery_status'] == 'Partial')
    not_delivered_count = sum(1 for item in po_items if item['delivery_status'] == 'Not Delivered')
    
    conn.close()
    
    return render_template('po_profile.html',
                           po=po,
                           po_number=po_number,
                           po_request_number=po_request_number,
                           po_items=po_items,
                           delivery_notes=delivery_notes,
                           vat_invoices=vat_invoices,
                           delivered_count=delivered_count,
                           partial_count=partial_count,
                           not_delivered_count=not_delivered_count)


@app.route('/add_po_item/<po_number>', methods=['POST'])
@login_required
def add_po_item(po_number):
    """Add new item to PO with comprehensive validation"""
    conn = sqlite3.connect('ProjectStatus.db')
    cursor = conn.cursor()
    
    try:
        part_number = request.form.get('part_number', '').strip()
        
        description = request.form.get('description', '').strip()
        if not description:
            flash('Description is required!', 'danger')
            conn.close()
            return redirect(url_for('po_profile', po_number=po_number))
        
        quantity_str = request.form.get('quantity', '').strip()
        if not quantity_str:
            flash('Quantity is required!', 'danger')
            conn.close()
            return redirect(url_for('po_profile', po_number=po_number))
        
        try:
            quantity = float(quantity_str)
            if quantity <= 0:
                flash('Quantity must be greater than zero!', 'danger')
                conn.close()
                return redirect(url_for('po_profile', po_number=po_number))
        except (ValueError, TypeError):
            flash('Invalid quantity value!', 'danger')
            conn.close()
            return redirect(url_for('po_profile', po_number=po_number))
        
        unit_price_str = request.form.get('unit_price', '0').strip()
        try:
            unit_price = float(unit_price_str) if unit_price_str else 0
            if unit_price < 0:
                flash('Unit price cannot be negative!', 'danger')
                conn.close()
                return redirect(url_for('po_profile', po_number=po_number))
        except (ValueError, TypeError):
            flash('Invalid unit price value!', 'danger')
            conn.close()
            return redirect(url_for('po_profile', po_number=po_number))
        
        total_price_str = request.form.get('total_price', '0').strip()
        try:
            total_price = float(total_price_str) if total_price_str else 0
            if total_price < 0:
                flash('Total price cannot be negative!', 'danger')
                conn.close()
                return redirect(url_for('po_profile', po_number=po_number))
        except (ValueError, TypeError):
            flash('Invalid total price value!', 'danger')
            conn.close()
            return redirect(url_for('po_profile', po_number=po_number))
        
        if total_price == 0 and unit_price > 0:
            total_price = quantity * unit_price
        
        notes = request.form.get('notes', '').strip()
        
        cursor.execute("""
            SELECT COALESCE(MAX(item_number), 0) + 1 
            FROM po_items 
            WHERE po_number = ?
        """, (po_number,))
        next_item_number = cursor.fetchone()[0]
        
        cursor.execute("""
            INSERT INTO po_items (
                po_number, item_number, part_number, description, 
                quantity, unit_price, total_price, notes
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (po_number, next_item_number, part_number, description, 
              quantity, unit_price, total_price, notes))
        
        conn.commit()
        flash('Item added successfully!', 'success')
        
    except KeyError as e:
        flash(f'Missing required field: {str(e)}', 'danger')
    except Exception as e:
        flash(f'Error adding item: {str(e)}', 'danger')
    finally:
        conn.close()
    
    return redirect(url_for('po_profile', po_number=po_number))


@app.route('/edit_po_item/<int:item_id>', methods=['POST'])
@login_required
def edit_po_item(item_id):
    """Edit existing PO item with comprehensive validation"""
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    cursor.execute("SELECT po_number FROM po_items WHERE id = ?", (item_id,))
    result = cursor.fetchone()
    
    if not result:
        flash('Item not found!', 'danger')
        conn.close()
        return redirect(url_for('view_po_status'))
    
    po_number = result['po_number']
    
    try:
        part_number = request.form.get('part_number', '').strip()
        
        description = request.form.get('description', '').strip()
        if not description:
            flash('Description is required!', 'danger')
            conn.close()
            return redirect(url_for('po_profile', po_number=po_number))
        
        quantity_str = request.form.get('quantity', '').strip()
        if not quantity_str:
            flash('Quantity is required!', 'danger')
            conn.close()
            return redirect(url_for('po_profile', po_number=po_number))
        
        try:
            quantity = float(quantity_str)
            if quantity <= 0:
                flash('Quantity must be greater than zero!', 'danger')
                conn.close()
                return redirect(url_for('po_profile', po_number=po_number))
        except (ValueError, TypeError):
            flash('Invalid quantity value!', 'danger')
            conn.close()
            return redirect(url_for('po_profile', po_number=po_number))
        
        unit_price_str = request.form.get('unit_price', '0').strip()
        try:
            unit_price = float(unit_price_str) if unit_price_str else 0
            if unit_price < 0:
                flash('Unit price cannot be negative!', 'danger')
                conn.close()
                return redirect(url_for('po_profile', po_number=po_number))
        except (ValueError, TypeError):
            flash('Invalid unit price value!', 'danger')
            conn.close()
            return redirect(url_for('po_profile', po_number=po_number))
        
        quantity_delivered_str = request.form.get('quantity_delivered', '0').strip()
        try:
            quantity_delivered = float(quantity_delivered_str) if quantity_delivered_str else 0
            if quantity_delivered < 0:
                flash('Quantity delivered cannot be negative!', 'danger')
                conn.close()
                return redirect(url_for('po_profile', po_number=po_number))
        except (ValueError, TypeError):
            flash('Invalid quantity delivered value!', 'danger')
            conn.close()
            return redirect(url_for('po_profile', po_number=po_number))
        
        delivery_status = request.form.get('delivery_status', 'Not Delivered')
        if delivery_status not in ['Not Delivered', 'Partial', 'Delivered']:
            flash('Invalid delivery status!', 'danger')
            conn.close()
            return redirect(url_for('po_profile', po_number=po_number))
        
        delivery_date = request.form.get('delivery_date', '').strip()
        notes = request.form.get('notes', '').strip()
        
        total_price = quantity * unit_price if unit_price > 0 else 0
        
        if quantity_delivered >= quantity:
            delivery_status = 'Delivered'
        elif quantity_delivered > 0:
            delivery_status = 'Partial'
        else:
            delivery_status = 'Not Delivered'
        
        cursor.execute("""
            UPDATE po_items SET
                part_number = ?,
                description = ?,
                quantity = ?,
                unit_price = ?,
                total_price = ?,
                quantity_delivered = ?,
                delivery_status = ?,
                delivery_date = ?,
                notes = ?,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
        """, (part_number, description, quantity, unit_price, total_price,
              quantity_delivered, delivery_status, delivery_date or None, notes, item_id))
        
        conn.commit()
        flash('Item updated successfully!', 'success')
        
    except KeyError as e:
        flash(f'Missing required field: {str(e)}', 'danger')
    except Exception as e:
        flash(f'Error updating item: {str(e)}', 'danger')
    finally:
        conn.close()
    
    return redirect(url_for('po_profile', po_number=po_number))


@app.route('/delete_po_item/<int:item_id>', methods=['POST'])
@login_required
def delete_po_item(item_id):
    """Delete PO item"""
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    cursor.execute("SELECT po_number FROM po_items WHERE id = ?", (item_id,))
    result = cursor.fetchone()
    
    if not result:
        flash('Item not found!', 'danger')
        conn.close()
        return redirect(url_for('view_po_status'))
    
    po_number = result['po_number']
    
    cursor.execute("DELETE FROM po_items WHERE id = ?", (item_id,))
    
    conn.commit()
    conn.close()
    
    flash('Item deleted successfully!', 'success')
    return redirect(url_for('po_profile', po_number=po_number))


@app.route('/update_po_vat/<po_number>', methods=['POST'])
@login_required
def update_po_vat(po_number):
    """Update VAT percentage for PO"""
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    vat_percentage = float(request.form['vat_percentage'])
    
    cursor.execute("SELECT total_amount FROM purchase_orders WHERE po_number = ?", (po_number,))
    result = cursor.fetchone()
    
    if not result:
        flash('Purchase Order not found!', 'danger')
        conn.close()
        return redirect(url_for('view_po_status'))
    
    total_amount = result['total_amount']
    vat_amount = total_amount * (vat_percentage / 100)
    total_with_vat = total_amount + vat_amount
    
    cursor.execute("""
        UPDATE purchase_orders SET
            vat_percentage = ?,
            vat_amount = ?,
            total_with_vat = ?
        WHERE po_number = ?
    """, (vat_percentage, vat_amount, total_with_vat, po_number))
    
    conn.commit()
    conn.close()
    
    flash('VAT updated successfully!', 'success')
    return redirect(url_for('po_profile', po_number=po_number))


@app.route('/import_po_items_excel/<po_request_number>', methods=['POST'])
@login_required
def import_po_items_excel(po_request_number):
    """Import PO items from Excel file"""
    if 'excel_file' not in request.files:
        flash('No file uploaded!', 'danger')
        return redirect(url_for('po_profile', po_request_number=po_request_number))
    
    file = request.files['excel_file']
    
    if file.filename == '':
        flash('No file selected!', 'danger')
        return redirect(url_for('po_profile', po_request_number=po_request_number))
    
    if not file.filename.endswith('.xlsx'):
        flash('Invalid file format! Please upload an Excel file (.xlsx only)', 'danger')
        return redirect(url_for('po_profile', po_request_number=po_request_number))
    
    try:
        import openpyxl
        from io import BytesIO
        
        file_content = BytesIO(file.read())
        workbook = openpyxl.load_workbook(file_content)
        sheet = workbook.active
        
        conn = sqlite3.connect('ProjectStatus.db')
        cursor = conn.cursor()
        
        cursor.execute("SELECT po_number FROM purchase_orders WHERE po_request_number = ?", (po_request_number,))
        result = cursor.fetchone()
        po_number = result[0] if result and result[0] else po_request_number
        
        cursor.execute("""
            SELECT COALESCE(MAX(item_number), 0) 
            FROM po_items 
            WHERE po_number = ?
        """, (po_number,))
        next_item_number = cursor.fetchone()[0] + 1
        
        items_added = 0
        items_updated = 0
        errors = []
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                    continue
                
                part_number = str(row[0]).strip() if row[0] else ''
                description = str(row[1]).strip() if row[1] else ''
                
                if not description:
                    errors.append(f"Row {row_idx}: Description is required")
                    continue
                
                try:
                    quantity = float(row[2]) if row[2] else 0
                    if quantity <= 0:
                        errors.append(f"Row {row_idx}: Quantity must be greater than zero")
                        continue
                except (ValueError, TypeError):
                    errors.append(f"Row {row_idx}: Invalid quantity value")
                    continue
                
                try:
                    unit_price = float(row[3]) if row[3] else 0
                    if unit_price < 0:
                        errors.append(f"Row {row_idx}: Unit price cannot be negative")
                        continue
                except (ValueError, TypeError):
                    errors.append(f"Row {row_idx}: Invalid unit price value")
                    continue
                
                try:
                    total_price = float(row[4]) if row[4] else 0
                    if total_price < 0:
                        errors.append(f"Row {row_idx}: Total price cannot be negative")
                        continue
                except (ValueError, TypeError):
                    errors.append(f"Row {row_idx}: Invalid total price value")
                    continue
                
                if total_price == 0 and unit_price > 0:
                    total_price = quantity * unit_price
                
                quantity_delivered_provided = False
                quantity_delivered = 0
                
                try:
                    if len(row) > 5 and row[5] is not None and str(row[5]).strip() != '':
                        quantity_delivered = float(row[5])
                        if quantity_delivered < 0:
                            quantity_delivered = 0
                        else:
                            quantity_delivered_provided = True
                except (ValueError, TypeError, IndexError):
                    pass
                
                existing_item = None
                
                if part_number:
                    cursor.execute("""
                        SELECT id, quantity_delivered, delivery_status FROM po_items 
                        WHERE po_number = ? AND LOWER(TRIM(part_number)) = LOWER(TRIM(?))
                    """, (po_number, part_number))
                    existing_item = cursor.fetchone()
                
                if not existing_item:
                    cursor.execute("""
                        SELECT id, quantity_delivered, delivery_status FROM po_items 
                        WHERE po_number = ? AND LOWER(TRIM(description)) = LOWER(TRIM(?))
                    """, (po_number, description))
                    existing_item = cursor.fetchone()
                
                if existing_item:
                    existing_qty_delivered = existing_item[1] or 0
                    existing_delivery_status = existing_item[2] or 'Not Delivered'
                    
                    if quantity_delivered_provided:
                        if quantity_delivered >= quantity:
                            final_delivery_status = 'Delivered'
                        elif quantity_delivered > 0:
                            final_delivery_status = 'Partial'
                        else:
                            final_delivery_status = 'Not Delivered'
                        final_qty_delivered = quantity_delivered
                    else:
                        final_delivery_status = existing_delivery_status
                        final_qty_delivered = existing_qty_delivered
                    
                    cursor.execute("""
                        UPDATE po_items SET
                            part_number = ?,
                            description = ?,
                            quantity = ?,
                            unit_price = ?,
                            total_price = ?,
                            quantity_delivered = ?,
                            delivery_status = ?,
                            updated_at = CURRENT_TIMESTAMP
                        WHERE id = ?
                    """, (part_number, description, quantity, unit_price, total_price, 
                          final_qty_delivered, final_delivery_status, existing_item[0]))
                    items_updated += 1
                else:
                    if quantity_delivered_provided:
                        if quantity_delivered >= quantity:
                            new_delivery_status = 'Delivered'
                        elif quantity_delivered > 0:
                            new_delivery_status = 'Partial'
                        else:
                            new_delivery_status = 'Not Delivered'
                    else:
                        new_delivery_status = 'Not Delivered'
                        quantity_delivered = 0
                    
                    cursor.execute("""
                        INSERT INTO po_items (
                            po_number, item_number, part_number, description, 
                            quantity, unit_price, total_price, 
                            quantity_delivered, delivery_status
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (po_number, next_item_number, part_number, description, 
                          quantity, unit_price, total_price, quantity_delivered, new_delivery_status))
                    next_item_number += 1
                    items_added += 1
                
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")
        
        conn.commit()
        conn.close()
        
        success_msg = []
        if items_added > 0:
            success_msg.append(f"{items_added} item(s) added")
        if items_updated > 0:
            success_msg.append(f"{items_updated} item(s) updated")
        
        if success_msg:
            flash(f"Excel import successful! {', '.join(success_msg)}", 'success')
        
        if errors:
            flash(f"Some rows had errors: {'; '.join(errors[:5])}", 'warning')
        
    except Exception as e:
        flash(f'Error importing Excel file: {str(e)}', 'danger')
    
    return redirect(url_for('po_profile', po_request_number=po_request_number))


@app.route('/export_po_items_excel/<po_number>')
@login_required
def export_po_items_excel(po_number):
    """Export PO items to Excel file"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        
        conn = sqlite3.connect('ProjectStatus.db')
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT part_number, description, quantity, unit_price, 
                   total_price, quantity_delivered
            FROM po_items
            WHERE po_number = ?
            ORDER BY item_number
        """, (po_number,))
        items = cursor.fetchall()
        conn.close()
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "PO Items"
        
        headers = ['Part Number', 'Description', 'Qty', 'Unit Price', 'Total', 'Delivered']
        
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        for row_idx, item in enumerate(items, start=2):
            sheet.cell(row=row_idx, column=1, value=item['part_number'] or '')
            sheet.cell(row=row_idx, column=2, value=item['description'] or '')
            sheet.cell(row=row_idx, column=3, value=item['quantity'] or 0)
            sheet.cell(row=row_idx, column=4, value=item['unit_price'] or 0)
            sheet.cell(row=row_idx, column=5, value=item['total_price'] or 0)
            sheet.cell(row=row_idx, column=6, value=item['quantity_delivered'] or 0)
        
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 50
        sheet.column_dimensions['C'].width = 12
        sheet.column_dimensions['D'].width = 15
        sheet.column_dimensions['E'].width = 15
        sheet.column_dimensions['F'].width = 15
        
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'PO_{po_number}_Items.xlsx'
        )
        
    except Exception as e:
        flash(f'Error exporting to Excel: {str(e)}', 'danger')
        return redirect(url_for('po_profile', po_number=po_number))


@app.route('/upload_vat_invoice/<po_request_number>', methods=['POST'])
@login_required
def upload_vat_invoice(po_request_number):
    """Upload VAT Invoice PDF for purchase order"""
    if 'vat_invoice' not in request.files:
        flash('No file uploaded!', 'danger')
        return redirect(url_for('po_profile', po_request_number=po_request_number))
    
    file = request.files['vat_invoice']
    
    if file.filename == '':
        flash('No file selected!', 'danger')
        return redirect(url_for('po_profile', po_request_number=po_request_number))
    
    if not file.filename.endswith('.pdf'):
        flash('Invalid file format! Please upload a PDF file', 'danger')
        return redirect(url_for('po_profile', po_request_number=po_request_number))
    
    invoice_name = request.form.get('invoice_name', '').strip()
    if not invoice_name:
        invoice_name = f"VAT Invoice - {file.filename}"
    
    try:
        pdf_data = file.read()
        file_size = len(pdf_data)
        
        conn = sqlite3.connect('ProjectStatus.db')
        cursor = conn.cursor()
        
        cursor.execute("""
            INSERT INTO vat_invoices 
            (po_request_number, invoice_name, file_name, file_blob, file_size, uploaded_by_user_id, uploaded_by_username)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (po_request_number, invoice_name, file.filename, pdf_data, file_size, 
              session.get('user_id'), session.get('username')))
        
        conn.commit()
        conn.close()
        
        flash('VAT Invoice uploaded successfully!', 'success')
    except Exception as e:
        flash(f'Error uploading VAT Invoice: {str(e)}', 'danger')
    
    return redirect(url_for('po_profile', po_request_number=po_request_number))


@app.route('/download_vat_invoice/<po_request_number>')
@login_required
def download_vat_invoice(po_request_number):
    """Download VAT Invoice PDF"""
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    cursor.execute("SELECT vat_invoice_pdf, po_number FROM purchase_orders WHERE po_request_number = ?", (po_request_number,))
    result = cursor.fetchone()
    conn.close()
    
    if result and result['vat_invoice_pdf']:
        from io import BytesIO
        output = BytesIO(result['vat_invoice_pdf'])
        po_num = result['po_number'] or po_request_number
        return send_file(
            output,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'VAT_Invoice_{po_num}.pdf'
        )
    else:
        flash('VAT Invoice not found!', 'danger')
        return redirect(url_for('po_profile', po_request_number=po_request_number))


@app.route('/add_po_delivery_note/<po_request_number>', methods=['POST'])
@login_required
def add_po_delivery_note(po_request_number):
    """Add delivery note with optional PDF"""
    status = request.form.get('status', '').strip()
    expected_delivery = request.form.get('expected_delivery', '').strip()
    notes = request.form.get('notes', '').strip()
    
    if not status:
        flash('Delivery status is required!', 'danger')
        return redirect(url_for('po_profile', po_request_number=po_request_number))
    
    delivery_note_pdf = None
    if 'delivery_note_pdf' in request.files:
        file = request.files['delivery_note_pdf']
        if file and file.filename and file.filename.endswith('.pdf'):
            delivery_note_pdf = file.read()
    
    try:
        from datetime import datetime
        conn = sqlite3.connect('ProjectStatus.db')
        cursor = conn.cursor()
        
        cursor.execute("SELECT po_number FROM purchase_orders WHERE po_request_number = ?", (po_request_number,))
        result = cursor.fetchone()
        po_number = result[0] if result and result[0] else po_request_number
        
        cursor.execute("""
            INSERT INTO purchase_order_monitoring 
            (po_number, status, expected_delivery, notes, delivery_note, registration_date, order_date)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (po_number, status, expected_delivery or None, notes, delivery_note_pdf,
              datetime.now().strftime('%Y-%m-%d %H:%M:%S'), datetime.now().strftime('%Y-%m-%d')))
        
        conn.commit()
        conn.close()
        
        flash('Delivery note added successfully!', 'success')
    except Exception as e:
        flash(f'Error adding delivery note: {str(e)}', 'danger')
    
    return redirect(url_for('po_profile', po_request_number=po_request_number))


@app.route('/view_vat_invoice/<int:invoice_id>')
@login_required
def view_vat_invoice(invoice_id):
    """View a specific VAT invoice PDF"""
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    cursor.execute("SELECT * FROM vat_invoices WHERE id = ?", (invoice_id,))
    invoice = cursor.fetchone()
    
    if not invoice:
        conn.close()
        flash('Invoice not found!', 'danger')
        return redirect(url_for('view_po_status'))
    
    conn.close()
    
    if invoice['file_blob']:
        from io import BytesIO
        output = BytesIO(invoice['file_blob'])
        return send_file(
            output,
            mimetype='application/pdf',
            as_attachment=False,
            download_name=invoice['file_name']
        )
    else:
        flash('Invoice not found!', 'danger')
        return redirect(url_for('view_po_status'))


@app.route('/download_vat_invoice_file/<int:invoice_id>')
@login_required
def download_vat_invoice_file(invoice_id):
    """Download a specific VAT invoice PDF"""
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    cursor.execute("SELECT * FROM vat_invoices WHERE id = ?", (invoice_id,))
    invoice = cursor.fetchone()
    
    if not invoice:
        conn.close()
        flash('Invoice not found!', 'danger')
        return redirect(url_for('view_po_status'))
    
    conn.close()
    
    if invoice['file_blob']:
        from io import BytesIO
        output = BytesIO(invoice['file_blob'])
        return send_file(
            output,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=invoice['file_name']
        )
    else:
        flash('Invoice not found!', 'danger')
        return redirect(url_for('view_po_status'))


@app.route('/delete_vat_invoice/<int:invoice_id>', methods=['POST'])
@login_required
@role_required('Technical Team Leader', 'General Manager', 'Presale Engineer', 'Project Manager')
def delete_vat_invoice(invoice_id):
    """Delete a VAT invoice - requires authorization"""
    try:
        conn = sqlite3.connect('ProjectStatus.db')
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        cursor.execute("SELECT po_request_number, uploaded_by_user_id FROM vat_invoices WHERE id = ?", (invoice_id,))
        result = cursor.fetchone()
        
        if not result:
            conn.close()
            flash('Invoice not found!', 'danger')
            return redirect(url_for('view_po_status'))
        
        po_request_number = result['po_request_number']
        uploaded_by = result['uploaded_by_user_id']
        current_user_id = session.get('user_id')
        current_user_role = session.get('role')
        
        if uploaded_by != current_user_id and current_user_role not in ['Technical Team Leader', 'General Manager']:
            conn.close()
            flash('You do not have permission to delete this invoice!', 'danger')
            return redirect(url_for('po_profile', po_request_number=po_request_number))
        
        cursor.execute("DELETE FROM vat_invoices WHERE id = ?", (invoice_id,))
        conn.commit()
        conn.close()
        flash('VAT Invoice deleted successfully!', 'success')
        return redirect(url_for('po_profile', po_request_number=po_request_number))
    except Exception as e:
        flash(f'Error deleting invoice: {str(e)}', 'danger')
        return redirect(url_for('view_po_status'))


@app.route('/edit_delivery_note/<int:note_id>', methods=['POST'])
@login_required
def edit_delivery_note(note_id):
    """Edit a delivery note status"""
    try:
        conn = sqlite3.connect('ProjectStatus.db')
        cursor = conn.cursor()
        
        cursor.execute("SELECT po_number FROM purchase_order_monitoring WHERE id = ? AND deleted_at IS NULL", (note_id,))
        result = cursor.fetchone()
        
        if not result:
            conn.close()
            flash('Delivery note not found!', 'danger')
            return redirect(url_for('view_po_status'))
        
        po_number = result[0]
        status = request.form.get('status', '').strip()
        expected_delivery = request.form.get('expected_delivery', '').strip() or None
        notes = request.form.get('notes', '').strip()
        
        if not status:
            flash('Status is required!', 'danger')
            conn.close()
            cursor.execute("SELECT po_request_number FROM purchase_orders WHERE po_number = ? OR po_request_number = ?", (po_number, po_number))
            po_result = cursor.fetchone()
            po_request_number = po_result[0] if po_result else po_number
            return redirect(url_for('po_profile', po_request_number=po_request_number))
        
        cursor.execute("""
            UPDATE purchase_order_monitoring 
            SET status = ?, expected_delivery = ?, notes = ?
            WHERE id = ?
        """, (status, expected_delivery, notes, note_id))
        
        cursor.execute("SELECT po_request_number FROM purchase_orders WHERE po_number = ? OR po_request_number = ?", (po_number, po_number))
        po_result = cursor.fetchone()
        po_request_number = po_result[0] if po_result else po_number
        
        conn.commit()
        conn.close()
        flash('Delivery note updated successfully!', 'success')
        return redirect(url_for('po_profile', po_request_number=po_request_number))
    except Exception as e:
        flash(f'Error updating delivery note: {str(e)}', 'danger')
        return redirect(url_for('view_po_status'))


@app.route('/delete_delivery_note/<int:note_id>', methods=['POST'])
@login_required
@role_required('Technical Team Leader', 'General Manager', 'Presale Engineer', 'Project Manager')
def delete_delivery_note(note_id):
    """Soft delete a delivery note - requires authorization"""
    try:
        from datetime import datetime
        conn = sqlite3.connect('ProjectStatus.db')
        cursor = conn.cursor()
        
        cursor.execute("SELECT po_number FROM purchase_order_monitoring WHERE id = ? AND deleted_at IS NULL", (note_id,))
        result = cursor.fetchone()
        
        if not result:
            conn.close()
            flash('Delivery note not found!', 'danger')
            return redirect(url_for('view_po_status'))
        
        po_number = result[0]
        
        cursor.execute("SELECT po_request_number FROM purchase_orders WHERE po_number = ? OR po_request_number = ?", (po_number, po_number))
        po_result = cursor.fetchone()
        po_request_number = po_result[0] if po_result else po_number
        
        cursor.execute("UPDATE purchase_order_monitoring SET deleted_at = ? WHERE id = ?", 
                      (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), note_id))
        conn.commit()
        conn.close()
        flash('Delivery note deleted successfully!', 'success')
        return redirect(url_for('po_profile', po_request_number=po_request_number))
    except Exception as e:
        flash(f'Error deleting delivery note: {str(e)}', 'danger')
        return redirect(url_for('view_po_status'))


##############333
@app.route('/request_po', methods=['GET', 'POST'])
@role_required('Sales Engineer', 'Presale Engineer', 'Project Manager', 'Technical Team Leader', 'General Manager')
def request_po():
    """Request PO from a quotation - requires approval from TTL/GM"""
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    
    if request.method == 'POST':
        # Handle PO Request submission
        quote_ref = request.form['quote_ref']
        project_manager = request.form.get('project_manager', '').strip()
        project_coordinator = request.form.get('project_coordinator', '').strip()
        vendor_id = request.form.get('vendor_id') or None
        distributor_id = request.form['distributor_id']
        notes = request.form.get('notes', '')
        supplier_quotation_id = request.form.get('supplier_quotation_id') or None
        
        # SECURITY: Validate project_manager if provided
        if project_manager:
            c.execute("""SELECT username FROM engineers 
                         WHERE username = ? AND role IN ('Project Manager', 'Implementation Engineer', 'Technical Team Leader')""",
                     (project_manager,))
            pm_result = c.fetchone()
            
            if not pm_result:
                flash('Invalid project manager selected!', 'danger')
                conn.close()
                return redirect(url_for('request_po', quote_ref=quote_ref))
        else:
            project_manager = None  # Set to NULL if not provided
        
        # SECURITY: Validate project_coordinator if provided
        if project_coordinator:
            c.execute("""SELECT username FROM engineers 
                         WHERE username = ? AND role IN ('Project Manager', 'Implementation Engineer', 'Project Coordinator', 'Technical Team Leader')""",
                     (project_coordinator,))
            pc_result = c.fetchone()
            
            if not pc_result:
                flash('Invalid project coordinator selected!', 'danger')
                conn.close()
                return redirect(url_for('request_po', quote_ref=quote_ref))
        else:
            project_coordinator = None  # Set to NULL if not provided
        
        # SECURITY: Fetch all authoritative data server-side, don't trust client inputs
        # Fetch quotation details from database
        c.execute("""SELECT project_name, system, presale_eng 
                     FROM projects WHERE quote_ref = ?""", (quote_ref,))
        quote_data = c.fetchone()
        
        if not quote_data:
            flash('Invalid quotation reference!', 'danger')
            conn.close()
            return redirect(url_for('project_summary'))
        
        project_name = quote_data['project_name']
        system = quote_data['system']
        presale_engineer = quote_data['presale_eng']
        
        # Fetch distributor name from database
        c.execute("SELECT name FROM distributors WHERE id = ?", (distributor_id,))
        dist_result = c.fetchone()
        
        if not dist_result:
            flash('Invalid distributor selected!', 'danger')
            conn.close()
            return redirect(url_for('request_po', quote_ref=quote_ref))
        
        distributor_name = dist_result['name']
        
        # Fetch vendor name if selected (with validation)
        vendor_name = None
        if vendor_id:
            try:
                vendor_id = int(vendor_id)  # Normalize to integer
                c.execute("SELECT name FROM vendors WHERE id = ?", (vendor_id,))
                vendor_result = c.fetchone()
                
                if not vendor_result:
                    flash('Invalid vendor selected!', 'danger')
                    conn.close()
                    return redirect(url_for('request_po', quote_ref=quote_ref))
                
                vendor_name = vendor_result['name']
            except (ValueError, TypeError):
                flash('Invalid vendor ID format!', 'danger')
                conn.close()
                return redirect(url_for('request_po', quote_ref=quote_ref))
        else:
            vendor_id = None  # Set to NULL if not provided
        
        # Generate unique PO Request Reference (RFPO-YYYYMMDD###)
        current_date = datetime.now().strftime('%Y%m%d')
        c.execute("SELECT COUNT(*) FROM po_requests WHERE po_request_reference LIKE ?", (f'RFPO-{current_date}%',))
        count = c.fetchone()[0] + 1
        po_request_reference = f'RFPO-{current_date}{count:03d}'
        
        # Check for duplicate pending requests for this quotation
        c.execute("""SELECT id FROM po_requests 
                     WHERE quote_ref = ? AND request_status = 'Pending Approval'""", (quote_ref,))
        existing = c.fetchone()
        
        if existing:
            flash('A PO request for this quotation is already pending approval!', 'warning')
            conn.close()
            return redirect(url_for('project_summary'))
        
        try:
            # Insert PO Request
            c.execute('''INSERT INTO po_requests (
                po_request_reference, quote_ref, project_name, system,
                presale_engineer, project_manager, project_coordinator, vendor_id, vendor_name,
                distributor_id, distributor_name, notes, request_status,
                requested_by_id, requested_by_name, requested_time, supplier_quotation_id
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                (po_request_reference, quote_ref, project_name, system,
                 presale_engineer, project_manager, project_coordinator, vendor_id, vendor_name,
                 distributor_id, distributor_name, notes, 'Pending Approval',
                 session['user_id'], session['username'], datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                 supplier_quotation_id))
            
            conn.commit()
            flash(f'PO Request {po_request_reference} submitted successfully! Awaiting approval from Technical Team Leader or General Manager.', 'success')
            
            # Get project_id to redirect to project detail
            c.execute("SELECT id FROM register_project WHERE project_name = ?", (project_name,))
            project_result = c.fetchone()
            project_id = project_result[0] if project_result else None
            
            # Send notifications to TTL and GM
            try:
                actor_id = session.get('user_id')
                actor_name = session.get('username', 'User')
                
                # Get TTL and GM recipients
                recipients = notification_service.get_admin_recipients()
                
                # Remove current user from recipients
                if actor_id in recipients:
                    recipients.remove(actor_id)
                
                if recipients:
                    notification_service.notify_activity(
                        event_code='po_request.created',
                        recipient_ids=recipients,
                        actor_id=actor_id,
                        context={
                            'actor_name': actor_name,
                            'po_request_reference': po_request_reference,
                            'quote_ref': quote_ref,
                            'project_name': project_name,
                            'distributor_name': distributor_name
                        },
                        url=url_for('po_requests_dashboard')
                    )
            except Exception as e:
                print(f"Notification error: {e}")
            
            # Redirect to project detail page if project_id found, otherwise to project summary
            if project_id:
                return redirect(url_for('project_detail', project_id=project_id))
            else:
                return redirect(url_for('project_summary'))
            
        except Exception as e:
            flash(f'Error submitting PO request: {str(e)}', 'danger')
            conn.rollback()
            return redirect(url_for('request_po', quote_ref=quote_ref))
        finally:
            conn.close()
    
    # GET Request - Display form
    quote_ref = request.args.get('quote_ref')
    
    if not quote_ref:
        flash('Quote reference is required!', 'danger')
        return redirect(url_for('project_summary'))
    
    # Fetch quotation details from projects table
    c.execute("""SELECT project_name, quote_ref, presale_eng, sales_eng, system
                 FROM projects WHERE quote_ref = ?""", (quote_ref,))
    quotation = c.fetchone()
    
    if not quotation:
        flash('Quotation not found!', 'danger')
        conn.close()
        return redirect(url_for('project_summary'))
    
    # Fetch distributors and vendors for dropdowns
    c.execute("SELECT id, name FROM distributors ORDER BY name")
    distributors = c.fetchall()
    c.execute("SELECT id, name FROM vendors ORDER BY name")
    vendors = c.fetchall()
    
    # Fetch project managers
    c.execute("SELECT username FROM engineers WHERE role IN ('Project Manager', 'Implementation Engineer', 'Technical Team Leader')")
    project_managers = c.fetchall()
    
    # Fetch project coordinators
    c.execute("SELECT username FROM engineers WHERE role IN ('Project Manager', 'Implementation Engineer', 'Project Coordinator', 'Technical Team Leader')")
    project_coordinators = c.fetchall()
    
    conn.close()
    
    return render_template('request_po.html',
                          quotation=quotation,
                          distributors=distributors,
                          vendors=vendors,
                          project_managers=project_managers,
                          project_coordinators=project_coordinators)

@app.route('/request_po_from_supplier_quotation/<int:quotation_id>', methods=['GET'])
@role_required('Sales Engineer', 'Presale Engineer', 'Project Manager', 'Technical Team Leader', 'General Manager')
def request_po_from_supplier_quotation(quotation_id):
    """Request PO from a supplier quotation - auto-fills form with supplier quotation data"""
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    
    # Fetch supplier quotation details
    c.execute("""SELECT * FROM supplier_quotations WHERE id = ?""", (quotation_id,))
    supplier_quotation = c.fetchone()
    
    if not supplier_quotation:
        flash('Supplier quotation not found!', 'danger')
        conn.close()
        return redirect(url_for('project_summary'))
    
    # Fetch quotation details from projects table using the quote_ref from supplier quotation
    quote_ref = supplier_quotation['quote_ref']
    c.execute("""SELECT project_name, quote_ref, presale_eng, sales_eng, system
                 FROM projects WHERE quote_ref = ?""", (quote_ref,))
    quotation = c.fetchone()
    
    if not quotation:
        flash('Quotation not found!', 'danger')
        conn.close()
        return redirect(url_for('project_summary'))
    
    # Fetch distributors and vendors for dropdowns
    c.execute("SELECT id, name FROM distributors ORDER BY name")
    distributors = c.fetchall()
    c.execute("SELECT id, name FROM vendors ORDER BY name")
    vendors = c.fetchall()
    
    # Fetch project managers
    c.execute("SELECT username FROM engineers WHERE role IN ('Project Manager', 'Implementation Engineer', 'Technical Team Leader')")
    project_managers = c.fetchall()
    
    # Fetch project coordinators
    c.execute("SELECT username FROM engineers WHERE role IN ('Project Manager', 'Implementation Engineer', 'Project Coordinator', 'Technical Team Leader')")
    project_coordinators = c.fetchall()
    
    conn.close()
    
    # Pre-fill data from supplier quotation
    prefill_data = {
        'distributor_id': supplier_quotation['distributor_id'],
        'vendor_id': supplier_quotation['vendor_id'],
        'system': supplier_quotation['system'],
        'notes': f"Request from supplier quotation: {supplier_quotation['filename']}\n{supplier_quotation['notes'] or ''}",
        'supplier_quotation_id': quotation_id
    }
    
    return render_template('request_po.html',
                          quotation=quotation,
                          distributors=distributors,
                          vendors=vendors,
                          project_managers=project_managers,
                          project_coordinators=project_coordinators,
                          prefill_data=prefill_data,
                          supplier_quotation=supplier_quotation)

@app.route('/create_po_from_request/<rfpo_ref>', methods=['GET'])
@login_required
def create_po_from_request(rfpo_ref):
    """Create PO from an approved RFPO request - pre-fills register_po form"""
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    
    # Fetch the RFPO request details
    c.execute("""SELECT * FROM po_requests WHERE po_request_reference = ?""", (rfpo_ref,))
    po_request = c.fetchone()
    
    if not po_request:
        flash('PO Request not found!', 'danger')
        conn.close()
        return redirect(url_for('po_requests_dashboard'))
    
    # Verify the request is approved
    if po_request['request_status'] != 'Approved':
        flash('Only approved PO requests can be converted to Purchase Orders!', 'warning')
        conn.close()
        return redirect(url_for('po_requests_dashboard'))
    
    # Fetch data for dropdowns (same as register_po)
    c.execute("SELECT id, project_name FROM register_project")
    projects = c.fetchall()
    c.execute("SELECT id, name FROM engineers WHERE role='Presale Engineer'")
    presale_engineers = c.fetchall()
    c.execute("SELECT id, name FROM engineers WHERE role IN ('Implementation Engineer', 'Project Manager')")
    project_managers = c.fetchall()
    c.execute("SELECT id, name FROM vendors ORDER BY name")
    vendors = c.fetchall()
    c.execute("SELECT id, name FROM distributors")
    distributors = c.fetchall()
    
    conn.close()
    
    # Render the register_po template with pre-filled data
    return render_template('register_po.html',
                          projects=projects,
                          presale_engineers=presale_engineers,
                          project_managers=project_managers,
                          vendors=vendors,
                          distributors=distributors,
                          rfpo_data=po_request)

####### PO Requests Dashboard #######
@app.route('/po_requests_dashboard')
@role_required('Sales Engineer', 'Presale Engineer', 'Project Manager', 'Technical Team Leader', 'General Manager')
def po_requests_dashboard():
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    
    # Get current user role
    user_role = None
    if 'user_id' in session:
        c.execute("SELECT role FROM users WHERE id = ?", (session['user_id'],))
        user_role_result = c.fetchone()
        if user_role_result:
            user_role = user_role_result[0]
    
    # Get filter values from URL
    filters = {
        'request_status': request.args.get('request_status'),
        'requested_by': request.args.get('requested_by'),
        'distributor': request.args.get('distributor'),
        'vendor': request.args.get('vendor'),
        'start_date': request.args.get('start_date'),
        'end_date': request.args.get('end_date'),
    }
    
    # Fetch data for filter dropdowns
    c.execute("SELECT DISTINCT request_status FROM po_requests ORDER BY request_status")
    status_options = [row[0] for row in c.fetchall()]
    
    c.execute("""SELECT DISTINCT u.username FROM po_requests pr
                 JOIN users u ON pr.requested_by_id = u.id
                 ORDER BY u.username""")
    requesters = [row[0] for row in c.fetchall()]
    
    c.execute("SELECT DISTINCT distributor_name FROM po_requests WHERE distributor_name IS NOT NULL ORDER BY distributor_name")
    distributors = [row[0] for row in c.fetchall()]
    
    c.execute("SELECT DISTINCT vendor_name FROM po_requests WHERE vendor_name IS NOT NULL ORDER BY vendor_name")
    vendors = [row[0] for row in c.fetchall()]
    
    # Base query with JOINs to get usernames and supplier quotation info
    query = """
        SELECT
            pr.id, pr.po_request_reference, pr.quote_ref, pr.project_name,
            pr.system, pr.presale_engineer, pr.distributor_name, pr.vendor_name,
            pr.project_manager, pr.request_status, pr.notes, pr.requested_time as created_at,
            pr.decision_time as approved_rejected_at, pr.rejection_reason,
            u_req.username as requested_by_username,
            u_appr.username as approved_rejected_by_username,
            pr.supplier_quotation_id,
            sq.filename as supplier_quotation_filename
        FROM po_requests pr
        LEFT JOIN users u_req ON pr.requested_by_id = u_req.id
        LEFT JOIN users u_appr ON pr.approved_by_id = u_appr.id
        LEFT JOIN supplier_quotations sq ON pr.supplier_quotation_id = sq.id
        WHERE 1=1
    """
    params = []
    
    # Apply filters
    if filters['request_status']:
        query += " AND pr.request_status = ?"
        params.append(filters['request_status'])
    if filters['requested_by']:
        query += " AND u_req.username = ?"
        params.append(filters['requested_by'])
    if filters['distributor']:
        query += " AND pr.distributor_name = ?"
        params.append(filters['distributor'])
    if filters['vendor']:
        query += " AND pr.vendor_name = ?"
        params.append(filters['vendor'])
    if filters['start_date']:
        query += " AND date(pr.created_at) >= ?"
        params.append(filters['start_date'])
    if filters['end_date']:
        query += " AND date(pr.created_at) <= ?"
        params.append(filters['end_date'])
    
    query += " ORDER BY pr.created_at DESC"
    
    c.execute(query, params)
    po_requests = c.fetchall()
    total_requests = len(po_requests)
    
    # Chart data - status distribution
    chart_query = "SELECT request_status, COUNT(*) FROM po_requests WHERE 1=1"
    if filters['request_status']:
        chart_query += " AND request_status = ?"
    if filters['requested_by']:
        chart_query += " AND requested_by_id IN (SELECT id FROM users WHERE username = ?)"
    if filters['distributor']:
        chart_query += " AND distributor_name = ?"
    if filters['vendor']:
        chart_query += " AND vendor_name = ?"
    if filters['start_date']:
        chart_query += " AND date(created_at) >= ?"
    if filters['end_date']:
        chart_query += " AND date(created_at) <= ?"
    chart_query += " GROUP BY request_status"
    
    c.execute(chart_query, params)
    status_counts_raw = c.fetchall()
    conn.close()
    
    status_labels = [row[0] for row in status_counts_raw]
    status_counts = [row[1] for row in status_counts_raw]
    
    return render_template('po_requests_dashboard.html',
                          po_requests=po_requests,
                          total_requests=total_requests,
                          user_role=user_role,
                          status_options=status_options,
                          requesters=requesters,
                          distributors=distributors,
                          vendors=vendors,
                          status_labels=status_labels,
                          status_counts=status_counts,
                          current_filters=filters)

@app.route('/approve_po_request/<rfpo_ref>', methods=['POST'])
@role_required('Technical Team Leader', 'General Manager')
def approve_po_request(rfpo_ref):
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    
    try:
        # Fetch the request with validation
        c.execute("""SELECT id, request_status, requested_by_id, project_name, quote_ref
                     FROM po_requests WHERE po_request_reference = ?""", (rfpo_ref,))
        po_request = c.fetchone()
        
        if not po_request:
            flash('PO request not found!', 'danger')
            return redirect(url_for('po_requests_dashboard'))
        
        if po_request['request_status'] != 'Pending Approval':
            flash(f'Cannot approve: Request is already {po_request["request_status"]}!', 'warning')
            return redirect(url_for('po_requests_dashboard'))
        
        # Update request status to Approved
        c.execute("""UPDATE po_requests 
                     SET request_status = 'Approved',
                         approved_by_id = ?,
                         decision_time = ?
                     WHERE po_request_reference = ?""",
                 (session['user_id'], datetime.now().strftime('%Y-%m-%d %H:%M:%S'), rfpo_ref))
        
        conn.commit()
        flash(f'PO Request {rfpo_ref} has been approved successfully!', 'success')
        
        # Send notification to requester
        try:
            actor_id = session.get('user_id')
            actor_name = session.get('username', 'Administrator')
            requester_id = po_request['requested_by_id']
            
            if requester_id and requester_id != actor_id:
                notification_service.notify_activity(
                    event_code='po_request.approved',
                    recipient_ids=[requester_id],
                    actor_id=actor_id,
                    context={
                        'actor_name': actor_name,
                        'po_request_reference': rfpo_ref,
                        'project_name': po_request['project_name'],
                        'quote_ref': po_request['quote_ref']
                    },
                    url=url_for('po_requests_dashboard')
                )
        except Exception as e:
            print(f"Notification error: {e}")
        
    except Exception as e:
        flash(f'Error approving request: {str(e)}', 'danger')
        conn.rollback()
    finally:
        conn.close()
    
    return redirect(url_for('po_requests_dashboard'))

@app.route('/reject_po_request/<rfpo_ref>', methods=['POST'])
@role_required('Technical Team Leader', 'General Manager')
def reject_po_request(rfpo_ref):
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    
    rejection_reason = request.form.get('rejection_reason', '').strip()
    
    if not rejection_reason:
        flash('Rejection reason is required!', 'danger')
        return redirect(url_for('po_requests_dashboard'))
    
    try:
        # Fetch the request with validation
        c.execute("""SELECT id, request_status, requested_by_id, project_name, quote_ref
                     FROM po_requests WHERE po_request_reference = ?""", (rfpo_ref,))
        po_request = c.fetchone()
        
        if not po_request:
            flash('PO request not found!', 'danger')
            return redirect(url_for('po_requests_dashboard'))
        
        if po_request['request_status'] != 'Pending Approval':
            flash(f'Cannot reject: Request is already {po_request["request_status"]}!', 'warning')
            return redirect(url_for('po_requests_dashboard'))
        
        # Update request status to Rejected
        c.execute("""UPDATE po_requests 
                     SET request_status = 'Rejected',
                         approved_by_id = ?,
                         decision_time = ?,
                         rejection_reason = ?
                     WHERE po_request_reference = ?""",
                 (session['user_id'], datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 
                  rejection_reason, rfpo_ref))
        
        conn.commit()
        flash(f'PO Request {rfpo_ref} has been rejected.', 'warning')
        
        # Send notification to requester
        try:
            actor_id = session.get('user_id')
            actor_name = session.get('username', 'Administrator')
            requester_id = po_request['requested_by_id']
            
            if requester_id and requester_id != actor_id:
                notification_service.notify_activity(
                    event_code='po_request.rejected',
                    recipient_ids=[requester_id],
                    actor_id=actor_id,
                    context={
                        'actor_name': actor_name,
                        'po_request_reference': rfpo_ref,
                        'project_name': po_request['project_name'],
                        'quote_ref': po_request['quote_ref'],
                        'rejection_reason': rejection_reason
                    },
                    url=url_for('po_requests_dashboard')
                )
        except Exception as e:
            print(f"Notification error: {e}")
        
    except Exception as e:
        flash(f'Error rejecting request: {str(e)}', 'danger')
        conn.rollback()
    finally:
        conn.close()
    
    return redirect(url_for('po_requests_dashboard'))

@app.route('/get_po_request_details/<int:request_id>')
@role_required('Sales Engineer', 'Presale Engineer', 'Project Manager', 'Technical Team Leader', 'General Manager')
def get_po_request_details(request_id):
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    
    c.execute("""SELECT
                    pr.po_request_reference, pr.quote_ref, pr.project_name,
                    pr.system, pr.presale_engineer, pr.distributor_name,
                    pr.vendor_name, pr.project_manager, pr.request_status,
                    pr.notes, pr.requested_time as created_at, pr.decision_time as approved_rejected_at,
                    pr.rejection_reason,
                    u_req.username as requested_by_username,
                    u_appr.username as approved_rejected_by_username
                 FROM po_requests pr
                 LEFT JOIN users u_req ON pr.requested_by_id = u_req.id
                 LEFT JOIN users u_appr ON pr.approved_by_id = u_appr.id
                 WHERE pr.id = ?""", (request_id,))
    
    po_request = c.fetchone()
    conn.close()
    
    if not po_request:
        return jsonify({'error': 'PO request not found'}), 404
    
    return jsonify(dict(po_request))

@app.route('/download_po_requests')
@role_required('Sales Engineer', 'Presale Engineer', 'Project Manager', 'Technical Team Leader', 'General Manager')
def download_po_requests():
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    
    # Apply same filters as dashboard
    filters = {
        'request_status': request.args.get('request_status'),
        'requested_by': request.args.get('requested_by'),
        'distributor': request.args.get('distributor'),
        'vendor': request.args.get('vendor'),
        'start_date': request.args.get('start_date'),
        'end_date': request.args.get('end_date'),
    }
    
    query = """
        SELECT
            pr.po_request_reference, pr.quote_ref, pr.project_name,
            pr.system, pr.presale_engineer, pr.distributor_name, pr.vendor_name,
            pr.project_manager, u_req.username as requested_by,
            pr.request_status, pr.requested_time as created_at, pr.decision_time as approved_rejected_at,
            u_appr.username as approved_rejected_by, pr.rejection_reason, pr.notes
        FROM po_requests pr
        LEFT JOIN users u_req ON pr.requested_by_id = u_req.id
        LEFT JOIN users u_appr ON pr.approved_by_id = u_appr.id
        WHERE 1=1
    """
    params = []
    
    if filters['request_status']:
        query += " AND pr.request_status = ?"
        params.append(filters['request_status'])
    if filters['requested_by']:
        query += " AND u_req.username = ?"
        params.append(filters['requested_by'])
    if filters['distributor']:
        query += " AND pr.distributor_name = ?"
        params.append(filters['distributor'])
    if filters['vendor']:
        query += " AND pr.vendor_name = ?"
        params.append(filters['vendor'])
    if filters['start_date']:
        query += " AND date(pr.created_at) >= ?"
        params.append(filters['start_date'])
    if filters['end_date']:
        query += " AND date(pr.created_at) <= ?"
        params.append(filters['end_date'])
    
    query += " ORDER BY pr.created_at DESC"
    
    c.execute(query, params)
    rows = c.fetchall()
    conn.close()
    
    # Convert to list of dicts for pandas
    data = [dict(row) for row in rows]
    
    # Create DataFrame
    df = pd.DataFrame(data)
    
    # Rename columns for Excel
    df.columns = ['RFPO Reference', 'Quote Reference', 'Project Name', 'System',
                  'Presale Engineer', 'Distributor', 'Vendor', 'Project Manager',
                  'Requested By', 'Status', 'Requested At', 'Approved/Rejected At',
                  'Approved/Rejected By', 'Rejection Reason', 'Notes']
    
    # Create Excel file in memory
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name='PO Requests', index=False)
        
        # Auto-adjust column widths
        worksheet = writer.sheets['PO Requests']
        for idx, col in enumerate(df.columns):
            max_length = max(df[col].astype(str).map(len).max(), len(col)) + 2
            worksheet.set_column(idx, idx, min(max_length, 50))
    
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'PO_Requests_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

##############333
@app.route('/register_po_follow_up', methods=['GET', 'POST'])
@login_required
def register_po_follow_up():
    # Handles GET request to display the form
    po_number = request.args.get('po_number')
    total_amount = request.args.get('total_amount', type=float)

    if request.method == 'POST':
        # --- Handles form submission ---

        # 1. Capture the current timestamp for the registration date
        registration_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # 2. Collect all data from the submitted form
        po_number = request.form['po_number']
        total_amount = request.form['total_amount']
        order_date = request.form['order_date']
        status = request.form['status']
        expected_delivery = request.form['expected_delivery']
        notes = request.form['notes']

        # 3. Handle optional file uploads
        invoice_file = request.files.get('invoice')
        delivery_note_file = request.files.get('delivery_note')
        delivered_boq_file = request.files.get('delivered_boq')

        invoice_data = invoice_file.read() if invoice_file and invoice_file.filename != '' else None
        delivery_note_data = delivery_note_file.read() if delivery_note_file and delivery_note_file.filename != '' else None
        delivered_boq_data = delivered_boq_file.read() if delivered_boq_file and delivered_boq_file.filename != '' else None

        # 4. Insert the new follow-up record into the database
        conn = sqlite3.connect('ProjectStatus.db')
        c = conn.cursor()
        try:
            c.execute('''INSERT INTO purchase_order_monitoring (
                           po_number, order_date, total_amount, status, expected_delivery,
                           notes, invoice, delivery_note, delivered_boq, registration_date
                       ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''', (
                po_number, order_date, total_amount, status, expected_delivery,
                notes, invoice_data, delivery_note_data, delivered_boq_data, registration_date
            ))
            conn.commit()
            flash('PO Follow-Up entry registered successfully!', 'success')
        except Exception as e:
            flash(f'An error occurred: {e}', 'danger')
        finally:
            conn.close()

        return redirect(url_for('view_po_status'))

    # --- For GET request, find the first order date to pre-fill the form ---
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    c.execute("""
        SELECT order_date FROM purchase_order_monitoring 
        WHERE po_number = ? 
        ORDER BY order_date ASC 
        LIMIT 1
    """, (po_number,))
    first_entry = c.fetchone()
    conn.close()

    # Use the first entry's date, or default to today if no entries exist
    first_order_date = first_entry[0] if first_entry else datetime.now().strftime('%Y-%m-%d')

    return render_template('register_po_follow_up.html',
                           po_number=po_number,
                           total_amount=total_amount,
                           first_order_date=first_order_date)
#######################
@app.route('/edit_po_follow_up/<int:entry_id>', methods=['GET', 'POST'])
@login_required
def edit_po_follow_up(entry_id):
    """Handles editing a specific PO follow-up entry."""
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    if request.method == 'POST':
        # --- Process the submitted form ---

        # Fetch the existing entry to get old file data and the po_number for redirecting
        c.execute("SELECT * FROM purchase_order_monitoring WHERE id = ?", (entry_id,))
        existing_entry = c.fetchone()
        if not existing_entry:
            flash('Follow-up entry not found.', 'danger')
            return redirect(url_for('view_po_status'))

        # Collect data from the form
        order_date = request.form['order_date']
        status = request.form['status']
        expected_delivery = request.form['expected_delivery']
        notes = request.form['notes']

        # Handle file uploads: use new file if provided, otherwise keep the existing one
        invoice_file = request.files.get('invoice')
        delivery_note_file = request.files.get('delivery_note')
        delivered_boq_file = request.files.get('delivered_boq')

        invoice_data = invoice_file.read() if invoice_file and invoice_file.filename else existing_entry['invoice']
        delivery_note_data = delivery_note_file.read() if delivery_note_file and delivery_note_file.filename else existing_entry['delivery_note']
        delivered_boq_data = delivered_boq_file.read() if delivered_boq_file and delivered_boq_file.filename else existing_entry['delivered_boq']

        try:
            c.execute("""
                UPDATE purchase_order_monitoring SET
                order_date = ?, status = ?, expected_delivery = ?, notes = ?,
                invoice = ?, delivery_note = ?, delivered_boq = ?
                WHERE id = ?
            """, (order_date, status, expected_delivery, notes, invoice_data,
                  delivery_note_data, delivered_boq_data, entry_id))
            conn.commit()
            flash('Follow-up entry updated successfully!', 'success')
        except Exception as e:
            flash(f'An error occurred: {e}', 'danger')
        finally:
            conn.close()

        return redirect(url_for('view_po_details', po_number=existing_entry['po_number']))

    # --- For a GET request, show the form pre-filled with data ---
    c.execute("SELECT * FROM purchase_order_monitoring WHERE id = ?", (entry_id,))
    entry = c.fetchone()
    conn.close()

    if not entry:
        flash('Follow-up entry not found.', 'danger')
        return redirect(url_for('view_po_status'))

    return render_template('edit_po_follow_up.html', entry=entry)


@app.route('/delete_po_follow_up/<int:entry_id>', methods=['POST'])
@login_required
def delete_po_follow_up(entry_id):
    """Deletes a specific PO follow-up entry."""
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    # First, get the po_number so we can redirect back to the correct page
    c.execute("SELECT po_number FROM purchase_order_monitoring WHERE id = ?", (entry_id,))
    entry = c.fetchone()

    if not entry:
        flash('Entry not found.', 'danger')
        conn.close()
        return redirect(url_for('view_po_status'))

    po_number_redirect = entry['po_number']

    try:
        c.execute("DELETE FROM purchase_order_monitoring WHERE id = ?", (entry_id,))
        conn.commit()
        flash('Follow-up entry deleted successfully.', 'success')
    except Exception as e:
        flash(f'Error deleting entry: {e}', 'danger')
    finally:
        conn.close()

    return redirect(url_for('view_po_details', po_number=po_number_redirect))
###################
@app.route('/view_document/<int:entry_id>/<doc_type>')
@login_required
def view_document(entry_id, doc_type):
    """Serves a document for inline viewing in the browser."""

    # Validate the document type to prevent security issues
    allowed_docs = ['invoice', 'delivery_note', 'delivered_boq']
    if doc_type not in allowed_docs:
        flash('Invalid document type requested.', 'danger')
        return redirect(request.referrer or url_for('view_po_status'))

    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    # Fetch the specific document BLOB from the database
    c.execute(f"SELECT {doc_type} FROM purchase_order_monitoring WHERE id = ?", (entry_id,))
    result = c.fetchone()
    conn.close()

    if result and result[0]:
        file_data = BytesIO(result[0])

        # Determine the correct MIME type based on the document
        mimetype = 'application/pdf'
        if doc_type == 'delivered_boq':
            mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        # Use as_attachment=False to show the file in the browser
        return send_file(file_data, mimetype=mimetype, as_attachment=False)
    else:
        flash('Document not found.', 'danger')
        return redirect(request.referrer or url_for('view_po_status'))
##################
import io
from flask import send_file, flash, redirect, url_for
import sqlite3

@app.route('/download_invoice/<int:id>')
@login_required
def download_invoice(id):
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    c.execute("SELECT invoice FROM purchase_order_monitoring WHERE id=?", (id,))
    result = c.fetchone()
    conn.close()

    if result and result[0]:
        output = io.BytesIO(result[0])  # Assuming invoice is stored as binary data
        return send_file(output, download_name=f'{id}_Invoice.pdf', as_attachment=True, mimetype='application/pdf')
    else:
        flash('Invoice not found!', 'danger')
        return redirect(url_for('view_po_status'))

@app.route('/download_delivery_note/<int:id>')
@login_required
def download_delivery_note(id):
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    c.execute("SELECT delivery_note FROM purchase_order_monitoring WHERE id=?", (id,))
    result = c.fetchone()
    conn.close()

    if result and result[0]:
        output = io.BytesIO(result[0])  # Assuming delivery note is stored as binary data
        return send_file(output, download_name=f'{id}_Delivery_Note.pdf', as_attachment=True, mimetype='application/pdf')
    else:
        flash('Delivery note not found!', 'danger')
        return redirect(url_for('view_po_status'))
####################
@app.route('/download_delivered_boq/<int:id>')
@login_required
def download_delivered_boq(id):
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    c.execute("SELECT delivered_boq FROM purchase_order_monitoring WHERE id=?", (id,))
    result = c.fetchone()
    conn.close()

    if result and result[0]:
        output = io.BytesIO(result[0])  # Assuming delivered_boq is stored as binary data
        return send_file(output, download_name=f'{id}_Delivered_Boq.xlsx', as_attachment=True)
    else:
        flash('Delivered BOQ not found!', 'danger')
        return redirect(url_for('view_po_status'))
#################
@app.route('/view_po_document/<int:po_id>/<doc_type>')
@login_required
def view_po_document(po_id, doc_type):
    """Serves a PO or Quotation PDF for inline browser viewing."""

    # Validate the requested document type
    if doc_type not in ['quotation', 'po_document']:
        flash('Invalid document type.', 'danger')
        return redirect(url_for('view_po_status'))

    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()

    # Fetch the specific document's binary data
    c.execute(f"SELECT {doc_type} FROM purchase_orders WHERE id = ?", (po_id,))
    result = c.fetchone()
    conn.close()

    if result and result[0]:
        file_data = BytesIO(result[0])
        # Serve the file for inline viewing (not as an attachment)
        return send_file(file_data, mimetype='application/pdf', as_attachment=False)
    else:
        flash('Document not found!', 'danger')
        return redirect(url_for('view_po_status'))
#############3
@app.route('/download_po_details_excel/<po_number>', methods=['GET'])
@login_required
def download_po_details_excel(po_number):
    conn = sqlite3.connect('ProjectStatus.db')

    # Query to fetch purchase order details
    query = '''
    SELECT
        po.po_number,
        po.order_date,
        po.total_amount,
        po.status,
        po.expected_delivery,
        po.partial_delivery,
        po.remaining_amount,
        po.payment_status,
        po.total_paid,
        po.remaining_balance,
        po.notes
    FROM
        purchase_order_monitoring po
    WHERE
        po.po_number = ?
    '''

    df = pd.read_sql_query(query, conn, params=(po_number,))
    conn.close()

    # Create an Excel file
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='PO Details')

    output.seek(0)
    return send_file(output, download_name=f'{po_number}_PO_Details.xlsx', as_attachment=True)


########################
@app.route('/solution_builder')
@permission_required('view_solution_builder')
def solution_builder():
    return render_template('solution_builder_hub.html')


import pandas as pd
import math
import pandas as pd
import math
from flask import flash, redirect, request, render_template, url_for


# Ensure you have other necessary imports like Flask, session, etc. at the top of your file.

@app.route('/passive_solution_builder', methods=['GET', 'POST'])
@login_required
def passive_solution_builder():
    if request.method == 'POST':
        try:
            # --- 1. Get All User Inputs ---
            if 'boq_file' not in request.files:
                flash('No file part in the request.', 'danger')
                return redirect(request.url)

            file = request.files['boq_file']
            if file.filename == '':
                flash('No selected file.', 'danger')
                return redirect(request.url)

            # Get selections from the form
            cable_category = request.form.get('cable_category', 'Cat6')
            avg_length = int(request.form.get('avg_copper_length') or 45)
            num_single_faceplates = int(request.form.get('num_single_faceplates', 0))
            num_double_faceplates = int(request.form.get('num_double_faceplates', 0))
            num_floor_boxes = int(request.form.get('num_floor_boxes', 0))
            num_devices_no_fp = int(request.form.get('num_devices_no_fp', 0))

            if not file.filename.endswith(('.xlsx', '.xls')):
                flash('Invalid file type. Please upload an Excel file (.xlsx or .xls).', 'danger')
                return redirect(request.url)

            # --- 2. Process the Excel File ---
            df = pd.read_excel(file)

            # Find columns, handling variations in naming (e.g., "Coper" vs "Copper")
            total_points_col = next((col for col in df.columns if 'total' in col.lower() and 'point' in col.lower()),
                                    None)
            pp_24_col = next((col for col in df.columns if '24' in col and 'panel' in col.lower()), None)
            pp_48_col = next((col for col in df.columns if '48' in col and 'panel' in col.lower()), None)

            if not total_points_col or not pp_24_col or not pp_48_col:
                flash(
                    "Error: Excel file must contain columns for 'Total Points', '24 port patch panel', and '48 port patch panel' to proceed.",
                    'danger')
                return redirect(request.url)

            # Clean and sum the required columns from the Excel file
            total_points_from_excel = pd.to_numeric(df[total_points_col], errors='coerce').fillna(0).sum()
            num_24_port_panels = int(pd.to_numeric(df[pp_24_col], errors='coerce').fillna(0).sum())
            num_48_port_panels = int(pd.to_numeric(df[pp_48_col], errors='coerce').fillna(0).sum())

            # --- 3. Generate the Bill of Quantities (BoQ) based on the Word document ---
            boq = []

            # Item 1: Cable Reel Calculation
            total_length = int(total_points_from_excel) * avg_length
            cable_divisor = 500 if cable_category == 'Cat6A' else 305
            total_cable_reels = math.ceil(total_length / cable_divisor)
            boq.append({'item': f'{cable_category} U/UTP Cable Reel', 'quantity': total_cable_reels,
                        'notes': f'Based on {int(total_points_from_excel)} points at ~{avg_length}m each'})

            # Items 2 & 3: Patch Panel Calculation (from Excel)
            boq.append({'item': f'24-Port {cable_category} Patch Panel', 'quantity': num_24_port_panels,
                        'notes': 'Summed from Excel file column'})
            boq.append({'item': f'48-Port {cable_category} Patch Panel', 'quantity': num_48_port_panels,
                        'notes': 'Summed from Excel file column'})

            # Item 4 & 5: Cable Organizer Calculation
            boq.append({'item': '1U Cable Organizer', 'quantity': num_24_port_panels,
                        'notes': 'Matches quantity of 24-port panels'})
            boq.append({'item': '2U Cable Organizer', 'quantity': num_48_port_panels,
                        'notes': 'Matches quantity of 48-port panels'})

            # Item 6: Module for Patch Panel Calculation
            total_pp_modules = (24 * num_24_port_panels) + (48 * num_48_port_panels)
            boq.append({'item': f'Module for Patch Panel ({cable_category})', 'quantity': total_pp_modules,
                        'notes': 'Total ports from all patch panels'})

            # Item 7 & 8: Patch Cord Calculation
            patch_cords_1m = math.ceil(total_points_from_excel * 1.1)
            patch_cords_3m = int(total_points_from_excel - num_devices_no_fp)
            boq.append({'item': f'1m {cable_category} Patch Cord', 'quantity': patch_cords_1m,
                        'notes': 'Total points + 10% spare'})
            boq.append({'item': f'3m {cable_category} Patch Cord', 'quantity': patch_cords_3m,
                        'notes': 'Total points minus devices without faceplates'})

            # Item 9 & 10: Faceplate & Connector Calculation
            boq.append({'item': '1-Port Faceplate', 'quantity': num_single_faceplates, 'notes': 'User input'})
            boq.append({'item': '2-Port Faceplate', 'quantity': num_double_faceplates, 'notes': 'User input'})
            boq.append({'item': 'Floor Box Outlets', 'quantity': num_floor_boxes, 'notes': 'User input'})
            boq.append({'item': f'RJ45-Plug Connector ({cable_category})', 'quantity': num_devices_no_fp,
                        'notes': 'For devices without faceplates'})

            modules_for_faceplates = int(total_points_from_excel - num_devices_no_fp)
            boq.append({'item': f'Module for Faceplates ({cable_category})', 'quantity': modules_for_faceplates,
                        'notes': 'Total points minus devices without faceplates'})

            flash('Bill of Quantities generated successfully!', 'success')
            return render_template('passive_solution_builder.html', results=boq, inputs=request.form)

        except Exception as e:
            flash(f"Error processing your request: {e}", "danger")
            return redirect(request.url)

    # For a GET request, just show the form
    return render_template('passive_solution_builder.html', results=None, inputs=None)

# ADD THIS NEW FUNCTION to app.py
import pandas as pd
import math
import base64
from io import BytesIO
from flask import send_file, flash, redirect, url_for, session


# Make sure you have all necessary imports at the top of your app.py file

@app.route('/download_boq')
@login_required
def download_boq():
    # Retrieve the saved inputs and file content from the session
    boq_inputs = session.pop('boq_inputs', None)
    if not boq_inputs:
        flash('No BoQ data available to download. Please generate a solution first.', 'warning')
        return redirect(url_for('passive_solution_builder'))

    try:
        # --- 1. Retrieve Saved Data ---
        form_data = boq_inputs['form_data']
        file_content_b64 = boq_inputs['file_content_b64']

        # Decode the file content and read it with pandas
        file_content = base64.b64decode(file_content_b64)
        df = pd.read_excel(BytesIO(file_content))

        # Get the form inputs again
        cable_category = form_data.get('cable_category', 'Cat6')
        avg_length = int(form_data.get('avg_copper_length') or 45)
        num_single_faceplates = int(form_data.get('num_single_faceplates', 0))
        num_double_faceplates = int(form_data.get('num_double_faceplates', 0))
        num_floor_boxes = int(form_data.get('num_floor_boxes', 0))
        num_devices_no_fp = int(form_data.get('num_devices_no_fp', 0))

        # --- 2. Re-run the BoQ Calculation Logic ---
        # (This is the same logic as in the passive_solution_builder function)
        total_points_col = next((col for col in df.columns if 'total' in col.lower() and 'point' in col.lower()), None)
        pp_24_col = next((col for col in df.columns if '24' in col and 'panel' in col.lower()), None)
        pp_48_col = next((col for col in df.columns if '48' in col and 'panel' in col.lower()), None)

        if not total_points_col or not pp_24_col or not pp_48_col:
            flash(
                "Could not generate download: Excel file must contain 'Total Points', '24 port', and '48 port' panel columns.",
                'danger')
            return redirect(url_for('passive_solution_builder'))

        total_points_from_excel = pd.to_numeric(df[total_points_col], errors='coerce').fillna(0).sum()
        num_24_port_panels = int(pd.to_numeric(df[pp_24_col], errors='coerce').fillna(0).sum())
        num_48_port_panels = int(pd.to_numeric(df[pp_48_col], errors='coerce').fillna(0).sum())

        boq = []
        total_length = int(total_points_from_excel) * avg_length
        cable_divisor = 500 if cable_category == 'Cat6A' else 305
        total_cable_reels = math.ceil(total_length / cable_divisor)
        boq.append({'Item': f'{cable_category} U/UTP Cable Reel', 'Quantity': total_cable_reels,
                    'Notes': f'Based on {int(total_points_from_excel)} points at ~{avg_length}m each'})
        boq.append({'Item': f'24-Port {cable_category} Patch Panel', 'Quantity': num_24_port_panels,
                    'Notes': 'Summed from Excel file column'})
        boq.append({'Item': f'48-Port {cable_category} Patch Panel', 'Quantity': num_48_port_panels,
                    'Notes': 'Summed from Excel file column'})
        boq.append({'Item': '1U Cable Organizer', 'Quantity': num_24_port_panels,
                    'Notes': 'Matches quantity of 24-port panels'})
        boq.append({'Item': '2U Cable Organizer', 'Quantity': num_48_port_panels,
                    'Notes': 'Matches quantity of 48-port panels'})
        total_pp_modules = (24 * num_24_port_panels) + (48 * num_48_port_panels)
        boq.append({'Item': f'Module for Patch Panel ({cable_category})', 'Quantity': total_pp_modules,
                    'Notes': 'Total ports from all patch panels'})
        patch_cords_1m = math.ceil(total_points_from_excel * 1.1)
        patch_cords_3m = int(total_points_from_excel - num_devices_no_fp)
        boq.append({'Item': f'1m {cable_category} Patch Cord', 'Quantity': patch_cords_1m,
                    'Notes': 'Total points + 10% spare'})
        boq.append({'Item': f'3m {cable_category} Patch Cord', 'Quantity': patch_cords_3m,
                    'Notes': 'Total points minus devices without faceplates'})
        boq.append({'Item': '1-Port Faceplate', 'Quantity': num_single_faceplates, 'Notes': 'User input'})
        boq.append({'Item': '2-Port Faceplate', 'Quantity': num_double_faceplates, 'Notes': 'User input'})
        boq.append({'Item': 'Floor Box Outlets', 'Quantity': num_floor_boxes, 'Notes': 'User input'})
        boq.append({'Item': f'RJ45-Plug Connector ({cable_category})', 'Quantity': num_devices_no_fp,
                    'Notes': 'For devices without faceplates'})
        modules_for_faceplates = int(total_points_from_excel - num_devices_no_fp)
        boq.append({'Item': f'Module for Faceplates ({cable_category})', 'Quantity': modules_for_faceplates,
                    'Notes': 'Total points minus devices without faceplates'})

        # --- 3. Generate and Send the Excel File ---
        if boq:
            df_boq = pd.DataFrame(boq)
            output = BytesIO()
            df_boq.to_excel(output, index=False, sheet_name='Generated_BoQ')
            output.seek(0)

            return send_file(
                output,
                download_name='Generated_BoQ.xlsx',
                as_attachment=True,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            flash('Failed to generate BoQ for download.', 'danger')
            return redirect(url_for('passive_solution_builder'))

    except Exception as e:
        flash(f"An error occurred while creating the download file: {e}", "danger")
        return redirect(url_for('passive_solution_builder'))
###########3
import pandas as pd

import math


def round_up_to_standard_cores(n):
    """
    Rounds a given number up to the next standard fiber core count.

    Args:
        n (int or float): The number of cores required.

    Returns:
        int: The next standard core count (e.g., 6, 12, 24, 48, 96, 144, 288).
    """
    # List of standard available core sizes
    standard_cores = [6, 12, 24, 48, 96, 144, 288]

    # If the required number is zero or less, no cores are needed.
    if n <= 0:
        return 0

    # Find the first standard core count that is greater than or equal to n
    for core_count in standard_cores:
        if n <= core_count:
            return core_count

    # If n is larger than the biggest standard size, return the number itself
    return math.ceil(n)


##########
import pandas as pd
import math


# Helper function to round up to the next standard fiber core count
def round_up_to_standard_cores(n):
    standard_cores = [6, 12,18, 24,36, 48,72, 96, 144,192,240, 288]
    if n <= 0: return 0
    # Find the first standard core count that is >= n
    for core_count in standard_cores:
        if n <= core_count:
            return core_count
    return math.ceil(n)


@app.route('/fiber_solution_builder', methods=['GET', 'POST'])
@login_required
def fiber_solution_builder():
    if request.method == 'POST':
        try:
            # --- 1. Get All User Inputs ---
            if 'boq_file' not in request.files:
                flash('No file part in the request.', 'danger')
                return redirect(request.url)

            file = request.files['boq_file']
            if file.filename == '':
                flash('No selected file.', 'danger')
                return redirect(request.url)

            # Get selections from the form
            mm_type_idf = request.form.get('mm_type_idf', 'OM4')
            mm_type_mdf = request.form.get('mm_type_mdf', 'OM4')
            mm_cores_idf = int(request.form.get('mm_cores_idf', 12))
            mm_cores_mdf = int(request.form.get('mm_cores_mdf', 12))
            sm_cores_idf = int(request.form.get('sm_cores_idf', 12))
            sm_cores_mdf = int(request.form.get('sm_cores_mdf', 12))

            is_redundant = 'is_redundant' in request.form
            is_dist_switch = 'is_dist_switch' in request.form
            dist_in_sep_cabinet = 'dist_in_sep_cabinet' in request.form
            is_core_switch = 'is_core_switch' in request.form
            core_in_sep_cabinet = 'core_in_sep_cabinet' in request.form
            mdf_connection_type = request.form.get('mdf_connection_type', 'MM')

            if not file.filename.endswith(('.xlsx', '.xls')):
                flash('Invalid file type. Please upload an Excel file (.xlsx or .xls).', 'danger')
                return redirect(request.url)

            # --- 2. Process the Excel File ---
            df = pd.read_excel(file)
            connection_col = next((col for col in df.columns if 'connection' in col.lower()), None)
            if not connection_col:
                flash("Error: Excel file must contain a 'Connection of fiber with the main rack' column.", 'danger')
                return redirect(request.url)

            num_mm_racks = df[df[connection_col] == 'MM']['Rack Name'].nunique()
            num_sm_racks = df[df[connection_col] == 'SM']['Rack Name'].nunique()

            # --- 3. Generate the Bill of Quantities (BoQ) ---
            boq = []

            # Item 1-4: Fiber Optic Cable
            boq.append({'item': f'Fiber Optic Cable MM ({mm_type_idf}) out from IDF', 'quantity': 0,
                        'notes': 'As per Site Survey / Customer BOQ'})
            boq.append({'item': f'Fiber Optic Cable MM ({mm_type_mdf}) between MDFs', 'quantity': 0,
                        'notes': 'As per Site Survey / Customer BOQ'})
            boq.append({'item': 'Fiber Optic Cable SM out from IDF', 'quantity': 0,
                        'notes': 'As per Site Survey / Customer BOQ'})
            boq.append({'item': 'Fiber Optic Cable SM between MDFs', 'quantity': 0,
                        'notes': 'As per Site Survey / Customer BOQ'})

            # Item 5 & 6: IDF Patch Panels
            boq.append({'item': f'Fiber Optic Patch Panel ({mm_cores_idf}-Core {mm_type_idf}) Fully Loaded for IDF',
                        'quantity': num_mm_racks, 'notes': f'One for each of the {num_mm_racks} MM racks'})
            boq.append({'item': f'Fiber Optic Patch Panel ({sm_cores_idf}-Core SM) Fully Loaded for IDF',
                        'quantity': num_sm_racks, 'notes': f'One for each of the {num_sm_racks} SM racks'})

            # Item 7 & 8: MDF (Distribution) Patch Panels
            if is_dist_switch:
                qty_mdf_dist = 2 if is_redundant and dist_in_sep_cabinet else 1

                if is_redundant and dist_in_sep_cabinet:
                    required_cores_mm = (mm_cores_idf * 0.5 * num_mm_racks) + (3 * mm_cores_mdf)
                elif is_core_switch:
                    required_cores_mm = (mm_cores_idf * num_mm_racks) + (1 * mm_cores_mdf)
                else:
                    required_cores_mm = (mm_cores_idf * num_mm_racks)
                panel_cores_mm = round_up_to_standard_cores(required_cores_mm)

                if is_redundant and dist_in_sep_cabinet:
                    required_cores_sm = (sm_cores_idf * 0.5 * num_sm_racks) + (3 * sm_cores_mdf)
                elif is_core_switch:
                    required_cores_sm = (sm_cores_idf * num_sm_racks) + (1 * sm_cores_mdf)
                else:
                    required_cores_sm = (sm_cores_idf * num_sm_racks)
                panel_cores_sm = round_up_to_standard_cores(required_cores_sm)

                if num_mm_racks > 0:
                    boq.append({
                                   'item': f'Fiber Optic Patch Panel ({panel_cores_mm}-Core {mm_type_idf}) Fully Loaded for MDF (Distribution)',
                                   'quantity': qty_mdf_dist, 'notes': 'For distribution switch'})
                if num_sm_racks > 0:
                    boq.append({
                                   'item': f'Fiber Optic Patch Panel ({panel_cores_sm}-Core SM) Fully Loaded for MDF (Distribution)',
                                   'quantity': qty_mdf_dist, 'notes': 'For distribution switch'})

            # Item 9 & 10: MDF (Core) Patch Panels
            if is_core_switch:
                qty_mdf_core = 2 if is_redundant and core_in_sep_cabinet else 1

                if is_dist_switch and is_redundant and dist_in_sep_cabinet and core_in_sep_cabinet:
                    required_cores_core_mm = 3 * mm_cores_mdf
                    required_cores_core_sm = 3 * sm_cores_mdf
                elif is_dist_switch:
                    required_cores_core_mm = 1 * mm_cores_mdf
                    required_cores_core_sm = 1 * sm_cores_mdf
                elif is_redundant and core_in_sep_cabinet:
                    required_cores_core_mm = (mm_cores_idf * 0.5 * num_mm_racks) + (1 * mm_cores_mdf)
                    required_cores_core_sm = (sm_cores_idf * 0.5 * num_sm_racks) + (1 * sm_cores_mdf)
                else:  # Not redundant and no distribution
                    required_cores_core_mm = mm_cores_idf * num_mm_racks
                    required_cores_core_sm = sm_cores_idf * num_sm_racks

                panel_cores_core_mm = round_up_to_standard_cores(required_cores_core_mm)
                panel_cores_core_sm = round_up_to_standard_cores(required_cores_core_sm)

                if num_mm_racks > 0:
                    boq.append({
                                   'item': f'Fiber Optic Patch Panel ({panel_cores_core_mm}-Core {mm_type_mdf}) Fully Loaded for MDF (Core)',
                                   'quantity': qty_mdf_core, 'notes': 'For core switch'})
                if num_sm_racks > 0:
                    boq.append(
                        {'item': f'Fiber Optic Patch Panel ({panel_cores_core_sm}-Core SM) Fully Loaded for MDF (Core)',
                         'quantity': qty_mdf_core, 'notes': 'For core switch'})

            # Item 11 & 12: Patch Cords
            patch_cord_qty_mm = num_mm_racks * 4
            patch_cord_qty_sm = num_sm_racks * 4
            if is_redundant and is_core_switch and core_in_sep_cabinet and is_dist_switch and dist_in_sep_cabinet:
                if mdf_connection_type == 'MM':
                    patch_cord_qty_mm += 36
                else:  # SM
                    patch_cord_qty_sm += 24
            elif is_redundant:
                if mdf_connection_type == 'MM':
                    patch_cord_qty_mm += 24
                else:  # SM
                    patch_cord_qty_sm += 12
            else:
                if mdf_connection_type == 'MM':
                    patch_cord_qty_mm += 6
                else:  # SM
                    patch_cord_qty_sm += 0






            boq.append({'item': f'Fiber Optic Patch Cord MM ({mm_type_idf})', 'quantity': patch_cord_qty_mm,
                        'notes': 'For IDF, MDF, and device connections'})
            boq.append({'item': f'Fiber Optic Patch Cord SM', 'quantity': patch_cord_qty_sm,
                        'notes': 'For IDF, MDF, and device connections'})

            flash('Fiber solution BoQ generated successfully!', 'success')
            return render_template('fiber_solution_builder.html', results=boq, inputs=request.form)

        except Exception as e:
            flash(f"Error processing your request: {e}", "danger")
            return redirect(request.url)

    # For a GET request, just show the form
    return render_template('fiber_solution_builder.html', results=None, inputs=None)
#################
@app.route('/view_passive_products')
@login_required
def view_passive_products():
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    c.execute("SELECT * FROM passive_products")
    products = c.fetchall()
    conn.close()
    return render_template('view_passive_products.html', products=products)

###############################3
@app.route('/register_passive_products', methods=['GET', 'POST'])
@login_required
def register_passive_products():
    if request.method == 'POST':
        if 'products_file' not in request.files:
            flash('No file part in the request.', 'danger')
            return redirect(request.url)

        file = request.files['products_file']
        if file.filename == '':
            flash('No selected file.', 'danger')
            return redirect(request.url)

        if file and file.filename.endswith('.xlsx'):
            try:
                df = pd.read_excel(file)
                conn = sqlite3.connect('ProjectStatus.db')
                c = conn.cursor()

                # Rename columns to match the database schema
                df.rename(columns={
                    'Part Number': 'part_number',
                    'Product Description': 'description',
                    'Vendor': 'vendor',
                    'Solution': 'solution_category',
                    'Specs': 'specs',
                    'Specific solution': 'specific_solution',
                    'Unit Price': 'unit_price'
                }, inplace=True)

                # Insert or replace data into the database
                for index, row in df.iterrows():
                    c.execute('''
                        INSERT OR REPLACE INTO passive_products 
                        (part_number, description, vendor, solution_category, specs, specific_solution, unit_price)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    ''', tuple(row))

                conn.commit()
                conn.close()
                flash(f'Successfully imported {len(df)} products!', 'success')
                return redirect(url_for('view_passive_products'))

            except Exception as e:
                flash(f"Error processing Excel file: {e}", "danger")
                return redirect(request.url)
        else:
            flash('Invalid file type. Please upload an Excel file (.xlsx).', 'danger')
            return redirect(request.url)

    return render_template('register_passive_products.html')
#########################3
@app.route('/get_specific_solutions/<solution_category>')
@login_required
def get_specific_solutions(solution_category):
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    c.execute("SELECT DISTINCT specific_solution FROM passive_products WHERE solution_category = ? ORDER BY specific_solution", (solution_category,))
    specific_solutions = [row[0] for row in c.fetchall()]
    conn.close()
    return jsonify(specific_solutions)
###############################
@app.route('/quotation_builder', methods=['GET'])
@login_required
@permission_required('view_quotation_builder')
def quotation_builder():
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()

    # Fetch data for the main project dropdown
    c.execute("SELECT project_name FROM register_project ORDER BY project_name")
    projects = c.fetchall()

    # Fetch distinct values for the product filters
    c.execute("SELECT DISTINCT vendor FROM passive_products ORDER BY vendor")
    vendors = c.fetchall()
    c.execute("SELECT DISTINCT solution_category FROM passive_products ORDER BY solution_category")
    solutions = c.fetchall()

    # Note: The 'Specific Solution' dropdown is now populated dynamically by JavaScript
    # so we do not need to fetch it here.

    conn.close()

    return render_template('quotation_builder.html',
                           projects=projects,
                           vendors=vendors,
                           solutions=solutions)

#################################
@app.route('/search_products')
@login_required
def search_products():
    # Get all filter values from the request
    vendor = request.args.get('vendor', '')
    part_number = request.args.get('part_number', '')
    specs = request.args.get('specs', '')
    solution = request.args.get('solution', '')
    specific_solution = request.args.get('specific_solution', '')

    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    # Build the query dynamically based on the provided filters
    query = "SELECT part_number, description, unit_price FROM passive_products WHERE 1=1"
    params = []

    if vendor:
        query += " AND vendor = ?"
        params.append(vendor)
    if part_number:
        query += " AND part_number LIKE ?"
        params.append(f'%{part_number}%')
    if specs:
        query += " AND specs LIKE ?"
        params.append(f'%{specs}%')
    if solution:
        query += " AND solution_category = ?"
        params.append(solution)
    if specific_solution:
        query += " AND specific_solution = ?"
        params.append(specific_solution)

    query += " LIMIT 20"  # Limit results to avoid overloading the page

    c.execute(query, tuple(params))
    products = [dict(row) for row in c.fetchall()]
    conn.close()

    return jsonify(products)

######################3
@app.route('/save_quotation', methods=['POST'])
@login_required
def save_quotation():
    json_data = request.form.get('quotation_data')
    if not json_data:
        flash('No quotation data received.', 'danger')
        return redirect(url_for('quotation_builder'))

    try:
        data = json.loads(json_data)
        project_name = data.get('projectName')
        quote_ref = data.get('quoteRef')
        items = data.get('items', [])
        total_value = data.get('totalValue')

        if not project_name or not quote_ref or not items:
            flash('Missing required data (Project Name, Quote Ref, or Items).', 'danger')
            return redirect(url_for('quotation_builder'))

        conn = sqlite3.connect('ProjectStatus.db')
        c = conn.cursor()

        # --- Step 1: Save to Database ---
        # Insert the main quotation record
        c.execute("""
            INSERT INTO quotations (project_name, quote_ref, created_by, created_date, total_value)
            VALUES (?, ?, ?, ?, ?)
        """, (project_name, quote_ref, session.get('username'), datetime.now().strftime('%Y-%m-%d'), total_value))

        quotation_id = c.lastrowid

        # Insert each line item
        for item in items:
            c.execute("""
                INSERT INTO quotation_items (quotation_id, part_number, description, quantity, unit_price, total_price)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (quotation_id, item['Part Number'], item['Description'], item['Quantity'], item['Unit Price'],
                  item['Total Price']))

        conn.commit()
        conn.close()
        flash('Quotation saved successfully!', 'success')
        
        # Send notifications to stakeholders
        try:
            actor_id = session.get('user_id')
            actor_name = session.get('username', 'Unknown')
            
            # Notify admins and project stakeholders
            recipients = notification_service.get_admin_recipients()
            
            notification_service.notify_activity(
                event_code='quotation.submitted',
                recipient_ids=recipients,
                actor_id=actor_id,
                context={
                    'actor_name': actor_name,
                    'project_name': project_name,
                    'quotation_ref': quote_ref,
                    'total_value': total_value
                },
                url=url_for('registered_quotations')
            )
        except Exception as e:
            print(f"Notification error: {e}")

        # --- Step 2: Generate Excel File for Download ---
        df = pd.DataFrame(items)
        output = BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            details_df = pd.DataFrame({
                'Project Name': [project_name],
                'Quote Reference': [quote_ref],
                'Total Value': [total_value]
            })
            details_df.to_excel(writer, index=False, sheet_name='Quotation', startrow=0)
            df.to_excel(writer, index=False, sheet_name='Quotation', startrow=4)

        output.seek(0)

        return send_file(
            output,
            download_name=f"{quote_ref}.xlsx",
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except sqlite3.IntegrityError:
        flash('Error: This Quote Reference already exists. Please use a unique reference.', 'danger')
        return redirect(url_for('quotation_builder'))
    except Exception as e:
        flash(f'An error occurred: {e}', 'danger')
        return redirect(url_for('quotation_builder'))
##########################3
import pandas as pd
from io import BytesIO
from flask import send_file
import json


@app.route('/download_generated_quotation', methods=['POST'])
@login_required
def download_generated_quotation():
    # Get the JSON string from the hidden form
    json_data = request.form.get('quotation_data')
    if not json_data:
        flash('No quotation data received.', 'danger')
        return redirect(url_for('quotation_builder'))

    try:
        # Convert the JSON string back into Python data
        data = json.loads(json_data)
        items = data.get('items', [])

        # Create a pandas DataFrame from the line items
        df = pd.DataFrame(items)

        # Prepare the Excel file in memory
        output = BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            # Add project details at the top
            details_df = pd.DataFrame({
                'Project Name': [data.get('projectName')],
                'Quote Reference': [data.get('quoteRef')],
                'Total Value': [data.get('totalValue')]
            })
            details_df.to_excel(writer, index=False, sheet_name='Quotation', startrow=0)

            # Write the main BoQ table below the details
            df.to_excel(writer, index=False, sheet_name='Quotation', startrow=4)

        output.seek(0)

        # Send the file to the user for download
        return send_file(
            output,
            download_name=f"{data.get('quoteRef', 'quotation')}.xlsx",
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        flash(f'An error occurred while generating the Excel file: {e}', 'danger')
        return redirect(url_for('quotation_builder'))
#####################SRM####################

@app.route('/register_srm_vendor', methods=['GET', 'POST'])
@login_required
def register_srm_vendor():
    """Register a new vendor with full SRM support"""
    if request.method == 'POST':
        name = request.form['name']
        address = request.form.get('address', '')
        contact_person = request.form.get('contact_person', '')
        phone = request.form.get('main_phone', '')
        email = request.form.get('main_email', '')
        status = request.form.get('status', 'active')
        category = request.form.get('category', '')
        website = request.form.get('website', '')
        notes = request.form.get('notes', '')

        conn = sqlite3.connect('ProjectStatus.db')
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        try:
            # Insert vendor with all SRM fields
            cursor.execute("""
                INSERT INTO vendors 
                (name, address, contact_person, phone, email, status, category, website, notes) 
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (name, address, contact_person, phone, email, status, category, website, notes))
            
            vendor_id = cursor.lastrowid
            
            # Log activity in SRM activity log
            cursor.execute("""
                INSERT INTO srm_activity_log (entity_type, entity_id, activity_type, description, user_id)
                VALUES ('vendor', ?, 'vendor_created', ?, ?)
            """, (vendor_id, f'Created new vendor: {name}', session.get('user_id')))
            
            conn.commit()
            flash(f'Vendor "{name}" registered successfully!', 'success')
            
            # Redirect to vendor detail page to add contacts, documents, etc.
            return redirect(url_for('vendor_detail', vendor_id=vendor_id))
            
        except sqlite3.IntegrityError:
            flash('A vendor with this name already exists.', 'danger')
            return redirect(url_for('register_srm_vendor'))
        except Exception as e:
            conn.rollback()
            flash(f'An error occurred: {e}', 'danger')
            return redirect(url_for('register_srm_vendor'))
        finally:
            conn.close()

    return render_template('register_vendor.html')


# OLD ROUTE - REMOVED - Using new SRM route at line ~2150
# @app.route('/add_vendor_contact/<int:vendor_id>', methods=['GET', 'POST'])
# @login_required
# def add_vendor_contact(vendor_id):
#     # Contacts feature not available in production database
#     flash('Vendor contacts feature is not available. Use the contact_person field in vendor details instead.', 'info')
#     return redirect(url_for('show_vendors'))

##########################
# NEW SRM-ENABLED VENDORS ROUTE
@app.route('/vendors')
@login_required
@permission_required('view_vendors')
def show_vendors():
    """Display all vendors with SRM enhancements (contacts, account managers)"""
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    search_query = request.args.get('search_query', '')

    # Query vendors with optional search
    query = "SELECT * FROM vendors"
    params = []
    if search_query:
        query += " WHERE name LIKE ?"
        params.append(f'%{search_query}%')
    query += " ORDER BY name"

    cursor.execute(query, params)
    vendors = cursor.fetchall()

    # Enrich with SRM data (contacts and account manager)
    vendors_data = []
    for vendor in vendors:
        # Get contacts for this vendor
        cursor.execute("""
            SELECT * FROM vendor_contacts 
            WHERE vendor_id = ? 
            ORDER BY is_primary DESC, name
        """, (vendor['id'],))
        contacts = cursor.fetchall()
        
        # Get account manager name
        cursor.execute("""
            SELECT u.username as manager_name
            FROM account_manager_assignments ama
            JOIN users u ON ama.user_id = u.id
            WHERE ama.entity_type = 'vendor' AND ama.entity_id = ?
            LIMIT 1
        """, (vendor['id'],))
        manager = cursor.fetchone()
        
        # Create enriched vendor dict
        vendor_dict = dict(vendor)
        vendor_dict['account_manager_name'] = manager['manager_name'] if manager else None
        
        vendors_data.append({
            'vendor': vendor_dict,
            'contacts': contacts
        })

    conn.close()
    return render_template('vendors.html', vendors_data=vendors_data, search_query=search_query)
############################################
# SRM-ENABLED EDIT VENDOR ROUTE
@app.route('/edit_vendor/<int:vendor_id>', methods=['GET', 'POST'])
@login_required
def edit_vendor(vendor_id):
    """Edit vendor basic information with SRM compatibility"""
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    if request.method == 'POST':
        name = request.form['name']
        address = request.form['address']
        contact_person = request.form.get('contact_person', '')
        phone = request.form.get('main_phone', '')
        email = request.form.get('main_email', '')
        status = request.form.get('status', 'active')
        category = request.form.get('category', '')
        website = request.form.get('website', '')

        try:
            # Update vendors table with SRM fields
            cursor.execute("""
                UPDATE vendors SET
                name = ?, address = ?, contact_person = ?, phone = ?, email = ?,
                status = ?, category = ?, website = ?
                WHERE id = ?
            """, (name, address, contact_person, phone, email, status, category, website, vendor_id))
            
            # Log activity in SRM activity log
            cursor.execute("""
                INSERT INTO srm_activity_log (entity_type, entity_id, activity_type, description, user_id)
                VALUES ('vendor', ?, 'vendor_updated', ?, ?)
            """, (vendor_id, f'Updated vendor details', session.get('user_id')))
            
            conn.commit()
            flash('Vendor details updated successfully!', 'success')
            
            # Redirect to vendor detail page to see all SRM features
            return redirect(url_for('vendor_detail', vendor_id=vendor_id))
            
        except Exception as e:
            flash(f'An error occurred: {e}', 'danger')
            conn.rollback()
        finally:
            conn.close()

        return redirect(url_for('show_vendors'))

    # For GET request, fetch the vendor with all users for account manager dropdown
    cursor.execute("SELECT * FROM vendors WHERE id = ?", (vendor_id,))
    vendor = cursor.fetchone()
    
    if not vendor:
        flash('Vendor not found.', 'danger')
        conn.close()
        return redirect(url_for('show_vendors'))
    
    # Get all users for account manager assignment
    cursor.execute("SELECT id, username, role FROM users ORDER BY username")
    all_users = cursor.fetchall()
    
    # Get current account manager
    cursor.execute("""
        SELECT user_id FROM account_manager_assignments
        WHERE entity_type = 'vendor' AND entity_id = ?
        LIMIT 1
    """, (vendor_id,))
    account_manager_row = cursor.fetchone()
    current_account_manager = account_manager_row['user_id'] if account_manager_row else None
    
    conn.close()

    return render_template('edit_vendor.html', 
                         vendor=vendor, 
                         all_users=all_users,
                         current_account_manager=current_account_manager)


# DELETE VENDOR ROUTE
@app.route('/delete_vendor/<int:vendor_id>', methods=['POST'])
@login_required
@role_required('General Manager', 'Technical Team Leader')
def delete_vendor(vendor_id):
    """Delete a vendor (Admin only) with SRM activity logging"""
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    try:
        # Get vendor details before deletion
        cursor.execute("SELECT name FROM vendors WHERE id = ?", (vendor_id,))
        vendor = cursor.fetchone()
        
        if not vendor:
            flash('Vendor not found.', 'danger')
            return redirect(url_for('show_vendors'))
        
        vendor_name = vendor['name']
        
        # Delete related records first (cascade delete)
        cursor.execute("DELETE FROM vendor_contacts WHERE vendor_id = ?", (vendor_id,))
        cursor.execute("DELETE FROM vendor_distributor WHERE vendor_id = ?", (vendor_id,))
        cursor.execute("DELETE FROM account_manager_assignments WHERE entity_type = 'vendor' AND entity_id = ?", (vendor_id,))
        cursor.execute("DELETE FROM performance_metrics WHERE entity_type = 'vendor' AND entity_id = ?", (vendor_id,))
        cursor.execute("DELETE FROM srm_documents WHERE entity_type = 'vendor' AND entity_id = ?", (vendor_id,))
        cursor.execute("DELETE FROM srm_activity_log WHERE entity_type = 'vendor' AND entity_id = ?", (vendor_id,))
        
        # Finally delete the vendor
        cursor.execute("DELETE FROM vendors WHERE id = ?", (vendor_id,))
        
        conn.commit()
        flash(f'Vendor "{vendor_name}" has been successfully deleted!', 'success')
        
    except Exception as e:
        conn.rollback()
        flash(f'Error deleting vendor: {e}', 'danger')
    finally:
        conn.close()
    
    return redirect(url_for('show_vendors'))


####################
# OLD ROUTES - REMOVED - Using new SRM routes at line ~2200+
# @app.route('/edit_vendor_contact/<int:contact_id>', methods=['GET', 'POST'])
# @login_required
# def edit_vendor_contact(contact_id):
#     # Contacts feature not available in production database
#     flash('Vendor contacts feature is not available. Use the contact_person field in vendor details instead.', 'info')
#     return redirect(url_for('show_vendors'))


# @app.route('/delete_vendor_contact/<int:contact_id>', methods=['POST'])
# @login_required
# def delete_vendor_contact(contact_id):
#     # Contacts feature not available in production database
#     flash('Vendor contacts feature is not available.', 'info')
#     return redirect(url_for('show_vendors'))
###########################################

# =================================================================
# =================== TASK MANAGEMENT ROUTES ======================
# =================================================================
@app.route('/tasks', methods=['GET', 'POST'])
@login_required
@permission_required('view_tasks')
def tasks():
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    user_id = session['user_id']
    user_role = session.get('user_role', '')

    # --- Handle POST Request (Creating a new task) ---
    if request.method == 'POST':
        # This part remains unchanged
        # ... (keep the existing POST logic here) ...
        title = request.form['title']
        description = request.form.get('description')
        assigned_to_id = request.form['assigned_to_id']
        due_date = request.form['due_date']
        priority = request.form['priority']
        created_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        c.execute('''
            INSERT INTO tasks (title, description, assigned_to_id, due_date, priority, assigned_by_id, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (title, description, assigned_to_id, due_date, priority, user_id, created_at))
        conn.commit()
        flash('New task created successfully!', 'success')
        return redirect(url_for('tasks'))

    # --- Handle GET Request (Viewing tasks & Filtering) ---

    # 1. Get filter values from the URL
    current_filters = {
        'created_from': request.args.get('created_from', ''),
        'created_to': request.args.get('created_to', ''),
        'due_from': request.args.get('due_from', ''),
        'due_to': request.args.get('due_to', ''),
        'assigned_to_filter': request.args.get('assigned_to_filter', '')
    }

    # 2. Get the list of users that can be assigned tasks (for the "New Task" modal)
    assignable_roles = {
        'General Manager': ['Presale Engineer', 'Sales Engineer', 'Project Coordinator', 'Implementation Engineer',
                            'Project Manager', 'Technical Team Leader', 'General Manager'],
        'Technical Team Leader': ['Sales Engineer', 'Presale Engineer', 'Implementation Engineer',
                                  'Project Coordinator', 'Project Manager'],
        'Project Manager': ['Implementation Engineer', 'Project Coordinator']
    }
    roles_to_query = assignable_roles.get(user_role, [])

    if roles_to_query:
        placeholders = ', '.join(['?'] * len(roles_to_query))
        c.execute(f"SELECT id, username, role FROM users WHERE role IN ({placeholders}) ORDER BY username",
                  roles_to_query)
    else:
        c.execute("SELECT id, username, role FROM users WHERE 1=0")

    assignable_users = c.fetchall()

    c.execute("SELECT id, username, role FROM users WHERE id = ?", (user_id,))
    current_user_details = c.fetchone()
    if current_user_details and not any(u['id'] == user_id for u in assignable_users):
        assignable_users.append(current_user_details)

    # 3. Build the main query for fetching tasks based on role visibility and filters
    query = '''
        SELECT t.*,
               assignee.username as assigned_to_name,
               assigner.username as assigned_by_name
        FROM tasks t
        JOIN users assignee ON t.assigned_to_id = assignee.id
        JOIN users assigner ON t.assigned_by_id = assigner.id
    '''
    params = []

    # Base visibility query
    where_clauses = []
    managed_user_ids = []

    if user_role == 'General Manager':
        pass  # No initial clause, sees all
    elif user_role in ['Technical Team Leader', 'Project Manager']:
        managed_roles = assignable_roles.get(user_role, [])
        if managed_roles:
            managed_roles_placeholders = ', '.join(['?'] * len(managed_roles))
            c.execute(f"SELECT id FROM users WHERE role IN ({managed_roles_placeholders})", managed_roles)
            managed_user_ids = [row['id'] for row in c.fetchall()]
        if user_id not in managed_user_ids:
            managed_user_ids.append(user_id)

        if managed_user_ids:
            id_placeholders = ', '.join(['?'] * len(managed_user_ids))
            where_clauses.append(
                f"(t.assigned_to_id IN ({id_placeholders}) OR t.assigned_by_id IN ({id_placeholders}))")
            params.extend(managed_user_ids * 2)
    else:
        where_clauses.append("(t.assigned_to_id = ? OR t.assigned_by_id = ?)")
        params.extend([user_id, user_id])

    # Add filter clauses
    if current_filters['created_from']:
        where_clauses.append("date(t.created_at) >= ?")
        params.append(current_filters['created_from'])
    if current_filters['created_to']:
        where_clauses.append("date(t.created_at) <= ?")
        params.append(current_filters['created_to'])
    if current_filters['due_from']:
        where_clauses.append("date(t.due_date) >= ?")
        params.append(current_filters['due_from'])
    if current_filters['due_to']:
        where_clauses.append("date(t.due_date) <= ?")
        params.append(current_filters['due_to'])
    if current_filters['assigned_to_filter']:
        where_clauses.append("t.assigned_to_id = ?")
        params.append(current_filters['assigned_to_filter'])

    if where_clauses:
        query += " WHERE " + " AND ".join(where_clauses)

    query += " ORDER BY t.due_date ASC"
    c.execute(query, params)
    all_tasks = c.fetchall()

    # 4. Get a list of users for the "Assigned To" filter dropdown
    # For managers, this is their team; for others, it's just them. GM sees all.
    filter_assignee_list = []
    if user_role == 'General Manager':
        c.execute("SELECT id, username FROM users ORDER BY username")
        filter_assignee_list = c.fetchall()
    elif managed_user_ids:  # This will be populated for TTL and PM
        id_placeholders = ', '.join(['?'] * len(managed_user_ids))
        c.execute(f"SELECT id, username FROM users WHERE id IN ({id_placeholders}) ORDER BY username", managed_user_ids)
        filter_assignee_list = c.fetchall()
    else:  # For regular users
        filter_assignee_list = [current_user_details]

    conn.close()

    return render_template('tasks.html',
                           tasks=all_tasks,
                           assignable_users=assignable_users,
                           current_user_id=user_id,
                           user_role=user_role,
                           current_filters=current_filters,
                           filter_assignee_list=filter_assignee_list)

###################################33
@app.route('/update_task_with_comment', methods=['POST'])
@login_required
def update_task_with_comment():
    """
    Handles updating a task's status and logging a mandatory comment.
    """
    task_id = request.form['task_id']
    new_status = request.form['new_status']
    comment = request.form['comment']
    user_id = session['user_id']
    created_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    if not comment:
        flash('A comment is required to change the task status.', 'danger')
        return redirect(url_for('tasks'))

    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    try:
        # 1. Update the task status
        c.execute("UPDATE tasks SET status = ?, updated_at = ? WHERE id = ?",
                  (new_status, created_at, task_id))

        # 2. Insert the new comment
        c.execute('''
            INSERT INTO task_comments (task_id, user_id, comment, created_at)
            VALUES (?, ?, ?, ?)
        ''', (task_id, user_id, f"Status changed to '{new_status}': {comment}", created_at))

        conn.commit()
        flash('Task status and comment saved successfully!', 'success')
    except Exception as e:
        flash(f"An error occurred: {e}", "danger")
    finally:
        conn.close()

    return redirect(url_for('tasks'))


@app.route('/get_task_comments/<int:task_id>')
@login_required
def get_task_comments(task_id):
    """
    Fetches all comments for a given task to display in the history modal.
    """
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute('''
        SELECT c.comment, c.created_at, u.username
        FROM task_comments c
        JOIN users u ON c.user_id = u.id
        WHERE c.task_id = ?
        ORDER BY c.created_at DESC
    ''', (task_id,))
    comments = [dict(row) for row in c.fetchall()]
    conn.close()
    return jsonify(comments)
##################################
@app.route('/add_task_comment', methods=['POST'])
@login_required
def add_task_comment():
    """
    Handles adding a comment to a task without changing its status.
    """
    task_id = request.form['task_id']
    comment = request.form['comment']
    user_id = session['user_id']
    created_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    if not comment:
        flash('Comment cannot be empty.', 'danger')
        return redirect(url_for('tasks'))

    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    try:
        # Insert the comment directly
        c.execute('''
            INSERT INTO task_comments (task_id, user_id, comment, created_at)
            VALUES (?, ?, ?, ?)
        ''', (task_id, user_id, comment, created_at))
        conn.commit()
        flash('Comment added successfully!', 'success')
    except Exception as e:
        flash(f"An error occurred: {e}", "danger")
    finally:
        conn.close()

    return redirect(url_for('tasks'))
###########################

@app.route('/delete_task/<int:task_id>', methods=['POST'])
@login_required
def delete_task(task_id):
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    # Permission Check: Only the assigner or a General Manager can delete a task
    c.execute("SELECT assigned_by_id FROM tasks WHERE id = ?", (task_id,))
    task = c.fetchone()

    if task and (task['assigned_by_id'] == session['user_id'] or session.get('user_role') == 'General Manager'):
        c.execute("DELETE FROM tasks WHERE id = ?", (task_id,))
        conn.commit()
        flash('Task deleted successfully.', 'success')
    else:
        flash('You do not have permission to delete this task.', 'danger')

    conn.close()
    return redirect(url_for('tasks'))

##############################################
# ============ USER MANAGEMENT SYSTEM ============
##############################################

@app.route('/manage_users')
@role_required('General Manager', 'Technical Team Leader')
@permission_required('manage_users')
def manage_users():
    """
    Admin panel to view and manage all system users
    Accessible by General Manager and Technical Team Leader
    """
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    
    c.execute("SELECT id, username, role FROM users ORDER BY username")
    users = c.fetchall()
    
    conn.close()
    return render_template('manage_users.html', users=users)


@app.route('/add_user', methods=['GET', 'POST'])
@role_required('General Manager', 'Technical Team Leader')
def add_user():
    """
    Add a new user to the system with encrypted password
    Accessible by General Manager and Technical Team Leader
    """
    if request.method == 'POST':
        username = request.form['username'].strip()
        password = request.form['password']
        confirm_password = request.form['confirm_password']
        role = request.form['role']
        email = request.form.get('email', '').strip()
        
        # Validation
        if not username or not password:
            flash('Username and password are required!', 'danger')
            return redirect(url_for('add_user'))
        
        if password != confirm_password:
            flash('Passwords do not match!', 'danger')
            return redirect(url_for('add_user'))
        
        if len(password) < 6:
            flash('Password must be at least 6 characters long!', 'danger')
            return redirect(url_for('add_user'))
        
        # Hash the password using werkzeug for security
        hashed_password = generate_password_hash(password)
        
        conn = sqlite3.connect('ProjectStatus.db')
        c = conn.cursor()
        
        try:
            # Insert new user with encrypted password and email
            c.execute("INSERT INTO users (username, password, role, email) VALUES (?, ?, ?, ?)",
                     (username, hashed_password, role, email if email else None))
            conn.commit()
            flash(f'User "{username}" added successfully with encrypted password!', 'success')
            return redirect(url_for('manage_users'))
        except sqlite3.IntegrityError:
            flash(f'Username "{username}" already exists!', 'danger')
            return redirect(url_for('add_user'))
        except Exception as e:
            flash(f'An error occurred: {e}', 'danger')
            return redirect(url_for('add_user'))
        finally:
            conn.close()
    
    # Available roles in the system
    roles = ['editor', 'General Manager', 'Technical Team Leader', 
             'Presale Engineer', 'Sales Engineer', 'Project Coordinator']
    
    return render_template('add_user.html', roles=roles)


@app.route('/edit_user/<int:user_id>', methods=['GET', 'POST'])
@role_required('General Manager', 'Technical Team Leader')
def edit_user(user_id):
    """
    Edit existing user details
    Accessible by General Manager and Technical Team Leader
    """
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    
    if request.method == 'POST':
        username = request.form['username'].strip()
        role = request.form['role']
        email = request.form.get('email', '').strip()
        new_password = request.form.get('new_password', '').strip()
        
        if not username:
            flash('Username is required!', 'danger')
            return redirect(url_for('edit_user', user_id=user_id))
        
        try:
            # Update user details
            if new_password:
                # User wants to change password - hash it
                if len(new_password) < 6:
                    flash('Password must be at least 6 characters long!', 'danger')
                    return redirect(url_for('edit_user', user_id=user_id))
                
                hashed_password = generate_password_hash(new_password)
                c.execute("UPDATE users SET username = ?, role = ?, email = ?, password = ? WHERE id = ?",
                         (username, role, email if email else None, hashed_password, user_id))
                flash(f'User "{username}" updated successfully with new encrypted password!', 'success')
            else:
                # Update without changing password
                c.execute("UPDATE users SET username = ?, role = ?, email = ? WHERE id = ?",
                         (username, role, email if email else None, user_id))
                flash(f'User "{username}" updated successfully!', 'success')
            
            conn.commit()
            return redirect(url_for('manage_users'))
        except sqlite3.IntegrityError:
            flash(f'Username "{username}" already exists!', 'danger')
            return redirect(url_for('edit_user', user_id=user_id))
        except Exception as e:
            flash(f'An error occurred: {e}', 'danger')
            return redirect(url_for('edit_user', user_id=user_id))
        finally:
            conn.close()
    
    # GET request - load user data
    c.execute("SELECT * FROM users WHERE id = ?", (user_id,))
    user = c.fetchone()
    conn.close()
    
    if not user:
        flash('User not found!', 'danger')
        return redirect(url_for('manage_users'))
    
    roles = ['editor', 'General Manager', 'Technical Team Leader', 
             'Presale Engineer', 'Sales Engineer', 'Project Coordinator']
    
    return render_template('edit_user.html', user=user, roles=roles)


@app.route('/admin_reset_password/<int:user_id>', methods=['GET', 'POST'])
@role_required('General Manager', 'Technical Team Leader')
def admin_reset_password(user_id):
    """
    Admin can reset user password directly without email OTP
    Accessible by General Manager and Technical Team Leader
    """
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    
    if request.method == 'POST':
        new_password = request.form.get('new_password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()
        
        # Validation
        if not new_password:
            flash('Password is required!', 'danger')
            return redirect(url_for('admin_reset_password', user_id=user_id))
        
        if len(new_password) < 6:
            flash('Password must be at least 6 characters long!', 'danger')
            return redirect(url_for('admin_reset_password', user_id=user_id))
        
        if new_password != confirm_password:
            flash('Passwords do not match!', 'danger')
            return redirect(url_for('admin_reset_password', user_id=user_id))
        
        try:
            # Get user info
            c.execute("SELECT username FROM users WHERE id = ?", (user_id,))
            user = c.fetchone()
            
            if not user:
                flash('User not found!', 'danger')
                return redirect(url_for('manage_users'))
            
            # Hash the new password
            hashed_password = generate_password_hash(new_password)
            
            # Update password in users table
            c.execute("UPDATE users SET password = ? WHERE id = ?", (hashed_password, user_id))
            
            # Also update in engineers table if the user exists there
            c.execute("UPDATE engineers SET password = ? WHERE name = ?", (hashed_password, user['username']))
            
            conn.commit()
            flash(f'Password for user "{user["username"]}" has been reset successfully by admin!', 'success')
            return redirect(url_for('manage_users'))
            
        except Exception as e:
            flash(f'An error occurred: {e}', 'danger')
            return redirect(url_for('admin_reset_password', user_id=user_id))
        finally:
            conn.close()
    
    # GET request - show reset form
    c.execute("SELECT * FROM users WHERE id = ?", (user_id,))
    user = c.fetchone()
    conn.close()
    
    if not user:
        flash('User not found!', 'danger')
        return redirect(url_for('manage_users'))
    
    return render_template('admin_reset_password.html', user=user)


@app.route('/delete_user/<int:user_id>', methods=['POST'])
@role_required('General Manager', 'Technical Team Leader')
def delete_user(user_id):
    """
    Delete a user from the system
    Accessible by General Manager and Technical Team Leader
    """
    # Prevent deleting yourself
    if user_id == session.get('user_id'):
        flash('You cannot delete your own account!', 'danger')
        return redirect(url_for('manage_users'))
    
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    
    try:
        c.execute("SELECT username FROM users WHERE id = ?", (user_id,))
        user = c.fetchone()
        
        if user:
            c.execute("DELETE FROM users WHERE id = ?", (user_id,))
            conn.commit()
            flash(f'User "{user[0]}" deleted successfully!', 'success')
        else:
            flash('User not found!', 'danger')
    except Exception as e:
        flash(f'An error occurred: {e}', 'danger')
    finally:
        conn.close()
    
    return redirect(url_for('manage_users'))

##############################################
# ============ PUBLIC REGISTRATION WITH ADMIN APPROVAL ============
##############################################

@app.route('/register', methods=['GET', 'POST'])
def register():
    """
    Public registration page - anyone can request an account
    Requires admin approval before account is activated
    """
    if request.method == 'POST':
        username = request.form['username'].strip()
        password = request.form['password']
        confirm_password = request.form['confirm_password']
        requested_role = request.form['role']
        email = request.form.get('email', '').strip()
        full_name = request.form.get('full_name', '').strip()
        reason = request.form.get('reason', '').strip()
        
        # Validation
        if not username or not password:
            flash('Username and password are required!', 'danger')
            return redirect(url_for('register'))
        
        if password != confirm_password:
            flash('Passwords do not match!', 'danger')
            return redirect(url_for('register'))
        
        if len(password) < 6:
            flash('Password must be at least 6 characters long!', 'danger')
            return redirect(url_for('register'))
        
        # Encrypt password immediately
        hashed_password = generate_password_hash(password)
        
        conn = sqlite3.connect('ProjectStatus.db')
        c = conn.cursor()
        
        try:
            # Check if username already exists in users table
            c.execute("SELECT username FROM users WHERE username = ?", (username,))
            if c.fetchone():
                flash(f'Username "{username}" is already taken!', 'danger')
                return redirect(url_for('register'))
            
            # Check if registration request already exists
            c.execute("SELECT username FROM registration_requests WHERE username = ? AND status = 'pending'", (username,))
            if c.fetchone():
                flash(f'A registration request for "{username}" is already pending approval!', 'info')
                return redirect(url_for('register'))
            
            # Insert registration request with encrypted password
            c.execute("""
                INSERT INTO registration_requests 
                (username, password_hash, requested_role, email, full_name, reason, status) 
                VALUES (?, ?, ?, ?, ?, ?, 'pending')
            """, (username, hashed_password, requested_role, email, full_name, reason))
            
            request_id = c.lastrowid
            conn.commit()
            flash(f'Registration request submitted successfully! Your account will be activated once approved by an administrator.', 'success')
            
            # Send notification to admins about new registration
            try:
                recipients = notification_service.get_admin_recipients()
                notification_service.notify_activity(
                    event_code='user.registered',
                    recipient_ids=recipients,
                    actor_id=None,
                    context={
                        'actor_name': 'System',
                        'username': username,
                        'full_name': full_name,
                        'requested_role': requested_role,
                        'email': email
                    },
                    url=url_for('pending_registrations')
                )
            except Exception as e:
                print(f"Notification error: {e}")
            
            return redirect(url_for('login'))
        except Exception as e:
            flash(f'An error occurred: {e}', 'danger')
            return redirect(url_for('register'))
        finally:
            conn.close()
    
    # Available roles for new users
    roles = ['Presale Engineer', 'Sales Engineer', 'Project Coordinator', 'editor']
    
    return render_template('register.html', roles=roles)


@app.route('/pending_registrations')
@role_required('General Manager', 'Technical Team Leader')
@permission_required('view_pending_registrations')
def pending_registrations():
    """
    Admin page to view and manage pending registration requests
    Accessible by General Manager and Technical Team Leader
    """
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    
    # Get all pending registration requests
    c.execute("""
        SELECT * FROM registration_requests 
        WHERE status = 'pending' 
        ORDER BY requested_at DESC
    """)
    pending_requests = c.fetchall()
    
    # Get recently reviewed requests (last 10)
    c.execute("""
        SELECT * FROM registration_requests 
        WHERE status IN ('approved', 'rejected') 
        ORDER BY reviewed_at DESC 
        LIMIT 10
    """)
    reviewed_requests = c.fetchall()
    
    conn.close()
    
    return render_template('pending_registrations.html', 
                          pending_requests=pending_requests,
                          reviewed_requests=reviewed_requests)


@app.route('/approve_registration/<int:request_id>', methods=['POST'])
@role_required('General Manager', 'Technical Team Leader')
def approve_registration(request_id):
    """
    Approve a registration request and create the user account
    Accessible by General Manager and Technical Team Leader
    """
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    
    try:
        # Get the registration request
        c.execute("SELECT * FROM registration_requests WHERE id = ?", (request_id,))
        request_data = c.fetchone()
        
        if not request_data:
            flash('Registration request not found!', 'danger')
            return redirect(url_for('pending_registrations'))
        
        if request_data['status'] != 'pending':
            flash('This registration has already been processed!', 'warning')
            return redirect(url_for('pending_registrations'))
        
        # Create the user account with the already-encrypted password
        c.execute("""
            INSERT INTO users (username, password, role) 
            VALUES (?, ?, ?)
        """, (request_data['username'], request_data['password_hash'], request_data['requested_role']))
        
        # Update the registration request status
        c.execute("""
            UPDATE registration_requests 
            SET status = 'approved', reviewed_by = ?, reviewed_at = CURRENT_TIMESTAMP 
            WHERE id = ?
        """, (session['username'], request_id))
        
        conn.commit()
        flash(f'User "{request_data["username"]}" has been approved and activated!', 'success')
        
        # Send notification to the approved user
        try:
            # Get the new user's ID
            c.execute("SELECT id FROM users WHERE username = ?", (request_data['username'],))
            new_user = c.fetchone()
            
            if new_user:
                notification_service.notify_activity(
                    event_code='user.approved',
                    recipient_ids=[new_user['id']],
                    actor_id=session.get('user_id'),
                    context={
                        'actor_name': session.get('username', 'Admin'),
                        'username': request_data['username']
                    },
                    url=url_for('login')
                )
        except Exception as e:
            print(f"Notification error: {e}")
            
    except sqlite3.IntegrityError:
        flash(f'Username "{request_data["username"]}" already exists!', 'danger')
    except Exception as e:
        flash(f'An error occurred: {e}', 'danger')
    finally:
        conn.close()
    
    return redirect(url_for('pending_registrations'))


@app.route('/reject_registration/<int:request_id>', methods=['POST'])
@role_required('General Manager', 'Technical Team Leader')
def reject_registration(request_id):
    """
    Reject a registration request
    Accessible by General Manager and Technical Team Leader
    """
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    
    try:
        # Get the registration request
        c.execute("SELECT username FROM registration_requests WHERE id = ?", (request_id,))
        request_data = c.fetchone()
        
        if not request_data:
            flash('Registration request not found!', 'danger')
            return redirect(url_for('pending_registrations'))
        
        # Update the registration request status
        c.execute("""
            UPDATE registration_requests 
            SET status = 'rejected', reviewed_by = ?, reviewed_at = CURRENT_TIMESTAMP 
            WHERE id = ?
        """, (session['username'], request_id))
        
        conn.commit()
        flash(f'Registration request for "{request_data[0]}" has been rejected.', 'info')
    except Exception as e:
        flash(f'An error occurred: {e}', 'danger')
    finally:
        conn.close()
    
    return redirect(url_for('pending_registrations'))


@app.route('/pending_otp_requests')
@role_required('General Manager', 'Technical Team Leader')
@permission_required('view_password_reset_otps')
def pending_otp_requests():
    """
    Admin page to view pending OTP requests for password resets
    Shows OTP codes that admins can share with users manually
    Accessible by General Manager and Technical Team Leader
    """
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    
    # Get all active (non-used, non-expired) OTP requests
    c.execute("""
        SELECT 
            prt.id,
            prt.user_id,
            prt.email,
            prt.otp_code,
            prt.created_at,
            prt.expires_at,
            prt.used,
            u.username,
            u.role
        FROM password_reset_tokens prt
        LEFT JOIN users u ON prt.user_id = u.id
        WHERE prt.used = 0
        ORDER BY prt.created_at DESC
        LIMIT 50
    """)
    otp_requests = c.fetchall()
    
    conn.close()
    
    return render_template('pending_otp_requests.html', otp_requests=otp_requests)


@app.route('/mark_otp_shared/<int:token_id>', methods=['POST'])
@role_required('General Manager', 'Technical Team Leader')
def mark_otp_shared(token_id):
    """
    Mark an OTP as shared with user (optional feature for tracking)
    """
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    
    try:
        c.execute("SELECT id FROM password_reset_tokens WHERE id = ?", (token_id,))
        if c.fetchone():
            # You could add a "shared_at" timestamp column if needed
            flash('OTP marked as shared. User can now verify it.', 'success')
        else:
            flash('OTP request not found!', 'danger')
    except Exception as e:
        flash(f'An error occurred: {e}', 'danger')
    finally:
        conn.close()
    
    return redirect(url_for('pending_otp_requests'))


@app.route('/access_control')
@role_required('General Manager', 'Technical Team Leader')
@permission_required('manage_permissions')
def access_control():
    """
    Admin page to manage user permissions
    Allows assigning/revoking specific page access for each user
    """
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    
    # Get all users
    c.execute("SELECT id, username, role FROM users ORDER BY username")
    users = c.fetchall()
    
    # Get all permissions grouped by category
    c.execute("SELECT * FROM permissions ORDER BY category, label")
    all_permissions = c.fetchall()
    
    # Group permissions by category
    permissions_by_category = {}
    for perm in all_permissions:
        category = perm['category'] or 'Other'
        if category not in permissions_by_category:
            permissions_by_category[category] = []
        permissions_by_category[category].append(perm)
    
    conn.close()
    
    return render_template('access_control.html', 
                          users=users, 
                          permissions_by_category=permissions_by_category)


@app.route('/get_user_permissions/<int:user_id>')
@role_required('General Manager', 'Technical Team Leader')
def get_user_permissions_api(user_id):
    """
    API endpoint to get user's effective permissions
    Returns JSON with permission IDs and grant types
    """
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    
    # Get user info
    c.execute("SELECT username, role FROM users WHERE id = ?", (user_id,))
    user = c.fetchone()
    
    if not user:
        return jsonify({'error': 'User not found'}), 404
    
    # Get role default permissions
    c.execute("""
        SELECT p.id, p.code, p.label, 'role_default' as source
        FROM role_permissions rp
        JOIN permissions p ON rp.permission_id = p.id
        WHERE rp.role = ?
    """, (user['role'],))
    role_perms = {row['id']: {'code': row['code'], 'label': row['label'], 'source': 'role_default', 'grant_type': 'allow'} 
                  for row in c.fetchall()}
    
    # Get user-specific overrides
    c.execute("""
        SELECT p.id, p.code, p.label, up.grant_type
        FROM user_permissions up
        JOIN permissions p ON up.permission_id = p.id
        WHERE up.user_id = ?
    """, (user_id,))
    
    for row in c.fetchall():
        role_perms[row['id']] = {
            'code': row['code'],
            'label': row['label'],
            'source': 'user_override',
            'grant_type': row['grant_type']
        }
    
    conn.close()
    
    return jsonify({
        'user': {'id': user_id, 'username': user['username'], 'role': user['role']},
        'permissions': role_perms
    })


@app.route('/update_user_permission', methods=['POST'])
@role_required('General Manager', 'Technical Team Leader')
def update_user_permission():
    """
    Update a specific user permission (allow/deny/inherit)
    """
    data = request.get_json()
    user_id = data.get('user_id')
    permission_id = data.get('permission_id')
    action = data.get('action')  # 'allow', 'deny', or 'inherit'
    
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    
    try:
        if action == 'inherit':
            # Remove user-specific override (inherit from role)
            c.execute("""
                DELETE FROM user_permissions 
                WHERE user_id = ? AND permission_id = ?
            """, (user_id, permission_id))
        else:
            # Add or update user-specific permission
            c.execute("""
                INSERT INTO user_permissions (user_id, permission_id, grant_type, updated_by)
                VALUES (?, ?, ?, ?)
                ON CONFLICT(user_id, permission_id) 
                DO UPDATE SET grant_type = ?, updated_at = CURRENT_TIMESTAMP, updated_by = ?
            """, (user_id, permission_id, action, session['username'], action, session['username']))
        
        conn.commit()
        
        # Refresh permissions in session if updating current user
        if user_id == session.get('user_id'):
            refresh_user_permissions()
        
        return jsonify({'success': True, 'message': 'Permission updated successfully'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


##############################################
# ============ PASSWORD RESET WITH OTP ============
##############################################

@app.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    """
    Step 1: Request password reset
    User enters username/email, system sends OTP to registered email
    """
    if request.method == 'POST':
        identifier = request.form['identifier'].strip()  # Can be username or email
        
        if not identifier:
            flash('Please enter your username or email!', 'danger')
            return redirect(url_for('forgot_password'))
        
        conn = sqlite3.connect('ProjectStatus.db')
        conn.row_factory = sqlite3.Row
        c = conn.cursor()
        
        try:
            # Find user by username
            c.execute("SELECT id, username, email FROM users WHERE username = ?", (identifier,))
            user = c.fetchone()
            
            if not user:
                flash('User not found! Please check your username.', 'danger')
                return redirect(url_for('forgot_password'))
            
            # Get email from multiple sources (priority order)
            recipient_email = None
            
            # 1. Check users table first
            if user['email']:
                recipient_email = user['email']
            
            # 2. Check registration_requests table
            if not recipient_email:
                c.execute("SELECT email FROM registration_requests WHERE username = ? AND email IS NOT NULL", (user['username'],))
                email_record = c.fetchone()
                if email_record and email_record['email']:
                    recipient_email = email_record['email']
            
            # 3. Check engineers table
            if not recipient_email:
                c.execute("SELECT email FROM engineers WHERE username = ? AND email IS NOT NULL", (user['username'],))
                engineer_record = c.fetchone()
                if engineer_record and engineer_record['email']:
                    recipient_email = engineer_record['email']
            
            # If still no email found, show error
            if not recipient_email:
                flash(f'No email address registered for user "{user["username"]}". Please contact administrator to add your email address.', 'warning')
                return redirect(url_for('forgot_password'))
            
            # Generate OTP
            otp_code = generate_otp(6)
            
            # Calculate expiration time (15 minutes from now)
            expires_at = datetime.now() + timedelta(minutes=15)
            
            # Save OTP to database
            c.execute("""
                INSERT INTO password_reset_tokens (user_id, email, otp_code, expires_at)
                VALUES (?, ?, ?, ?)
            """, (user['id'], recipient_email, otp_code, expires_at))
            conn.commit()
            
            # Don't send email - admin will share OTP manually
            # This is a workaround until domain is verified for Resend
            flash(f'Password reset request submitted for "{user["username"]}". Please contact your administrator to get the OTP code.', 'info')
            
            # Store user_id in session for next step
            session['reset_user_id'] = user['id']
            session['reset_email'] = recipient_email
            
            return redirect(url_for('verify_otp'))
            
        except Exception as e:
            flash(f'An error occurred: {e}', 'danger')
            return redirect(url_for('forgot_password'))
        finally:
            conn.close()
    
    return render_template('forgot_password.html')


@app.route('/verify_otp', methods=['GET', 'POST'])
def verify_otp():
    """
    Step 2: Verify OTP code
    User enters the OTP received via email
    """
    if 'reset_user_id' not in session:
        flash('Please request a password reset first!', 'warning')
        return redirect(url_for('forgot_password'))
    
    if request.method == 'POST':
        otp_code = request.form['otp_code'].strip()
        
        if not otp_code:
            flash('Please enter the OTP code!', 'danger')
            return redirect(url_for('verify_otp'))
        
        conn = sqlite3.connect('ProjectStatus.db')
        conn.row_factory = sqlite3.Row
        c = conn.cursor()
        
        try:
            # Verify OTP
            c.execute("""
                SELECT * FROM password_reset_tokens
                WHERE user_id = ? AND otp_code = ? AND used = 0
                AND datetime(expires_at) > datetime('now')
                ORDER BY created_at DESC
                LIMIT 1
            """, (session['reset_user_id'], otp_code))
            
            token = c.fetchone()
            
            if not token:
                flash('Invalid or expired OTP code! Please request a new one.', 'danger')
                return redirect(url_for('verify_otp'))
            
            # OTP is valid - proceed to password reset
            session['reset_token_id'] = token['id']
            flash('OTP verified successfully! Please set your new password.', 'success')
            return redirect(url_for('reset_password_with_otp'))
            
        except Exception as e:
            flash(f'An error occurred: {e}', 'danger')
            return redirect(url_for('verify_otp'))
        finally:
            conn.close()
    
    return render_template('verify_otp.html', email=session.get('reset_email'))


@app.route('/reset_password_with_otp', methods=['GET', 'POST'])
def reset_password_with_otp():
    """
    Step 3: Set new password
    After OTP verification, user sets a new password
    """
    if 'reset_token_id' not in session or 'reset_user_id' not in session:
        flash('Please verify OTP first!', 'warning')
        return redirect(url_for('forgot_password'))
    
    if request.method == 'POST':
        new_password = request.form['new_password']
        confirm_password = request.form['confirm_password']
        
        # Validation
        if not new_password or not confirm_password:
            flash('Both password fields are required!', 'danger')
            return redirect(url_for('reset_password_with_otp'))
        
        if new_password != confirm_password:
            flash('Passwords do not match!', 'danger')
            return redirect(url_for('reset_password_with_otp'))
        
        if len(new_password) < 6:
            flash('Password must be at least 6 characters long!', 'danger')
            return redirect(url_for('reset_password_with_otp'))
        
        conn = sqlite3.connect('ProjectStatus.db')
        c = conn.cursor()
        
        try:
            # Hash the new password
            hashed_password = generate_password_hash(new_password)
            
            # Update user password in both users and engineers tables
            c.execute("UPDATE users SET password = ? WHERE id = ?", 
                     (hashed_password, session['reset_user_id']))
            
            # Update engineers table if exists
            c.execute("SELECT username FROM users WHERE id = ?", (session['reset_user_id'],))
            username = c.fetchone()[0]
            c.execute("UPDATE engineers SET password = ? WHERE username = ?", 
                     (hashed_password, username))
            
            # Mark OTP as used
            c.execute("UPDATE password_reset_tokens SET used = 1 WHERE id = ?", 
                     (session['reset_token_id'],))
            
            conn.commit()
            
            # Clear session data
            session.pop('reset_user_id', None)
            session.pop('reset_email', None)
            session.pop('reset_token_id', None)
            
            flash('Password reset successful! You can now login with your new password.', 'success')
            return redirect(url_for('login'))
            
        except Exception as e:
            flash(f'An error occurred: {e}', 'danger')
            return redirect(url_for('reset_password_with_otp'))
        finally:
            conn.close()
    
    return render_template('reset_password.html')


@app.route('/resend_otp', methods=['POST'])
def resend_otp():
    """
    Resend OTP code if user didn't receive it or it expired
    """
    if 'reset_user_id' not in session or 'reset_email' not in session:
        flash('Session expired! Please start over.', 'warning')
        return redirect(url_for('forgot_password'))
    
    conn = sqlite3.connect('ProjectStatus.db')
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    
    try:
        # Get user info
        c.execute("SELECT username FROM users WHERE id = ?", (session['reset_user_id'],))
        user = c.fetchone()
        
        # Generate new OTP
        otp_code = generate_otp(6)
        expires_at = datetime.now() + timedelta(minutes=15)
        
        # Save new OTP
        c.execute("""
            INSERT INTO password_reset_tokens (user_id, email, otp_code, expires_at)
            VALUES (?, ?, ?, ?)
        """, (session['reset_user_id'], session['reset_email'], otp_code, expires_at))
        conn.commit()
        
        # Send email
        email_sent = send_otp_email(session['reset_email'], otp_code, user['username'])
        
        if email_sent:
            flash('New OTP code sent to your email!', 'success')
        else:
            flash('Email service not configured. Check console for OTP.', 'warning')
        
    except Exception as e:
        flash(f'Error resending OTP: {e}', 'danger')
    finally:
        conn.close()
    
    return redirect(url_for('verify_otp'))

##############################################
# AI DEAL VALUE MANAGEMENT ROUTES
##############################################

@app.route('/admin/update_deal_values', methods=['GET', 'POST'])
@login_required
@role_required('General Manager', 'Technical Team Leader')
def admin_update_deal_values():
    """
    Admin route to update all deal values using AI logic
    Handles quote revisions intelligently to avoid duplicates
    """
    if request.method == 'POST':
        conn = sqlite3.connect('ProjectStatus.db')
        cursor = conn.cursor()
        
        # Get all projects from register_project
        cursor.execute("SELECT id, project_name, deal_value FROM register_project")
        registered_projects = cursor.fetchall()
        
        updates_made = 0
        total_value_changed = 0
        update_details = []
        
        for project_id, project_name, current_deal_value in registered_projects:
            # Calculate new deal value using AI logic
            new_deal_value, included_quotes = calculate_deal_value_for_project(project_name, conn)
            
            # Check if update is needed
            current_deal_value = current_deal_value or 0.0
            
            if abs(new_deal_value - current_deal_value) > 0.01:
                # Update the database
                cursor.execute("""
                    UPDATE register_project
                    SET deal_value = ?
                    WHERE id = ?
                """, (new_deal_value, project_id))
                
                updates_made += 1
                total_value_changed += abs(new_deal_value - current_deal_value)
                
                update_details.append({
                    'project_name': project_name,
                    'old_value': current_deal_value,
                    'new_value': new_deal_value,
                    'change': new_deal_value - current_deal_value,
                    'quotes': included_quotes
                })
        
        conn.commit()
        conn.close()
        
        if updates_made > 0:
            flash(f'Successfully updated {updates_made} projects! Total value adjusted: ${total_value_changed:,.2f}', 'success')
        else:
            flash('All deal values are already up to date!', 'info')
        
        return render_template('admin_deal_value_results.html', 
                             update_details=update_details,
                             total_projects=len(registered_projects),
                             updates_made=updates_made,
                             total_value_changed=total_value_changed)
    
    # GET request - show confirmation page
    conn = sqlite3.connect('ProjectStatus.db')
    cursor = conn.cursor()
    
    cursor.execute("SELECT COUNT(*) FROM register_project")
    total_projects = cursor.fetchone()[0]
    
    conn.close()
    
    return render_template('admin_deal_value_update.html', total_projects=total_projects)

@app.route('/api/calculate_deal_value/<project_name>', methods=['GET'])
@login_required
def api_calculate_deal_value(project_name):
    """
    API endpoint to calculate deal value for a specific project
    Returns JSON with calculated value and included quotes
    """
    try:
        deal_value, included_quotes = calculate_deal_value_for_project(project_name)
        
        return jsonify({
            'success': True,
            'deal_value': deal_value,
            'quotes': [{
                'quote_ref': q['quote_ref'],
                'price': q['price'],
                'system': q['system'],
                'revision': q['revision']
            } for q in included_quotes]
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

##############################################
# ============ SMART CCTV PRODUCT SELECTOR ============
##############################################

@app.route('/cctv_smart_selector', methods=['GET'])
@login_required
def cctv_smart_selector():
    """
    AI-Powered CCTV Product Selection Tool
    Helps users find the perfect camera based on their requirements
    """
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    
    # Get filter parameters
    resolution_filter = request.args.get('resolution')
    camera_type_filter = request.args.get('camera_type')
    min_price = request.args.get('min_price')
    max_price = request.args.get('max_price')
    features_filter = request.args.get('features')
    search_query = request.args.get('search')
    
    # Build query
    query = """
        SELECT id, category, item_code, item_name, price, resolution, camera_type, features
        FROM cctv_hikvision_products
        WHERE 1=1
    """
    params = []
    
    if resolution_filter and resolution_filter != 'all':
        query += " AND resolution = ?"
        params.append(resolution_filter)
    
    if camera_type_filter and camera_type_filter != 'all':
        query += " AND camera_type = ?"
        params.append(camera_type_filter)
    
    if min_price:
        query += " AND price >= ?"
        params.append(float(min_price))
    
    if max_price:
        query += " AND price <= ?"
        params.append(float(max_price))
    
    if features_filter:
        query += " AND features LIKE ?"
        params.append(f"%{features_filter}%")
    
    if search_query:
        query += " AND (item_name LIKE ? OR item_code LIKE ? OR category LIKE ?)"
        search_param = f"%{search_query}%"
        params.extend([search_param, search_param, search_param])
    
    query += " ORDER BY price ASC"
    
    c.execute(query, params)
    products = c.fetchall()
    
    # Get filter options
    c.execute("SELECT DISTINCT resolution FROM cctv_hikvision_products ORDER BY resolution")
    resolutions = [r[0] for r in c.fetchall()]
    
    c.execute("SELECT DISTINCT camera_type FROM cctv_hikvision_products ORDER BY camera_type")
    camera_types = [ct[0] for ct in c.fetchall()]
    
    c.execute("SELECT MIN(price), MAX(price) FROM cctv_hikvision_products")
    price_range = c.fetchone()
    
    # Get statistics
    c.execute("SELECT COUNT(*) FROM cctv_hikvision_products")
    total_products = c.fetchone()[0]
    
    conn.close()
    
    return render_template('cctv_smart_selector.html',
                         products=products,
                         resolutions=resolutions,
                         camera_types=camera_types,
                         price_range=price_range,
                         total_products=total_products,
                         selected_resolution=resolution_filter,
                         selected_camera_type=camera_type_filter,
                         min_price=min_price,
                         max_price=max_price,
                         selected_features=features_filter,
                         search_query=search_query)


@app.route('/api/cctv/recommend', methods=['POST'])
@login_required
def api_cctv_recommend():
    """
    AI-powered recommendation API
    Returns recommended products based on user requirements
    """
    data = request.get_json()
    
    budget = data.get('budget', 1000)
    purpose = data.get('purpose', 'general')  # indoor, outdoor, general
    priority = data.get('priority', 'balanced')  # price, quality, features
    
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    
    # Build recommendation query based on purpose
    if purpose == 'indoor':
        query = """
            SELECT * FROM cctv_hikvision_products 
            WHERE (features LIKE '%Indoor%' OR camera_type = 'Dome/Turret')
            AND price <= ?
            ORDER BY price ASC
            LIMIT 10
        """
    elif purpose == 'outdoor':
        query = """
            SELECT * FROM cctv_hikvision_products 
            WHERE (features LIKE '%Outdoor%' OR features LIKE '%Night Vision%')
            AND price <= ?
            ORDER BY price DESC
            LIMIT 10
        """
    else:
        if priority == 'price':
            query = "SELECT * FROM cctv_hikvision_products WHERE price <= ? ORDER BY price ASC LIMIT 10"
        elif priority == 'quality':
            query = """
                SELECT * FROM cctv_hikvision_products 
                WHERE price <= ?
                ORDER BY 
                    CASE 
                        WHEN resolution = '8MP (4K)' THEN 1
                        WHEN resolution = '6MP' THEN 2
                        WHEN resolution = '5MP' THEN 3
                        ELSE 4
                    END,
                    price DESC
                LIMIT 10
            """
        else:
            query = "SELECT * FROM cctv_hikvision_products WHERE price <= ? ORDER BY price ASC LIMIT 10"
    
    c.execute(query, (budget,))
    products = c.fetchall()
    conn.close()
    
    recommendations = []
    for p in products:
        recommendations.append({
            'id': p[0],
            'name': p[3],
            'code': p[2],
            'price': p[4],
            'resolution': p[5],
            'type': p[6],
            'features': p[7]
        })
    
    return jsonify({'success': True, 'recommendations': recommendations})


@app.route('/api/cctv/compare', methods=['POST'])
@login_required
def api_cctv_compare():
    """
    Product comparison API
    Compare up to 4 products side-by-side
    """
    data = request.get_json()
    product_ids = data.get('product_ids', [])
    
    if len(product_ids) == 0:
        return jsonify({'success': False, 'error': 'No products selected'}), 400
    
    if len(product_ids) > 4:
        return jsonify({'success': False, 'error': 'Maximum 4 products can be compared'}), 400
    
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    
    placeholders = ','.join('?' * len(product_ids))
    query = f"SELECT * FROM cctv_hikvision_products WHERE id IN ({placeholders})"
    
    c.execute(query, product_ids)
    products = c.fetchall()
    conn.close()
    
    comparison = []
    for p in products:
        comparison.append({
            'id': p[0],
            'category': p[1],
            'code': p[2],
            'name': p[3],
            'price': p[4],
            'resolution': p[5],
            'type': p[6],
            'features': p[7]
        })
    
    return jsonify({'success': True, 'products': comparison})


##############################################
# ============ USER PROFILE API ENDPOINTS ============
##############################################

@app.route('/api/user/timezone', methods=['GET', 'POST'])
@login_required
def api_user_timezone():
    """
    Get or update user's timezone preference
    GET: Returns user's current timezone
    POST: Updates user's timezone (Body: { "timezone": "Asia/Riyadh" })
    """
    conn = sqlite3.connect('ProjectStatus.db')
    c = conn.cursor()
    user_id = session.get('user_id')
    
    try:
        if request.method == 'POST':
            data = request.get_json()
            timezone = data.get('timezone')
            
            if not timezone:
                return jsonify({'success': False, 'error': 'Timezone is required'}), 400
            
            c.execute("UPDATE users SET timezone = ? WHERE id = ?", (timezone, user_id))
            conn.commit()
            
            return jsonify({
                'success': True,
                'message': 'Timezone updated successfully',
                'timezone': timezone
            })
        else:
            # GET request
            c.execute("SELECT timezone FROM users WHERE id = ?", (user_id,))
            result = c.fetchone()
            timezone = result[0] if result and result[0] else 'UTC'
            
            return jsonify({
                'success': True,
                'timezone': timezone
            })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


##############################################
# ============ NOTIFICATION API ENDPOINTS ============
##############################################

@app.route('/api/notifications/feed', methods=['GET'])
@login_required
def api_notifications_feed():
    """
    Get notifications for the current user
    Query params: limit (default 10), offset (default 0), unread_only (default false)
    """
    try:
        user_id = session.get('user_id')
        limit = int(request.args.get('limit', 10))
        offset = int(request.args.get('offset', 0))
        unread_only = request.args.get('unread_only', 'false').lower() == 'true'
        
        notifications = notification_service.get_notifications(
            user_id=user_id,
            unread_only=unread_only,
            limit=limit,
            offset=offset
        )
        
        return jsonify({
            'success': True,
            'notifications': notifications
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/notifications/mark_read', methods=['POST'])
@login_required
def api_notifications_mark_read():
    """
    Mark notifications as read
    Body: { "notification_ids": [1, 2, 3] } or { "all": true }
    """
    try:
        user_id = session.get('user_id')
        data = request.get_json()
        
        if data.get('all'):
            success = notification_service.mark_all_as_read(user_id)
        else:
            notification_ids = data.get('notification_ids', [])
            success = notification_service.mark_as_read(notification_ids, user_id)
        
        return jsonify({
            'success': success,
            'message': 'Notifications marked as read' if success else 'Failed to mark notifications as read'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/notifications/count', methods=['GET'])
@login_required
def api_notifications_count():
    """
    Get unread notification count for the current user
    """
    try:
        user_id = session.get('user_id')
        count = notification_service.get_unread_count(user_id)
        
        return jsonify({
            'success': True,
            'unread_count': count
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


if __name__ == '__main__':
    init_db()
    seed_permissions()
    seed_default_role_permissions()
    host = os.getenv('HOST', '0.0.0.0')
    port = int(os.getenv('PORT', 5000))
    debug = os.getenv('FLASK_DEBUG', '0') == '1'
    app.run(host=host, port=port, debug=debug)
